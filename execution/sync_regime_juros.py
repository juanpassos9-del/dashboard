import json
import os
try:
    import tomllib
except Exception:
    tomllib = None
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from supabase import create_client


DEFAULT_EXCEL_PATH = r"C:\Users\Mini PC\Documents\ANALISE JUROS\Curva_DI_RTD_Monitor_PrecoTempo.xlsx"
APP_STATE_KEY = "regime_juros"


def get_secret_value(*names: str) -> str:
    for name in names:
        value = os.environ.get(name, "")
        if value:
            return value

    secrets_path = Path(".streamlit") / "secrets.toml"
    if not secrets_path.exists():
        return ""

    try:
        if tomllib is not None:
            with secrets_path.open("rb") as fp:
                secrets = tomllib.load(fp)
        else:
            secrets = {}
            for raw_line in secrets_path.read_text(encoding="utf-8").splitlines():
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                secrets[key.strip()] = value.strip().strip('"').strip("'")
    except Exception:
        return ""

    for name in names:
        value = secrets.get(name, "")
        if value:
            return str(value)
    return ""


def read_regime_juros_open_excel(path: str = DEFAULT_EXCEL_PATH) -> dict | None:
    """Read live RTD values from the workbook already open in desktop Excel."""
    try:
        import win32com.client
    except Exception:
        return None

    target = Path(path)
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        return None

    try:
        workbooks = excel.Workbooks
    except Exception:
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            workbooks = excel.Workbooks
        except Exception:
            return None

    for wb in workbooks:
        try:
            same_file = Path(str(wb.FullName)).resolve().samefile(target)
        except Exception:
            same_file = str(wb.Name).lower() == target.name.lower()
        if not same_file:
            continue

        ws = wb.Worksheets("Indice Atual")
        return {
            "taxa_sintetica": ws.Range("E2").Value,
            "variacao_bps": ws.Range("H2").Value,
            "regime_estrutural": ws.Range("K2").Value,
            "updated_at": datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%Y-%m-%d %H:%M:%S"),
            "source": "excel_rtd_live",
        }
    return None


def read_regime_juros_excel(path: str = DEFAULT_EXCEL_PATH) -> dict:
    excel_path = Path(path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel nao encontrado: {excel_path}")

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb["Indice Atual"]
    return {
        "taxa_sintetica": ws["E2"].value,
        "variacao_bps": ws["H2"].value,
        "regime_estrutural": ws["K2"].value,
        "updated_at": datetime.fromtimestamp(
            excel_path.stat().st_mtime,
            ZoneInfo("America/Sao_Paulo"),
        ).strftime("%Y-%m-%d %H:%M:%S"),
        "source": "excel_rtd_saved",
    }


def sync_to_supabase(payload: dict) -> None:
    url = get_secret_value("SUPABASE_URL", "SUPABASE")
    key = get_secret_value("SUPABASE_SERVICE_ROLE", "SUPABASE_KEY", "SUPABASE_SERVICE")
    if not url or not key:
        raise RuntimeError("Configure SUPABASE e SUPABASE_SERVICE no ambiente.")

    client = create_client(url, key)
    client.table("app_state").upsert(
        {
            "key": APP_STATE_KEY,
            "value": payload,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        },
        on_conflict="key",
    ).execute()


def main() -> None:
    excel_path = os.environ.get("REGIME_JUROS_EXCEL_PATH", DEFAULT_EXCEL_PATH)
    payload = read_regime_juros_open_excel(excel_path) or read_regime_juros_excel(excel_path)
    sync_to_supabase(payload)
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

import streamlit as st
import streamlit.components.v1 as components

import os
import base64
import html
import json
import time
import pandas as pd
from datetime import datetime, timedelta, timezone
from typing import Any
from zoneinfo import ZoneInfo
from supabase import create_client, Client

try:
    from execution.source_health import get_source_health, mark_source
except Exception:
    def mark_source(*args, **kwargs):
        return None

    def get_source_health(*args, **kwargs):
        return {}


# ── Configuração da Página ──────────────────────────────────────────────────
st.set_page_config(page_title="Terminal TTS | Inteligência", layout="wide")

# ── Supabase ────────────────────────────────────────────────────────────────
@st.cache_resource
def init_supabase() -> Client:
    try:
        url = st.secrets.get("SUPABASE_URL", os.environ.get("SUPABASE_URL", ""))
        key = st.secrets.get("SUPABASE_SERVICE_ROLE", os.environ.get("SUPABASE_SERVICE_ROLE", ""))
        if not key:
            key = st.secrets.get("SUPABASE_KEY", os.environ.get("SUPABASE_KEY", ""))
        if not url or not key:
            st.error("🔑 Credenciais do Supabase não encontradas. Verifique os segredos.")
            return None
        return create_client(url, key)
    except Exception as e:
        st.error(f"🌐 Falha na conexão com o Banco de Dados: {e}")
        return None

supabase = init_supabase()


LOCAL_TMP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".tmp")


MAX_AUTH_USERS = 1000
AUTH_REQUIRED = False
APP_STATE_ALLOWED_KEYS = {
    "ai_insight",
    "ai_insight_history",
    "boletim_focus",
    "calendario_economico",
    "dados_mercado",
    "financial_juice_news",
    "fluxo_estrangeiro_b3",
    "manual_trades",
    "market_report",
    "market_report_daily",
    "mercados_globais",
    "regime_juros",
    "risk_manual_trades",
}


def safe_external_url(value: Any, fallback: str = "#") -> str:
    """Return an escaped http(s) URL for HTML anchors."""
    raw = str(value or "").strip()
    if raw.startswith(("http://", "https://")):
        return html.escape(raw, quote=True)
    return html.escape(fallback, quote=True)


def _auth_rerun():
    try:
        st.rerun()
    except Exception:
        st.experimental_rerun()


def _auth_user_email(user) -> str:
    if not user:
        return ""
    if isinstance(user, dict):
        return user.get("email", "") or ""
    return getattr(user, "email", "") or ""


def _auth_user_id(user) -> str:
    if not user:
        return ""
    if isinstance(user, dict):
        return user.get("id", "") or ""
    return getattr(user, "id", "") or ""


def _auth_user_phone(user) -> str:
    if not user:
        return ""
    metadata = user.get("user_metadata", {}) if isinstance(user, dict) else getattr(user, "user_metadata", {}) or {}
    return metadata.get("phone", "") or metadata.get("celular", "") or ""


def _auth_user_role(user) -> str:
    if not user:
        return ""
    metadata = user.get("user_metadata", {}) if isinstance(user, dict) else getattr(user, "user_metadata", {}) or {}
    return metadata.get("role", "") or ""


def count_registered_profiles():
    if not supabase:
        return None
    try:
        response = supabase.table("profiles").select("id", count="exact").limit(1).execute()
        return getattr(response, "count", None)
    except Exception:
        return None


def count_admin_profiles():
    if not supabase:
        return None
    try:
        response = supabase.table("profiles").select("id", count="exact").eq("role", "admin").limit(1).execute()
        return getattr(response, "count", None)
    except Exception:
        return None


def count_registered_auth_users():
    if not supabase:
        return None
    try:
        response = supabase.auth.admin.list_users()
        users = getattr(response, "users", None)
        if users is not None:
            return len(users)
        if isinstance(response, list):
            return len(response)
    except Exception:
        return None
    return None


def _auth_user_matches_email(user, email: str) -> bool:
    return _auth_user_email(user).strip().lower() == email.strip().lower()


def find_auth_user_by_email(email: str):
    if not supabase or not email:
        return None
    try:
        response = supabase.auth.admin.list_users(page=1, per_page=1000)
        users = getattr(response, "users", None)
        if users is None:
            users = response if isinstance(response, list) else []
        for user in users or []:
            if _auth_user_matches_email(user, email):
                return user
    except Exception:
        return None
    return None


def confirm_auth_user_email(email: str):
    user = find_auth_user_by_email(email)
    user_id = _auth_user_id(user)
    if not user_id:
        return False, "Nao consegui localizar esse email no Supabase Auth para confirmar."
    try:
        response = supabase.auth.admin.update_user_by_id(user_id, {"email_confirm": True})
        confirmed_user = getattr(response, "user", None) or user
        return True, confirmed_user
    except Exception as e:
        return False, f"Nao consegui confirmar o email automaticamente: {e}"


def count_registered_users():
    profile_count = count_registered_profiles()
    if profile_count is not None:
        return profile_count
    return count_registered_auth_users()


def get_existing_profile(user_id: str):
    if not supabase or not user_id:
        return None
    try:
        response = supabase.table("profiles").select("role").eq("user_id", user_id).limit(1).execute()
        if response.data:
            return response.data[0]
    except Exception:
        return None
    return None


def resolve_new_user_role(total_profiles):
    return "admin" if total_profiles == 0 else "member"


def upsert_auth_profile(user, phone: str = "", role: str = ""):
    if not supabase or not user:
        return None
    user_id = _auth_user_id(user)
    email = _auth_user_email(user)
    if not user_id or not email:
        return None
    try:
        existing_profile = get_existing_profile(user_id)
        existing_role = (existing_profile or {}).get("role")
        admin_count = count_admin_profiles()
        if admin_count == 0 and existing_role != "admin":
            resolved_role = "admin"
        else:
            resolved_role = existing_role or role or _auth_user_role(user) or "member"
        payload = {
            "user_id": user_id,
            "email": email,
            "phone": phone or _auth_user_phone(user),
            "role": resolved_role,
            "is_active": True,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        response = supabase.table("profiles").upsert(payload, on_conflict="user_id").execute()
        return response
    except Exception as e:
        return f"Tabela profiles indisponivel: {e}"


def auth_sign_in(email: str, password: str):
    if not supabase:
        return None, "Conexao Supabase indisponivel."
    try:
        response = supabase.auth.sign_in_with_password({"email": email, "password": password})
        user = getattr(response, "user", None)
        session = getattr(response, "session", None)
        if not user:
            return None, "Login nao retornou usuario. Verifique email/senha."
        st.session_state["auth_user"] = user
        st.session_state["auth_session"] = session
        profile_warning = upsert_auth_profile(user)
        user_id = _auth_user_id(user)
        profile = get_existing_profile(user_id)
        st.session_state["auth_role"] = (profile or {}).get("role") or _auth_user_role(user) or "member"
        return profile_warning, None
    except Exception as e:
        error_text = str(e)
        if "email not confirmed" in error_text.lower():
            confirmed, result = confirm_auth_user_email(email)
            if confirmed:
                try:
                    response = supabase.auth.sign_in_with_password({"email": email, "password": password})
                    user = getattr(response, "user", None)
                    session = getattr(response, "session", None)
                    if not user:
                        return None, "Email confirmado, mas login nao retornou usuario. Tente novamente."
                    st.session_state["auth_user"] = user
                    st.session_state["auth_session"] = session
                    profile_warning = upsert_auth_profile(user)
                    profile = get_existing_profile(_auth_user_id(user))
                    st.session_state["auth_role"] = (profile or {}).get("role") or _auth_user_role(user) or "member"
                    return profile_warning, None
                except Exception as retry_error:
                    return None, f"Email confirmado, mas o login ainda falhou: {retry_error}"
            return None, result
        return None, f"Falha no login: {e}"


def auth_sign_up(email: str, password: str, phone: str):
    if not supabase:
        return None, "Conexao Supabase indisponivel."
    total_users = count_registered_users()
    if total_users is not None and total_users >= MAX_AUTH_USERS:
        return None, "Limite de 1000 usuarios atingido."
    role = resolve_new_user_role(total_users)
    try:
        try:
            response = supabase.auth.admin.create_user({
                "email": email,
                "password": password,
                "email_confirm": True,
                "user_metadata": {"phone": phone, "celular": phone, "role": role},
            })
            user = getattr(response, "user", None)
            if user:
                upsert_auth_profile(user, phone, role)
            return auth_sign_in(email, password)
        except Exception as admin_error:
            if "already" in str(admin_error).lower() or "registered" in str(admin_error).lower() or "exists" in str(admin_error).lower():
                confirmed, result = confirm_auth_user_email(email)
                if confirmed:
                    return auth_sign_in(email, password)
            response = supabase.auth.sign_up({
                "email": email,
                "password": password,
                "options": {"data": {"phone": phone, "celular": phone, "role": role}},
            })
        user = getattr(response, "user", None)
        session = getattr(response, "session", None)
        if not user:
            return None, "Cadastro nao retornou usuario. Verifique os dados."
        st.session_state["auth_user"] = user
        st.session_state["auth_session"] = session
        profile_warning = upsert_auth_profile(user, phone, role)
        st.session_state["auth_role"] = role
        if session:
            return profile_warning, None
        return profile_warning, "Cadastro criado. Se a confirmacao por email estiver ativa no Supabase, confirme o email antes de entrar."
    except Exception as e:
        return None, f"Falha no cadastro: {e}"


@st.cache_data(show_spinner=False)
def load_asset_data_uri(path: str) -> str:
    try:
        if not os.path.exists(path):
            return ""
        ext = os.path.splitext(path)[1].lower().lstrip(".") or "png"
        mime = "jpeg" if ext in {"jpg", "jpeg"} else ext
        with open(path, "rb") as f:
            encoded = base64.b64encode(f.read()).decode("ascii")
        return f"data:image/{mime};base64,{encoded}"
    except Exception:
        return ""


def render_auth_loading(message: str = "Validando acesso...", submessage: str = "Carregando Terminal Global"):
    logo_path = os.path.join(os.path.dirname(__file__), "assets", "trading_strategy_logo_login.png")
    logo_data_uri = load_asset_data_uri(logo_path)
    logo_html = (
        f'<img class="tts-loading-logo" src="{logo_data_uri}" alt="Trading Strategy">'
        if logo_data_uri else
        '<div class="tts-loading-logo-fallback">TTS</div>'
    )
    st.markdown(
        """
        <style>
          .tts-loading-box {
            margin: 14px 0 4px;
            border: 1px solid rgba(56,189,248,.24);
            border-radius: 12px;
            background: linear-gradient(135deg, rgba(15,23,42,.96), rgba(2,8,23,.96));
            padding: 16px;
            display: flex;
            align-items: center;
            gap: 14px;
            box-shadow: inset 0 1px 0 rgba(255,255,255,.06), 0 12px 28px rgba(0,0,0,.24);
          }
          .tts-loading-mark {
            width: 58px;
            height: 58px;
            border-radius: 999px;
            overflow: hidden;
            border: 1px solid rgba(183,186,183,.25);
            background: #1b1b29;
            position: relative;
            flex: 0 0 auto;
          }
          .tts-loading-mark::after {
            content: "";
            position: absolute;
            inset: -2px;
            border-radius: 999px;
            border: 2px solid transparent;
            border-top-color: #38BDF8;
            border-right-color: rgba(56,189,248,.38);
            animation: tts-spin .9s linear infinite;
          }
          .tts-loading-logo {
            width: 100%;
            height: 100%;
            object-fit: cover;
            display: block;
          }
          .tts-loading-logo-fallback {
            width: 100%;
            height: 100%;
            display: flex;
            align-items: center;
            justify-content: center;
            color: #CBD5E1;
            font-weight: 950;
          }
          .tts-loading-title {
            color: #F8FAFC;
            font-size: .92rem;
            font-weight: 950;
          }
          .tts-loading-sub {
            color: #94A3B8;
            font-size: .76rem;
            font-weight: 750;
            margin-top: 3px;
          }
          .tts-loading-bar {
            margin-top: 9px;
            width: 100%;
            height: 4px;
            border-radius: 999px;
            overflow: hidden;
            background: rgba(148,163,184,.16);
          }
          .tts-loading-bar span {
            display: block;
            height: 100%;
            width: 42%;
            border-radius: 999px;
            background: linear-gradient(90deg, #38BDF8, #22C55E);
            animation: tts-slide 1.2s ease-in-out infinite;
          }
          @keyframes tts-spin { to { transform: rotate(360deg); } }
          @keyframes tts-slide {
            0% { transform: translateX(-115%); }
            55% { transform: translateX(95%); }
            100% { transform: translateX(245%); }
          }
        </style>
        <div class="tts-loading-box">
          <div class="tts-loading-mark">__LOGO_HTML__</div>
          <div style="min-width:0; flex:1;">
            <div class="tts-loading-title">__MESSAGE__</div>
            <div class="tts-loading-sub">__SUBMESSAGE__</div>
            <div class="tts-loading-bar"><span></span></div>
          </div>
        </div>
        """.replace("__LOGO_HTML__", logo_html)
        .replace("__MESSAGE__", html.escape(message))
        .replace("__SUBMESSAGE__", html.escape(submessage)),
        unsafe_allow_html=True,
    )


def render_auth_top_loading(message: str = "Carregando dashboard..."):
    st.markdown(
        """
        <style>
          .tts-top-loading {
            margin: 0 0 8px;
            border: 1px solid rgba(56,189,248,.16);
            border-radius: 8px;
            background: rgba(15,23,42,.72);
            padding: 8px 10px;
            display: flex;
            align-items: center;
            gap: 9px;
            color: #CBD5E1;
            font-size: .76rem;
            font-weight: 850;
          }
          .tts-top-loading-dot {
            width: 8px;
            height: 8px;
            border-radius: 999px;
            background: #22C55E;
            box-shadow: 0 0 10px rgba(34,197,94,.75);
            flex: 0 0 auto;
            animation: tts-top-pulse 1s ease-in-out infinite;
          }
          .tts-top-loading-bar {
            width: 80px;
            height: 3px;
            border-radius: 999px;
            overflow: hidden;
            background: rgba(148,163,184,.16);
            margin-left: auto;
          }
          .tts-top-loading-bar span {
            display: block;
            width: 45%;
            height: 100%;
            border-radius: 999px;
            background: #38BDF8;
            animation: tts-top-slide 1s ease-in-out infinite;
          }
          @keyframes tts-top-pulse {
            0%, 100% { opacity: .45; transform: scale(.9); }
            50% { opacity: 1; transform: scale(1.08); }
          }
          @keyframes tts-top-slide {
            0% { transform: translateX(-110%); }
            100% { transform: translateX(230%); }
          }
        </style>
        <div class="tts-top-loading">
          <span class="tts-top-loading-dot"></span>
          <span>__MESSAGE__</span>
          <span class="tts-top-loading-bar"><span></span></span>
        </div>
        """.replace("__MESSAGE__", html.escape(message)),
        unsafe_allow_html=True,
    )


def render_auth_screen():
    logo_path = os.path.join(os.path.dirname(__file__), "assets", "trading_strategy_logo_login.png")
    logo_data_uri = load_asset_data_uri(logo_path)
    logo_html = (
        f'<img class="tts-auth-logo" src="{logo_data_uri}" alt="Trading Strategy">'
        if logo_data_uri else
        '<div class="tts-auth-logo-fallback">TTS</div>'
    )
    st.markdown(
        """
        <style>
          [data-testid="stSidebar"] { display:none; }
          .stApp {
            background:
              radial-gradient(circle at 50% 18%, rgba(148,163,184,.12), transparent 28rem),
              linear-gradient(135deg, #151522 0%, #080b13 58%, #05070d 100%);
          }
          .tts-auth-wrap {
            min-height: 82vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 34px 0;
          }
          .tts-auth-card {
            width: min(560px, 100%);
            border: 1px solid rgba(148,163,184,.18);
            border-radius: 14px;
            background:
              linear-gradient(180deg, rgba(31,31,45,.96) 0%, rgba(10,14,25,.98) 100%);
            box-shadow: 0 30px 80px rgba(0,0,0,.42), inset 0 1px 0 rgba(255,255,255,.07);
            padding: 30px;
          }
          .tts-auth-logo-box {
            display:flex;
            justify-content:center;
            align-items:center;
            margin:0 auto 20px;
            width:178px;
            height:178px;
            border:1px solid rgba(183,186,183,.22);
            border-radius:999px;
            background:#1b1b29;
            box-shadow:
              0 20px 50px rgba(0,0,0,.32),
              inset 0 1px 0 rgba(255,255,255,.06);
            overflow:hidden;
          }
          .tts-auth-logo {
            width:100%;
            height:100%;
            object-fit:cover;
            display:block;
          }
          .tts-auth-logo-fallback {
            width:112px;
            height:112px;
            border-radius:999px;
            display:flex;
            align-items:center;
            justify-content:center;
            border:2px solid #b7bab7;
            color:#b7bab7;
            font-size:1.6rem;
            font-weight:950;
          }
          .tts-auth-brand {
            color:#f8fafc;
            font-size:1.65rem;
            font-weight:950;
            letter-spacing:.02em;
            margin-bottom:4px;
            text-align:center;
          }
          .tts-auth-sub {
            color:#94a3b8;
            font-size:.9rem;
            font-weight:700;
            margin-bottom:20px;
            text-align:center;
          }
          .tts-auth-pill {
            display:flex;
            width:max-content;
            margin:0 auto 14px;
            color:#38bdf8;
            background:rgba(56,189,248,.09);
            border:1px solid rgba(56,189,248,.22);
            border-radius:999px;
            padding:5px 9px;
            font-size:.72rem;
            font-weight:900;
            letter-spacing:.05em;
            text-transform:uppercase;
          }
          div[data-testid="stForm"] {
            border:1px solid rgba(148,163,184,.14);
            background:rgba(2,6,23,.24);
            border-radius:10px;
            padding:14px;
          }
          .stTabs [data-baseweb="tab-list"] { gap:8px; justify-content:center; }
          .stTabs [data-baseweb="tab"] {
            border-radius:8px;
            padding:8px 16px;
            color:#CBD5E1;
            font-weight:900;
          }
          .stButton > button, .stFormSubmitButton > button {
            border-radius:8px;
            font-weight:950;
          }
        </style>
        <div class="tts-auth-wrap">
          <div class="tts-auth-card">
            <div class="tts-auth-logo-box">__LOGO_HTML__</div>
            <div class="tts-auth-pill">Acesso restrito</div>
            <div class="tts-auth-brand">Terminal TTS</div>
            <div class="tts-auth-sub">Trading Strategy | Mercado como voce nunca viu</div>
        """.replace("__LOGO_HTML__", logo_html),
        unsafe_allow_html=True,
    )
    tab_login, tab_signup = st.tabs(["Entrar", "Criar conta"])
    with tab_login:
        with st.form("auth_login_form"):
            email = st.text_input("Email", key="auth_login_email").strip().lower()
            password = st.text_input("Senha", type="password", key="auth_login_password")
            submitted = st.form_submit_button("Entrar", use_container_width=True)
        if submitted:
            if not email or not password:
                st.error("Informe email e senha.")
            else:
                render_auth_loading("Validando acesso...", "Conectando ao Supabase e preparando seu terminal.")
                with st.spinner("Abrindo dashboard..."):
                    warning, error = auth_sign_in(email, password)
                if error:
                    st.session_state.pop("auth_loading_message", None)
                    st.error(error)
                else:
                    if warning:
                        st.warning(warning)
                    st.session_state["auth_loading_message"] = "Carregando dashboard..."
                    st.session_state["auth_loading_until"] = time.time() + 8.0
                    _auth_rerun()
    with tab_signup:
        with st.form("auth_signup_form"):
            email = st.text_input("Email", key="auth_signup_email").strip().lower()
            phone = st.text_input("Celular", key="auth_signup_phone").strip()
            password = st.text_input("Senha", type="password", key="auth_signup_password")
            submitted = st.form_submit_button("Criar conta", use_container_width=True)
        if submitted:
            if not email or not phone or not password:
                st.error("Informe email, celular e senha.")
            elif len(password) < 6:
                st.error("Use uma senha com pelo menos 6 caracteres.")
            else:
                render_auth_loading("Criando acesso...", "Registrando usuario e preparando seu terminal.")
                with st.spinner("Criando conta..."):
                    warning, message = auth_sign_up(email, password, phone)
                if warning:
                    st.warning(warning)
                if message:
                    st.session_state.pop("auth_loading_message", None)
                    st.info(message)
                else:
                    st.session_state["auth_loading_message"] = "Carregando dashboard..."
                    st.session_state["auth_loading_until"] = time.time() + 8.0
                    _auth_rerun()
    st.stop()


def require_authenticated_user():
    if not AUTH_REQUIRED:
        st.session_state.pop("auth_loading_message", None)
        st.session_state.pop("auth_loading_until", None)
        return {"id": "public", "email": "acesso.publico@tts.local", "user_metadata": {"role": "public"}}
    user = st.session_state.get("auth_user")
    if user:
        if "auth_role" not in st.session_state:
            profile = get_existing_profile(_auth_user_id(user))
            st.session_state["auth_role"] = (profile or {}).get("role") or _auth_user_role(user) or "member"
        return user
    render_auth_screen()


def start_post_auth_loading():
    loading_message = st.session_state.get("auth_loading_message", "")
    loading_until = float(st.session_state.get("auth_loading_until", 0) or 0)
    if not loading_message or time.time() > loading_until:
        st.session_state.pop("auth_loading_message", None)
        st.session_state.pop("auth_loading_until", None)
        return None
    placeholder = st.empty()
    with placeholder:
        render_auth_top_loading(loading_message)
    return placeholder


def stop_post_auth_loading(placeholder):
    if placeholder is not None:
        placeholder.empty()
    st.session_state.pop("auth_loading_message", None)
    st.session_state.pop("auth_loading_until", None)


def fetch_app_state(key: str):
    if key not in APP_STATE_ALLOWED_KEYS:
        print(f"[SECURITY] app_state read blocked for unexpected key: {key}")
        return None
    """Busca dados no Supabase com tratamento de erro e redundância."""
    if not supabase: return None
    try:
        response = supabase.table("app_state").select("value").eq("key", key).execute()
        if response.data and len(response.data) > 0:
            return response.data[0]["value"]
    except Exception as e:
        print(f"[ERROR] Fetch {key}: {e}")
    return None


def has_supabase_service_role() -> bool:
    try:
        return bool(st.secrets.get("SUPABASE_SERVICE_ROLE", os.environ.get("SUPABASE_SERVICE_ROLE", "")))
    except Exception:
        return bool(os.environ.get("SUPABASE_SERVICE_ROLE", ""))


def sync_app_state_value(key: str, value) -> tuple[bool, str]:
    """Sincroniza app_state sem derrubar a UI quando o Supabase bloquear RLS."""
    if key not in APP_STATE_ALLOWED_KEYS:
        return False, f"Chave app_state nao permitida para sincronizacao: {key}"
    if not supabase:
        return False, "Supabase indisponivel."
    try:
        supabase.table("app_state").upsert({
            "key": key,
            "value": value,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }).execute()
        return True, ""
    except Exception as e:
        message = str(e)
        if "row-level security" in message.lower() or "42501" in message:
            if not has_supabase_service_role():
                return False, "RLS bloqueou o salvamento online. Configure SUPABASE_SERVICE_ROLE no Streamlit Secrets ou aplique a policy de app_state."
            return False, "RLS bloqueou o salvamento online mesmo com Supabase configurado. Verifique policies da tabela app_state."
        return False, message


@st.cache_data(ttl=2, show_spinner=False)
def fetch_app_state_fast(key: str):
    """Cache curto para dados RTD usados por mais de um fragmento."""
    return fetch_app_state(key)

@st.cache_data(ttl=30, show_spinner=False)
def fetch_app_state_cached(key: str):
    """Cache de fallback para evitar consultas repetidas ao Supabase."""
    return fetch_app_state(key)

@st.cache_data(ttl=30, show_spinner=False)
def fetch_live_global_markets():
    """Busca cotacoes globais direto da fonte com cache curto para o Streamlit Cloud."""
    try:
        from execution.fetch_global_markets import fetch_global_data
        data = fetch_global_data(save_file=False)
        if data:
            rows = sum(len(v) for v in (data.get("categories", data) or {}).values() if isinstance(v, list))
            mark_source("Yahoo Finance", "ok", message="Mercados globais ao vivo.", rows=rows, source="execution.fetch_global_markets")
        else:
            mark_source("Yahoo Finance", "error", message="Mercados globais retornaram vazio.", source="execution.fetch_global_markets")
        return data
    except Exception as e:
        print(f"[ERROR] Live global markets: {e}")
        mark_source("Yahoo Finance", "error", message=str(e), source="execution.fetch_global_markets")
        return None

def get_global_markets_data():
    """Usa Supabase/cache primeiro para nao travar o boot do Streamlit Cloud."""
    cached_data = fetch_app_state_cached("mercados_globais")
    if cached_data:
        rows = sum(len(v) for v in (cached_data.get("categories", cached_data) or {}).values() if isinstance(v, list))
        mark_source("Mercados Globais Cache", "stale", message="Usando app_state/Supabase como fallback rapido.", rows=rows, source="Supabase app_state")
        return cached_data
    live_data = fetch_live_global_markets()
    if live_data:
        return live_data
    mark_source("Mercados Globais Cache", "error", message="Sem cache e sem fonte ao vivo.", source="Supabase/Yahoo")
    return None

@st.cache_data(ttl=300, show_spinner=False)
def fetch_investing_calendar_live():
    try:
        from execution.fetch_calendar import _fetch_investing_calendar
        events = _fetch_investing_calendar()
        if events:
            mark_source("Investing Calendar", "ok", message="Calendario economico ao vivo.", rows=len(events), source="Investing")
        else:
            mark_source("Investing Calendar", "error", message="Investing retornou calendario vazio.", source="Investing")
        return events
    except Exception as e:
        print(f"[WARN] Investing calendar unavailable: {e}")
        mark_source("Investing Calendar", "error", message=str(e), source="Investing")
        return None

def get_calendar_data():
    """Usa Supabase/cache primeiro para evitar spinner longo no boot."""
    cached_data = fetch_app_state_cached("calendario_economico")
    if cached_data:
        mark_source("Calendario Cache", "stale", message="Usando app_state/Supabase como fallback rapido.", rows=len(cached_data) if isinstance(cached_data, list) else None, source="Supabase app_state")
        return cached_data
    live_events = fetch_investing_calendar_live()
    if live_events:
        return live_events
    for path in [
        os.path.join(os.path.dirname(__file__), "calendario_economico.json"),
        os.path.join(os.path.dirname(__file__), "execution", "calendario_economico.json"),
    ]:
        try:
            if os.path.exists(path):
                import json as _json
                with open(path, "r", encoding="utf-8") as f:
                    local_events = _json.load(f)
                if local_events:
                    mark_source("Calendario Cache", "stale", message=f"Usando arquivo local {os.path.basename(path)}.", rows=len(local_events) if isinstance(local_events, list) else None, source="arquivo local")
                    return local_events
        except Exception as e:
            print(f"[WARN] Local calendar unavailable: {e}")
            mark_source("Calendario Cache", "error", message=str(e), source="arquivo local")
    mark_source("Calendario Cache", "error", message="Sem calendario em cache/local.", source="fallback")
    return None

@st.cache_data(ttl=30, show_spinner=False)
def load_bloomberg_news_feed(refresh_nonce: int = 0):
    """Monta o feed pesado com cache para evitar travamentos no rerender."""
    del refresh_nonce
    news_sources = []
    warnings = []
    news_list = fetch_app_state_cached("financial_juice_news") or []
    if news_list:
        news_sources.append("Historico Supabase")
        mark_source("Financial Juice Cache", "stale", message="Usando historico Supabase.", rows=len(news_list), source="Supabase app_state")

    try:
        from execution.fetch_financial_juice import cached_financial_juice_news
        cached_news = cached_financial_juice_news(limit=10)
        if cached_news:
            news_list.extend(cached_news)
            news_sources.append("Historico local")
            mark_source("Financial Juice Cache", "stale", message="Usando historico local.", rows=len(cached_news), source="cache local")
    except Exception as e:
        warnings.append(f"Historico local indisponivel: {e}")

    try:
        from execution.fetch_financial_juice import fetch_financial_juice_news
        live_news = fetch_financial_juice_news(
            limit=35,
            min_network_interval=30,
            fast_mode=True,
            translate=False,
        )
        if live_news:
            news_list.extend(live_news)
            news_sources.append("Financial Juice RSS direto")
            mark_source("Financial Juice RSS", "ok", message="Feed RSS direto atualizado.", rows=len(live_news), source="Financial Juice")
        elif news_list:
            warnings.append("Buscando noticias novas; exibindo historico recente.")
            mark_source("Financial Juice RSS", "stale", message="Fonte ao vivo sem novas noticias; exibindo historico.", rows=len(news_list), source="Financial Juice/cache")
    except Exception as e:
        if news_list:
            warnings.append("Fonte ao vivo indisponivel; exibindo historico recente.")
            mark_source("Financial Juice RSS", "stale", message=f"Fonte ao vivo indisponivel: {e}", rows=len(news_list), source="Financial Juice/cache")
        else:
            warnings.append(f"Financial Juice direto indisponivel: {e}")
            mark_source("Financial Juice RSS", "error", message=str(e), source="Financial Juice")

    try:
        from execution.fetch_news_api import fetch_news_api_news
        news_api_items = fetch_news_api_news(limit=25, max_age_seconds=900)
        if news_api_items:
            news_list.extend(news_api_items)
            news_sources.append("News API")
            mark_source("News API", "ok", message="Fonte complementar carregada.", rows=len(news_api_items), source="News API")
    except Exception as e:
        warnings.append(f"News API indisponivel: {e}")
        mark_source("News API", "stale", message=str(e), source="News API")

    seen_news = set()
    unique_news = []
    for item in news_list:
        key = (item.get("link") or item.get("title_en") or item.get("title_pt") or "").strip().lower()[:160]
        if not key or key in seen_news:
            continue
        seen_news.add(key)
        unique_news.append(item)

    def news_sort_key(item):
        try:
            return float(item.get("timestamp") or 0)
        except Exception:
            return 0

    unique_news = sorted(unique_news, key=news_sort_key, reverse=True)
    if not unique_news:
        try:
            from execution.fetch_financial_juice import cached_financial_juice_news
            unique_news = cached_financial_juice_news(limit=10)
            if unique_news:
                news_sources.append("Historico local")
                warnings.append("Aguardando noticias novas; exibindo ultimas 10 do historico.")
                mark_source("Financial Juice Cache", "stale", message="Aguardando noticias novas; exibindo historico local.", rows=len(unique_news), source="cache local")
        except Exception as e:
            warnings.append(f"Historico de noticias indisponivel: {e}")
            mark_source("Financial Juice Cache", "error", message=str(e), source="cache local")
    try:
        from execution.fetch_financial_juice import ensure_brazil_time
        for item in unique_news:
            ensure_brazil_time(item)
    except Exception as e:
        warnings.append(f"Normalizacao rapida indisponivel: {e}")

    return unique_news, news_sources, warnings, datetime.now().strftime("%H:%M:%S")


def news_ticker_impact(item):
    text = " ".join([
        str(item.get("title_en") or item.get("title") or item.get("title_pt") or ""),
        str(item.get("summary") or item.get("description") or ""),
        str(item.get("source") or ""),
    ]).lower()
    score = 0
    rules = [
        (5, ["fed", "fomc", "powell", "ecb", "boj", "boe", "copom", "bcb", "interest rate", "juros"]),
        (5, ["cpi", "pce", "ppi", "inflation", "inflacao", "inflação"]),
        (4, ["treasury", "treasuries", "yield", "yields", "dxy", "dollar", "oil", "crude", "brent", "wti"]),
        (4, ["payroll", "jobs", "jobless", "gdp", "retail sales", "pmi", "ism"]),
        (4, ["iran", "israel", "china", "russia", "war", "guerra", "sanctions", "attack", "ataque"]),
        (3, ["s&p", "nasdaq", "dow", "stocks", "futures", "bitcoin", "crypto", "ibovespa", "ewz"]),
    ]
    for weight, keywords in rules:
        if any(keyword in text for keyword in keywords):
            score += weight
    if any(word in text for word in ["breaking", "urgent", "alert", "unexpected", "surprise"]):
        score += 3
    if score >= 12:
        return "URGENTE", "#FF2D20", score
    if score >= 8:
        return "ALTO", "#FF4B4B", score
    if score >= 4:
        return "MEDIO", "#FF9800", score
    return "BAIXO", "#94A3B8", score


def render_high_impact_news_ticker():
    try:
        news_list, _, _, feed_loaded_at = load_bloomberg_news_feed(0)
    except Exception as e:
        print(f"[WARN] High impact news ticker unavailable: {e}")
        return
    if not news_list:
        return

    high_impact = []
    for item in news_list:
        label, color, score = news_ticker_impact(item)
        if label not in {"URGENTE", "ALTO"}:
            continue
        title = item.get("title_en") or item.get("title") or item.get("title_pt") or ""
        if not title:
            continue
        high_impact.append({
            "label": label,
            "color": color,
            "score": score,
            "title": html.escape(str(title)),
            "time": html.escape(str(item.get("published_str") or item.get("time") or "")),
        })
        if len(high_impact) >= 8:
            break

    if not high_impact:
        return

    ticker_items = []
    for item in high_impact:
        time_part = f"<span class='tts-news-time'>{item['time']}</span>" if item["time"] else ""
        ticker_items.append(
            f"<span class='tts-news-item'><span class='tts-news-badge' style='border-color:{item['color']}66;color:{item['color']};'>{item['label']}</span>{time_part}<span>{item['title']}</span></span>"
        )
    ticker_html = "<span class='tts-news-seq'>" + "<span class='tts-news-dot'>•</span>".join(ticker_items) + "</span>"
    st.markdown(
        f"""
        <style>
          .tts-news-ticker-wrap {{
            position: sticky;
            top: 0;
            z-index: 1000;
            margin: 0 0 10px;
            border: 1px solid rgba(255,75,75,.28);
            border-left: 4px solid #FF4B4B;
            border-radius: 8px;
            background: linear-gradient(90deg, rgba(17,24,39,.96), rgba(8,13,20,.96));
            box-shadow: 0 8px 20px rgba(0,0,0,.20);
            overflow: hidden;
          }}
          .tts-news-ticker-head {{
            display:flex;
            align-items:center;
            gap:8px;
            padding:7px 10px 0;
            color:#F8FAFC;
            font-size:.70rem;
            font-weight:950;
            letter-spacing:.06em;
            text-transform:uppercase;
          }}
          .tts-news-live {{
            width:7px;
            height:7px;
            border-radius:999px;
            background:#FF4B4B;
            box-shadow:0 0 11px rgba(255,75,75,.90);
          }}
          .tts-news-loaded {{
            margin-left:auto;
            color:#64748B;
            font-size:.62rem;
            font-weight:800;
          }}
          .tts-news-marquee {{
            white-space:nowrap;
            overflow:hidden;
            padding:7px 0 9px;
          }}
          .tts-news-track {{
            display:inline-block;
            min-width:100%;
            animation: tts-news-scroll 38s linear infinite;
          }}
          .tts-news-marquee:hover .tts-news-track {{
            animation-play-state: paused;
          }}
          .tts-news-item {{
            display:inline-flex;
            align-items:center;
            gap:8px;
            margin:0 18px;
            color:#E5E7EB;
            font-size:.78rem;
            font-weight:800;
          }}
          .tts-news-badge {{
            border:1px solid;
            border-radius:5px;
            padding:1px 5px;
            font-size:.62rem;
            font-weight:950;
          }}
          .tts-news-time {{
            color:#94A3B8;
            font-size:.68rem;
          }}
          .tts-news-dot {{
            color:#475569;
            font-weight:950;
          }}
          @keyframes tts-news-scroll {{
            0% {{ transform: translateX(12%); }}
            100% {{ transform: translateX(-100%); }}
          }}
        </style>
        <div class="tts-news-ticker-wrap">
          <div class="tts-news-ticker-head">
            <span class="tts-news-live"></span>
            <span>NEWS High Impact</span>
            <span class="tts-news-loaded">Atualizado {html.escape(str(feed_loaded_at))}</span>
          </div>
          <div class="tts-news-marquee">
            <div class="tts-news-track">{ticker_html}{ticker_html}</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


@st.cache_data(ttl=900, show_spinner=False)
def get_macro_news_hub_cached(schema_version="macro_news_hub_v1"):
    del schema_version
    from execution.news_macro_hub import build_macro_news_hub

    return build_macro_news_hub(limit=24)


def render_macro_news_hub():
    try:
        hub = get_macro_news_hub_cached("macro_news_hub_v1")
    except Exception as e:
        st.warning(f"Hub de noticias macro indisponivel agora: {e}")
        return

    items = hub.get("items") or []
    if not items:
        st.info("Hub de noticias macro aguardando manchetes relevantes das fontes prioritarias.")
        return

    def chip_list(values):
        return "".join(f"<span class='mnh-chip'>{sanitize_text(str(value))}</span>" for value in values[:4])

    cards = []
    for item in items[:12]:
        impact = sanitize_text(str(item.get("impact", "BAIXO")))
        impact_class = "high" if impact == "ALTO" else ("medium" if impact == "MEDIO" else "low")
        link = safe_external_url(item.get("link"))
        title = sanitize_text(str(item.get("title", "")))
        source = sanitize_text(str(item.get("source", "")))
        provider = sanitize_text(str(item.get("provider", "")))
        published = sanitize_text(str(item.get("published_str", "")))
        bias = sanitize_text(str(item.get("bias", "Neutro")))
        score = sanitize_text(str(item.get("score", "")))
        themes = chip_list(item.get("themes") or [])
        assets = chip_list(item.get("assets") or [])
        cards.append(f"""
            <a class="mnh-card {impact_class}" href="{link}" target="_blank" rel="noopener noreferrer">
              <div class="mnh-top">
                <span class="mnh-impact">{impact}</span>
                <span class="mnh-source">{source} · {provider} · {published}</span>
              </div>
              <div class="mnh-title">{title}</div>
              <div class="mnh-row"><b>Vies:</b> {bias} <b>Score:</b> {score}</div>
              <div class="mnh-tags">{themes}{assets}</div>
            </a>
        """)

    counts = hub.get("counts") or {}
    style = """
    <style>
      .mnh-wrap{border:1px solid #1F2937;background:#070B12;border-radius:9px;padding:14px;margin:14px 0 18px;}
      .mnh-head{display:flex;justify-content:space-between;gap:12px;align-items:flex-end;flex-wrap:wrap;margin-bottom:11px;}
      .mnh-head h3{margin:0;color:#E5E7EB;font-size:1.05rem;}
      .mnh-meta{color:#94A3B8;font-size:.75rem;font-weight:800;}
      .mnh-grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:9px;}
      .mnh-card{display:block;text-decoration:none;background:#0B1220;border:1px solid #263244;border-left:5px solid #64748B;border-radius:7px;padding:10px;color:#E5E7EB;min-height:132px;}
      .mnh-card.high{border-left-color:#FF3333;background:#1A080B;}
      .mnh-card.medium{border-left-color:#FF9800;}
      .mnh-card.low{border-left-color:#38BDF8;}
      .mnh-top{display:flex;justify-content:space-between;gap:8px;align-items:center;margin-bottom:7px;}
      .mnh-impact{font-size:.68rem;font-weight:950;color:#FFF;background:#263244;border-radius:999px;padding:3px 7px;}
      .mnh-card.high .mnh-impact{background:#B91C1C;}
      .mnh-card.medium .mnh-impact{background:#A16207;}
      .mnh-source{color:#94A3B8;font-size:.68rem;font-weight:800;text-align:right;}
      .mnh-title{font-size:.88rem;font-weight:900;line-height:1.25;color:#F8FAFC;margin-bottom:8px;}
      .mnh-row{font-size:.72rem;color:#CBD5E1;margin-bottom:7px;}
      .mnh-row b{color:#94A3B8;}
      .mnh-tags{display:flex;gap:5px;flex-wrap:wrap;}
      .mnh-chip{border:1px solid #334155;background:#111827;color:#BFDBFE;border-radius:999px;padding:3px 6px;font-size:.66rem;font-weight:850;}
      @media(max-width:1200px){.mnh-grid{grid-template-columns:repeat(2,minmax(0,1fr));}}
      @media(max-width:760px){.mnh-grid{grid-template-columns:1fr;}}
    </style>
    """
    markup = f"""
    <div class="mnh-wrap">
      <div class="mnh-head">
        <div>
          <h3>Hub Macro de Noticias</h3>
          <div class="mnh-meta">Fontes prioritarias abertas · atualizado {sanitize_text(str(hub.get('updated_at', '---')))}{' · cache' if hub.get('stale') else ''}</div>
        </div>
        <div class="mnh-meta">ALTO {counts.get('alto', 0)} · MEDIO {counts.get('medio', 0)} · BAIXO {counts.get('baixo', 0)}</div>
      </div>
      <div class="mnh-grid">{''.join(cards)}</div>
    </div>
    """
    components.html(style + markup, height=620, scrolling=True)


@st.fragment(run_every=30)
def render_bloomberg_news_feed_fragment():
    """Atualiza somente o feed de noticias, sem redesenhar o terminal inteiro."""
    if "bb_translate_news_fast" not in st.session_state:
        st.session_state.bb_translate_news_fast = False
    if "bb_translation_cache" not in st.session_state:
        st.session_state.bb_translation_cache = {}
    if "bb_news_history" not in st.session_state:
        st.session_state.bb_news_history = []

    def esc(value) -> str:
        return html.escape(str(value or ""), quote=True)

    def translate_news_item(item: dict) -> dict:
        translated_item = dict(item)
        title_original = item.get("title_en") or item.get("title") or item.get("title_pt") or ""
        summary_original = item.get("summary") or item.get("description") or ""
        try:
            from execution.fetch_financial_juice import translate_text_google
            cache = st.session_state.bb_translation_cache
            if title_original:
                title_key = f"title::{title_original}"
                if title_key not in cache:
                    cache[title_key] = translate_text_google(title_original)
                translated_item["title_pt"] = cache[title_key]
            if summary_original and summary_original != title_original:
                summary_key = f"summary::{summary_original}"
                if summary_key not in cache:
                    cache[summary_key] = translate_text_google(summary_original)
                translated_item["summary_pt"] = cache[summary_key]
            elif title_original:
                translated_item["summary_pt"] = translated_item.get("title_pt", "")
        except Exception:
            try:
                from execution.fetch_financial_juice import ensure_portuguese_fields
                ensure_portuguese_fields(translated_item)
            except Exception:
                pass
        return translated_item

    def news_title(item, translated: bool | None = None) -> str:
        if translated is None:
            translated = bool(st.session_state.get("bb_translate_news_fast", False))
        if translated:
            return item.get("title_pt") or item.get("title_en") or item.get("title") or "---"
        return item.get("title_en") or item.get("title") or item.get("title_pt") or "---"

    def news_summary(item, translated: bool | None = None) -> str:
        if translated is None:
            translated = bool(st.session_state.get("bb_translate_news_fast", False))
        if translated:
            summary = item.get("summary_pt") or item.get("summary") or item.get("description") or ""
        else:
            summary = item.get("summary") or item.get("description") or ""
        title = news_title(item, translated=translated)
        if summary == title:
            return ""
        return summary

    def news_original_text(item) -> str:
        summary = item.get("summary") or item.get("description") or ""
        title = item.get("title_en") or item.get("title") or item.get("title_pt") or ""
        return f"{title} {summary}".lower()

    def infer_tags(item) -> list[str]:
        text = news_original_text(item)
        rules = [
            ("Fed", ["fed", "fomc", "powell"]),
            ("Inflacao", ["inflacao", "inflação", "inflation", "cpi", "pce"]),
            ("Treasuries", ["treasury", "treasuries", "yield", "yields", "titulos", "títulos"]),
            ("USD", ["dolar", "dólar", "dollar", "usd", "dxy"]),
            ("Energia", ["petroleo", "petróleo", "oil", "crude", "brent", "wti"]),
            ("Geopolitica", ["ira", "irã", "iran", "israel", "ataque", "war", "guerra"]),
            ("China", ["china", "pboc", "yuan"]),
            ("Brasil", ["brasil", "bcb", "copom", "real", "ibovespa"]),
        ]
        tags = [label for label, needles in rules if any(needle in text for needle in needles)]
        return tags[:4] or ["Macro"]

    def market_impact(item):
        text = news_original_text(item)
        source = str(item.get("source", "")).lower()
        score = 0
        reasons = []
        rules = [
            (5, "Banco Central", ["fed", "fomc", "powell", "ecb", "bce", "boj", "boe", "copom", "bcb", "juros", "interest rate"]),
            (5, "Inflacao", ["cpi", "pce", "ppi", "inflacao", "inflação", "inflation", "core prices"]),
            (4, "Treasuries", ["treasury", "treasuries", "yield", "yields", "titulos", "títulos", "rendimentos"]),
            (4, "USD", ["dolar", "dólar", "dollar", "usd", "dxy", "forex", "cambio", "câmbio"]),
            (4, "Energia", ["petroleo", "petróleo", "oil", "crude", "brent", "wti", "opep", "opec"]),
            (4, "Geopolitica", ["ira", "irã", "iran", "israel", "china", "russia", "rússia", "guerra", "war", "ataque", "sanctions", "sanções"]),
            (4, "Dados Macro", ["payroll", "emprego", "jobs", "jobless", "gdp", "pib", "retail sales", "pmi", "ism"]),
            (3, "Bolsas", ["s&p", "nasdaq", "dow", "stocks", "acoes", "ações", "indices", "índices", "futuros"]),
            (3, "Emergentes", ["brazil", "brasil", "real", "ibovespa", "ewz", "eem", "china", "yuan"]),
        ]
        for weight, label, keywords in rules:
            if any(keyword in text for keyword in keywords):
                score += weight
                reasons.append(label)
        if any(word in text for word in ["breaking", "urgente", "alerta", "unexpected", "surpresa", "forecast", "previsao", "previsão"]):
            score += 3
            reasons.append("Surpresa")
        if any(name in source for name in ["financial", "reuters", "bloomberg", "cnbc"]):
            score += 1

        unique_reasons = []
        for reason in reasons:
            if reason not in unique_reasons:
                unique_reasons.append(reason)
        if score >= 12:
            return "critical", "URGENTE", unique_reasons[:4]
        if score >= 8:
            return "high", "ALTO IMPACTO", unique_reasons[:3]
        if score >= 4:
            return "medium", "IMPACTO MEDIO", unique_reasons[:3]
        return "low", "BAIXO IMPACTO", unique_reasons[:2]

    st.caption("Somente este feed atualiza a cada 30s. O restante do terminal permanece estavel.")
    refresh_col, translate_col = st.columns([1, 1])
    with refresh_col:
        if st.button("Atualizar feed agora", use_container_width=True, key="bb_refresh_news_fast"):
            load_bloomberg_news_feed.clear()
    with translate_col:
        translate_label = "Ver em ingles" if st.session_state.bb_translate_news_fast else "Traduzir noticias"
        if st.button(translate_label, use_container_width=True, key="bb_translate_news_button_fast"):
            st.session_state.bb_translate_news_fast = not st.session_state.bb_translate_news_fast

    filter_term = st.text_input(
        "Filtrar noticias",
        placeholder="Digite Fed, dolar, petroleo, Brasil...",
        label_visibility="collapsed",
        key="bb_news_filter_fast",
    ).strip()

    news_list, news_sources, news_warnings, feed_loaded_at = load_bloomberg_news_feed(0)
    if news_list:
        st.session_state.bb_news_history = news_list[:10]
    elif st.session_state.bb_news_history:
        news_list = st.session_state.bb_news_history
        news_warnings = ["Fonte ao vivo carregando; exibindo historico da sessao."]
    for warning in news_warnings[:2]:
        st.warning(warning)
    if news_list:
        st.session_state.bb_last_news_history = news_list[:10]
    elif st.session_state.get("bb_last_news_history"):
        news_list = st.session_state.bb_last_news_history
        news_sources = ["Historico da sessao"]
        feed_loaded_at = "historico"
    if not news_list:
        st.info("Carregando noticias novas. Assim que houver historico, as ultimas 10 permanecem visiveis aqui.")
        return

    if filter_term:
        term = filter_term.lower()
        filtered_news = [
            item for item in news_list
            if term in news_title(item).lower()
            or term in news_summary(item).lower()
        ]
    else:
        filtered_news = news_list

    impact_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    filtered_news = sorted(
        filtered_news,
        key=lambda item: (
            impact_order.get(market_impact(item)[0], 9),
            -(item.get("timestamp") or 0),
        ),
    )

    if "selected_news_id" not in st.session_state:
        st.session_state.selected_news_id = None
    if not st.session_state.selected_news_id and filtered_news:
        st.session_state.selected_news_id = filtered_news[0].get("id")

    if not filtered_news:
        st.info("Nenhuma manchete correspondente encontrada.")
        return

    translate_enabled = bool(st.session_state.get("bb_translate_news_fast", False))
    visible_news = filtered_news[:45]
    if translate_enabled:
        with st.spinner("Traduzindo feed para portugues do Brasil..."):
            translated_feed = [translate_news_item(item) for item in filtered_news]
        visible_news = translated_feed[:45]

    cards = []
    for idx, item in enumerate(visible_news):
        is_featured = item.get("id") == st.session_state.selected_news_id or idx == 0
        impact_level, impact_label, impact_reasons = market_impact(item)
        title = esc(news_title(item))
        summary_raw = news_summary(item)
        summary = esc(summary_raw)
        published = esc(item.get("published_str", "00:00"))
        source = esc(item.get("source", "Financial Juice"))
        link = safe_external_url(item.get("link"))
        icon_text = esc("FJ" if source == "Financial Juice" else source[:2].upper())
        tags_html = "".join(f'<span class="bb-news-tag">{esc(tag)}</span>' for tag in infer_tags(item))
        impact_badge = (
            f'<span class="bb-impact-badge {impact_level}">{esc(impact_label)}</span>'
            if impact_label
            else ""
        )
        reason_tags = "".join(f'<span class="bb-news-tag">{esc(reason)}</span>' for reason in impact_reasons)
        featured_class = " bb-featured" if is_featured else ""
        impact_class = f" bb-impact-{impact_level}" if impact_level in ["critical", "high", "medium"] else ""
        close_html = '<span class="bb-news-close">x</span>' if is_featured else ""
        summary_html = (
            f'<div class="bb-news-summary">{summary}</div>'
            if summary and summary != title
            else ""
        )
        cards.append(
            f'<div class="bb-news-card{featured_class}{impact_class}">'
            f'{close_html}'
            f'<div class="bb-news-rail"></div>'
            f'<div class="bb-news-icon">{icon_text}</div>'
            f'<div class="bb-news-content">'
            f'<div class="bb-news-title">{title}</div>'
            f'{summary_html}'
            f'<div class="bb-news-meta">'
            f'<span>{published}</span><span>{source}</span>{impact_badge}{reason_tags}{tags_html}'
            f'</div>'
            f'</div>'
            f'<a class="bb-news-link" href="{link}" target="_blank" rel="noopener noreferrer">↗</a>'
            f'</div>'
        )

    feed_header = (
        f'<div class="bb-feed-header">'
        f'<span>Feed de Noticias em Tempo Real</span>'
        f'<span class="bb-live-pill"><span class="bb-status-led"></span>LIVE 30s - {esc(" + ".join(news_sources) or "Fontes")} - {len(filtered_news)} noticias - {"PT-BR" if translate_enabled else "EN"}</span>'
        f'</div>'
    )
    st.markdown(f'<div class="bb-news-feed">{feed_header}{"".join(cards)}</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="bb-status-footer">
        <div>
            <span class="bb-status-led"></span>
            <span style="color: #00FFA3; font-weight: bold;">LIVE FEED</span>
            &nbsp;|&nbsp; {'Traducao manual ativada nos cards visiveis' if translate_enabled else 'Feed em ingles, sem traducao, atualiza a cada 30s'}
            &nbsp;|&nbsp; Origem: {esc(" + ".join(news_sources) or "Fontes")}
        </div>
        <div>
            Ultimo Refresh: {feed_loaded_at}
            &nbsp;|&nbsp; Fonte critica: Financial Juice RSS + cache Supabase
        </div>
    </div>
    """, unsafe_allow_html=True)

def fetch_app_state_with_time(key: str):
    if key not in APP_STATE_ALLOWED_KEYS:
        print(f"[SECURITY] app_state timed read blocked for unexpected key: {key}")
        return None, "Bloqueado", None
    """Busca dados no Supabase e retorna uma tupla (valor, data_atualizacao_formatada_local, dt_utc)."""
    if not supabase: return None, "Sem conexão", None
    try:
        response = supabase.table("app_state").select("value, updated_at").eq("key", key).execute()
        if response.data and len(response.data) > 0:
            val = response.data[0]["value"]
            upd_raw = response.data[0].get("updated_at")
            if upd_raw:
                try:
                    s = upd_raw.replace("Z", "+00:00")
                    dt_utc = datetime.fromisoformat(s)
                    tz_br = timezone(timedelta(hours=-3))
                    dt_local = dt_utc.astimezone(tz_br)
                    return val, dt_local.strftime("%d/%m/%Y %H:%M:%S"), dt_utc
                except Exception as ex:
                    print(f"Erro formatar data {upd_raw}: {ex}")
                    return val, str(upd_raw), None
            return val, "---", None
    except Exception as e:
        print(f"[ERROR] Fetch {key} with time: {e}")
    return None, "Erro", None

def save_credentials(creds):
    print("[SECURITY] save_credentials bloqueado: nao grave credenciais em app_state.")
    return None

# ── Persistência de Trades Manuais ─────────────────────────────────────────
_TRADES_KEY  = "risk_manual_trades"
_TRADES_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".tmp", "manual_trades.json")

def load_manual_trades() -> list:
    """Carrega trades do Supabase; usa cache local como fallback."""
    # Tenta Supabase primeiro
    if supabase:
        try:
            resp = supabase.table("app_state").select("value").eq("key", _TRADES_KEY).execute()
            if resp.data and len(resp.data) > 0:
                val = resp.data[0]["value"]
                if isinstance(val, list):
                    return val
                if isinstance(val, str):
                    import json as _json
                    return _json.loads(val)
        except Exception as e:
            print(f"[WARN] load_manual_trades supabase: {e}")
    # Fallback local
    try:
        if os.path.exists(_TRADES_FILE):
            import json as _json
            with open(_TRADES_FILE, "r", encoding="utf-8") as f:
                return _json.load(f)
    except Exception as e:
        print(f"[WARN] load_manual_trades local: {e}")
    return []

def save_manual_trades(trades: list):
    """Salva trades no Supabase e em cache local (.tmp/manual_trades.json)."""
    import json as _json
    # Salva local (backup)
    try:
        os.makedirs(os.path.dirname(_TRADES_FILE), exist_ok=True)
        with open(_TRADES_FILE, "w", encoding="utf-8") as f:
            _json.dump(trades, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[WARN] save_manual_trades local: {e}")
    # Salva Supabase
    if supabase:
        ok, warning = sync_app_state_value(_TRADES_KEY, trades)
        if not ok and warning:
            print(f"[WARN] save_manual_trades supabase: {warning}")

def sanitize_text(text):
    """Proteção básica contra injeção de scripts."""
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    text = text.replace("javascript:", "")
    return html.escape(text, quote=False)

def clean_val(val):
    """Limpa strings com formatações variadas de milhar/decimal para float robusto."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = s.replace(" ", "")
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s:
        parts = s.split(".")
        if len(parts) == 2 and len(parts[1]) == 3:
            s = s.replace(".", "")
    elif "," in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) == 3:
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    try:
        return float(s)
    except:
        return 0.0



# ── Estilo Global ───────────────────────────────────────────────────────────
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Roboto+Mono:wght@400;700&family=Inter:wght@400;700&display=swap');
    
    html, body, [class*="css"] { font-family: 'Roboto Mono', monospace; background-color: #050505; color: #E0E0E0; }
    
    /* Price Cards */
    .main-card { background: #111111; border-left: 5px solid #FF9800; padding: 20px; border-radius: 5px; box-shadow: 0 4px 15px rgba(0,0,0,0.5); margin-bottom: 20px; }
    .price-large { font-size: 3.2rem; font-weight: bold; color: #FFFFFF; line-height: 1; }
    .label-small { font-size: 0.75rem; color: #888; text-transform: uppercase; margin-bottom: 5px; }
    
    /* Semáforo */
    .status-box { padding: 12px; border-radius: 4px; text-align: center; font-weight: bold; background: #1A1A1A; border: 1px solid #333; font-size: 0.85rem; }
    
    /* Escada Moderna */
    .ladder-container { background: #0A0A0A; border: 1px solid #222; border-radius: 4px; overflow: hidden; margin-top: 15px; font-size: 0.85rem; }
    .ladder-row { display: grid; grid-template-columns: 1fr 1.5fr 1fr 1.5fr; padding: 6px 12px; border-bottom: 1px solid #1a1a1a; align-items: center; }
    .ladder-row:last-child { border-bottom: none; }
    .ladder-header { background: #151515; color: #666; font-weight: bold; text-transform: uppercase; font-size: 0.7rem; letter-spacing: 1px; }
    .level-col { font-weight: bold; }
    .price-col { font-family: 'Roboto Mono', monospace; color: #FFF; text-align: right; font-weight: bold; }
    .delta-col { font-size: 0.75rem; text-align: right; }
    .ajuste-row { background: #1A1A1A !important; border: 1px solid #FF980088; color: #FFF; }
    .pos-row { background: #1F0A0A; color: #FF4B4B; }
    .neg-row { background: #0A1F13; color: #00FFA3; }
    .highlight-row { background: #FFC107 !important; color: #000 !important; font-weight: bold; }
    .highlight-row .price-col, .highlight-row .level-col, .highlight-row .delta-col { color: #000 !important; }
    
    /* IA Banner */
    .ia-banner { background: #111; border: 1px solid #333; border-left: 8px solid #444; padding: 20px; border-radius: 8px; margin: 20px 0; }
    </style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# FRAGMENTS
# ══════════════════════════════════════════════════════════════════════════════

@st.fragment(run_every=30)
def painel_tickers_topo():
    """Mini cards de cotações globais no topo do terminal."""
    global_data = get_global_markets_data()
    if not global_data: return

    # Suporte para formato com ou sem chave 'categories'
    categories = global_data.get("categories", global_data)
    
    # Mapeamento dos ativos solicitados pelo usuário
    targets = {
        "EWZ": "EWZ (Brazil ETF)",
        "EEM": "EEM (Emerging Markets)",
        "6L": "6L (Real CME)",
        "PBR": "PETR4 (ADR)",
        "VALE": "VALE (ADR)",
        "BRENT": "BRENT OIL"
    }
    
    found_assets = {}
    if isinstance(categories, dict):
        for cat_assets in categories.values():
            if not isinstance(cat_assets, list): continue
            for asset in cat_assets:
                for key, target_name in targets.items():
                    if asset.get('name') == target_name:
                        found_assets[key] = asset

    # Renderização em colunas
    cols = st.columns(len(targets))
    for i, key in enumerate(targets.keys()):
        with cols[i]:
            asset = found_assets.get(key)
            if asset:
                change = asset.get('change', 0)
                color = "#00FFA3" if change >= 0 else "#FF4B4B"
                price = asset.get('price', 0)
                # Formatação compacta para o topo
                price_fmt = f"{price:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                st.markdown(f"""
                    <div style="background: #111; border: 1px solid #222; border-top: 2px solid {color}; padding: 8px; border-radius: 4px; text-align: center; box-shadow: 0 2px 5px rgba(0,0,0,0.3);">
                        <div style="font-size: 0.6rem; color: #888; font-weight: bold; text-transform: uppercase; letter-spacing: 1px;">{key}</div>
                        <div style="font-size: 1.1rem; font-weight: bold; color: #FFF; margin: 2px 0;">{price_fmt}</div>
                        <div style="font-size: 0.75rem; color: {color}; font-weight: bold;">{change:+.2f}%</div>
                    </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                    <div style="background: #0A0A0A; border: 1px solid #222; padding: 8px; border-radius: 4px; text-align: center; color: #444;">
                        <div style="font-size: 0.6rem;">{key}</div>
                        <div style="font-size: 1.1rem;">---</div>
                    </div>
                """, unsafe_allow_html=True)


BR_TOP_MOVERS_TICKERS = {
    "PETR4": "PETR4.SA",
    "PETR3": "PETR3.SA",
    "VALE3": "VALE3.SA",
    "ITUB4": "ITUB4.SA",
    "BBDC4": "BBDC4.SA",
    "BBAS3": "BBAS3.SA",
    "B3SA3": "B3SA3.SA",
    "WEGE3": "WEGE3.SA",
    "ABEV3": "ABEV3.SA",
    "RENT3": "RENT3.SA",
    "LREN3": "LREN3.SA",
    "MGLU3": "MGLU3.SA",
    "PRIO3": "PRIO3.SA",
    "SUZB3": "SUZB3.SA",
    "JBSS3": "JBSS3.SA",
    "RAIL3": "RAIL3.SA",
    "ELET3": "ELET3.SA",
    "HAPV3": "HAPV3.SA",
    "CSNA3": "CSNA3.SA",
    "GGBR4": "GGBR4.SA",
    "CMIG4": "CMIG4.SA",
    "VIVT3": "VIVT3.SA",
    "RADL3": "RADL3.SA",
    "EQTL3": "EQTL3.SA",
    "EMBR3": "EMBR3.SA",
    "BRFS3": "BRFS3.SA",
    "BPAC11": "BPAC11.SA",
    "SANB11": "SANB11.SA",
}


def _fmt_br_money(value) -> str:
    try:
        num = float(value)
        return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "---"


@st.cache_data(ttl=180, show_spinner=False)
def get_top_movers_brasil(limit: int = 6) -> dict:
    """Ranking B3 do dia via yfinance, com mini-universo liquido para nao pesar o app."""
    try:
        import yfinance as yf

        symbols = list(BR_TOP_MOVERS_TICKERS.values())
        df = yf.download(
            symbols,
            period="5d",
            interval="5m",
            group_by="ticker",
            progress=False,
            auto_adjust=False,
            prepost=False,
            threads=False,
            timeout=12,
        )
        if df is None or df.empty:
            mark_source("Top Movers Brasil", "error", message="Yahoo Finance retornou vazio.", source="yfinance")
            return {"gainers": [], "losers": [], "updated_at": "", "source": "yfinance"}

        rows = []
        now_label = datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%H:%M:%S")
        for ticker, yf_symbol in BR_TOP_MOVERS_TICKERS.items():
            try:
                ticker_df = df[yf_symbol] if isinstance(df.columns, pd.MultiIndex) else df
                ticker_df = ticker_df.dropna(subset=["Close"])
                if ticker_df.empty:
                    continue

                last_session_date = ticker_df.index[-1].date()
                intraday = ticker_df[ticker_df.index.date == last_session_date]
                if intraday.empty:
                    intraday = ticker_df.tail(1)

                last = float(intraday["Close"].iloc[-1])
                open_price = float(intraday["Open"].dropna().iloc[0]) if "Open" in intraday else last
                previous_sessions = ticker_df[ticker_df.index.date < last_session_date]
                prev_close = float(previous_sessions["Close"].dropna().iloc[-1]) if not previous_sessions.empty else open_price
                reference = prev_close if prev_close > 0 else open_price
                day_change = ((last - reference) / reference) * 100 if reference else 0.0
                day_change_abs = last - reference

                five_min_change = None
                if len(intraday) >= 2:
                    prev_5m = float(intraday["Close"].iloc[-2])
                    if prev_5m:
                        five_min_change = ((last - prev_5m) / prev_5m) * 100

                volume = None
                if "Volume" in intraday:
                    vol_series = intraday["Volume"].dropna()
                    if not vol_series.empty:
                        volume = float(vol_series.sum())

                rows.append({
                    "ticker": ticker,
                    "symbol": yf_symbol,
                    "price": last,
                    "change": day_change,
                    "change_abs": day_change_abs,
                    "change_5m": five_min_change,
                    "volume": volume,
                })
            except Exception:
                continue

        rows.sort(key=lambda item: item.get("change", 0), reverse=True)
        gainers = [row for row in rows if row.get("change", 0) >= 0][:limit]
        losers = sorted([row for row in rows if row.get("change", 0) < 0], key=lambda item: item.get("change", 0))[:limit]
        mark_source("Top Movers Brasil", "ok" if rows else "error", rows=len(rows), message="Ranking B3 via yfinance.", source="yfinance")
        return {"gainers": gainers, "losers": losers, "updated_at": now_label, "source": "yfinance"}
    except Exception as e:
        mark_source("Top Movers Brasil", "error", message=str(e), source="yfinance")
        return {"gainers": [], "losers": [], "updated_at": "", "source": "yfinance"}


def render_top_movers_brasil():
    movers = get_top_movers_brasil()
    gainers = movers.get("gainers") or []
    losers = movers.get("losers") or []
    if not gainers and not losers:
        st.info("Top Movers Brasil indisponivel agora. Tentando novamente no proximo ciclo de cache.")
        return

    def render_row(item, positive: bool) -> str:
        color = "#00FFA3" if positive else "#FF4B4B"
        bg = "rgba(0,255,163,.07)" if positive else "rgba(255,75,75,.08)"
        five = item.get("change_5m")
        accel = ""
        if isinstance(five, (int, float)):
            five_color = "#00FFA3" if five >= 0 else "#FF4B4B"
            arrow = "▲" if five >= 0 else "▼"
            accel = f"<span class='tm-five' style='color:{five_color};'>{arrow} 5m {five:+.2f}%</span>"
        return f"""
        <div class="tm-row" style="background:{bg};">
          <div>
            <strong>{html.escape(str(item.get('ticker', '---')))}</strong>
            <span>{html.escape(str(item.get('symbol', '')))}</span>
          </div>
          <div class="tm-price">{_fmt_br_money(item.get('price'))}</div>
          <div class="tm-change" style="color:{color};">{float(item.get('change') or 0):+.2f}%</div>
          <div class="tm-abs" style="color:{color};">{float(item.get('change_abs') or 0):+.2f}</div>
          {accel}
        </div>
        """

    css = """
    <style>
    .tm-wrap{border:1px solid #1E293B;background:#07111F;border-radius:10px;padding:14px;margin:14px 0 18px 0;}
    .tm-head{display:flex;justify-content:space-between;gap:12px;align-items:end;margin-bottom:10px;}
    .tm-title{color:#F8FAFC;font-weight:950;font-size:1.05rem;}
    .tm-sub{color:#94A3B8;font-size:.72rem;font-weight:800;}
    .tm-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;}
    .tm-box{border:1px solid #172338;border-radius:8px;padding:10px;background:#0B1220;}
    .tm-box h4{margin:0 0 8px 0;font-size:.78rem;text-transform:uppercase;letter-spacing:.04em;color:#CBD5E1;}
    .tm-row{display:grid;grid-template-columns:1.15fr .8fr .62fr .6fr .75fr;gap:8px;align-items:center;border:1px solid #1F2A3D;border-radius:7px;padding:8px 9px;margin-bottom:7px;}
    .tm-row strong{display:block;color:#FFF;font-size:.88rem;line-height:1.05;}
    .tm-row span{display:block;color:#64748B;font-size:.62rem;font-weight:800;margin-top:2px;}
    .tm-price,.tm-change,.tm-abs{font-weight:950;color:#F8FAFC;text-align:right;}
    .tm-five{text-align:right;font-size:.65rem!important;font-weight:950!important;}
    @media(max-width:900px){.tm-grid{grid-template-columns:1fr}.tm-row{grid-template-columns:1fr .8fr .65fr}.tm-abs,.tm-five{display:none!important;}}
    </style>
    """
    html_block = f"""
    {css}
    <section class="tm-wrap">
      <div class="tm-head">
        <div>
          <div class="tm-title">Top Movers Brasil</div>
          <div class="tm-sub">Altas e baixas do dia | Fonte: dados do dashboard + yfinance | Atualizado {html.escape(str(movers.get('updated_at') or '---'))}</div>
        </div>
      </div>
      <div class="tm-grid">
        <div class="tm-box">
          <h4>Maiores altas</h4>
          {''.join(render_row(item, True) for item in gainers)}
        </div>
        <div class="tm-box">
          <h4>Maiores baixas</h4>
          {''.join(render_row(item, False) for item in losers)}
        </div>
      </div>
    </section>
    """
    st.markdown(html_block, unsafe_allow_html=True)

@st.fragment(run_every=2)
def painel_topo_rtd():
    """Parte superior em tempo real (2s): Preços, Métricas e Semáforo."""
    dados = fetch_app_state_fast("dados_mercado")
    if not dados or (isinstance(dados, list) and len(dados) == 0):
        st.info("⏳ Aguardando dados do Terminal Bridge...")
        return

    data = dados[0] if isinstance(dados, list) else dados

    def safe_fmt(v):
        try: return f"{float(v):.2f}" if v is not None else "---"
        except: return str(v)


    # 1. Cabeçalho
    bridge_time = data.get('updated_at', '')[-8:] or "---"
    st.markdown(f"### 📡 {data['symbol']} | <span style='color:#888;'>{bridge_time}</span>", unsafe_allow_html=True)

    # 2. Preços
    c1, c2, c3 = st.columns([2, 1, 1])
    change_val = data['change_percent']
    change_str = f"{change_val:.2%}" if isinstance(change_val, float) else str(change_val)
    color_hex  = '#00FFA3' if (isinstance(change_val, float) and change_val >= 0) else '#FF4B4B'

    with c1:
        last_price_val = clean_val(data.get('last_price', 0))
        st.markdown(f"""
            <div class="main-card">
                <div class="label-small">Último Preço</div>
                <div class="price-large">{last_price_val:,.2f}</div>
                <div style="color:{color_hex}; font-weight:bold; margin-top:5px;">{change_str}</div>
            </div>
        """, unsafe_allow_html=True)
    with c2:
        st.metric("VWAP",   safe_fmt(data['vwap']))
        st.metric("Ajuste", safe_fmt(data['adjustment']))
    with c3:
        saldo_agr = data.get('saldo_agressao')
        if saldo_agr is not None:
            try:
                saldo_num = clean_val(saldo_agr)
                color = "#00FFA3" if saldo_num > 0 else ("#FF4B4B" if saldo_num < 0 else "#E0E0E0")
                # Formata com sinal (+/-) e separadores de milhar no padrão brasileiro
                saldo_fmt = f"{saldo_num:+,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except:
                color = "#E0E0E0"
                saldo_fmt = str(saldo_agr)
            st.markdown(f"<div class='label-small'>Saldo Agressão</div><div style='color:{color}; font-size:1.1rem; font-weight:bold;'>{saldo_fmt}</div>", unsafe_allow_html=True)
            
    # --- INTERPRETAÇÃO DE DIVERGÊNCIA DE DELTA ---
    last_price_val = data.get('last_price')
    vwap_val = data.get('vwap')
    saldo_agr = data.get('saldo_agressao')
    
    if last_price_val is not None and vwap_val is not None and saldo_agr is not None:
        try:
            np_price = clean_val(last_price_val)
            np_vwap = clean_val(vwap_val)
            
            # Ajuste de magnitude robusto
            if np_price > 0 and np_vwap > 0:
                ratio = np_price / np_vwap
                if ratio > 500:
                    np_vwap *= 1000
                elif ratio < 0.002:
                    np_price *= 1000
            
            saldo_num = clean_val(saldo_agr)
            
            if np_price > np_vwap and saldo_num < 0:
                st.markdown("""
                    <div style="background: rgba(255, 75, 75, 0.1); border: 2px solid #FF4B4B; padding: 15px; border-radius: 8px; text-align: center; margin-top: 15px; margin-bottom: 20px; box-shadow: 0 4px 15px rgba(255,75,75,0.2);">
                        <span style="color: #FF4B4B; font-weight: bold; font-size: 1.2rem; letter-spacing: 1px;">⚠️ DIVERGÊNCIA DE DELTA (VENDA)</span><br>
                        <span style="color: #DDD; font-size: 0.85rem;">Preço atual acima da VWAP com Saldo Acumulado Vendedor. Alerta de absorção na venda ou exaustão de compra!</span>
                    </div>
                """, unsafe_allow_html=True)
            elif np_price < np_vwap and saldo_num > 0:
                st.markdown("""
                    <div style="background: rgba(0, 255, 163, 0.1); border: 2px solid #00FFA3; padding: 15px; border-radius: 8px; text-align: center; margin-top: 15px; margin-bottom: 20px; box-shadow: 0 4px 15px rgba(0,255,163,0.2);">
                        <span style="color: #00FFA3; font-weight: bold; font-size: 1.2rem; letter-spacing: 1px;">🚀 DIVERGÊNCIA DE DELTA (COMPRA)</span><br>
                        <span style="color: #DDD; font-size: 0.85rem;">Preço atual abaixo da VWAP com Saldo Acumulado Comprador. Alerta de absorção na compra ou exaustão de venda!</span>
                    </div>
                """, unsafe_allow_html=True)
        except Exception as e:
            pass

    # 3. Semáforo
    st.markdown("#### 🚥 SEMÁFORO DIRECIONAL")
    if "semaforo" in data:
        sem = data["semaforo"]
        def sig_style(text):
            t = str(text).upper() if text else ""
            if any(x in t for x in ["VENDA","VENDER","VENDIDO"]): return "background-color:#400000;color:#FF4B4B;border:1px solid #FF4B4B;"
            if any(x in t for x in ["COMPRA","COMPRAR","COMPRADO"]): return "background-color:#002611;color:#00FFA3;border:1px solid #00FFA3;"
            return "background-color:#1A1A1A;color:#E0E0E0;border:1px solid #333;"

        s1, s2, s3 = st.columns(3)
        with s1: st.markdown(f"<div class='status-box' style='{sig_style(sem.get('direcao'))}'>DIREÇÃO DO DIA<br>{sem.get('direcao','---')}</div>", unsafe_allow_html=True)
        with s2: st.markdown(f"<div class='status-box' style='{sig_style(sem.get('correlacao_rtd'))}'>CORRELAÇÕES RTD<br>{sem.get('correlacao_rtd','---')}</div>", unsafe_allow_html=True)
        with s3: st.markdown(f"<div class='status-box' style='{sig_style(sem.get('correlacao_interna'))}'>CORRELAÇÃO INTERNA<br>{sem.get('correlacao_interna','---')}</div>", unsafe_allow_html=True)

    # 4. Histograma de Variação % do Dia (Movido para baixo do Semáforo a pedido do usuário)
    if data.get("correlacoes") or data.get("acoes_peso"):
        st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
        st.markdown("#### 📊 HISTOGRAMA DE VARIAÇÃO % DO DIA")
        try:
            bar_data = []
            
            # Filtra e adiciona ativos macro das correlações
            if data.get("correlacoes"):
                for row in data["correlacoes"]:
                    fator = row[0]
                    if "CORR" in str(fator).upper():
                        continue
                    var_val = clean_val(row[2])
                    bar_data.append({"Ativo": fator, "Variação %": var_val})
            
            # Adiciona ações de maior peso
            if data.get("acoes_peso"):
                for row in data["acoes_peso"]:
                    fator = row[0]
                    var_val = clean_val(row[3])
                    bar_data.append({"Ativo": fator, "Variação %": var_val})
            
            df_bar = pd.DataFrame(bar_data)
            
            if not df_bar.empty:
                # Garante arredondamento de 2 casas decimais no dataframe para evitar floats longos
                df_bar['Variação %'] = df_bar['Variação %'].round(2)
                
                import plotly.express as px
                df_bar['Cor'] = df_bar['Variação %'].apply(lambda x: '#00FFA3' if x >= 0 else '#FF4B4B')
                
                fig = px.bar(
                    df_bar,
                    x='Variação %',
                    y='Ativo',
                    orientation='h',
                    text='Variação %',
                    color='Cor',
                    color_discrete_map="identity"
                )
                
                num_items = len(df_bar)
                chart_height = max(180, num_items * 28 + 20)
                
                fig.update_layout(
                    paper_bgcolor='rgba(0,0,0,0)',
                    plot_bgcolor='rgba(0,0,0,0)',
                    font_family='"Roboto Mono", monospace',
                    font_color='#E0E0E0',
                    margin=dict(l=10, r=10, t=10, b=10),
                    height=chart_height,
                    xaxis=dict(
                        showgrid=True, 
                        gridcolor='#1a1a1a', 
                        zeroline=True, 
                        zerolinecolor='#444',
                        title=None,
                        ticksuffix="%"
                    ),
                    yaxis=dict(
                        title=None,
                        autorange="reversed"
                    ),
                    showlegend=False
                )
                
                fig.update_traces(
                    texttemplate='%{text:+.2f}%',
                    textposition='inside',
                    insidetextanchor='middle',
                    marker_line_color='#050505',
                    marker_line_width=1,
                    opacity=0.85
                )
                
                st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
        except Exception as e:
            st.error(f"Erro ao gerar histograma de variação: {e}")

@st.fragment(run_every=60)
def secao_ia_fragment():
    """Seção de IA isolada para evitar atualizações constantes (60s)."""
    st.markdown("---")
    
    # Histórico da IA (Tendência)
    history_data = fetch_app_state("ai_insight_history")
    if history_data:
        st.markdown("<div style='font-size: 0.7rem; color: #666; margin-bottom: 8px; font-weight: bold; letter-spacing: 1px;'>⏳ HISTÓRICO DE DIREÇÃO (ÚLTIMAS 5)</div>", unsafe_allow_html=True)
        cols_h = st.columns(5)
        for i in range(5):
            with cols_h[i]:
                if i < len(history_data):
                    h = history_data[i]
                    h_sent = h.get('sentiment', 'NEUTRO')
                    h_time = h.get('updated_at', '')
                    if h_sent == "COMPRA":   h_color, h_bg = "#00FFA3", "#002611"
                    elif h_sent == "VENDA":  h_color, h_bg = "#FF4B4B", "#400000"
                    else:                  h_color, h_bg = "#E0E0E0", "#111"
                    
                    st.markdown(f"""
                        <div style="background:{h_bg}; border: 1px solid {h_color}44; border-radius: 4px; padding: 4px; text-align: center; height: 45px; display: flex; flex-direction: column; justify-content: center;">
                            <div style="font-size: 0.6rem; color: #888;">{h_time}</div>
                            <div style="font-size: 0.7rem; font-weight: bold; color: {h_color};">{h_sent}</div>
                        </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown('<div style="background:#050505; border: 1px dashed #222; border-radius: 4px; height: 45px;"></div>', unsafe_allow_html=True)
        st.markdown("<div style='margin-bottom:15px;'></div>", unsafe_allow_html=True)

    ai_data = fetch_app_state("ai_insight")
    if ai_data:
        sent = ai_data.get('sentiment', 'NEUTRO')
        ibg, itext = ("#002611", "#00FFA3") if sent == "COMPRA" else (("#400000", "#FF4B4B") if sent == "VENDA" else ("#111", "#E0E0E0"))
        
        st.markdown(f"""
            <div style="background:{ibg}; border: 1px solid {itext}44; padding: 20px; border-radius: 8px; border-left: 8px solid {itext}; margin-bottom: 25px;">
                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:12px;">
                    <b style="color:{itext}; font-size: 1rem;">🤖 ANALISTA IA: {sent}</b>
                    <span style="color: #666; font-size: 0.75rem;">{ai_data.get('updated_at', '')}</span>
                </div>
                <div style="color: #E0E0E0; font-size: 0.95rem; line-height: 1.5;">{sanitize_text(ai_data.get('insight', '')).replace(chr(10), '<br>')}</div>
            </div>
        """, unsafe_allow_html=True)

@st.fragment(run_every=300)
def secao_market_report_fragment():
    """Renderiza o Market Report com os tres registros do dia."""
    daily_data = fetch_app_state("market_report_daily")
    latest_report = fetch_app_state("market_report")
    reports = []
    if isinstance(daily_data, dict):
        reports = daily_data.get("reports") or []
    if not reports and latest_report:
        reports = [latest_report]
    st.markdown("---")
    st.markdown("### Market Report")
    st.caption("Atualizacao automatica: 07:05, 13:05 e 19:05 (Sao Paulo). Os reports ficam registrados ate virar o dia.")

    def prime_market_report_files():
        snapshots = {
            "mercados_globais.json": get_global_markets_data(),
            "dados_mercado.json": fetch_app_state_cached("dados_mercado"),
            "calendario_economico.json": get_calendar_data(),
        }
        for path, value in snapshots.items():
            if value:
                try:
                    with open(path, "w", encoding="utf-8") as f:
                        _json.dump(value, f, ensure_ascii=False, indent=2)
                except Exception as e:
                    print(f"[WARN] Falha ao preparar {path} para Market Report: {e}")

    if st.button("Atualizar analise agora", type="primary", use_container_width=True, key="market_report_refresh_now"):
        with st.spinner("Atualizando Market Report..."):
            try:
                import json as _json
                from execution.market_report import generate_market_report

                try:
                    for secret_key in ("GOOGLE_API_KEY", "GEMINI_API_KEY", "OPENAI_API_KEY"):
                        secret_value = st.secrets.get(secret_key, "")
                        if secret_value:
                            os.environ[secret_key] = secret_value
                except Exception:
                    pass

                prime_market_report_files()
                generated = generate_market_report(force=True)
                if not generated:
                    st.warning("Nao foi possivel gerar uma nova analise agora. O fallback local tambem falhou; verifique os logs do Streamlit e as fontes de dados.")
                else:
                    st.session_state["market_report_last_generated"] = generated
                    sync_warnings = []
                    saved_online = False
                    if supabase:
                        for key, paths in {
                            "market_report": ["market_report.json", "execution/market_report.json"],
                            "market_report_daily": ["market_report_daily.json", "execution/market_report_daily.json"],
                        }.items():
                            for path in paths:
                                if os.path.exists(path):
                                    with open(path, "r", encoding="utf-8") as f:
                                        ok, warning = sync_app_state_value(key, _json.load(f))
                                        saved_online = saved_online or ok
                                        if warning:
                                            sync_warnings.append(f"{key}: {warning}")
                                    break
                        fetch_app_state_cached.clear()
                        fetch_app_state_fast.clear()
                        if sync_warnings:
                            st.session_state["market_report_sync_warning"] = "Analise gerada e exibida, mas nao foi possivel salvar tudo no historico online: " + " | ".join(sync_warnings[:2])
                        elif saved_online:
                            st.session_state["market_report_sync_success"] = "Analise atualizada e salva."
                        else:
                            st.session_state["market_report_sync_warning"] = "Analise gerada e exibida nesta sessao, mas nao foi salva no historico online."
                    else:
                        st.session_state["market_report_sync_warning"] = "Analise gerada, mas Supabase esta indisponivel para salvar no historico online."
                    st.rerun()
            except Exception as e:
                st.error(f"Erro ao atualizar Market Report: {e}")

    if st.button("Atualizar Analista IA Macro", use_container_width=True, key="ai_macro_refresh_now"):
        with st.spinner("Atualizando Analista IA Macro..."):
            try:
                import json as _json
                from execution.ai_analyst import generate_macro_insight

                try:
                    for secret_key in ("GOOGLE_API_KEY", "GEMINI_API_KEY"):
                        secret_value = st.secrets.get(secret_key, "")
                        if secret_value:
                            os.environ[secret_key] = secret_value
                except Exception:
                    pass

                prime_market_report_files()
                generate_macro_insight()
                ai_warnings = []
                saved_online = False
                new_insight = None
                for path in ["ai_insight.json", "execution/ai_insight.json"]:
                    if os.path.exists(path):
                        with open(path, "r", encoding="utf-8") as f:
                            new_insight = _json.load(f)
                        break
                if not new_insight:
                    st.warning("Nao foi possivel gerar a analise do Analista IA agora.")
                elif supabase:
                    ok, warning = sync_app_state_value("ai_insight", new_insight)
                    saved_online = saved_online or ok
                    if warning:
                        ai_warnings.append(f"ai_insight: {warning}")
                    try:
                        history = fetch_app_state("ai_insight_history") or []
                        if not isinstance(history, list):
                            history = []
                        history.append({
                            "sentiment": new_insight.get("sentiment", "NEUTRO"),
                            "updated_at": new_insight.get("updated_at", ""),
                            "insight": new_insight.get("insight", ""),
                            "macro_regime": new_insight.get("macro_regime", ""),
                            "confidence": new_insight.get("confidence", ""),
                            "macro_score": new_insight.get("macro_score", 0),
                            "curve_regime": new_insight.get("curve_regime", ""),
                            "curve_bias": new_insight.get("curve_bias", ""),
                            "id": int(time.time()),
                        })
                        ok, warning = sync_app_state_value("ai_insight_history", history[-5:])
                        saved_online = saved_online or ok
                        if warning:
                            ai_warnings.append(f"ai_insight_history: {warning}")
                    except Exception as hist_error:
                        ai_warnings.append(f"ai_insight_history: {hist_error}")

                    fetch_app_state_cached.clear()
                    fetch_app_state_fast.clear()
                    if ai_warnings:
                        st.session_state["ai_macro_sync_warning"] = "Analista IA atualizado, mas houve aviso ao salvar historico online: " + " | ".join(ai_warnings[:2])
                    elif saved_online:
                        st.session_state["ai_macro_sync_success"] = "Analista IA atualizado e salvo."
                    else:
                        st.session_state["ai_macro_sync_warning"] = "Analista IA gerado nesta sessao, mas nao foi salvo no historico online."
                    st.rerun()
                else:
                    st.session_state["ai_macro_sync_warning"] = "Analista IA gerado, mas Supabase esta indisponivel para salvar no historico online."
                    st.rerun()
            except Exception as e:
                st.error(f"Erro ao atualizar Analista IA Macro: {e}")

    if st.session_state.pop("market_report_sync_success", ""):
        st.success("Analise atualizada e salva.")
    sync_warning = st.session_state.pop("market_report_sync_warning", "")
    if sync_warning:
        st.warning(sync_warning)
    if st.session_state.pop("ai_macro_sync_success", ""):
        st.success("Analista IA atualizado e salvo.")
    ai_sync_warning = st.session_state.pop("ai_macro_sync_warning", "")
    if ai_sync_warning:
        st.warning(ai_sync_warning)

    render_macro_news_hub()

    ai_card_html = ""
    ai_data = fetch_app_state_cached("ai_insight")
    ai_history = fetch_app_state_cached("ai_insight_history") or []
    if ai_data:
        sent = ai_data.get("sentiment", "NEUTRO")
        if sent == "COMPRA":
            ai_bg, ai_color, ai_label = "#061F14", "#00FFA3", "COMPRA"
        elif sent == "VENDA":
            ai_bg, ai_color, ai_label = "#260606", "#FF4B4B", "VENDA"
        else:
            ai_bg, ai_color, ai_label = "#171717", "#FFD166", "NEUTRO"
        history_items = []
        if isinstance(ai_history, list):
            ai_history = ai_history[-5:]
        else:
            ai_history = []
        for idx in range(5):
            if idx < len(ai_history):
                item = ai_history[idx]
                h_sent = sanitize_text(str(item.get("sentiment", "NEUTRO")))
                h_time = sanitize_text(str(item.get("updated_at", "---")))
                h_regime = sanitize_text(str(item.get("macro_regime", "")))
                h_color = "#00FFA3" if h_sent == "COMPRA" else ("#FF4B4B" if h_sent == "VENDA" else "#FFD166")
                display_regime = h_regime or ("Risk-on" if h_sent == "COMPRA" else ("Risk-off" if h_sent == "VENDA" else "Neutro"))
                subtitle = f"<small style='display:block; color:#CBD5E1; margin-top:2px; font-size:0.66rem;'>{display_regime}</small>"
                history_items.append(
                    f"<span style='min-width:128px; border:1px solid {h_color}55; color:{h_color}; background:#0B0F16; border-radius:5px; padding:7px 9px; font-size:0.72rem; font-weight:900; text-align:center;'>{h_time} {h_sent}{subtitle}</span>"
                )
            else:
                history_items.append(
                    "<span style='min-width:128px; border:1px dashed #334155; color:#64748B; background:#0B0F16; border-radius:5px; padding:7px 9px; font-size:0.72rem; font-weight:850; text-align:center;'>aguardando<small style='display:block; color:#475569; margin-top:2px; font-size:0.66rem;'>regime</small></span>"
                )
        meta_chips = []
        for label, value in [
            ("Regime", ai_data.get("macro_regime")),
            ("Confianca", ai_data.get("confidence")),
            ("Score", ai_data.get("macro_score")),
            ("Curva", ai_data.get("curve_regime")),
            ("Curva vies", ai_data.get("curve_bias")),
        ]:
            if value not in (None, "", "---"):
                meta_chips.append(
                    f"<span style='border:1px solid #334155; background:#0B0F16; color:#CBD5E1; border-radius:5px; padding:5px 7px; font-size:0.72rem; font-weight:850;'><b style='color:#94A3B8;'>{sanitize_text(str(label))}</b> {sanitize_text(str(value))}</span>"
                )

        panel_height = 920
        ai_card_html = f"""
            <section style="background:{ai_bg}; border:1px solid {ai_color}55; border-left:7px solid {ai_color}; border-radius:8px; padding:18px 20px; margin:16px 0 18px; height:{panel_height}px; overflow-y:auto; box-sizing:border-box;">
                <div style="display:flex; justify-content:space-between; gap:14px; align-items:flex-start; flex-wrap:wrap;">
                    <div>
                        <div style="color:#94A3B8; font-size:0.72rem; font-weight:900; text-transform:uppercase; letter-spacing:.04em;">Analista IA Macro Global</div>
                        <div style="color:{ai_color}; font-size:1.35rem; font-weight:950; margin-top:3px;">{ai_label}</div>
                    </div>
                    <div style="text-align:right; color:#94A3B8; font-size:0.75rem;">{sanitize_text(str(ai_data.get('updated_at', '---')))}</div>
                </div>
                <div style="color:#94A3B8; font-size:0.7rem; font-weight:900; text-transform:uppercase; letter-spacing:.04em; margin-top:14px;">Historico das ultimas 5 analises</div>
                <div style="display:flex; gap:6px; flex-wrap:wrap; margin-top:7px;">{''.join(history_items)}</div>
                <div style="display:flex; gap:6px; flex-wrap:wrap; margin-top:10px;">{''.join(meta_chips)}</div>
                <div style="color:#E5E7EB; font-size:0.94rem; line-height:1.52; margin-top:12px;">{sanitize_text(ai_data.get('insight', '')).replace(chr(10), '<br>')}</div>
            </section>
        """

    if st.session_state.get("market_report_last_generated"):
        generated = st.session_state["market_report_last_generated"]
        if isinstance(generated, dict) and generated not in reports:
            reports.append(generated)

    if not reports:
        if ai_card_html:
            st.markdown(ai_card_html, unsafe_allow_html=True)
        st.info("Nenhum Market Report registrado para hoje ainda.")
        return

    slot_order = {"manha": 1, "tarde": 2, "noite": 3}
    reports = sorted(reports, key=lambda item: slot_order.get(item.get("slot"), 99))
    latest = reports[-1]

    panel_height = 920
    report_header_html = f"""
        <div style="background:#0A0A0A; border:1px solid #1a1a1a; border-top:4px solid #FF9800; padding:18px 20px; border-radius:8px; margin:0 0 14px;">
            <div style="display:flex; justify-content:space-between; gap:15px; align-items:flex-start; margin-bottom:12px; flex-wrap:wrap;">
                <div>
                    <div style="color:#94A3B8; font-size:0.72rem; font-weight:900; text-transform:uppercase; letter-spacing:.04em;">Ultimo report</div>
                    <h3 style="margin:3px 0 0; color:#FF9800; font-family:'Inter', sans-serif;">{sanitize_text(latest.get('slot_label', 'Market Report')).upper()}</h3>
                </div>
                <span style="color:#777; font-size:0.75rem; font-family:'Roboto Mono', monospace;">{sanitize_text(latest.get('updated_at', '---'))} | {sanitize_text(latest.get('provider', 'IA'))}</span>
            </div>
        </div>
    """

    if ai_card_html:
        ai_col, report_col = st.columns([0.48, 0.52], gap="medium")
        with ai_col:
            st.markdown(ai_card_html, unsafe_allow_html=True)
        with report_col:
            with st.container(height=panel_height):
                st.markdown(report_header_html, unsafe_allow_html=True)
                st.markdown(latest.get("report", ""))
    else:
        st.markdown(report_header_html, unsafe_allow_html=True)
        with st.container():
            st.markdown(latest.get("report", ""))


@st.fragment(run_every=2)
def painel_inferior_rtd():
    """Parte inferior em tempo real (2s): Correlações e Escada."""
    dados = fetch_app_state_fast("dados_mercado")
    if not dados or (isinstance(dados, list) and len(dados) == 0):
        return

    data = dados[0] if isinstance(dados, list) else dados



    # 5. Escada de Níveis

    if data.get("escada"):
        st.markdown("""<div style="background:#FF9800; color:#000; padding:5px 15px; font-weight:bold; border-radius:4px 4px 0; font-size:0.8rem; display:flex; justify-content:space-between;">
            <span>ESCADA DE NÍVEIS</span>
            <span>0,5% EM 0,5%</span>
        </div>""", unsafe_allow_html=True)
        
        html = '<div class="ladder-container" style="margin-top:0; border-top:none;">'
        html += '<div class="ladder-row ladder-header" style="grid-template-columns: 0.8fr 1fr 1.2fr 1.2fr 1fr;"><div>NÍVEL</div><div style="text-align:right;">Δ %</div><div style="text-align:right;">PREÇO</div><div style="text-align:right;">Δ P/ ÚLTIMO</div><div style="text-align:right;">MARCADOR</div></div>'
        
        last_price = clean_val(data.get('last_price', 0))
        precos_niveis = [float(row[2]) for row in data["escada"]]
        preco_mais_proximo = min(precos_niveis, key=lambda x: abs(x - last_price))

        for row in data["escada"]:
            nivel, var_pct, preco, dist = row
            preco_val = float(preco)
            is_highlight = abs(preco_val - preco_mais_proximo) < 0.1
            is_ajuste = "AJUSTE" in str(nivel).upper()
            row_class = "highlight-row" if is_highlight else ("ajuste-row" if is_ajuste else ("pos-row" if any(x in str(nivel) for x in ["+","1","2","3","4","5"]) else "neg-row"))
            marcador = "◀ PREÇO" if is_highlight else ""
            p_fmt = f"{preco_val:,.0f}".replace(",", ".")
            d_fmt = f"{float(dist):+,.0f}".replace(",", ".")
            v_fmt = f"{float(var_pct):+.1f}%".replace(".", ",")
            
            html += f'<div class="ladder-row {row_class}" style="grid-template-columns: 0.8fr 1fr 1.2fr 1.2fr 1fr;">'
            html += f'<div class="level-col">{nivel}</div>'
            html += f'<div class="delta-col" style="text-align:right;">{v_fmt}</div>'
            html += f'<div class="price-col" style="text-align:right;">{p_fmt}</div>'
            html += f'<div class="delta-col" style="text-align:right;">{d_fmt}</div>'
            html += f'<div style="text-align:right; font-size:0.7rem;">{marcador}</div>'
            html += '</div>'
        html += '</div>'
        st.markdown(html, unsafe_allow_html=True)

@st.fragment(run_every=30)
def painel_topo_global():
    """Heatmap de indicadores globais no Terminal Global."""
    global_data = get_global_markets_data()
    if not global_data: return
    
    categories = global_data.get("categories", global_data)
    esc_html = html.escape

    st.markdown("""
    <style>
        .tg-heatmap-tabs {
            display: flex;
            align-items: center;
            gap: 6px;
            margin: 4px 0 8px;
            color: #94a3b8;
            font-family: "Roboto Mono", "Consolas", monospace;
        }
        .tg-heatmap-tab {
            border: 1px solid #334155;
            border-radius: 5px;
            padding: 5px 9px;
            font-size: 0.73rem;
            font-weight: 800;
        }
        .tg-heatmap-tab.active {
            background: #1e3a5f;
            color: #f8fafc;
            border-color: #60a5fa;
        }
        .tg-heatmap-tab.muted { border-color: transparent; }
        .tg-heatmap-grid {
            display: grid;
            gap: 4px;
            margin: 10px 0 18px;
        }
        .tg-heatmap-row {
            display: grid;
            grid-template-columns: 18px 1fr;
            gap: 4px;
            align-items: stretch;
            min-width: 0;
        }
        .tg-heatmap-label {
            writing-mode: vertical-rl;
            transform: rotate(180deg);
            display: flex;
            align-items: center;
            justify-content: center;
            background: #111827;
            color: #94a3b8;
            font-size: 0.72rem;
            font-weight: 900;
            border: 1px solid #273142;
            border-radius: 4px;
            text-transform: uppercase;
        }
        .tg-heatmap-card-grid {
            display: grid;
            grid-template-columns: repeat(8, minmax(112px, 1fr));
            gap: 4px;
            min-width: 0;
        }
        .tg-heatmap-card {
            min-height: 86px;
            border: 1px solid rgba(255,255,255,0.15);
            border-radius: 3px;
            padding: 6px 6px 5px;
            color: #fff;
            font-family: "Roboto Mono", "Consolas", monospace;
            overflow: hidden;
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.08);
        }
        .tg-heatmap-card.up { background: linear-gradient(180deg, #21c45a 0%, #15803d 100%); }
        .tg-heatmap-card.down { background: linear-gradient(180deg, #ff3138 0%, #b91c1c 100%); }
        .tg-heatmap-card.flat {
            background: linear-gradient(180deg, #ffe600 0%, #d6b600 100%);
            color: #111827;
        }
        .tg-heatmap-symbol {
            font-size: 0.72rem;
            line-height: 1;
            font-weight: 900;
            text-transform: uppercase;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            text-shadow: 0 1px 1px rgba(0,0,0,0.25);
        }
        .tg-heatmap-price {
            font-size: 1.35rem;
            line-height: 1.05;
            font-weight: 950;
            margin-top: 4px;
            white-space: nowrap;
        }
        .tg-heatmap-range {
            display: grid;
            grid-template-columns: 1fr auto;
            gap: 4px;
            margin-top: 5px;
            font-size: 0.66rem;
            line-height: 1.1;
            font-weight: 800;
        }
        .tg-heatmap-low {
            display: inline-block;
            background: #ffe600;
            color: #111827;
            border: 1px solid rgba(0,0,0,0.22);
            padding: 0 3px;
            border-radius: 2px;
            margin-top: 2px;
        }
        .tg-heatmap-change {
            text-align: right;
            font-size: 0.72rem;
            font-weight: 950;
            white-space: nowrap;
        }
        @media (max-width: 900px) {
            .tg-heatmap-card-grid { grid-template-columns: repeat(3, minmax(98px, 1fr)); }
        }
        @media (max-width: 640px) {
            .tg-heatmap-card-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
            .tg-heatmap-tabs { overflow-x: auto; padding-bottom: 4px; }
        }
    </style>
    """, unsafe_allow_html=True)

    def find_category(name):
        if name in categories:
            return categories.get(name)
        normalized = name.split(" ", 1)[-1].lower()
        for key, value in categories.items():
            if str(key).split(" ", 1)[-1].lower() == normalized:
                return value
        return []

    def assets_from(category_name):
        assets = find_category(category_name)
        return assets if isinstance(assets, list) else []

    def fmt_num(value):
        try:
            value_float = float(value)
            if abs(value_float) >= 10:
                return f"{value_float:.2f}"
            return f"{value_float:.4f}"
        except Exception:
            return "---"

    indices_assets = assets_from("📊 ÍNDICES")
    fx_assets = assets_from("💱 MOEDAS / FOREX")
    etf_assets = assets_from("🇺🇸 ETFs SETORIAIS")
    emerg_assets = assets_from("🌏 EMERGENTES & BRASIL")
    bond_assets = assets_from("🇺🇸 TREASURIES (YIELDS)")
    commodity_assets = assets_from("🛢️ COMMODITIES & CRIPTO")
    commodity_names = {str(asset.get("name", "")).upper(): asset for asset in commodity_assets}
    display_groups = [
        ("Indices", indices_assets),
        ("Energy", [commodity_names[name] for name in ["BRENT OIL", "WTI OIL", "NATURAL GAS"] if name in commodity_names]),
        ("Sectors", etf_assets),
        ("Bonds", bond_assets),
        ("Metals", [commodity_names[name] for name in ["GOLD", "SILVER", "COPPER", "PLATINUM", "PALLADIUM"] if name in commodity_names]),
        ("Crypto", [commodity_names[name] for name in ["BITCOIN", "ETHEREUM", "SOLANA"] if name in commodity_names]),
        ("Currencies", fx_assets),
        ("Emerg", emerg_assets),
    ]
    tabs = ["All", "Indices", "Energy", "Bonds", "Sectors", "Metals", "Crypto", "Currencies"]
    tab_html = "".join(
        f"<span class='tg-heatmap-tab {'active' if tab == 'All' else 'muted'}'>{esc_html(tab)}</span>"
        for tab in tabs
    )
    rows = []
    for label, assets in display_groups:
        if not assets:
            continue
        cards = []
        for asset in assets:
            name = esc_html(str(asset.get("name", "---")))
            price = fmt_num(asset.get("price"))
            high = fmt_num(asset.get("high", asset.get("price")))
            low = fmt_num(asset.get("low", asset.get("price")))
            try:
                change_value = float(asset.get("change", 0))
                change = f"{change_value:+.2f}%"
                change_class = "up" if change_value > 0 else "down" if change_value < 0 else "flat"
            except Exception:
                change = "---"
                change_class = "flat"
            cards.append(
                f"<article class='tg-heatmap-card {change_class}'>"
                f"<div class='tg-heatmap-symbol'>{name}</div>"
                f"<div class='tg-heatmap-price'>{price}</div>"
                f"<div class='tg-heatmap-range'>"
                f"<div><div>H {high}</div><div class='tg-heatmap-low'>L {low}</div></div>"
                f"<div class='tg-heatmap-change'>{change}</div>"
                f"</div>"
                f"</article>"
            )
        rows.append(
            f"<section class='tg-heatmap-row'>"
            f"<div class='tg-heatmap-label'>{esc_html(label)}</div>"
            f"<div class='tg-heatmap-card-grid'>{''.join(cards)}</div>"
            f"</section>"
        )

    st.markdown("#### 🌎 INDICADORES GLOBAIS")
    st.markdown(
        f"<div class='tg-heatmap-tabs'>{tab_html}</div><div class='tg-heatmap-grid'>{''.join(rows)}</div>",
        unsafe_allow_html=True,
    )


def render_source_health_panel():
    """Mostra a saude das fontes sem disparar novas chamadas externas."""
    health = get_source_health(max_age_seconds=1800)
    if not health:
        return

    order = [
        "Yahoo Finance",
        "Mercados Globais Cache",
        "Investing Calendar",
        "Calendario Cache",
        "Financial Juice RSS",
        "Financial Juice Cache",
        "FRED",
        "Lightweight Yahoo",
        "Lightweight BCB",
    ]
    labels = {
        "Yahoo Finance": "Yahoo",
        "Mercados Globais Cache": "Mercados cache",
        "Investing Calendar": "Investing",
        "Calendario Cache": "Calendario cache",
        "Financial Juice RSS": "NEWS",
        "Financial Juice Cache": "NEWS cache",
        "FRED": "FRED",
        "Lightweight Yahoo": "LW Yahoo",
        "Lightweight BCB": "LW BCB",
    }
    status_label = {"ok": "online", "stale": "cache", "error": "erro", "disabled": "off"}

    def fmt_age(seconds):
        try:
            seconds = int(seconds)
        except Exception:
            return "--"
        if seconds < 60:
            return f"{seconds}s"
        if seconds < 3600:
            return f"{seconds // 60}min"
        return f"{seconds // 3600}h"

    items = []
    for name in order:
        item = health.get(name)
        if not item:
            continue
        status = str(item.get("status", "error"))
        rows = item.get("rows")
        rows_text = f" | {rows}" if rows is not None else ""
        msg = html.escape(str(item.get("message") or ""))
        label = html.escape(labels.get(name, name))
        status_safe = html.escape(status)
        state_text = html.escape(status_label.get(status, status))
        rows_safe = html.escape(rows_text)
        age_text = fmt_age(item.get("age_seconds", 0))
        items.append(
            f'<div class="source-health-pill {status_safe}" title="{msg}">'
            f'<span class="source-dot"></span>'
            f'<strong>{label}</strong>'
            f'<em>{state_text} | {age_text}{rows_safe}</em>'
            f'</div>'
        )
    if not items:
        return

    st.markdown(
        f"""<style>
          .source-health-wrap {{
            display:flex; align-items:center; gap:8px; flex-wrap:wrap;
            margin:4px 0 14px; padding:8px 10px;
            border:1px solid rgba(148,163,184,.20); border-radius:8px;
            background:rgba(15,23,42,.56);
            font-family:"Roboto Mono","Consolas",monospace;
          }}
          .source-health-title {{
            color:#94a3b8; font-size:.68rem; font-weight:900;
            letter-spacing:.08em; text-transform:uppercase; margin-right:3px;
          }}
          .source-health-pill {{
            display:flex; align-items:center; gap:6px;
            border:1px solid rgba(148,163,184,.22); border-radius:999px;
            padding:4px 8px; background:rgba(2,6,23,.54);
            color:#cbd5e1; font-size:.68rem; line-height:1;
          }}
          .source-health-pill strong {{ color:#f8fafc; font-size:.68rem; }}
          .source-health-pill em {{ color:#94a3b8; font-style:normal; }}
          .source-health-pill.ok {{ border-color:rgba(34,197,94,.36); }}
          .source-health-pill.stale {{ border-color:rgba(255,152,0,.42); }}
          .source-health-pill.error {{ border-color:rgba(255,75,75,.48); }}
          .source-health-pill.disabled {{ opacity:.72; }}
          .source-dot {{ width:7px; height:7px; border-radius:999px; background:#94a3b8; }}
          .source-health-pill.ok .source-dot {{ background:#22c55e; box-shadow:0 0 9px rgba(34,197,94,.7); }}
          .source-health-pill.stale .source-dot {{ background:#ff9800; box-shadow:0 0 9px rgba(255,152,0,.65); }}
          .source-health-pill.error .source-dot {{ background:#ff4b4b; box-shadow:0 0 9px rgba(255,75,75,.7); }}
        </style><div class="source-health-wrap"><span class="source-health-title">Saude das fontes</span>{''.join(items)}</div>""",
        unsafe_allow_html=True,
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_yield_curve_regime_cached(global_data):
    from execution.yield_curve_regime import analyze_yield_curve_regime

    return analyze_yield_curve_regime(global_data)


def render_yield_curve_regime_panel():
    """Painel local de interpretacao da curva de juros americana."""
    global_data = get_global_markets_data()
    if not global_data:
        return

    try:
        analysis = get_yield_curve_regime_cached(global_data)
    except Exception as exc:
        st.info(f"Curva de juros americana indisponivel no momento: {exc}")
        return

    data = analysis.get("data", {})
    regime = analysis.get("regime", "Neutro")
    confidence = analysis.get("confidence", "Baixo")
    bias = analysis.get("operational_bias", "Neutro")
    impacts = analysis.get("impacts", {})
    esc = html.escape

    regime_class = "neutral"
    if "Bear" in regime or "Parallel Up" in regime or "Risk-off" in bias:
        regime_class = "off"
    elif "Bull" in regime or "Parallel Down" in regime or "Risk-on" in bias:
        regime_class = "on"

    def fmt_yield(key):
        val = data.get(key)
        chg = data.get(f"{key}_change_bps")
        if val is None:
            return "---", "---"
        chg_txt = "---" if chg is None else f"{float(chg):+.1f} bps"
        return f"{float(val):.2f}%", chg_txt

    def fmt_spread(key):
        val = data.get(key)
        chg = data.get(f"{key}_change_bps")
        if val is None:
            return "---", "---"
        chg_txt = "---" if chg is None else f"{float(chg):+.1f} bps"
        return f"{float(val):+.2f} pp", chg_txt

    yield_cards = []
    for label, key in [("2Y", "us02y"), ("5Y", "us05y"), ("10Y", "us10y"), ("30Y", "us30y")]:
        value, change = fmt_yield(key)
        yield_cards.append(
            f"<div class='yc-mini'><span>{label}</span><strong>{value}</strong><small>{change}</small></div>"
        )

    spread_10_2, spread_10_2_chg = fmt_spread("spread_10y_2y")
    spread_30_5, spread_30_5_chg = fmt_spread("spread_30y_5y")
    impact_html = "".join(
        f"<span class='yc-chip'><b>{esc(str(asset))}</b> {esc(str(view))}</span>"
        for asset, view in impacts.items()
    )

    st.markdown(
        f"""
        <style>
            .yc-panel {{
                border: 1px solid #263247;
                background: linear-gradient(180deg, #0b1220 0%, #090d14 100%);
                border-radius: 8px;
                padding: 16px 18px;
                margin: 2px 0 18px;
                color: #E5E7EB;
                box-shadow: 0 10px 28px rgba(0,0,0,0.22);
            }}
            .yc-head {{
                display: flex;
                justify-content: space-between;
                gap: 12px;
                align-items: flex-start;
                flex-wrap: wrap;
                margin-bottom: 12px;
            }}
            .yc-kicker {{
                color: #93A4B8;
                font: 800 0.72rem "Roboto Mono", monospace;
                text-transform: uppercase;
                letter-spacing: .03em;
            }}
            .yc-title {{
                font-size: 2rem;
                line-height: 1.02;
                font-weight: 950;
                color: #F8FAFC;
                margin-top: 5px;
                text-transform: uppercase;
                letter-spacing: 0;
                text-shadow: 0 0 22px rgba(96,165,250,.22);
            }}
            .yc-badge {{
                border-radius: 8px;
                padding: 11px 14px;
                font-weight: 950;
                text-align: right;
                min-width: 230px;
                border: 2px solid #334155;
                background: #111827;
                box-shadow: inset 0 1px 0 rgba(255,255,255,.06), 0 0 22px rgba(0,0,0,.20);
            }}
            .yc-badge.on {{
                color: #00FFA3;
                border-color: rgba(0,255,163,.72);
                background: linear-gradient(180deg, rgba(0,255,163,.16), #0f1724);
            }}
            .yc-badge.off {{
                color: #FF4B4B;
                border-color: rgba(255,75,75,.8);
                background: linear-gradient(180deg, rgba(255,75,75,.18), #0f1724);
            }}
            .yc-badge.neutral {{
                color: #FFD166;
                border-color: rgba(255,209,102,.75);
                background: linear-gradient(180deg, rgba(255,209,102,.16), #0f1724);
            }}
            .yc-badge-main {{
                font-size: 1.45rem;
                line-height: 1.1;
                margin-top: 3px;
                text-transform: uppercase;
            }}
            .yc-grid {{
                display: grid;
                grid-template-columns: repeat(6, minmax(0, 1fr));
                gap: 8px;
                margin: 10px 0 12px;
            }}
            .yc-mini {{
                border: 1px solid #1f2a3a;
                background: #0f1724;
                border-radius: 7px;
                padding: 8px 10px;
                min-height: 72px;
            }}
            .yc-mini span, .yc-mini small {{
                display:block;
                color:#94A3B8;
                font-size:.72rem;
                font-weight:800;
            }}
            .yc-mini strong {{
                display:block;
                color:#F8FAFC;
                font-size:1.08rem;
                margin:4px 0 2px;
            }}
            .yc-reading {{
                color: #CBD5E1;
                line-height: 1.45;
                font-size: 1rem;
                margin: 10px 0 12px;
                border-left: 4px solid #60A5FA;
                padding: 9px 12px;
                background: rgba(15,23,36,.78);
                border-radius: 6px;
            }}
            .yc-chips {{
                display:flex;
                flex-wrap:wrap;
                gap:6px;
                margin-top: 10px;
            }}
            .yc-chip {{
                border:1px solid #2b3b52;
                background:#111827;
                color:#CBD5E1;
                border-radius:5px;
                padding:5px 7px;
                font-size:.75rem;
                font-weight:800;
            }}
            .yc-footer {{
                margin-top: 10px;
                border-top: 1px solid #1f2a3a;
                padding-top: 9px;
                color:#F8FAFC;
                font-size:.86rem;
                font-weight:800;
            }}
            @media (max-width: 980px) {{
                .yc-grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
            }}
        </style>
        <section class="yc-panel">
            <div class="yc-head">
                <div>
                    <div class="yc-kicker">Curva de juros americana | IA local TTS</div>
                    <div class="yc-title">{esc(regime)}</div>
                    <div style="color:#94A3B8; font-size:.78rem; margin-top:3px;">Fonte: {esc(str(analysis.get('source', '---')))} | confianca {esc(confidence)}</div>
                </div>
                <div class="yc-badge {regime_class}">
                    <div style="font-size:.72rem; color:#94A3B8;">Vies operacional</div>
                    <div class="yc-badge-main">{esc(bias)}</div>
                </div>
            </div>
            <div class="yc-grid">
                {''.join(yield_cards)}
                <div class="yc-mini"><span>10Y - 2Y</span><strong>{spread_10_2}</strong><small>{spread_10_2_chg}</small></div>
                <div class="yc-mini"><span>30Y - 5Y</span><strong>{spread_30_5}</strong><small>{spread_30_5_chg}</small></div>
            </div>
            <div class="yc-reading">{esc(str(analysis.get('macro_reading', '')))}</div>
            <div class="yc-chips">{impact_html}</div>
            <div class="yc-footer">{esc(str(analysis.get('trader_sentence', '')))}</div>
        </section>
        """,
        unsafe_allow_html=True,
    )

@st.fragment(run_every=30)
def painel_corpo_global():
    """Tabelas detalhadas de mercados globais."""
    global_data = get_global_markets_data()
    if not global_data: return
    
    categories = global_data.get("categories", global_data)

    def styled_change_dataframe(df):
        def color_change(val):
            try:
                v = float(val)
                color = '#00FFA3' if v >= 0 else '#FF4B4B'
                return f'color: {color}; font-weight: bold'
            except Exception:
                return ''

        styler = df.style
        if hasattr(styler, "map"):
            return styler.map(color_change, subset=['Var %'])
        return styler.applymap(color_change, subset=['Var %'])
    
    st.markdown("---")
    # Organizar em colunas de 2 para economizar espaço
    cat_names = list(categories.keys())
    for i in range(0, len(cat_names), 2):
        c1, c2 = st.columns(2)
        with c1:
            cat = cat_names[i]
            st.markdown(f"##### {cat}")
            assets = categories[cat]
            df = pd.DataFrame(assets)[['name', 'price', 'change']]
            df.columns = ['Ativo', 'Preço', 'Var %']
            def color_change(val):
                try: 
                    v = float(val)
                    color = '#00FFA3' if v >= 0 else '#FF4B4B'
                    return f'color: {color}; font-weight: bold'
                except: return ''
            st.dataframe(styled_change_dataframe(df), hide_index=True, use_container_width=True)
        
        if i + 1 < len(cat_names):
            with c2:
                cat = cat_names[i+1]
                st.markdown(f"##### {cat}")
                assets = categories[cat]
                df = pd.DataFrame(assets)[['name', 'price', 'change']]
                df.columns = ['Ativo', 'Preço', 'Var %']
                st.dataframe(styled_change_dataframe(df), hide_index=True, use_container_width=True)

def pagina_terminal_bloomberg():
    """Pagina do Terminal Bloomberg de noticias com atualizacao leve e cacheada."""
    def esc(value) -> str:
        return html.escape(str(value or ""), quote=True)

    def news_title(item) -> str:
        return item.get("title_en") or item.get("title") or item.get("title_pt") or "---"

    def news_summary(item) -> str:
        summary = item.get("summary") or item.get("description") or ""
        return "" if summary == news_title(item) else summary

    def infer_tags(item) -> list[str]:
        text = f"{news_title(item)} {news_summary(item)}".lower()
        rules = [
            ("Fed", ["fed", "fomc", "powell"]),
            ("Inflação", ["inflação", "inflation", "cpi", "pce"]),
            ("Títulos dos EUA", ["treasury", "treasuries", "yield", "yields", "títulos"]),
            ("Índices dos EUA", ["s&p", "nasdaq", "dow", "índices", "stocks", "ações"]),
            ("USD", ["dólar", "dollar", "usd", "dxy"]),
            ("Energia", ["petróleo", "oil", "crude", "brent", "wti"]),
            ("Geopolítica", ["irã", "iran", "israel", "ataque", "war", "guerra"]),
            ("China", ["china", "pboc", "yuan"]),
            ("Brasil", ["brasil", "bcb", "copom", "real", "ibovespa"]),
        ]
        tags = [label for label, needles in rules if any(needle in text for needle in needles)]
        return tags[:4] or ["Macro"]

    def market_impact(item):
        text = f"{news_title(item)} {news_summary(item)}".lower()
        source = str(item.get("source", "")).lower()
        score = 0
        reasons = []

        rules = [
            (5, "Banco Central", ["fed", "fomc", "powell", "williams", "jefferson", "musalem", "bce", "ecb", "boj", "boe", "copom", "bcb", "juros", "taxa de juros", "interest rate"]),
            (5, "Inflação", ["cpi", "pce", "ppi", "inflação", "inflation", "núcleo", "core prices"]),
            (4, "Treasuries", ["treasury", "treasuries", "yield", "yields", "títulos dos eua", "rendimentos"]),
            (4, "USD", ["dólar", "dollar", "usd", "dxy", "forex", "câmbio"]),
            (4, "Energia", ["petróleo", "oil", "crude", "brent", "wti", "opep", "opec", "hormuz"]),
            (4, "Geopolítica", ["irã", "iran", "israel", "china", "russia", "rússia", "guerra", "war", "ataque", "missile", "sanções"]),
            (4, "Dados Macro", ["payroll", "emprego", "jobs", "jobless", "gdp", "pib", "retail sales", "pmi", "ism"]),
            (3, "Bolsas", ["s&p", "nasdaq", "dow", "stocks", "ações", "índices", "futuros"]),
            (3, "Emergentes", ["brazil", "brasil", "real", "ibovespa", "ewz", "eem", "china", "yuan"]),
            (3, "Cripto", ["bitcoin", "crypto", "ethereum", "cripto"]),
        ]

        for weight, label, keywords in rules:
            if any(keyword in text for keyword in keywords):
                score += weight
                reasons.append(label)

        if any(word in text for word in ["breaking", "urgente", "alerta", "unexpected", "surpresa", "forecast", "previsão", "acima do esperado", "abaixo do esperado"]):
            score += 3
            reasons.append("Surpresa")

        if any(name in source for name in ["financial", "reuters", "bloomberg", "cnbc"]):
            score += 1

        unique_reasons = []
        for reason in reasons:
            if reason not in unique_reasons:
                unique_reasons.append(reason)

        if score >= 12:
            return "critical", "URGENTE", unique_reasons[:4]
        if score >= 8:
            return "high", "ALTO IMPACTO", unique_reasons[:3]
        if score >= 4:
            return "medium", "IMPACTO MEDIO", unique_reasons[:3]
        return "low", "BAIXO IMPACTO", unique_reasons[:2]
    
    # CSS Customizado Exclusivo para o Terminal Bloomberg
    st.markdown("""
    <style>
        .bb-terminal-container {
            background-color: #000000 !important;
            color: #00FF00 !important;
            font-family: 'Consolas', 'Courier New', monospace !important;
            padding: 1.5rem;
            border-radius: 8px;
            border: 2px solid #222222;
            margin-bottom: 20px;
        }
        
        .bb-ticker-bar {
            background-color: #000000;
            border: 1px solid #222222;
            padding: 0.5rem;
            border-radius: 6px;
            display: flex;
            overflow-x: auto;
            white-space: nowrap;
            margin-bottom: 15px;
            font-family: 'Consolas', monospace;
            font-size: 0.8rem;
        }
        
        .bb-ticker-item {
            margin-right: 1.5rem;
            display: inline-block;
        }
        
        .bb-ticker-up {
            color: #00FFA3 !important;
            font-weight: bold;
        }
        
        .bb-ticker-down {
            color: #FF4B4B !important;
            font-weight: bold;
        }

        .bb-quote-grid {
            display: grid;
            gap: 4px;
            margin: 10px 0 18px;
        }

        .bb-quote-panel {
            display: grid;
            grid-template-columns: 18px 1fr;
            gap: 4px;
            align-items: stretch;
            min-width: 0;
        }

        .bb-quote-title {
            writing-mode: vertical-rl;
            transform: rotate(180deg);
            display: flex;
            align-items: center;
            justify-content: center;
            background: #111827;
            color: #94a3b8;
            font-family: "Inter", "Segoe UI", Arial, sans-serif;
            font-size: 0.72rem;
            font-weight: 900;
            letter-spacing: 0;
            margin: 0;
            border: 1px solid #273142;
            border-radius: 4px;
            text-transform: uppercase;
        }

        .bb-quote-card-grid {
            display: grid;
            grid-template-columns: repeat(8, minmax(112px, 1fr));
            gap: 4px;
            min-width: 0;
        }

        .bb-quote-card {
            min-height: 86px;
            border: 1px solid rgba(255,255,255,0.15);
            border-radius: 3px;
            padding: 6px 6px 5px;
            color: #fff;
            font-family: "Roboto Mono", "Consolas", monospace;
            overflow: hidden;
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.08);
        }

        .bb-quote-card.up {
            background: linear-gradient(180deg, #21c45a 0%, #15803d 100%);
        }

        .bb-quote-card.down {
            background: linear-gradient(180deg, #ff3138 0%, #b91c1c 100%);
        }

        .bb-quote-card.flat {
            background: linear-gradient(180deg, #5f5362 0%, #3f3947 100%);
        }

        .bb-quote-symbol {
            font-size: 0.72rem;
            line-height: 1;
            font-weight: 900;
            text-transform: uppercase;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            text-shadow: 0 1px 1px rgba(0,0,0,0.25);
        }

        .bb-quote-price {
            font-size: 1.35rem;
            line-height: 1.05;
            font-weight: 950;
            letter-spacing: -0.02em;
            margin-top: 4px;
            white-space: nowrap;
        }

        .bb-quote-range {
            display: grid;
            grid-template-columns: 1fr auto;
            gap: 4px;
            margin-top: 5px;
            font-size: 0.66rem;
            line-height: 1.1;
            font-weight: 800;
        }

        .bb-quote-low {
            display: inline-block;
            background: #ffe600;
            color: #111827;
            border: 1px solid rgba(0,0,0,0.22);
            padding: 0 3px;
            border-radius: 2px;
            margin-top: 2px;
        }

        .bb-quote-change {
            text-align: right;
            font-size: 0.72rem;
            font-weight: 950;
            white-space: nowrap;
        }

        .bb-quote-tabs {
            display: flex;
            align-items: center;
            gap: 6px;
            margin: 4px 0 8px;
            color: #94a3b8;
            font-family: "Roboto Mono", "Consolas", monospace;
        }

        .bb-quote-tab {
            border: 1px solid #334155;
            border-radius: 5px;
            padding: 5px 9px;
            font-size: 0.73rem;
            font-weight: 800;
        }

        .bb-quote-tab.active {
            background: #1e3a5f;
            color: #f8fafc;
            border-color: #60a5fa;
        }

        .bb-quote-tab.muted {
            border-color: transparent;
        }

        .bb-news-feed {
            position: sticky;
            top: 0.75rem;
            max-height: calc(100vh - 225px);
            min-height: 420px;
            overflow-y: auto;
            background: #090d12;
            border: 1px solid #263443;
            border-radius: 7px;
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.03);
        }

        .bb-feed-header {
            position: sticky;
            top: 0;
            z-index: 2;
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 7px 10px;
            background: #151f2a;
            border-bottom: 1px solid #283544;
            color: #9aa6b2;
            font-family: "Consolas", monospace;
            font-size: 0.75rem;
            text-transform: uppercase;
        }

        .bb-live-pill {
            display: inline-flex;
            align-items: center;
            gap: 5px;
            color: #00FFA3;
            font-weight: 800;
        }

        .bb-news-card {
            position: relative;
            display: grid;
            grid-template-columns: 6px 34px minmax(0, 1fr) 22px;
            gap: 9px;
            padding: 9px 10px 8px 0;
            min-height: 64px;
            background: #18222d;
            border-bottom: 1px solid #0c1218;
            color: #d8dee7;
            font-family: "Inter", "Segoe UI", Arial, sans-serif;
        }

        .bb-news-card.bb-featured {
            background: #1f3141;
            min-height: 124px;
        }

        .bb-news-rail {
            width: 6px;
            align-self: stretch;
            background: #34495e;
        }

        .bb-news-card.bb-impact-high {
            background: #2a181b;
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.03), 0 0 0 1px rgba(255,59,48,0.18);
        }

        .bb-news-card.bb-impact-critical {
            background: linear-gradient(90deg, #3a090b 0%, #211216 100%);
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.05), 0 0 0 1px rgba(255,59,48,0.45), 0 0 18px rgba(255,59,48,0.12);
        }

        .bb-news-card.bb-impact-medium {
            background: #261f12;
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.03), 0 0 0 1px rgba(255,153,0,0.14);
        }

        .bb-news-card.bb-impact-critical .bb-news-rail {
            background: #ff1f1f;
        }

        .bb-news-card.bb-impact-high .bb-news-rail {
            background: #ff3b30;
        }

        .bb-news-card.bb-impact-medium .bb-news-rail {
            background: #ff9900;
        }

        .bb-news-card.bb-impact-critical .bb-news-title {
            color: #ffeded;
            font-weight: 950;
            text-shadow: 0 0 10px rgba(255,59,48,0.25);
        }

        .bb-news-card.bb-impact-high .bb-news-title {
            color: #ff6b5f;
            font-weight: 800;
        }

        .bb-news-card.bb-impact-medium .bb-news-title {
            color: #ffb24a;
        }

        .bb-news-icon {
            width: 28px;
            height: 28px;
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            margin-top: 4px;
            background: #18222d;
            border: 3px solid #36d5f5;
            color: #d8f8ff;
            font-size: 0.62rem;
            font-weight: 900;
            letter-spacing: 0.2px;
        }

        .bb-news-title {
            color: #edf2f7;
            font-size: 0.95rem;
            font-weight: 700;
            line-height: 1.3;
            margin-bottom: 3px;
        }

        .bb-news-card:not(.bb-featured) .bb-news-title {
            font-weight: 500;
        }

        .bb-news-summary {
            color: #d1d8e0;
            font-size: 0.88rem;
            line-height: 1.42;
            margin-top: 4px;
        }

        .bb-news-card:not(.bb-featured) .bb-news-summary {
            display: -webkit-box;
            -webkit-line-clamp: 2;
            -webkit-box-orient: vertical;
            overflow: hidden;
        }

        .bb-news-meta {
            display: flex;
            flex-wrap: wrap;
            align-items: center;
            gap: 6px;
            margin-top: 6px;
            color: #9aa6b2;
            font-size: 0.76rem;
            line-height: 1.25;
        }

        .bb-news-tag {
            display: inline-flex;
            align-items: center;
            border-radius: 4px;
            padding: 1px 6px;
            background: #303946;
            color: #b7c0ca;
            font-size: 0.7rem;
            line-height: 1.45;
        }

        .bb-impact-badge {
            display: inline-flex;
            align-items: center;
            border-radius: 4px;
            padding: 1px 6px;
            font-size: 0.68rem;
            line-height: 1.45;
            font-weight: 900;
            letter-spacing: 0.3px;
        }

        .bb-impact-badge.high {
            background: #4a1111;
            color: #ff6b5f;
            border: 1px solid rgba(255,59,48,0.35);
        }

        .bb-impact-badge.critical {
            background: #ff2d20;
            color: #fff;
            border: 1px solid rgba(255,255,255,0.28);
            box-shadow: 0 0 14px rgba(255,59,48,0.22);
        }

        .bb-impact-badge.medium {
            background: #3d2804;
            color: #ffb24a;
            border: 1px solid rgba(255,153,0,0.35);
        }

        .bb-impact-badge.low {
            background: #1f2937;
            color: #94a3b8;
            border: 1px solid rgba(148,163,184,0.18);
        }

        .bb-news-link {
            color: #aeb8c4 !important;
            text-decoration: none !important;
            align-self: end;
            justify-self: center;
            font-size: 0.9rem;
            opacity: 0.85;
        }

        .bb-news-link:hover {
            color: #36d5f5 !important;
            opacity: 1;
        }

        .bb-news-close {
            position: absolute;
            top: 4px;
            right: 6px;
            color: #bac4ce;
            font-size: 1.2rem;
            font-weight: 800;
            line-height: 1;
        }
        
        .bb-status-footer {
            background-color: #050505;
            border: 1px solid #222222;
            padding: 0.4rem 1rem;
            border-radius: 4px;
            font-family: 'Consolas', monospace;
            font-size: 0.75rem;
            color: #6b7280;
            margin-top: 15px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        @media (max-width: 900px) {
            .bb-quote-card-grid {
                grid-template-columns: repeat(3, minmax(98px, 1fr));
            }
            .bb-status-footer {
                display: block;
            }
        }

        @media (max-width: 640px) {
            .bb-quote-card-grid {
                grid-template-columns: repeat(2, minmax(0, 1fr));
            }
            .bb-quote-tabs {
                overflow-x: auto;
                padding-bottom: 4px;
            }
        }
        
        .bb-status-led {
            display: inline-block;
            width: 7px;
            height: 7px;
            border-radius: 50%;
            background-color: #00FFA3;
            box-shadow: 0 0 6px #00FFA3;
            margin-right: 5px;
            animation: bb-led-pulse 1.2s infinite;
        }
        
        @keyframes bb-led-pulse {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.3; }
        }
    </style>
    """, unsafe_allow_html=True)

    global_data = get_global_markets_data()

    def render_quote_grids(data):
        if not data:
            st.info("Aguardando grades de cotacoes globais.")
            return

        categories = data.get("categories", data)
        preferred_order = [
            "📊 ÍNDICES",
            "💱 MOEDAS / FOREX",
            "🇺🇸 ETFs SETORIAIS",
            "🌏 EMERGENTES & BRASIL",
            "🇺🇸 TREASURIES (YIELDS)",
            "🛢️ COMMODITIES & CRIPTO",
        ]

        def find_category(name):
            if name in categories:
                return categories.get(name)
            normalized = name.split(" ", 1)[-1].lower()
            for key, value in categories.items():
                if str(key).split(" ", 1)[-1].lower() == normalized:
                    return value
            return None

        def fmt_num(value):
            try:
                value_float = float(value)
                if abs(value_float) >= 10:
                    return f"{value_float:.2f}"
                return f"{value_float:.4f}"
            except Exception:
                return "---"

        def assets_from(category_name):
            assets = find_category(category_name)
            return assets if isinstance(assets, list) else []

        indices_assets = assets_from("📊 ÍNDICES")
        fx_assets = assets_from("💱 MOEDAS / FOREX")
        etf_assets = assets_from("🇺🇸 ETFs SETORIAIS")
        emerg_assets = assets_from("🌏 EMERGENTES & BRASIL")
        bond_assets = assets_from("🇺🇸 TREASURIES (YIELDS)")
        commodity_assets = assets_from("🛢️ COMMODITIES & CRIPTO")

        commodity_names = {str(asset.get("name", "")).upper(): asset for asset in commodity_assets}
        display_groups = [
            ("Indices", indices_assets),
            ("Energy", [commodity_names[name] for name in ["BRENT OIL", "WTI OIL", "NATURAL GAS"] if name in commodity_names]),
            ("Sectors", etf_assets),
            ("Bonds", bond_assets),
            ("Metals", [commodity_names[name] for name in ["GOLD", "SILVER", "COPPER", "PLATINUM", "PALLADIUM"] if name in commodity_names]),
            ("Crypto", [commodity_names[name] for name in ["BITCOIN", "ETHEREUM", "SOLANA"] if name in commodity_names]),
            ("Currencies", fx_assets),
            ("Emerg", emerg_assets),
        ]

        panels = []
        tabs = ["All", "Indices", "Energy", "Bonds", "Sectors", "Metals", "Crypto", "Currencies"]
        for category_name, assets in display_groups:
            if not assets:
                continue
            cards = []
            for asset in assets:
                name = esc(asset.get("name", "---"))
                price = fmt_num(asset.get("price"))
                try:
                    change_value = float(asset.get("change", 0))
                    change = f"{change_value:+.2f}%"
                    change_class = "up" if change_value > 0 else "down" if change_value < 0 else "flat"
                except Exception:
                    change = "---"
                    change_class = "flat"
                high = fmt_num(asset.get("high", asset.get("price")))
                low = fmt_num(asset.get("low", asset.get("price")))
                cards.append(
                    f"<article class='bb-quote-card {change_class}'>"
                    f"<div class='bb-quote-symbol'>{name}</div>"
                    f"<div class='bb-quote-price'>{price}</div>"
                    f"<div class='bb-quote-range'>"
                    f"<div><div>H {high}</div><div class='bb-quote-low'>L {low}</div></div>"
                    f"<div class='bb-quote-change'>{change}</div>"
                    f"</div>"
                    f"</article>"
                )

            panels.append(
                f"<section class='bb-quote-panel'>"
                f"<h3 class='bb-quote-title'>{esc(category_name)}</h3>"
                f"<div class='bb-quote-card-grid'>{''.join(cards)}</div>"
                f"</section>"
            )

        if panels:
            tab_html = "".join(
                f"<span class='bb-quote-tab {'active' if tab == 'All' else 'muted'}'>{tab}</span>"
                for tab in tabs
            )
            st.markdown(f"<div class='bb-quote-tabs'>{tab_html}</div><div class='bb-quote-grid'>{''.join(panels)}</div>", unsafe_allow_html=True)

    render_bloomberg_news_feed_fragment()
    render_quote_grids(global_data)
    return

    st.caption("Atualizacao automatica reduzida para 60s para manter a pagina responsiva. Use o filtro para focar nas manchetes relevantes.")
    if st.button("Atualizar feed agora", use_container_width=True, key="bb_refresh_news"):
        load_bloomberg_news_feed.clear()
    filter_term = st.text_input(
        "Filtrar noticias",
        placeholder="Digite Fed, dolar, petroleo, Brasil...",
        label_visibility="collapsed",
        key="bb_news_filter",
    ).strip()

    news_list, news_sources, news_warnings, feed_loaded_at = load_bloomberg_news_feed(0)
    for warning in news_warnings[:2]:
        st.warning(warning)

    if not news_list:
        st.info("Carregando noticias novas. Assim que houver historico, as ultimas 10 permanecem visiveis aqui.")
        return

    # Filtra notícias se houver termo ativo
    if filter_term:
        filtered_news = [
            item for item in news_list
            if filter_term.lower() in news_title(item).lower()
            or filter_term.lower() in news_summary(item).lower()
        ]
    else:
        filtered_news = news_list

    impact_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    filtered_news = sorted(
        filtered_news,
        key=lambda item: (
            impact_order.get(market_impact(item)[0], 9),
            -(item.get("timestamp") or 0),
        ),
    )

    critical_count = sum(1 for item in filtered_news if market_impact(item)[0] == "critical")
    high_count = sum(1 for item in filtered_news if market_impact(item)[0] == "high")
    medium_count = sum(1 for item in filtered_news if market_impact(item)[0] == "medium")
    latest_time = esc(filtered_news[0].get("published_str", "--:--")) if filtered_news else "--:--"
    st.markdown(
        f'<div class="bb-news-toolbar">'
        f'<div class="bb-news-stat"><span>Noticias</span><strong>{len(filtered_news)}</strong></div>'
        f'<div class="bb-news-stat"><span>Urgente</span><strong style="color:#ff2d20;">{critical_count}</strong></div>'
        f'<div class="bb-news-stat"><span>Alto impacto</span><strong style="color:#ff6b5f;">{high_count}</strong></div>'
        f'<div class="bb-news-stat"><span>Impacto medio</span><strong style="color:#ffb24a;">{medium_count}</strong></div>'
        f'<div class="bb-news-stat"><span>Mais recente</span><strong>{latest_time}</strong></div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # Notícia ativa
    if "selected_news_id" not in st.session_state:
        st.session_state.selected_news_id = None
    if not st.session_state.selected_news_id and filtered_news:
        st.session_state.selected_news_id = filtered_news[0]["id"]

    # 3. Feed de Notícias fixo
    if filtered_news:
        cards = []
        for idx, item in enumerate(filtered_news[:45]):
            is_featured = item.get("id") == st.session_state.selected_news_id or idx == 0
            impact_level, impact_label, impact_reasons = market_impact(item)
            title = esc(news_title(item))
            summary_raw = news_summary(item)
            summary = esc(summary_raw)
            published = esc(item.get("published_str", "00:00"))
            source = esc(item.get("source", "Financial Juice"))
            link = safe_external_url(item.get("link"))
            icon_text = esc("FJ" if source == "Financial Juice" else source[:2].upper())
            tags_html = "".join(f'<span class="bb-news-tag">{esc(tag)}</span>' for tag in infer_tags(item))
            impact_badge = (
                f'<span class="bb-impact-badge {impact_level}">{esc(impact_label)}</span>'
                if impact_label
                else ""
            )
            reason_tags = "".join(f'<span class="bb-news-tag">{esc(reason)}</span>' for reason in impact_reasons)
            featured_class = " bb-featured" if is_featured else ""
            impact_class = f" bb-impact-{impact_level}" if impact_level in ["critical", "high", "medium"] else ""
            close_html = '<span class="bb-news-close">×</span>' if is_featured else ""
            summary_html = (
                f'<div class="bb-news-summary">{summary}</div>'
                if summary and summary != title
                else ""
            )

            cards.append(
                f'<div class="bb-news-card{featured_class}{impact_class}">'
                f'{close_html}'
                f'<div class="bb-news-rail"></div>'
                f'<div class="bb-news-icon">{icon_text}</div>'
                f'<div class="bb-news-content">'
                f'<div class="bb-news-title">{title}</div>'
                f'{summary_html}'
                f'<div class="bb-news-meta">'
                f'<span>{published}</span><span>{source}</span>{impact_badge}{reason_tags}{tags_html}'
                f'</div>'
                f'</div>'
                f'<a class="bb-news-link" href="{link}" target="_blank" rel="noopener noreferrer">↗</a>'
                f'</div>'
            )
        feed_header = (
            f'<div class="bb-feed-header">'
            f'<span>Feed de Noticias em Tempo Real</span>'
            f'<span class="bb-live-pill"><span class="bb-status-led"></span>LIVE • {esc(" + ".join(news_sources) or "Fontes")} • {len(filtered_news)} noticias</span>'
            f'</div>'
        )
        st.markdown(f'<div class="bb-news-feed">{feed_header}{"".join(cards)}</div>', unsafe_allow_html=True)
    else:
        st.info("Nenhuma manchete correspondente encontrada.")

    # 4. Status Bar inferior
    st.markdown(f"""
    <div class="bb-status-footer">
        <div>
            <span class="bb-status-led"></span>
            <span style="color: #00FFA3; font-weight: bold;">LIVE FEED</span>
            &nbsp;|&nbsp; Feed em ingles, sem traducao, atualiza a cada 30s
            &nbsp;|&nbsp; Origem: {esc(" + ".join(news_sources) or "Fontes")}
        </div>
        <div>
            Ultimo Refresh: {feed_loaded_at}
            &nbsp;|&nbsp; Fonte critica: Financial Juice RSS + cache Supabase
        </div>
    </div>
    """, unsafe_allow_html=True)

def render_terminal_global_correlation_panel():
    interval = "1"
    st.markdown(
        """
        <div style="border:1px solid #243244; border-radius:8px 8px 0 0; padding:10px 12px; background:#0b1220;">
            <div style="font-size:0.72rem; color:#94A3B8; font-weight:900; text-transform:uppercase;">Correlação Macro</div>
            <div style="font-size:0.82rem; color:#E5E7EB; font-weight:800; margin-top:2px;">USA500 | GOLD | UKOIL | US10Y | US30Y | DXY</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    tv_html = f"""
    <div class="tradingview-widget-container" style="height: 1085px; width: 100%; background:#0b0f17;">
      <div id="tg_side_correlation_tv" style="height: 100%; width: 100%;"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
      <script type="text/javascript">
      new TradingView.widget(
      {{
        "autosize": true,
        "symbol": "USA500",
        "interval": "{interval}",
        "timezone": "America/Sao_Paulo",
        "theme": "dark",
        "style": "2",
        "locale": "br",
        "toolbar_bg": "#0b0f17",
        "enable_publishing": false,
        "hide_top_toolbar": false,
        "hide_side_toolbar": true,
        "allow_symbol_change": true,
        "save_image": false,
        "details": false,
        "hotlist": false,
        "calendar": false,
        "hide_volume": true,
        "container_id": "tg_side_correlation_tv",
        "overrides": {{
            "mainSeriesProperties.lineStyle.color": "#B026FF",
            "mainSeriesProperties.lineStyle.linewidth": 3,
            "paneProperties.background": "#0b0f17",
            "paneProperties.vertGridProperties.color": "#1f2937",
            "paneProperties.horzGridProperties.color": "#1f2937"
        }},
        "studies": [
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "TVC:GOLD" }}, "plots": {{ "Plot": {{ "color": "#FFD166" }} }} }},
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "TVC:UKOIL" }}, "plots": {{ "Plot": {{ "color": "#2F80ED" }} }} }},
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "OTCB:US10Y" }}, "plots": {{ "Plot": {{ "color": "#FF9800" }} }} }},
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "OTCB:US30Y" }}, "plots": {{ "Plot": {{ "color": "#00BFFF" }} }} }},
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "CAPITALCOM:DXY" }}, "plots": {{ "Plot": {{ "color": "#F8FAFC" }} }} }}
        ]
      }}
      );
      </script>
    </div>
    """
    components.html(tv_html, height=1100)


def render_terminal_global_layout_css():
    st.markdown(
        """
        <style>
        div[data-testid="column"]:has(#tg-side-correlation-anchor) {
            position: sticky;
            top: 0.75rem;
            align-self: flex-start;
        }
        @media (max-width: 1100px) {
            div[data-testid="column"]:has(#tg-side-correlation-anchor) {
                display: none;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_terminal_global_macro_class_comparatives():
    """Render four compact macro class comparison charts below Terminal Global charts."""
    def render_macro_class_chart(title, description, container_id, main_symbol, main_color, overlays):
        chart_height = 860
        interval = "5"
        studies = ",\n          ".join(
            [
                '{{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "{}" }}, "plots": {{ "Plot": {{ "color": "{}" }} }} }}'.format(symbol, color)
                for symbol, color, _label in overlays
            ]
        )
        legend_items = "".join(
            [
                f"<span style='display:inline-flex; align-items:center; gap:5px; margin-right:9px; margin-bottom:5px; color:#CBD5E1; font-size:0.68rem; font-weight:800;'><i style='width:8px; height:8px; border-radius:50%; background:{color}; display:inline-block;'></i>{label}</span>"
                for _symbol, color, label in [(main_symbol, main_color, "Base")] + overlays
            ]
        )
        st.markdown(f"#### {title}")
        st.markdown(
            f"<div style='color:#94A3B8; font-size:0.76rem; line-height:1.25; min-height:34px; margin-bottom:7px;'>{description}</div>"
            f"<div style='margin-bottom:8px;'>{legend_items}</div>",
            unsafe_allow_html=True,
        )
        tv_html_class = f"""
        <div class="tradingview-widget-container" style="height: {chart_height}px; width: 100%; background:#0b0f17;">
          <div id="{container_id}" style="height: 100%; width: 100%;"></div>
          <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
          <script type="text/javascript">
          new TradingView.widget(
          {{
            "autosize": true,
            "symbol": "{main_symbol}",
            "interval": "{interval}",
            "timezone": "America/Sao_Paulo",
            "theme": "dark",
            "style": "2",
            "locale": "br",
            "toolbar_bg": "#0b0f17",
            "enable_publishing": false,
            "hide_top_toolbar": false,
            "hide_side_toolbar": true,
            "allow_symbol_change": true,
            "save_image": true,
            "details": false,
            "hotlist": false,
            "calendar": false,
            "hide_volume": true,
            "container_id": "{container_id}",
            "overrides": {{
                "mainSeriesProperties.lineStyle.color": "{main_color}",
                "mainSeriesProperties.lineStyle.linewidth": 3,
                "scalesProperties.scaleMode": 2,
                "paneProperties.background": "#0b0f17",
                "paneProperties.vertGridProperties.color": "#1f2937",
                "paneProperties.horzGridProperties.color": "#1f2937"
            }},
            "studies": [
              {studies}
            ]
          }}
          );
          </script>
        </div>
        """
        components.html(tv_html_class, height=chart_height + 20)

    st.markdown("---")
    st.markdown("### Comparativo por Classe Macro")
    st.markdown(
        "<p style='color:#94A3B8; font-size:0.88rem; margin-top:-4px;'>Quatro leituras lado a lado em 5 minutos: commodities, moedas, equity e bonds.</p>",
        unsafe_allow_html=True,
    )
    commodity_col, fx_col, equity_col, bonds_col = st.columns(4, gap="medium")
    with commodity_col:
        render_macro_class_chart(
            "Commodities",
            "Energia, metais industriais e metais preciosos pela fonte ActivTrades.",
            "tg_macro_commodities",
            "ACTIVTRADES:BRENT",
            "#22C55E",
            [
                ("ACTIVTRADES:LCRUDE", "#F97316", "Petroleo WTI"),
                ("ACTIVTRADES:NGAS", "#60A5FA", "Gas natural"),
                ("ACTIVTRADES:COPPERN2026", "#D97706", "Cobre"),
                ("ACTIVTRADES:GOLD", "#FACC15", "Ouro"),
                ("ACTIVTRADES:SILVER", "#E2E8F0", "Prata"),
            ],
        )
    with fx_col:
        render_macro_class_chart(
            "FX",
            "Moedas de commodities, safe havens/majors, emergentes e carry.",
            "tg_macro_fx",
            "CAPITALCOM:DXY",
            "#F8FAFC",
            [
                ("OANDA:AUDUSD", "#22C55E", "AUDUSD"),
                ("OANDA:USDCAD", "#F97316", "USDCAD"),
                ("OANDA:GBPUSD", "#A855F7", "GBPUSD"),
                ("OANDA:EURUSD", "#38BDF8", "EURUSD"),
                ("FX_IDC:USDBRL", "#FACC15", "USDBRL"),
                ("OANDA:USDJPY", "#EF4444", "USDJPY"),
            ],
        )
    with equity_col:
        render_macro_class_chart(
            "Equity",
            "Indices globais: volatilidade, EUA, Brasil, Europa e Japao.",
            "tg_macro_equity",
            "ACTIVTRADES:USA500",
            "#A855F7",
            [
                ("ACTIVTRADES:VXX.US", "#EF4444", "VIX/VXX"),
                ("ACTIVTRADES:JP225", "#38BDF8", "Nikkei"),
                ("ACTIVTRADES:BRA50", "#22C55E", "IBOV/BRA50"),
                ("ACTIVTRADES:EURO50", "#F97316", "EuroStoxx"),
                ("ACTIVTRADES:USARUS", "#FACC15", "RTY/Russell"),
                ("ACTIVTRADES:USATEC", "#60A5FA", "Nasdaq"),
            ],
        )
    with bonds_col:
        render_macro_class_chart(
            "Bonds",
            "Curvas globais: EUA, Brasil e Alemanha.",
            "tg_macro_bonds",
            "OTCB:US10Y",
            "#FF9800",
            [
                ("OTCB:US02Y", "#FACC15", "2Y USA"),
                ("OTCB:US30Y", "#00BFFF", "30Y USA"),
                ("BMFBOVESPA:DI1F2029", "#22C55E", "DI1F2029 BR"),
                ("BMFBOVESPA:DI1F2032", "#14B8A6", "DI1F2032 BR"),
                ("BMFBOVESPA:DI1F2035", "#84CC16", "DI1F2035 BR"),
                ("OANDA:DE10YBEUR", "#38BDF8", "10Y Alemanha"),
            ],
        )


def pagina_terminal_global():
    """Página de Terminal Global."""
    render_terminal_global_layout_css()
    st.markdown("<div id='tg-top'></div>", unsafe_allow_html=True)
    painel_topo_global()
    render_source_health_panel()
    render_yield_curve_regime_panel()
    secao_calendario_global_fragment()
    
    body_col, corr_col = st.columns([0.74, 0.26], gap="medium")

    with body_col:
        st.markdown("---")
        st.markdown("<div id='tg-graficos'></div>", unsafe_allow_html=True)
    
        global_chart_assets = {
            "USA500": {"tv": "USA500", "yf": "^GSPC"},
            "UKOIL": {"tv": "UKOIL", "yf": "BZ=F"},
            "BRA50": {"tv": "BRA50", "yf": "^BVSP"},
            "BTCUSDT": {"tv": "BTCUSDT", "yf": "BTC-USD"},
            "S&P 500": {"tv": "USA500", "yf": "^GSPC"},
            "NASDAQ": {"tv": "ACTIVTRADES:USATEC", "yf": "^IXIC"},
            "USATEC": {"tv": "ACTIVTRADES:USATEC", "yf": "^IXIC"},
            "BRENT OIL": {"tv": "TVC:UKOIL", "yf": "BZ=F"},
            "WTI OIL": {"tv": "TVC:USOIL", "yf": "CL=F"},
            "GOLD": {"tv": "TVC:GOLD", "yf": "GC=F"},
            "BITCOIN": {"tv": "BINANCE:BTCUSDT", "yf": "BTC-USD"},
            "ETHUSDT": {"tv": "BINANCE:ETHUSDT", "yf": "ETH-USD"},
            "DXY (Dólar Index)": {"tv": "CAPITALCOM:DXY", "yf": "DX-Y.NYB"},
            "US10Y OTCB": {"tv": "OTCB:US10Y", "yf": "^TNX"},
            "US30Y OTCB": {"tv": "OTCB:US30Y", "yf": "^TYX"},
            "EWZ (Brazil ETF)": {"tv": "AMEX:EWZ", "yf": "EWZ"},
            "EEM (Emerging Markets)": {"tv": "AMEX:EEM", "yf": "EEM"},
        }
    
        chart_col_1, chart_col_2 = st.columns(2)
    
        with chart_col_1:
            st.markdown("#### 📈 Gráfico Global")
            col_sel, col_int = st.columns([2, 1])
            with col_sel:
                sym = st.selectbox("Ativo", list(global_chart_assets.keys()), index=0, key="global_sym")
            with col_int:
                interval = st.selectbox("Intervalo", ["5"], index=0, key="global_int")
            
            tv_html = f"""
            <div class="tradingview-widget-container" style="height: 480px; width: 100%;">
              <div id="tv_global" style="height: 100%; width: 100%;"></div>
              <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
              <script type="text/javascript">
              new TradingView.widget({{
                "autosize": true,
                "symbol": "{global_chart_assets[sym]['tv']}",
                "interval": "{interval}",
                "timezone": "America/Sao_Paulo",
                "theme": "dark",
                "style": "3",
                "locale": "br",
                "toolbar_bg": "#f1f3f6",
                "enable_publishing": false,
                "hide_top_toolbar": false,
                "hide_side_toolbar": false,
                "save_image": true,
                "hide_volume": true,
                "container_id": "tv_global",
                "studies": [
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Session" }}, "plots": {{ "VWAP": {{ "color": "#FFD166" }} }} }},
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Week" }}, "plots": {{ "VWAP": {{ "color": "#06D6A0" }} }} }},
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Month" }}, "plots": {{ "VWAP": {{ "color": "#118AB2" }} }} }}
                ]
              }});
              </script>
            </div>
            """
            components.html(tv_html, height=500)

        with chart_col_2:
            st.markdown("#### Grafico Global 2")
            col_sel_2, col_int_2 = st.columns([2, 1])
            with col_sel_2:
                sym_2 = st.selectbox("Ativo", list(global_chart_assets.keys()), index=1, key="global_sym_2")
            with col_int_2:
                interval_2 = st.selectbox("Intervalo", ["5"], index=0, key="global_int_2")

            tv_html_2 = f"""
            <div class="tradingview-widget-container" style="height: 480px; width: 100%;">
              <div id="tv_global_2" style="height: 100%; width: 100%;"></div>
              <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
              <script type="text/javascript">
              new TradingView.widget({{
                "autosize": true,
                "symbol": "{global_chart_assets[sym_2]['tv']}",
                "interval": "{interval_2}",
                "timezone": "America/Sao_Paulo",
                "theme": "dark",
                "style": "3",
                "locale": "br",
                "toolbar_bg": "#f1f3f6",
                "enable_publishing": false,
                "hide_top_toolbar": false,
                "hide_side_toolbar": false,
                "save_image": true,
                "hide_volume": true,
                "container_id": "tv_global_2",
                "studies": [
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Session" }}, "plots": {{ "VWAP": {{ "color": "#FFD166" }} }} }},
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Week" }}, "plots": {{ "VWAP": {{ "color": "#06D6A0" }} }} }},
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Month" }}, "plots": {{ "VWAP": {{ "color": "#118AB2" }} }} }}
                ]
              }});
              </script>
            </div>
            """
            components.html(tv_html_2, height=500)

        chart_col_3, chart_col_4 = st.columns(2)

        with chart_col_3:
            st.markdown("#### Grafico Global 3")
            col_sel_3, col_int_3 = st.columns([2, 1])
            with col_sel_3:
                sym_3 = st.selectbox("Ativo", list(global_chart_assets.keys()), index=2, key="global_sym_3")
            with col_int_3:
                interval_3 = st.selectbox("Intervalo", ["5"], index=0, key="global_int_3")

            tv_html_3 = f"""
            <div class="tradingview-widget-container" style="height: 480px; width: 100%;">
              <div id="tv_global_3" style="height: 100%; width: 100%;"></div>
              <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
              <script type="text/javascript">
              new TradingView.widget({{
                "autosize": true,
                "symbol": "{global_chart_assets[sym_3]['tv']}",
                "interval": "{interval_3}",
                "timezone": "America/Sao_Paulo",
                "theme": "dark",
                "style": "3",
                "locale": "br",
                "toolbar_bg": "#f1f3f6",
                "enable_publishing": false,
                "hide_top_toolbar": false,
                "hide_side_toolbar": false,
                "save_image": true,
                "hide_volume": true,
                "container_id": "tv_global_3",
                "studies": [
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Session" }}, "plots": {{ "VWAP": {{ "color": "#FFD166" }} }} }},
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Week" }}, "plots": {{ "VWAP": {{ "color": "#06D6A0" }} }} }},
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Month" }}, "plots": {{ "VWAP": {{ "color": "#118AB2" }} }} }}
                ]
              }});
              </script>
            </div>
            """
            components.html(tv_html_3, height=500)

        with chart_col_4:
            st.markdown("#### Grafico Global 4")
            col_sel_4, col_int_4 = st.columns([2, 1])
            with col_sel_4:
                sym_4 = st.selectbox("Ativo", list(global_chart_assets.keys()), index=3, key="global_sym_4")
            with col_int_4:
                interval_4 = st.selectbox("Intervalo", ["5"], index=0, key="global_int_4")

            tv_html_4 = f"""
            <div class="tradingview-widget-container" style="height: 480px; width: 100%;">
              <div id="tv_global_4" style="height: 100%; width: 100%;"></div>
              <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
              <script type="text/javascript">
              new TradingView.widget({{
                "autosize": true,
                "symbol": "{global_chart_assets[sym_4]['tv']}",
                "interval": "{interval_4}",
                "timezone": "America/Sao_Paulo",
                "theme": "dark",
                "style": "3",
                "locale": "br",
                "toolbar_bg": "#f1f3f6",
                "enable_publishing": false,
                "hide_top_toolbar": false,
                "hide_side_toolbar": false,
                "save_image": true,
                "hide_volume": true,
                "container_id": "tv_global_4",
                "studies": [
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Session" }}, "plots": {{ "VWAP": {{ "color": "#FFD166" }} }} }},
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Week" }}, "plots": {{ "VWAP": {{ "color": "#06D6A0" }} }} }},
                  {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Month" }}, "plots": {{ "VWAP": {{ "color": "#118AB2" }} }} }}
                ]
              }});
              </script>
            </div>
            """
            components.html(tv_html_4, height=500)

        render_terminal_global_macro_class_comparatives()

        with st.container():
            st.markdown("#### 🤖 Analista Técnico IA")
            st.info(f"Análise baseada no histórico diário (OHLC) de {sym}.")
            if st.button("Gerar Análise Técnica (IA)", use_container_width=True):
                with st.spinner("Lendo o gráfico via IA..."):
                    import sys
                    exec_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'execution'))
                    if exec_path not in sys.path:
                        sys.path.append(exec_path)
                    from tech_analyst import get_technical_analysis
                
                    yf_ticker = global_chart_assets[sym]['yf']
                    insight = get_technical_analysis(sym, yf_ticker)
                
                    st.session_state[f'tech_insight_{sym}'] = insight
                
            if f'tech_insight_{sym}' in st.session_state:
                insight_text = st.session_state[f'tech_insight_{sym}']
                if "Erro" in insight_text or "429" in insight_text:
                    border_color = "#FF4B4B"
                    if "quota" in insight_text.lower() or "429" in insight_text:
                        insight_text = "⚠️ Limite de requisições da IA (Quota Exceeded) atingido. Tente novamente em alguns minutos."
                else:
                    border_color = "#00FFA3"
                
                st.markdown(f"""
                    <div style="background: #111; border: 1px solid #333; padding: 15px; border-radius: 8px; border-left: 5px solid {border_color}; font-size: 0.85rem; color: #E0E0E0; max-height: 380px; overflow-y: auto;">
                        {insight_text.replace(chr(10), '<br>')}
                    </div>
                """, unsafe_allow_html=True)
            
    with corr_col:
        st.markdown("---")
        st.markdown("<div id='tg-side-correlation-anchor'></div>", unsafe_allow_html=True)
        render_terminal_global_correlation_panel()

@st.fragment(run_every=30)
def sidebar_mercados():
    global_data = get_global_markets_data()
    if not global_data: 
        st.info("Carregando mercados...")
        return
    
    # Suporte tanto para o formato antigo quanto para o novo
    if "categories" in global_data:
        categories = global_data["categories"]
        metadata = global_data.get("metadata", {})
        last_upd = metadata.get("last_updated", "---")
    else:
        categories = global_data
        last_upd = "---"

    def remember_sidebar_quotes(category_map):
        now_ts = time.time()
        history = st.session_state.setdefault("sidebar_quote_history", {})
        cutoff = now_ts - (30 * 60)
        for assets in category_map.values():
            if not isinstance(assets, list):
                continue
            for item in assets:
                if not isinstance(item, dict):
                    continue
                name = str(item.get("name", "")).strip()
                if not name:
                    continue
                try:
                    price = float(item.get("price", 0))
                except (TypeError, ValueError):
                    continue
                if price <= 0:
                    continue
                points = [(ts, px) for ts, px in history.get(name, []) if ts >= cutoff]
                if not points or abs(points[-1][0] - now_ts) >= 20:
                    points.append((now_ts, price))
                history[name] = points[-80:]

    def quote_5m_momentum(name, current_price):
        points = st.session_state.get("sidebar_quote_history", {}).get(str(name), [])
        if not points:
            return ("5m ...", "#64748B", "")
        target_ts = time.time() - (5 * 60)
        older_points = [(ts, px) for ts, px in points if ts <= target_ts]
        if not older_points:
            return ("5m ...", "#64748B", "")
        ref_ts, ref_price = min(older_points, key=lambda point: abs(point[0] - target_ts))
        if ref_price <= 0:
            return ("5m ...", "#64748B", "")
        pct = ((current_price - ref_price) / ref_price) * 100
        if abs(pct) < 0.01:
            return ("5m 0.00%", "#94A3B8", "")
        arrow = "&#9650;" if pct > 0 else "&#9660;"
        color = "#00FFA3" if pct > 0 else "#FF4B4B"
        accel_label = ""
        if abs(pct) >= 0.35:
            accel_label = "FORTE"
        elif abs(pct) >= 0.12:
            accel_label = "ACELERA"
        return (f"{arrow} 5m {pct:+.2f}%", color, accel_label)

    remember_sidebar_quotes(categories)

    st.markdown(f"<div style='text-align:right; font-size:0.65rem; color:#666; margin-bottom:10px;'>ATUALIZADO ÀS: {last_upd}</div>", unsafe_allow_html=True)

    for cat_name, assets in categories.items():
        st.markdown(f"<div style='font-size:0.75rem; font-weight:bold; color:#FF9800; margin-bottom:5px;'>{cat_name}</div>", unsafe_allow_html=True)
        for item in assets:
            if not isinstance(item, dict): continue
            change_val = item.get('change', 0)
            if not isinstance(change_val, (int, float)):
                try: change_val = float(change_val)
                except (ValueError, TypeError): change_val = 0.0

            color = "#00FFA3" if change_val >= 0 else "#FF4B4B"
            
            price_val = item.get('price', 0)
            if not isinstance(price_val, (int, float)):
                try: price_val = float(price_val)
                except (ValueError, TypeError): price_val = 0.0

            # Formatação de preço: 4 casas se for pequeno (moedas), 2 se for grande
            price_fmt = f"{price_val:.4f}" if price_val < 10 else f"{price_val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            mom_5m, mom_5m_color, accel_label = quote_5m_momentum(item.get('name', '---'), price_val)
            accel_badge = (
                f"<span style='display:inline-block; margin-top:2px; padding:1px 5px; border-radius:999px; "
                f"background:{mom_5m_color}22; border:1px solid {mom_5m_color}88; color:{mom_5m_color}; "
                f"font-size:0.56rem; font-weight:900; letter-spacing:0.02em;'>{accel_label}</span>"
                if accel_label else ""
            )
            
            item_name = html.escape(str(item.get('name', '---')))
            quote_html = (
                "<div style='display:flex; justify-content:space-between; border-bottom:1px solid #1a1a1a; "
                "padding:4px 0; align-items:center;'>"
                f"<span style='font-size:0.75rem; color:#AAA; max-width:60%;'>{item_name}</span>"
                "<div style='text-align:right; min-width:78px;'>"
                f"<div style='font-size:0.96rem; font-weight:900; line-height:1.08;'>{price_fmt}</div>"
                f"<div style='color:{color}; font-weight:900; font-size:0.76rem; line-height:1.12;'>{change_val:+.2f}%</div>"
                f"<div style='color:{mom_5m_color}; font-weight:800; font-size:0.64rem; line-height:1.08;'>{mom_5m}</div>"
                f"{accel_badge}"
                "</div></div>"
            )
            st.markdown(quote_html, unsafe_allow_html=True)
        st.markdown("<div style='margin-bottom:15px;'></div>", unsafe_allow_html=True)


@st.fragment(run_every=3600)
def sidebar_calendario():
    calendar_data = get_calendar_data()
    if not calendar_data: 
        st.info("Carregando calendário...")
        return
    
    # 1. Seletor de Dia (Filtro Semanal)
    try:
        dates = sorted(list(set([e.get('date', '') for e in calendar_data if e.get('date')])))
        if not dates:
            st.warning("Nenhuma data disponível no calendário.")
            return
            
        today_str = datetime.now().strftime("%Y-%m-%d")
        default_idx = dates.index(today_str) if today_str in dates else 0
        selected_date = st.selectbox("📅 Selecione o Dia", dates, index=default_idx)
    except Exception as e:
        st.error(f"Erro ao processar datas do calendário: {e}")
        return

    
    st.markdown("---")
    
    # 2. Filtragem e Exibição
    filtered_events = [e for e in calendar_data if e['date'] == selected_date]
    
    if not filtered_events:
        st.write("Nenhum evento importante para este dia.")
        return

    for event in filtered_events:
        impact_color = "#FF4B4B" if event['impact'] == "HIGH" else ("#FF9800" if event['impact'] == "MEDIUM" else "#888")
        
        actual = event.get('actual', '---')
        forecast = event.get('forecast', '---')
        
        # Design do Evento
        st.markdown(f"""
            <div style='border-bottom:1px solid #222; padding:10px 0;'>
                <div style='display:flex; justify-content:space-between; align-items:center;'>
                    <span style='font-size:0.7rem; color:#888;'>{event['time']} | {event['currency']}</span>
                    <span style='font-size:0.65rem; background:#1a1a1a; padding:2px 6px; border-radius:10px; color:{impact_color}; border:1px solid {impact_color}44;'>{event['impact']} | {event.get('bull_count', 1)} touro(s)</span>
                </div>
                <div style='font-size:0.85rem; font-weight:bold; margin:4px 0;'>{event['icon']} {event['event']}</div>
                <div style='display:flex; gap:15px; margin-top:5px;'>
                    <div style='font-size:0.7rem;'>
                        <span style='color:#666;'>PREV:</span> <b style='color:#DDD;'>{forecast}</b>
                    </div>
                    <div style='font-size:0.7rem;'>
                        <span style='color:#666;'>ATUAL:</span> <b style='color:#FFF;'>{actual}</b>
                    </div>
                </div>
            </div>
        """, unsafe_allow_html=True)

@st.fragment(run_every=1800)
def secao_calendario_global_fragment():
    """Calendario economico compacto dentro do Terminal Global."""
    calendar_data = get_calendar_data() or []

    now_br = datetime.now(ZoneInfo("America/Sao_Paulo"))
    today_str = now_br.strftime("%Y-%m-%d")
    selected_currencies = ["USD", "BRL", "EUR", "GBP", "JPY", "CNY", "CAD", "AUD", "NZD", "CHF"]
    impact_rank = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "HOLIDAY": 3}

    events = [
        event for event in calendar_data
        if event.get("date") == today_str and event.get("currency") in selected_currencies
    ]
    events.sort(key=lambda item: (item.get("time", "99:99"), impact_rank.get(item.get("impact", ""), 9)))

    st.markdown("---")
    st.markdown("<div id='tg-calendario'></div>", unsafe_allow_html=True)
    st.markdown("#### Calendario Economico")
    source_label = next((event.get("source") for event in events if event.get("source")), "Supabase")
    st.caption(f"Eventos de hoje ({today_str}) | Horario de Brasilia | Foco IA: proximos eventos EUA/USD e Brasil/BRL | Fonte: {source_label}")

    def event_datetime(event):
        try:
            return datetime.strptime(
                f"{event.get('date')} {event.get('time')}",
                "%Y-%m-%d %H:%M",
            ).replace(tzinfo=ZoneInfo("America/Sao_Paulo"))
        except Exception:
            return None

    upcoming_events = []
    next_event = None
    for event in events:
        event_dt = event_datetime(event)
        if event_dt and event_dt >= now_br:
            upcoming_events.append(event)

    if upcoming_events:
        next_event = upcoming_events[0]

    investing_events = [event for event in events if event.get("source") == "Investing.com"]
    analysis_pool = investing_events if investing_events else events
    macro_global_data = get_global_markets_data()

    def has_actual(event):
        return str(event.get("actual", "---")).strip() not in ["", "---", "-"]

    def has_projection(event):
        return (
            str(event.get("actual", "---")).strip() in ["", "---", "-"]
            and str(event.get("forecast", "---")).strip() not in ["", "---", "-"]
            and str(event.get("previous", "---")).strip() not in ["", "---", "-"]
        )

    def is_us_or_brl_event(event):
        currency = str(event.get("currency", "")).upper()
        text = f"{event.get('event', '')} {event.get('country', '')}".lower()
        br_terms = ["brasil", "brazil", "brl", "bcb", "copom", "selic", "ipca", "igp", "real", "fiscal"]
        us_terms = ["fed", "fomc", "powell", "treasury", "payroll", "jobless", "jolts", "pce", "cpi", "ppi", "ism", "pmi"]
        return currency in {"USD", "BRL"} or any(term in text for term in br_terms + us_terms)

    def macro_focus_rank(event):
        currency = str(event.get("currency", "")).upper()
        text = f"{event.get('event', '')} {event.get('country', '')}".lower()
        if currency == "USD":
            region_rank = 0
        elif currency == "BRL" or any(term in text for term in ["brasil", "brazil", "bcb", "copom", "selic", "ipca", "real"]):
            region_rank = 1
        elif is_us_or_brl_event(event):
            region_rank = 2
        else:
            region_rank = 5
        event_dt = event_datetime(event)
        minutes_until = int((event_dt - now_br).total_seconds() // 60) if event_dt else 9999
        bull_rank = -int(event.get("bull_count", 0) or 0)
        return (region_rank, impact_rank.get(event.get("impact", ""), 9), bull_rank, max(minutes_until, 0), event.get("time", "99:99"))

    released_investing_events = [event for event in analysis_pool if has_actual(event)]
    upcoming_analyzable_events = [
        event for event in upcoming_events
        if event in analysis_pool and (has_projection(event) or has_actual(event))
    ]
    focused_upcoming_events = sorted([event for event in upcoming_events if is_us_or_brl_event(event)], key=macro_focus_rank)
    focused_upcoming_analyzable_events = sorted([event for event in upcoming_analyzable_events if is_us_or_brl_event(event)], key=macro_focus_rank)
    focused_released_events = [event for event in released_investing_events if is_us_or_brl_event(event)]
    display_upcoming_events = focused_upcoming_events if focused_upcoming_events else upcoming_events
    analysis_event = (
        focused_upcoming_analyzable_events[0]
        if focused_upcoming_analyzable_events
        else (
            upcoming_analyzable_events[0]
            if upcoming_analyzable_events
            else (focused_released_events[-1] if focused_released_events else (released_investing_events[-1] if released_investing_events else next_event))
        )
    )

    analysis_col, widget_col = st.columns([1.05, 1], gap="large")

    with analysis_col:
            if not events:
                st.info("Calendario economico temporariamente sem eventos carregados. O widget Investing continua disponivel ao lado.")
                st.markdown(
                    """
                    <div style="border:1px solid #334155; border-left:5px solid #64748B; border-radius:8px; padding:16px 18px; margin:10px 0 16px 0; background:#0b1220;">
                        <div style="font-size:0.72rem; color:#94A3B8; font-weight:800; text-transform:uppercase;">IA Macro TTS</div>
                        <div style="font-size:1rem; color:#F8FAFC; font-weight:900; margin-top:6px;">Aguardando dados do calendario</div>
                        <div style="font-size:0.88rem; color:#CBD5E1; margin-top:8px; line-height:1.5;">Assim que Supabase, Investing ou cache local retornar eventos do dia, a leitura macro volta a analisar surpresa, juros, inflacao, DXY, petroleo e indices americanos.</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            next_event_key = None
            if next_event or analysis_event:
                if display_upcoming_events:
                    display_next_event = display_upcoming_events[0]
                    next_event_key = (
                        display_next_event.get("date"),
                        display_next_event.get("time"),
                        display_next_event.get("currency"),
                        display_next_event.get("event"),
                    )
                    top_impact = display_next_event.get("impact", "")
                    top_impact_color = "#FF4B4B" if top_impact == "HIGH" else ("#FF9800" if top_impact == "MEDIUM" else "#888")
                    upcoming_rows = []
                    for idx, event in enumerate(display_upcoming_events[:3], start=1):
                        impact = event.get("impact", "")
                        impact_color = "#FF4B4B" if impact == "HIGH" else ("#FF9800" if impact == "MEDIUM" else "#888")
                        border_top = "border-top:1px solid #242b36;" if idx > 1 else ""
                        focus_badge = "FOCO USD/BRL" if is_us_or_brl_event(event) else "GLOBAL"
                        upcoming_rows.append(
                            f"<div style='{border_top} padding:{'10px' if idx > 1 else '0'} 0 0 0; margin-top:{'10px' if idx > 1 else '0'};'>"
                            f"<div style='display:flex; justify-content:space-between; gap:18px; align-items:center; flex-wrap:wrap;'>"
                            f"<div style='font-size:0.72rem; color:#888; font-weight:800; text-transform:uppercase;'>Proximo evento #{idx} | {focus_badge}</div>"
                            f"<div style='color:{impact_color}; font-weight:900; font-size:0.84rem;'>{impact or '---'} | {event.get('bull_count', 1)} touro(s)</div>"
                            f"</div>"
                            f"<div style='font-size:1rem; font-weight:800; color:#FFF; margin-top:4px;'>{event.get('time', '---')} | {event.get('currency', '---')} | {event.get('event', '---')}</div>"
                            f"<div style='font-size:0.78rem; color:#AAA; margin-top:6px;'>Atual: <b style='color:#FFF;'>{event.get('actual', '---') or '---'}</b> &nbsp;|&nbsp; Projecao: <b>{event.get('forecast', '---') or '---'}</b> &nbsp;|&nbsp; Anterior: <b>{event.get('previous', '---') or '---'}</b></div>"
                            f"</div>"
                        )
                    st.markdown(
                        f"<div style='border:1px solid {top_impact_color}; border-left:5px solid {top_impact_color}; border-radius:8px; padding:12px 14px; margin:10px 0 14px 0; background:#111;'>{''.join(upcoming_rows)}</div>",
                        unsafe_allow_html=True,
                    )

                st.markdown("<div id='tg-ia-macro'></div>", unsafe_allow_html=True)
                try:
                    from execution.macro_calendar_ai import interpret_event
                    if not analysis_event:
                        st.info("IA Macro TTS aguardando evento com Atual, Projecao ou Anterior para analisar.")
                        macro_ai = None
                    else:
                        macro_ai = interpret_event(analysis_event, macro_global_data)
                    if not macro_ai:
                        raise StopIteration
                    score = macro_ai.get("risk_score", 0)
                    effect_color = "#00FFA3" if score > 20 else ("#FF4B4B" if score < -20 else "#FF9800")
                    impacts = macro_ai.get("asset_impacts", {})
                    impacts_html = "".join(
                        f"<span style='display:inline-block; border:1px solid #334155; background:#111827; border-radius:6px; padding:5px 9px; margin:4px 6px 0 0; color:#CBD5E1; font-size:0.78rem;'>{asset}: <b style='color:#FFF;'>{bias}</b></span>"
                        for asset, bias in impacts.items()
                    )
                    if macro_ai.get("status") == "Projecao analisada":
                        panel_label = "Projecao do proximo evento"
                    else:
                        panel_label = "Ultimo dado divulgado analisado" if analysis_event in released_investing_events else "Proximo evento aguardando divulgacao"
                    actual = analysis_event.get("actual", "---") or "---"
                    forecast = analysis_event.get("forecast", "---") or "---"
                    previous = analysis_event.get("previous", "---") or "---"
                    st.markdown(
                        f"""
                        <div style="border:1px solid #334155; border-left:5px solid {effect_color}; border-radius:8px; padding:18px 20px; margin:0 0 16px 0; background:#0b1220;">
                            <div style="display:flex; justify-content:space-between; gap:16px; align-items:flex-start; flex-wrap:wrap;">
                                <div>
                                    <div style="font-size:0.72rem; color:#94A3B8; font-weight:800; text-transform:uppercase;">IA Macro TTS - foco EUA/BRL + intermercados</div>
                                    <div style="font-size:0.74rem; color:#64748B; margin-top:6px; text-transform:uppercase;">{panel_label}</div>
                                    <div style="font-size:1.15rem; color:#FFF; font-weight:900; margin-top:4px;">{analysis_event.get('time', '---')} | {analysis_event.get('currency', '---')} | {analysis_event.get('event', '---')}</div>
                                </div>
                                <div style="text-align:right;">
                                    <div style="font-size:0.72rem; color:#94A3B8; font-weight:700;">Efeito da surpresa</div>
                                    <div style="font-size:1.35rem; color:{effect_color}; font-weight:900; line-height:1.15; margin-top:4px;">{macro_ai.get('risk_classification', 'Neutro')}</div>
                                </div>
                            </div>
                            <div style="display:grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap:10px; margin-top:16px;">
                                <div style="background:#111827; border:1px solid #1F2937; border-radius:8px; padding:10px;"><span style="color:#94A3B8; font-size:0.78rem;">Atual</span><br><b style="font-size:1rem;">{actual}</b></div>
                                <div style="background:#111827; border:1px solid #1F2937; border-radius:8px; padding:10px;"><span style="color:#94A3B8; font-size:0.78rem;">Projecao</span><br><b style="font-size:1rem;">{forecast}</b></div>
                                <div style="background:#111827; border:1px solid #1F2937; border-radius:8px; padding:10px;"><span style="color:#94A3B8; font-size:0.78rem;">Anterior</span><br><b style="font-size:1rem;">{previous}</b></div>
                                <div style="background:#111827; border:1px solid #1F2937; border-radius:8px; padding:10px;"><span style="color:#94A3B8; font-size:0.78rem;">Surpresa</span><br><b style="font-size:1rem;">{macro_ai.get('surprise_label', '---')}</b></div>
                            </div>
                            <div style="display:grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap:10px; margin-top:10px;">
                                <div><span style="color:#64748B;">Status</span><br><b>{macro_ai.get('status', '---')}</b></div>
                                <div><span style="color:#64748B;">Categoria</span><br><b>{macro_ai.get('category', '---')}</b></div>
                                <div><span style="color:#64748B;">Impacto Investing</span><br><b>{analysis_event.get('bull_count', 1)} touro(s)</b></div>
                            </div>
                            <div style="margin-top:14px; color:#E5E7EB; line-height:1.55; font-size:0.94rem;">{macro_ai.get('operational_summary', '')}</div>
                            <div style="margin-top:12px;">{impacts_html}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                except StopIteration:
                    pass
                except Exception as e:
                    st.warning(f"IA Macro TTS indisponivel: {e}")

            try:
                from execution.macro_calendar_ai import interpret_event

                interpreted_history = []
                pending_history = []
                focus_history_events = [
                    event for event in investing_events
                    if str(event.get("currency", "")).upper() in {"USD", "BRL"}
                ]
                for event in reversed(focus_history_events):
                    result = interpret_event(event, macro_global_data)
                    if result.get("status") == "Interpretado":
                        interpreted_history.append((event, result))
                    else:
                        pending_history.append((event, result))
                    if len(interpreted_history) >= 5:
                        break

                if not interpreted_history:
                    interpreted_history = pending_history[:5]

                if interpreted_history:
                    history_cards = []
                    for event, result in interpreted_history:
                        score = int(result.get("risk_score", 0))
                        effect_color = "#00FFA3" if score > 20 else ("#FF4B4B" if score < -20 else "#FF9800")
                        status_color = "#00FFA3" if result.get("status") == "Interpretado" else "#FF9800"
                        actual = event.get("actual", "---") or "---"
                        forecast = event.get("forecast", "---") or "---"
                        previous = event.get("previous", "---") or "---"
                        history_cards.append(
                            f"<div style='border-bottom:1px solid #1F2937; padding:13px 0;'>"
                            f"<div style='display:flex; justify-content:space-between; gap:14px; align-items:flex-start;'>"
                            f"<div style='min-width:0;'>"
                            f"<div style='font-size:0.82rem; color:#94A3B8;'>{event.get('time', '---')} | {event.get('currency', '---')} | <b style='color:#E5E7EB;'>{event.get('event', '---')}</b> <span style='color:{status_color}; font-weight:800;'>- {result.get('status', '---')}</span></div>"
                            f"<div style='font-size:0.78rem; color:#CBD5E1; margin-top:5px;'>Atual <b>{actual}</b> | Projecao <b>{forecast}</b> | Anterior <b>{previous}</b> | Surpresa <b>{result.get('surprise_label', '---')}</b></div>"
                            f"<div style='font-size:0.9rem; color:#F8FAFC; margin-top:8px; line-height:1.5;'>{result.get('operational_summary', '')}</div>"
                            f"</div>"
                            f"<div style='text-align:right; min-width:130px;'>"
                            f"<div style='font-size:0.72rem; color:#94A3B8;'>Efeito</div>"
                            f"<div style='color:{effect_color}; font-weight:900; font-size:0.9rem;'>{result.get('risk_classification', 'Neutro')}</div>"
                            f"</div>"
                            f"</div>"
                            f"</div>"
                        )
                    st.markdown("<div id='tg-historico'></div>", unsafe_allow_html=True)
                    st.markdown(
                        f"<div style='border:1px solid #334155; border-radius:8px; padding:16px; margin:0 0 16px 0; background:#0b1220;'>"
                        f"<div style='font-size:0.78rem; color:#94A3B8; font-weight:800; text-transform:uppercase; margin-bottom:8px;'>Historico IA Macro TTS - ultimos 5 eventos USD/BRL divulgados</div>"
                        f"{''.join(history_cards)}"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
            except Exception as e:
                st.warning(f"Historico IA Macro TTS indisponivel: {e}")


    with widget_col:
            investing_calendar_url = (
                "https://sslecal2.investing.com?"
                "ecoDayBackground=%230b0f17&"
                "defaultFont=%23000000&"
                "innerBorderColor=%23242b36&"
                "borderColor=%23242b36&"
                "ecoDayFontColor=%23ffffff&"
                "columns=exc_flags,exc_currency,exc_importance,exc_actual,exc_forecast,exc_previous&"
                "importance=1,2,3&"
                "features=datepicker,timezone,timeselector,filters&"
                "countries=25,6,37,72,22,17,35,43,11,12,4,5&"
                "calType=day&"
                "timeZone=12&"
                "lang=12"
            )
            st.markdown("##### Calendario Investing em tempo real")
            st.caption("Widget oficial do Investing.com com Atual, Projecao e Anterior. Ajuste o fuso no proprio widget se necessario.")
            components.iframe(investing_calendar_url, height=760, scrolling=True)

@st.fragment(run_every=300)
def secao_fluxo_estrangeiro_fragment():
    """Seção que exibe o Fluxo do Investidor Estrangeiro na B3."""
    fluxo_data = fetch_app_state("fluxo_estrangeiro_b3")
    if not fluxo_data or "records" not in fluxo_data:
        return

    st.markdown("---")
    st.markdown("#### 🌍 FLUXO ESTRANGEIRO NA B3")
    st.markdown(f"<span style='color: #666; font-size: 0.75rem; font-family: \"Roboto Mono\", monospace;'>ÚLTIMA ATUALIZAÇÃO: {fluxo_data.get('updated_at', '---')[:10]}</span>", unsafe_allow_html=True)
    
    records = fluxo_data["records"][:30] # Últimos 30 dias
    if not records:
        return
        
    df = pd.DataFrame(records)
    df = df.sort_values("date") # Ordem cronológica para o gráfico
    
    import plotly.graph_objects as go
    
    # Cores baseadas no valor (positivo = verde, negativo = vermelho)
    colors = ['#00FFA3' if val >= 0 else '#FF4B4B' for val in df['foreigners']]
    
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df['date'],
        y=df['foreigners'],
        marker_color=colors,
        text=df['foreigners'].apply(lambda x: f"R$ {x/1000:,.1f}M".replace(",", "X").replace(".", ",").replace("X", ".")),
        textposition='outside',
        textfont=dict(color='#E0E0E0', size=9)
    ))
    
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font_family='"Roboto Mono", monospace',
        font_color='#E0E0E0',
        margin=dict(l=10, r=10, t=30, b=10),
        height=350,
        xaxis=dict(showgrid=False, title=None, tickangle=-45),
        yaxis=dict(showgrid=True, gridcolor='#1a1a1a', zeroline=True, zerolinecolor='#444', title="R$ Milhares"),
        showlegend=False
    )
    
    st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})

@st.fragment(run_every=300)
def secao_boletim_focus_fragment():
    """Seção que exibe as expectativas do Boletim Focus no formato histórico."""
    focus_data = fetch_app_state("boletim_focus")
    if not focus_data or "years" not in focus_data:
        return
        
    st.markdown("---")
    st.markdown("#### 🇧🇷 BOLETIM FOCUS (BCB)")
    st.markdown(f"<span style='color: #666; font-size: 0.75rem; font-family: \"Roboto Mono\", monospace;'>DATA BASE: {focus_data.get('publish_date', '---')}</span>", unsafe_allow_html=True)
    
    years = sorted(list(focus_data["years"].keys()))
    if not years:
        return
        
    # CSS para a tabela do Focus
    st.markdown("""
    <style>
    .focus-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.75rem;
        font-family: "Roboto Mono", monospace;
    }
    .focus-table th {
        text-align: right;
        color: #888;
        padding: 4px;
        border-bottom: 1px solid #333;
        font-weight: normal;
    }
    .focus-table th:first-child {
        text-align: left;
    }
    .focus-table td {
        text-align: right;
        color: #FFF;
        padding: 4px;
        border-bottom: 1px solid #222;
    }
    .focus-table td:first-child {
        text-align: left;
        color: #AAA;
    }
    </style>
    """, unsafe_allow_html=True)
    
    cols = st.columns(len(years))
    
    for i, year in enumerate(years):
        with cols[i]:
            data_year = focus_data["years"][year]
            st.markdown(f"<div style='text-align:center; padding:5px; background-color:#1a1a1a; border-radius:5px; margin-bottom:5px;'><b style='color:#00FFA3;'>{year}</b></div>", unsafe_allow_html=True)
            
            html = "<table class='focus-table'><tr><th>Indicador</th><th>Há 4 sem</th><th>Há 1 sem</th><th>Hoje</th><th></th></tr>"
            
            indicadores = [
                ("IPCA", data_year.get("IPCA", {})),
                ("PIB", data_year.get("PIB", {})),
                ("Câmbio", data_year.get("Cambio", {})),
                ("Selic", data_year.get("Selic", {}))
            ]
            
            for nome, vals in indicadores:
                if not vals or not isinstance(vals, dict):
                    html += f"<tr><td>{nome}</td><td>---</td><td>---</td><td>---</td><td></td></tr>"
                    continue
                    
                v4 = vals.get("4_sem")
                v1 = vals.get("1_sem")
                v0 = vals.get("hoje")
                
                # Formatters
                def f_val(x):
                    return f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if x is not None else "---"
                
                v4_str = f_val(v4)
                v1_str = f_val(v1)
                v0_str = f_val(v0)
                
                # Setas de tendência
                seta = ""
                if v0 is not None and v1 is not None:
                    if v0 > v1: seta = "<span style='color:#FF4B4B'>▲</span>" # Subiu (Vermelho se IPCA/Selic, mas mantendo simples)
                    elif v0 < v1: seta = "<span style='color:#00FFA3'>▼</span>" # Caiu
                    else: seta = "<span style='color:#888'>=</span>"
                
                html += f"<tr><td>{nome}</td><td>{v4_str}</td><td>{v1_str}</td><td><b>{v0_str}</b></td><td>{seta}</td></tr>"
                
            html += "</table>"
            st.markdown(html, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PÁGINAS DO DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

REGIME_JUROS_EXCEL_PATH = r"C:\Users\Mini PC\Documents\ANALISE JUROS\Curva_DI_RTD_Monitor_PrecoTempo.xlsx"
REGIME_JUROS_SNAPSHOT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "regime_juros_snapshot.json")


@st.cache_data(ttl=30, show_spinner=False)
def get_regime_juros_indice_atual():
    """Le apenas os campos principais da aba Indice Atual do monitor RTD de DI."""
    try:
        from openpyxl import load_workbook

        online_snapshot = fetch_app_state_fast("regime_juros")
        if isinstance(online_snapshot, dict):
            online_snapshot["source"] = online_snapshot.get("source") or "supabase"
            return online_snapshot

        if os.path.exists(REGIME_JUROS_EXCEL_PATH):
            wb = load_workbook(REGIME_JUROS_EXCEL_PATH, data_only=True, read_only=True)
            ws = wb["Indice Atual"]
            return {
                "taxa_sintetica": ws["E2"].value,
                "variacao_bps": ws["H2"].value,
                "regime_estrutural": ws["K2"].value,
                "updated_at": datetime.fromtimestamp(
                    os.path.getmtime(REGIME_JUROS_EXCEL_PATH),
                    ZoneInfo("America/Sao_Paulo"),
                ).strftime("%H:%M:%S"),
                "source": "excel_rtd_local",
            }

        if os.path.exists(REGIME_JUROS_SNAPSHOT_PATH):
            with open(REGIME_JUROS_SNAPSHOT_PATH, "r", encoding="utf-8") as fp:
                snapshot = json.load(fp)
            snapshot["source"] = snapshot.get("source") or "snapshot"
            return snapshot

        return {"error": "Arquivo Excel de juros e snapshot nao encontrados."}
    except Exception as exc:
        return {"error": str(exc)}


def render_regime_juros_section():
    data = get_regime_juros_indice_atual()
    if data.get("error"):
        st.info(f"REGIME DE JUROS indisponivel: {data['error']}")
        return

    try:
        taxa = float(data.get("taxa_sintetica") or 0)
    except (TypeError, ValueError):
        taxa = 0.0
    try:
        variacao = float(data.get("variacao_bps") or 0)
    except (TypeError, ValueError):
        variacao = 0.0

    regime = html.escape(str(data.get("regime_estrutural") or "---"))
    var_color = "#00FFA3" if variacao >= 0 else "#FF4B4B"
    regime_color = "#FF9800" if "bear" in regime.lower() else ("#00FFA3" if "bull" in regime.lower() else "#94A3B8")
    taxa_fmt = f"{taxa:.3f}".replace(".", ",")
    var_fmt = f"{variacao:+.2f}".replace(".", ",")
    updated = html.escape(str(data.get("updated_at") or "---"))
    source_map = {
        "excel_rtd_local": "Excel RTD ProfitChart",
        "supabase": "Supabase RTD",
        "snapshot_excel_local": "Snapshot local",
        "snapshot": "Snapshot local",
    }
    source = source_map.get(str(data.get("source") or ""), "Snapshot local")

    st.markdown(
        f"""
        <section style="margin:10px 0 14px; padding:12px 14px; border:1px solid #243244; border-radius:8px; background:#0B1220;">
          <div style="display:flex; justify-content:space-between; align-items:center; gap:12px; flex-wrap:wrap;">
            <div>
              <div style="font-size:0.72rem; color:#93C5FD; font-weight:900; letter-spacing:.08em; text-transform:uppercase;">Regime de Juros</div>
              <div style="font-size:0.68rem; color:#64748B; margin-top:2px;">Indice Atual | {source} | Atualizado {updated}</div>
            </div>
            <div style="display:flex; gap:10px; flex-wrap:wrap;">
              <div style="min-width:150px; padding:9px 12px; border:1px solid #1E293B; border-radius:8px; background:#0F172A;">
                <div style="font-size:0.66rem; color:#94A3B8; font-weight:800;">Taxa Sintetica</div>
                <div style="font-size:1.35rem; color:#FFFFFF; font-weight:950; line-height:1.1;">{taxa_fmt}</div>
              </div>
              <div style="min-width:150px; padding:9px 12px; border:1px solid #1E293B; border-radius:8px; background:#0F172A;">
                <div style="font-size:0.66rem; color:#94A3B8; font-weight:800;">Variacao bps</div>
                <div style="font-size:1.35rem; color:{var_color}; font-weight:950; line-height:1.1;">{var_fmt}</div>
              </div>
              <div style="min-width:210px; padding:9px 12px; border:1px solid {regime_color}88; border-radius:8px; background:{regime_color}18;">
                <div style="font-size:0.66rem; color:#CBD5E1; font-weight:800;">Regime Estrutural</div>
                <div style="font-size:1.05rem; color:#FFFFFF; font-weight:950; line-height:1.15;">{regime}</div>
              </div>
            </div>
          </div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def pagina_terminal():
    """Renderiza o terminal principal de trading."""
    painel_tickers_topo()   # Indicadores Globais no Topo
    render_regime_juros_section()
    render_top_movers_brasil()
    painel_topo_rtd()       # Tempo Real (1s)
    secao_ia_fragment()     # Estático/Lento (60s)
    painel_inferior_rtd()   # Tempo Real (1s) - Escada de Níveis
    secao_boletim_focus_fragment() # Estático/Lento (300s)
    secao_fluxo_estrangeiro_fragment() # Fluxo B3 (300s)

def pagina_market_report():
    """Página dedicada ao Market Report Institucional."""
    painel_tickers_topo()
    secao_market_report_fragment()


@st.cache_data(ttl=900, show_spinner=False)
def get_watchlist_payload_cached(global_data, schema_version="watchlist_v4_position_only"):
    from execution.watchlist_ai import generate_watchlist

    return generate_watchlist(global_data)


@st.cache_data(ttl=1800, show_spinner=False)
def get_watchlist_chart_candles(ticker: str, period: str = "6mo", interval: str = "1h"):
    if not ticker:
        return []
    try:
        import yfinance as yf

        df = yf.download(
            ticker,
            period=period,
            interval=interval,
            progress=False,
            auto_adjust=False,
            threads=False,
            timeout=10,
        )
        if df is None or df.empty:
            return []
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df = df.dropna(subset=["Open", "High", "Low", "Close"])
        df = df.resample("4h").agg({
            "Open": "first",
            "High": "max",
            "Low": "min",
            "Close": "last",
        }).dropna(subset=["Open", "High", "Low", "Close"]).tail(180)
        candles = []
        for ts, row in df.iterrows():
            candles.append({
                "time": int(pd.Timestamp(ts).timestamp()),
                "open": float(row["Open"]),
                "high": float(row["High"]),
                "low": float(row["Low"]),
                "close": float(row["Close"]),
            })
        return candles
    except Exception:
        return []


def _watchlist_activation_info(item: dict, candles: list[dict]) -> dict:
    try:
        entry_price = float(item.get("entrada") or item.get("entrada_ideal"))
        for candle in reversed(candles or []):
            if float(candle["low"]) <= entry_price <= float(candle["high"]):
                timestamp = int(candle["time"])
                dt_br = datetime.fromtimestamp(timestamp, timezone.utc).astimezone(ZoneInfo("America/Sao_Paulo"))
                return {
                    "price": entry_price,
                    "time": timestamp,
                    "label": dt_br.strftime("%d/%m/%Y %H:%M"),
                }
    except Exception:
        pass
    return {"price": None, "time": None, "label": "Aguardando entrada"}


def _watchlist_chart_html(item: dict, uid: str) -> str:
    ticker = str(item.get("ticker") or "").strip()
    candles = get_watchlist_chart_candles(ticker)
    if not candles:
        return "<div class='wl-chart-empty'>Grafico indisponivel para este ativo.</div>"

    def level(label, key, color, style=0):
        value = item.get(key)
        try:
            value = float(value)
            if value <= 0:
                return None
            return {"label": label, "price": value, "color": color, "style": style}
        except Exception:
            return None

    levels = [
        level("Atual", "preco_atual", "#F8FAFC", 2),
        level("Entrada", "entrada", "#22D3EE", 0),
        level("Gain parcial", "gain_1", "#00FFA3", 0),
        level("Gain final", "gain_final", "#10B981", 0),
        level("Loss", "loss", "#FF4B4B", 0),
    ]
    activation = _watchlist_activation_info(item, candles)
    payload = json.dumps({
        "candles": candles,
        "levels": [item for item in levels if item],
        "entryPrice": activation.get("price"),
        "activationTime": activation.get("time"),
    }, ensure_ascii=False)
    return f"""
    <div id="wl-chart-{uid}" class="wl-chart"></div>
    <script>
    (function() {{
      const payload = {payload};
      const root = document.getElementById("wl-chart-{uid}");
      if (!root || !window.LightweightCharts || !payload.candles || !payload.candles.length) return;
      const chart = LightweightCharts.createChart(root, {{
        layout: {{ background: {{ color: "#020617" }}, textColor: "#CBD5E1" }},
        grid: {{ vertLines: {{ color: "rgba(148,163,184,.10)" }}, horzLines: {{ color: "rgba(148,163,184,.10)" }} }},
        rightPriceScale: {{ borderColor: "rgba(148,163,184,.20)" }},
        timeScale: {{ borderColor: "rgba(148,163,184,.20)", timeVisible: false, rightOffset: 5, barSpacing: 5 }},
        crosshair: {{ mode: 1 }},
        handleScroll: false,
        handleScale: true,
      }});
      let series;
      if (chart.addSeries && LightweightCharts.CandlestickSeries) {{
        series = chart.addSeries(LightweightCharts.CandlestickSeries, {{
          upColor: "#00C896", downColor: "#FF4B4B", borderVisible: false,
          wickUpColor: "#00C896", wickDownColor: "#FF4B4B"
        }});
      }} else {{
        series = chart.addCandlestickSeries({{
          upColor: "#00C896", downColor: "#FF4B4B", borderVisible: false,
          wickUpColor: "#00C896", wickDownColor: "#FF4B4B"
        }});
      }}
      series.setData(payload.candles);
      if (payload.activationTime && series.setMarkers) {{
        series.setMarkers([{{
          time: payload.activationTime,
          position: "aboveBar",
          color: "#22D3EE",
          shape: "circle",
          text: "ENTRADA"
        }}]);
      }}
      (payload.levels || []).forEach(function(level) {{
        series.createPriceLine({{
          price: level.price,
          color: level.color,
          lineWidth: 2,
          lineStyle: level.style || 0,
          axisLabelVisible: true,
          title: level.label,
        }});
      }});
      chart.timeScale().fitContent();
      new ResizeObserver(function() {{
        chart.applyOptions({{ width: root.clientWidth }});
      }}).observe(root);
    }})();
    </script>
    """


def pagina_watchlist():
    """Radar IA TTS Position Trade."""
    st.title("WATCHLIST")
    st.caption("Radar IA TTS para Position Trade: Brasil, EUA, Cripto, Moedas, Commodities e Metais.")

    global_data = get_global_markets_data()
    if st.button("Atualizar Watchlist agora", type="primary", use_container_width=True, key="watchlist_refresh_now"):
        get_watchlist_payload_cached.clear()

    try:
        payload = get_watchlist_payload_cached(global_data, "watchlist_v4_position_only")
    except Exception as e:
        st.error(f"Nao foi possivel gerar a WATCHLIST agora: {e}")
        return

    raw_recs = [rec for rec in payload.get("recommendations", []) if rec.get("tipo") == "Position"]
    macro = payload.get("macro", {})
    quality = payload.get("data_quality", {})

    def fmt_price(value):
        if value is None or value == "":
            return "---"
        try:
            return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return str(value)

    def as_float(value):
        try:
            if value is None or value == "":
                return None
            return float(value)
        except Exception:
            return None

    def fmt_pct(value):
        if value is None:
            return "---"
        try:
            sign = "+" if float(value) > 0 else ""
            return f"{sign}{float(value):.2f}%".replace(".", ",")
        except Exception:
            return "---"

    def pct_from_base(target, base):
        target_f = as_float(target)
        base_f = as_float(base)
        if target_f is None or base_f is None or base_f == 0:
            return None
        return ((target_f / base_f) - 1) * 100

    def trade_pct(target, base, direction="compra"):
        target_f = as_float(target)
        base_f = as_float(base)
        if target_f is None or base_f is None or base_f == 0 or target_f == 0:
            return None
        if str(direction).lower() == "venda":
            return ((base_f / target_f) - 1) * 100
        return ((target_f / base_f) - 1) * 100

    def pct_class(value, neutral_band=0.05):
        if value is None:
            return "neutral"
        if value > neutral_band:
            return "positive"
        if value < -neutral_band:
            return "negative"
        return "flat"

    def enrich_watchlist_recommendation(item):
        enriched = dict(item)
        activation = _watchlist_activation_info(
            item,
            get_watchlist_chart_candles(str(item.get("ticker") or "").strip()),
        )
        enriched["entrada_ativada"] = bool(activation.get("time"))
        enriched["entrada_ativada_em"] = activation.get("label")
        enriched["entrada_ativada_ts"] = activation.get("time")
        return enriched

    recs = [enrich_watchlist_recommendation(rec) for rec in raw_recs]

    def score_color(score):
        try:
            score = float(score)
        except Exception:
            score = 0
        if score >= 72:
            return "#00FFA3"
        if score >= 62:
            return "#FFB020"
        if score >= 52:
            return "#94A3B8"
        return "#FF4B4B"

    st.markdown(
        f"""
        <style>
          .wl-kpi-grid {{display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px; margin:10px 0 16px;}}
          .wl-panel-grid {{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:16px;}}
          .wl-kpi {{background:#0B1220; border:1px solid #1F2937; border-radius:8px; padding:12px;}}
          .wl-kpi span {{display:block; color:#94A3B8; font-size:.72rem; font-weight:800; text-transform:uppercase;}}
          .wl-kpi strong {{display:block; color:#F8FAFC; font-size:1.1rem; margin-top:5px;}}
          .wl-category {{--cat:#38BDF8; --cat-rgb:56,189,248; border:1px solid rgba(var(--cat-rgb),.42); border-top:3px solid var(--cat); border-radius:10px; background:linear-gradient(180deg,rgba(var(--cat-rgb),.13),rgba(11,18,32,.48) 24%,rgba(11,18,32,.12)); padding:14px; margin:0 0 18px; box-shadow:0 12px 32px rgba(0,0,0,.18);}}
          .wl-category-header {{display:flex; justify-content:space-between; align-items:center; gap:12px; margin-bottom:10px;}}
          .wl-category-name {{display:flex; align-items:center; gap:8px; color:#F8FAFC; font-size:1rem; font-weight:950; letter-spacing:.02em; text-transform:uppercase;}}
          .wl-category-dot {{width:9px; height:9px; border-radius:999px; background:var(--cat); box-shadow:0 0 16px rgba(var(--cat-rgb),.82); flex:0 0 auto;}}
          .wl-category-subtitle {{color:var(--cat); font-size:.7rem; font-weight:950; text-transform:uppercase; text-align:right;}}
          .wl-card {{border:1px solid #263244; border-radius:8px; background:#0B1220; padding:13px 14px; margin-bottom:10px; position:relative;}}
          .wl-card.selected {{border-color:rgba(0,255,163,.72); border-left:5px solid #00FFA3; background:linear-gradient(90deg, rgba(0,255,163,.10), #0B1220 38%); box-shadow:0 0 0 1px rgba(0,255,163,.10), 0 10px 28px rgba(0,0,0,.24);}}
          .wl-card.selected.short {{border-color:rgba(255,75,75,.72); border-left-color:#FF4B4B; background:linear-gradient(90deg, rgba(255,75,75,.12), #0B1220 38%);}}
          .wl-head {{display:flex; justify-content:space-between; gap:12px; align-items:flex-start;}}
          .wl-symbol {{font-size:1.05rem; color:#FFF; font-weight:950;}}
          .wl-direction {{display:inline-block; margin-left:6px; border-radius:999px; padding:2px 7px; font-size:.66rem; font-weight:950; vertical-align:middle;}}
          .wl-direction.compra {{color:#00FFA3; border:1px solid rgba(0,255,163,.52); background:rgba(0,255,163,.08);}}
          .wl-direction.venda {{color:#FFB4A8; border:1px solid rgba(255,75,75,.55); background:rgba(255,75,75,.10);}}
          .wl-meta {{color:#94A3B8; font-size:.76rem; margin-top:3px;}}
          .wl-selected-badge {{display:inline-block; margin-left:8px; border:1px solid rgba(0,255,163,.55); border-radius:999px; padding:2px 7px; color:#00FFA3; background:rgba(0,255,163,.08); font-size:.66rem; font-weight:950; vertical-align:middle;}}
          .wl-score {{font-size:1.35rem; font-weight:950; text-align:right;}}
          .wl-grid {{display:grid; grid-template-columns:repeat(6,minmax(0,1fr)); gap:8px; margin-top:12px;}}
          .wl-box {{background:#111827; border:1px solid #1F2937; border-radius:7px; padding:8px; min-height:58px;}}
          .wl-box span {{color:#94A3B8; font-size:.68rem; display:block;}}
          .wl-box b {{color:#F8FAFC; font-size:.88rem; display:block; margin-top:3px;}}
          .wl-box small {{display:block; color:#64748B; font-size:.64rem; margin-top:2px; line-height:1.2;}}
          .wl-box.price {{border-color:#334155; background:linear-gradient(180deg,#111827,#0B1220);}}
          .wl-box.entry {{border-color:rgba(56,189,248,.40); background:rgba(14,116,144,.12);}}
          .wl-box.gain {{border-color:rgba(0,255,163,.36); background:rgba(0,255,163,.08);}}
          .wl-box.loss {{border-color:rgba(255,75,75,.38); background:rgba(255,75,75,.08);}}
          .wl-box.result.positive {{border-color:rgba(0,255,163,.62); background:linear-gradient(180deg,rgba(0,255,163,.16),rgba(0,255,163,.06));}}
          .wl-box.result.negative {{border-color:rgba(255,75,75,.62); background:linear-gradient(180deg,rgba(255,75,75,.16),rgba(255,75,75,.06));}}
          .wl-box.result.flat {{border-color:rgba(255,176,32,.58); background:linear-gradient(180deg,rgba(255,176,32,.14),rgba(255,176,32,.05));}}
          .wl-box.result.neutral {{border-color:#334155;}}
          .wl-pct-positive {{color:#00FFA3 !important;}}
          .wl-pct-negative {{color:#FF4B4B !important;}}
          .wl-pct-flat {{color:#FFB020 !important;}}
          .wl-pct-neutral {{color:#94A3B8 !important;}}
          .wl-text {{color:#CBD5E1; font-size:.86rem; line-height:1.45; margin-top:10px;}}
          .wl-action {{display:inline-block; border:1px solid #334155; border-radius:999px; padding:4px 8px; font-size:.72rem; font-weight:900; color:#F8FAFC; background:#111827; margin-top:8px;}}
          .wl-chart {{height:230px; width:100%; margin-top:12px; border:1px solid rgba(148,163,184,.18); border-radius:8px; background:#020617; overflow:hidden;}}
          .wl-chart-empty {{margin-top:12px; border:1px solid rgba(148,163,184,.18); border-radius:8px; background:#020617; color:#94A3B8; padding:14px; font-size:.78rem; font-weight:800;}}
          .wl-panel-title {{display:flex; justify-content:space-between; align-items:center; gap:10px; margin:4px 0 10px;}}
          .wl-panel-title h3 {{margin:0; color:#F8FAFC; font-size:1.05rem;}}
          .wl-panel-title span {{color:#94A3B8; font-size:.72rem; font-weight:800; text-transform:uppercase;}}
          .wl-comment {{border:1px solid rgba(var(--cat-rgb),.34); background:rgba(9,15,26,.68); border-radius:8px; padding:11px 12px; color:#CBD5E1; font-size:.86rem; line-height:1.45; margin-bottom:10px;}}
          .wl-empty {{border:1px solid rgba(var(--cat-rgb),.26); background:rgba(var(--cat-rgb),.10); color:#BAE6FD; border-radius:8px; padding:13px 14px; font-size:.86rem;}}
          @media(max-width:900px) {{.wl-kpi-grid {{grid-template-columns:1fr 1fr;}} .wl-panel-grid {{grid-template-columns:1fr;}} .wl-grid {{grid-template-columns:1fr 1fr;}}}}
        </style>
        <div class="wl-kpi-grid">
          <div class="wl-kpi"><span>Regime Macro</span><strong>{html.escape(str(macro.get('regime', '---')))}</strong></div>
          <div class="wl-kpi"><span>Score Macro</span><strong>{html.escape(str(macro.get('score', '---')))}</strong></div>
          <div class="wl-kpi"><span>Ativos carregados</span><strong>{quality.get('assets_loaded', 0)}</strong><small style="color:#94A3B8;">Hist {quality.get('historical_assets', 0)} | Dash {quality.get('dashboard_assets', 0)}</small></div>
          <div class="wl-kpi"><span>Atualizado</span><strong>{html.escape(str(payload.get('generated_at', '---'))[11:])}</strong></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    wl_component_css = """
    <style>
      body {margin:0; background:transparent; font-family:"Inter","Segoe UI",Arial,sans-serif; color:#F8FAFC;}
      .wl-category {--cat:#38BDF8; --cat-rgb:56,189,248; border:1px solid rgba(var(--cat-rgb),.42); border-top:3px solid var(--cat); border-radius:10px; background:linear-gradient(180deg,rgba(var(--cat-rgb),.13),rgba(11,18,32,.48) 24%,rgba(11,18,32,.12)); padding:14px; margin:0 0 18px; box-shadow:0 12px 32px rgba(0,0,0,.18);}
      .wl-category-header {display:flex; justify-content:space-between; align-items:center; gap:12px; margin-bottom:10px;}
      .wl-category-name {display:flex; align-items:center; gap:8px; color:#F8FAFC; font-size:1rem; font-weight:950; letter-spacing:.02em; text-transform:uppercase;}
      .wl-category-dot {width:9px; height:9px; border-radius:999px; background:var(--cat); box-shadow:0 0 16px rgba(var(--cat-rgb),.82); flex:0 0 auto;}
      .wl-category-subtitle {color:var(--cat); font-size:.7rem; font-weight:950; text-transform:uppercase; text-align:right;}
      .wl-comment {border:1px solid rgba(var(--cat-rgb),.34); background:rgba(9,15,26,.68); border-radius:8px; padding:11px 12px; color:#CBD5E1; font-size:.86rem; line-height:1.45; margin-bottom:10px;}
      .wl-card {border:1px solid #263244; border-radius:8px; background:#0B1220; padding:13px 14px; margin-bottom:10px; position:relative;}
      .wl-card.selected {border-color:rgba(0,255,163,.72); border-left:5px solid #00FFA3; background:linear-gradient(90deg, rgba(0,255,163,.10), #0B1220 38%); box-shadow:0 0 0 1px rgba(0,255,163,.10), 0 10px 28px rgba(0,0,0,.24);}
      .wl-card.selected.short {border-color:rgba(255,75,75,.72); border-left-color:#FF4B4B; background:linear-gradient(90deg, rgba(255,75,75,.12), #0B1220 38%);}
      .wl-head {display:flex; justify-content:space-between; gap:12px; align-items:flex-start;}
      .wl-symbol {font-size:1.05rem; color:#FFF; font-weight:950;}
      .wl-direction {display:inline-block; margin-left:6px; border-radius:999px; padding:2px 7px; font-size:.66rem; font-weight:950; vertical-align:middle;}
      .wl-direction.compra {color:#00FFA3; border:1px solid rgba(0,255,163,.52); background:rgba(0,255,163,.08);}
      .wl-direction.venda {color:#FFB4A8; border:1px solid rgba(255,75,75,.55); background:rgba(255,75,75,.10);}
      .wl-meta {color:#94A3B8; font-size:.76rem; margin-top:3px;}
      .wl-selected-badge {display:inline-block; margin-left:8px; border:1px solid rgba(0,255,163,.55); border-radius:999px; padding:2px 7px; color:#00FFA3; background:rgba(0,255,163,.08); font-size:.66rem; font-weight:950; vertical-align:middle;}
      .wl-score {font-size:1.35rem; font-weight:950; text-align:right;}
      .wl-grid {display:grid; grid-template-columns:repeat(6,minmax(0,1fr)); gap:8px; margin-top:12px;}
      .wl-box {background:#111827; border:1px solid #1F2937; border-radius:7px; padding:8px; min-height:58px;}
      .wl-box span {color:#94A3B8; font-size:.68rem; display:block;}
      .wl-box b {color:#F8FAFC; font-size:.88rem; display:block; margin-top:3px;}
      .wl-box small {display:block; color:#64748B; font-size:.64rem; margin-top:2px; line-height:1.2;}
      .wl-box.price {border-color:#334155; background:linear-gradient(180deg,#111827,#0B1220);}
      .wl-box.entry {border-color:rgba(56,189,248,.40); background:rgba(14,116,144,.12);}
      .wl-box.gain {border-color:rgba(0,255,163,.36); background:rgba(0,255,163,.08);}
      .wl-box.loss {border-color:rgba(255,75,75,.38); background:rgba(255,75,75,.08);}
      .wl-box.result.positive {border-color:rgba(0,255,163,.62); background:linear-gradient(180deg,rgba(0,255,163,.16),rgba(0,255,163,.06));}
      .wl-box.result.negative {border-color:rgba(255,75,75,.62); background:linear-gradient(180deg,rgba(255,75,75,.16),rgba(255,75,75,.06));}
      .wl-box.result.flat {border-color:rgba(255,176,32,.58); background:linear-gradient(180deg,rgba(255,176,32,.14),rgba(255,176,32,.05));}
      .wl-pct-positive {color:#00FFA3 !important;} .wl-pct-negative {color:#FF4B4B !important;} .wl-pct-flat {color:#FFB020 !important;} .wl-pct-neutral {color:#94A3B8 !important;}
      .wl-text {color:#CBD5E1; font-size:.86rem; line-height:1.45; margin-top:10px;}
      .wl-action {display:inline-block; border:1px solid #334155; border-radius:999px; padding:4px 8px; font-size:.72rem; font-weight:900; color:#F8FAFC; background:#111827; margin-top:8px;}
      .wl-chart {height:230px; width:100%; margin-top:12px; border:1px solid rgba(148,163,184,.18); border-radius:8px; background:#020617; overflow:hidden;}
      .wl-chart-empty {margin-top:12px; border:1px solid rgba(148,163,184,.18); border-radius:8px; background:#020617; color:#94A3B8; padding:14px; font-size:.78rem; font-weight:800;}
      .wl-empty {border:1px solid rgba(var(--cat-rgb),.26); background:rgba(var(--cat-rgb),.10); color:#BAE6FD; border-radius:8px; padding:13px 14px; font-size:.86rem;}
      @media(max-width:900px){.wl-grid{grid-template-columns:1fr 1fr;}}
    </style>
    """

    def build_recommendation_cards(items, limit=10, uid_prefix="wl"):
        if not items:
            return '<div class="wl-empty">Sem recomendacoes com dados suficientes para este bloco.</div>'
        html_cards = []
        for idx, item in enumerate(sorted(items, key=lambda r: r.get("score_atual", 0), reverse=True)[:limit]):
            score = item.get("score_atual", 0)
            color = score_color(score)
            action_text = str(item.get("acao", "---"))
            direction = str(item.get("direcao", "compra")).lower()
            direction = "venda" if direction == "venda" else "compra"
            direction_label = "SHORT" if direction == "venda" else "LONG"
            selected = action_text in {"comprar", "vender"} or float(score or 0) >= 72
            card_class = "wl-card selected short" if selected and direction == "venda" else "wl-card selected" if selected else "wl-card"
            selected_badge = '<span class="wl-selected-badge">SELECIONADA</span>' if selected else ""
            ref_entry = item.get("entrada") or item.get("entrada_ideal")
            is_active = bool(item.get("entrada_ativada"))
            result_pct = trade_pct(item.get("preco_atual"), ref_entry, direction) if is_active else None
            result_class = pct_class(result_pct)
            result_text = fmt_pct(result_pct) if is_active else "Aguardando"
            result_hint = "pos entrada" if is_active else "sem ativacao"
            gain_1_pct = trade_pct(item.get("gain_1"), ref_entry, direction)
            gain_final_pct = trade_pct(item.get("gain_final"), ref_entry, direction)
            loss_pct = trade_pct(item.get("loss"), ref_entry, direction)
            activation_label = item.get("entrada_ativada_em") or "Aguardando entrada"
            chart_html = _watchlist_chart_html(item, f"{uid_prefix}-{idx}")
            html_cards.append(
                f"""
                <div class="{card_class}">
                  <div class="wl-head">
                    <div>
                      <div class="wl-symbol">{html.escape(str(item.get('ativo', '---')))} <span class="wl-direction {direction}">{direction_label}</span> <span style="color:#94A3B8;font-size:.78rem;">{html.escape(str(item.get('tipo', '---')))}</span>{selected_badge}</div>
                      <div class="wl-meta">{html.escape(str(item.get('bloco', '---')))} | {html.escape(str(item.get('setor', '---')))} | {html.escape(str(item.get('status', '---')))}</div>
                    </div>
                    <div>
                      <div class="wl-score" style="color:{color};">{score}</div>
                      <div style="color:#94A3B8;font-size:.7rem;text-align:right;">score atual</div>
                    </div>
                  </div>
                  <div class="wl-grid">
                    <div class="wl-box price"><span>Preco atual</span><b>{fmt_price(item.get('preco_atual'))}</b><small>referencia viva</small></div>
                    <div class="wl-box result {result_class}"><span>Resultado</span><b class="wl-pct-{result_class}">{html.escape(str(result_text))}</b><small>{result_hint}</small></div>
                    <div class="wl-box entry"><span>Entrada</span><b>{fmt_price(ref_entry)}</b><small>preco unico</small></div>
                    <div class="wl-box gain"><span>Gain parcial</span><b>{fmt_price(item.get('gain_1'))}</b><small>{fmt_pct(gain_1_pct)}</small></div>
                    <div class="wl-box gain"><span>Gain final</span><b>{fmt_price(item.get('gain_final'))}</b><small>{fmt_pct(gain_final_pct)}</small></div>
                    <div class="wl-box loss"><span>Loss</span><b>{fmt_price(item.get('loss'))}</b><small>{fmt_pct(loss_pct)}</small></div>
                  </div>
                  <div class="wl-text"><b>Ativacao:</b> {html.escape(str(activation_label))}</div>
                  <div class="wl-text"><b>Tese:</b> {html.escape(str(item.get('tese_principal', '')))}</div>
                  <div class="wl-text"><b>Confirmacoes:</b> {html.escape(str(item.get('confirmacoes', '')))}</div>
                  <div class="wl-text"><b>Filtros:</b> {html.escape(str(item.get('filtros', '---')))}</div>
                  <span class="wl-action">{html.escape(str(item.get('acao', '---')).upper())} | RR {html.escape(str(item.get('risco_retorno', '---')))} | {html.escape(str(item.get('tamanho_sugerido', '---')))} | {html.escape(str(item.get('fonte_descricao', 'historico yfinance')))}</span>
                  {chart_html}
                </div>
                """
            )
        return "".join(line.strip() for line in "".join(html_cards).splitlines())

    def render_recommendations(items, limit=10, uid_prefix="wl-detail"):
        selected_items = sorted(items, key=lambda r: r.get("score_atual", 0), reverse=True)[:limit]
        html_cards = (
            '<script src="https://unpkg.com/lightweight-charts/dist/lightweight-charts.standalone.production.js"></script>'
            + wl_component_css
            + build_recommendation_cards(selected_items, limit, uid_prefix)
        )
        components.html(html_cards, height=max(150, len(selected_items) * 485 + 40), scrolling=True)

    def rec_filter(tipo=None, bloco=None):
        out = recs
        if tipo:
            out = [r for r in out if r.get("tipo") == tipo]
        if bloco:
            out = [r for r in out if str(r.get("bloco", "")).startswith(bloco)]
        return out

    def render_watchlist_results(rows):
        st.markdown("#### Registro de takes e stops")
        if not rows:
            st.caption("Ainda nao houve recomendacao com take ou stop atingido.")
            return
        df_results = pd.DataFrame(rows).tail(30).iloc[::-1].copy()
        cols = ["data", "ativo", "direcao", "tipo", "bloco", "evento", "entrada", "saida", "preco_atual", "resultado_pct", "score"]
        df_results = df_results[[c for c in cols if c in df_results.columns]]
        rename = {
            "data": "Data",
            "ativo": "Ativo",
            "direcao": "Direcao",
            "tipo": "Tipo",
            "bloco": "Bloco",
            "evento": "Evento",
            "entrada": "Entrada",
            "saida": "Saida",
            "preco_atual": "Preco atual",
            "resultado_pct": "Resultado %",
            "score": "Score",
        }
        df_results = df_results.rename(columns=rename)
        if "Resultado %" in df_results.columns:
            df_results["Resultado %"] = pd.to_numeric(df_results["Resultado %"], errors="coerce")

        def style_result(row):
            result = row.get("Resultado %", 0)
            event = str(row.get("Evento", ""))
            color = "color:#00FFA3; font-weight:900;" if result >= 0 else "color:#FF4B4B; font-weight:900;"
            bg = "background-color:rgba(0,255,163,.08);" if "TAKE" in event else "background-color:rgba(255,75,75,.08);"
            return [bg + (color if col == "Resultado %" else "") for col in row.index]

        styled = (
            df_results.style
            .apply(style_result, axis=1)
            .format({
                "Entrada": "{:,.2f}",
                "Saida": "{:,.2f}",
                "Preco atual": "{:,.2f}",
                "Resultado %": "{:+.2f}%",
            }, na_rep="---")
        )
        st.dataframe(styled, hide_index=True, use_container_width=True, height=min(430, 38 + len(df_results) * 35))

    comments = payload.get("commentary", {})
    watchlist_blocks = [
        ("Brasil", "Brasil - Acoes", "Position Trade", "brasil", 5, "#00FFA3", "0,255,163"),
        ("EUA", "EUA - ETFs Setoriais", "Rotacao setorial", "eua", 5, "#38BDF8", "56,189,248"),
        ("Cripto", "Cripto", "Liquidez + beta", "cripto", 4, "#F59E0B", "245,158,11"),
        ("Moedas", "Moedas / Forex", "FX macro", "moedas", 4, "#22D3EE", "34,211,238"),
        ("Commodities", "Commodities", "Energia + graos", "commodities", 4, "#F97316", "249,115,22"),
        ("Metais", "Metais", "Preciosos + industriais", "metais", 4, "#FACC15", "250,204,21"),
    ]

    def block_comment(comment_key, title):
        text = comments.get(comment_key)
        if text:
            return text
        return f"Radar {title} aguardando novo ciclo de dados. Clique em Atualizar Watchlist agora se o cache antigo ainda estiver ativo."

    def render_panel(block_prefix, title, subtitle, comment_key, limit, color, rgb):
        panel_items = rec_filter(bloco=block_prefix)
        cards_html = build_recommendation_cards(panel_items, limit=limit, uid_prefix=f"panel-{block_prefix}")
        panel_html = (
            '<script src="https://unpkg.com/lightweight-charts/dist/lightweight-charts.standalone.production.js"></script>'
            + wl_component_css
            + f'<div class="wl-category" style="--cat:{html.escape(color)}; --cat-rgb:{html.escape(rgb)};">'
            f'<div class="wl-category-header">'
            f'<div class="wl-category-name"><span class="wl-category-dot"></span>{html.escape(title)}</div>'
            f'<div class="wl-category-subtitle">{html.escape(subtitle)}</div>'
            f'</div>'
            f'<div class="wl-comment">{html.escape(str(block_comment(comment_key, title)))}</div>'
            f'{cards_html}'
            f'</div>'
        )
        panel_count = min(len(panel_items), limit)
        components.html(panel_html, height=max(210, panel_count * 515 + 95), scrolling=True)

    st.markdown("#### Mesa WATCHLIST")
    for idx in range(0, len(watchlist_blocks), 2):
        left, right = st.columns(2, gap="large")
        with left:
            render_panel(*watchlist_blocks[idx])
        if idx + 1 < len(watchlist_blocks):
            with right:
                render_panel(*watchlist_blocks[idx + 1])

    st.markdown("---")
    try:
        from execution.watchlist_ai import update_watchlist_results

        watchlist_results = update_watchlist_results(recs)
    except Exception as e:
        watchlist_results = []
        st.caption(f"Registro de takes/stops indisponivel agora: {e}")
    render_watchlist_results(watchlist_results)

    st.markdown("---")

    tabs = st.tabs([
        "Visao Macro Global",
        "Brasil - Acoes",
        "EUA - Rotacao Setorial",
        "Cripto",
        "Moedas",
        "Commodities",
        "Metais",
        "Position",
        "Comentario da IA",
    ])

    with tabs[0]:
        st.markdown("#### Visao Macro Global")
        st.write(f"Regime: **{macro.get('regime', '---')}** | SPX: `{macro.get('spx')}` | Nasdaq: `{macro.get('nasdaq')}` | VIX: `{macro.get('vix')}` | DXY: `{macro.get('dxy')}` | EWZ: `{macro.get('ewz')}` | IBOV: `{macro.get('ibov')}`")
        st.caption(f"Fonte: {quality.get('source', '---')}. Takes e stops acionados ficam registrados no historico local da WATCHLIST.")
    with tabs[1]:
        render_recommendations(rec_filter(bloco="Brasil"), limit=16, uid_prefix="tab-brasil")
    with tabs[2]:
        render_recommendations(rec_filter(bloco="EUA"), limit=10, uid_prefix="tab-eua")
    with tabs[3]:
        render_recommendations(rec_filter(bloco="Cripto"), limit=10, uid_prefix="tab-cripto")
    with tabs[4]:
        render_recommendations(rec_filter(bloco="Moedas"), limit=10, uid_prefix="tab-moedas")
    with tabs[5]:
        render_recommendations(rec_filter(bloco="Commodities"), limit=10, uid_prefix="tab-commodities")
    with tabs[6]:
        render_recommendations(rec_filter(bloco="Metais"), limit=10, uid_prefix="tab-metais")
    with tabs[7]:
        st.markdown("#### Radar Position")
        for block_prefix, title, *_rest in watchlist_blocks:
            st.markdown(f"##### {title}")
            render_recommendations(rec_filter(tipo="Position", bloco=block_prefix), limit=6, uid_prefix=f"tab-position-{block_prefix}")
    with tabs[8]:
        for _block_prefix, title, _subtitle, comment_key, _limit, _color, _rgb in watchlist_blocks:
            st.markdown(f"#### {title}\n{block_comment(comment_key, title)}")


@st.cache_data(ttl=1800, show_spinner=False)
def get_watchlist_quant_cached(global_data, schema_version="watchlist_quant_v2"):
    from execution.watchlist_quant import build_watchlist_quant

    return build_watchlist_quant(global_data)


def pagina_watchlist_quant():
    """Quant screening using dashboard/watchlist universe."""
    st.title("WATCHLIST QUANT")
    st.caption("Screening quantitativo com Momentum, Reversao a Media, Pairs/StatArb, Event-driven, Volatility e Crypto Quant.")

    global_data = get_global_markets_data()
    if st.button("Atualizar WATCHLIST QUANT", type="primary", use_container_width=True, key="watchlist_quant_refresh"):
        get_watchlist_quant_cached.clear()

    try:
        payload = get_watchlist_quant_cached(global_data, "watchlist_quant_v2")
    except Exception as e:
        st.error(f"Nao foi possivel gerar o screening quant agora: {e}")
        return

    st.markdown(
        f"""
        <style>
          .quant-kpis {{display:grid; grid-template-columns:repeat(6,minmax(0,1fr)); gap:10px; margin:10px 0 18px;}}
          .quant-kpi {{background:#0B1220; border:1px solid #1F2937; border-radius:8px; padding:12px;}}
          .quant-kpi span {{display:block; color:#94A3B8; font-size:.72rem; font-weight:900; text-transform:uppercase;}}
          .quant-kpi strong {{display:block; color:#F8FAFC; font-size:1.05rem; margin-top:5px;}}
          .quant-note {{border:1px solid #263244; border-left:4px solid #22D3EE; background:#07111F; border-radius:8px; padding:11px 13px; color:#CBD5E1; margin-bottom:14px;}}
          @media (max-width: 1100px) {{.quant-kpis {{grid-template-columns:repeat(2,minmax(0,1fr));}}}}
        </style>
        <div class="quant-kpis">
          <div class="quant-kpi"><span>Ativos carregados</span><strong>{payload.get('assets_loaded', 0)}</strong></div>
          <div class="quant-kpi"><span>Top Momentum</span><strong>{html.escape(str(payload.get('summary', {}).get('top_momentum', '---')))}</strong></div>
          <div class="quant-kpi"><span>Top Reversao</span><strong>{html.escape(str(payload.get('summary', {}).get('top_reversion', '---')))}</strong></div>
          <div class="quant-kpi"><span>Top Event</span><strong>{html.escape(str(payload.get('summary', {}).get('top_event', '---')))}</strong></div>
          <div class="quant-kpi"><span>Top Volatility</span><strong>{html.escape(str(payload.get('summary', {}).get('top_volatility', '---')))}</strong></div>
          <div class="quant-kpi"><span>Top Crypto</span><strong>{html.escape(str(payload.get('summary', {}).get('top_crypto', '---')))}</strong></div>
        </div>
        <div class="quant-note">Modelo local e deterministico. Os sinais sao rankings quantitativos para triagem, nao execucao automatica.</div>
        """,
        unsafe_allow_html=True,
    )

    def fmt_quant_df(rows: list[dict[str, Any]], cols: list[str], rename: dict[str, str]):
        if not rows:
            st.info("Sem sinais suficientes neste bloco agora.")
            return
        df = pd.DataFrame(rows)
        df = df[[col for col in cols if col in df.columns]].rename(columns=rename)
        st.dataframe(df, hide_index=True, use_container_width=True, height=min(520, 38 + len(df) * 36))

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["Momentum", "Reversao a Media", "Pairs / StatArb", "Event-driven", "Volatility", "Crypto Quant"])
    with tab1:
        st.markdown("#### Momentum / Trend Following")
        fmt_quant_df(
            payload.get("momentum", []),
            ["symbol", "block", "direction", "score", "price", "entry", "stop", "target", "ret20", "ret60", "vol20", "setup"],
            {
                "symbol": "Ativo", "block": "Classe", "direction": "Direcao", "score": "Score",
                "price": "Preco", "entry": "Entrada", "stop": "Stop", "target": "Alvo",
                "ret20": "Ret 20d %", "ret60": "Ret 60d %", "vol20": "Vol 20d %", "setup": "Setup",
            },
        )
    with tab2:
        st.markdown("#### Mean Reversion / Z-score")
        fmt_quant_df(
            payload.get("mean_reversion", []),
            ["symbol", "block", "direction", "score", "price", "entry", "stop", "target", "z20", "vol20", "setup"],
            {
                "symbol": "Ativo", "block": "Classe", "direction": "Direcao", "score": "Score",
                "price": "Preco", "entry": "Entrada", "stop": "Stop", "target": "Alvo",
                "z20": "Z 20d", "vol20": "Vol 20d %", "setup": "Setup",
            },
        )
    with tab3:
        st.markdown("#### Pairs Trading / StatArb")
        fmt_quant_df(
            payload.get("pairs", []),
            ["pair", "block", "long", "short", "corr", "zscore", "score", "setup"],
            {
                "pair": "Par", "block": "Classe", "long": "Comprar", "short": "Vender",
                "corr": "Correlacao", "zscore": "Z Spread", "score": "Score", "setup": "Setup",
            },
        )
    with tab4:
        st.markdown("#### Event-driven")
        fmt_quant_df(
            payload.get("event_driven", []),
            ["symbol", "block", "direction", "score", "price", "entry", "stop", "target", "ret20", "z20", "vol20", "setup"],
            {
                "symbol": "Ativo", "block": "Classe", "direction": "Direcao", "score": "Score",
                "price": "Preco", "entry": "Entrada", "stop": "Stop", "target": "Alvo",
                "ret20": "Impulso 20d %", "z20": "Z 20d", "vol20": "Vol 20d %", "setup": "Setup",
            },
        )
    with tab5:
        st.markdown("#### Volatility")
        fmt_quant_df(
            payload.get("volatility", []),
            ["symbol", "block", "direction", "score", "price", "entry", "stop", "target", "vol20", "atr14", "z20", "setup"],
            {
                "symbol": "Ativo", "block": "Classe", "direction": "Regime", "score": "Score",
                "price": "Preco", "entry": "Referencia", "stop": "Stop", "target": "Alvo",
                "vol20": "Vol 20d %", "atr14": "ATR 14", "z20": "Z 20d", "setup": "Setup",
            },
        )
    with tab6:
        st.markdown("#### Crypto Quant")
        fmt_quant_df(
            payload.get("crypto_quant", []),
            ["symbol", "block", "direction", "score", "price", "entry", "stop", "target", "ret20", "ret60", "vol20", "setup"],
            {
                "symbol": "Ativo", "block": "Classe", "direction": "Direcao", "score": "Score",
                "price": "Preco", "entry": "Entrada", "stop": "Stop", "target": "Alvo",
                "ret20": "Ret 20d %", "ret60": "Ret 60d %", "vol20": "Vol 20d %", "setup": "Setup",
            },
        )


@st.cache_data(ttl=60, show_spinner=False)
def load_crypto_terminal_payload(refresh_key: int = 0):
    """Load public crypto sources and deterministic regime snapshot."""
    from execution.crypto_bgeometrics import fetch_bgeometrics_snapshot
    from execution.crypto_binance import fetch_binance_crypto_snapshot
    from execution.crypto_coingecko import fetch_coingecko_crypto_snapshot
    from execution.crypto_defillama import fetch_defillama_crypto_snapshot
    from execution.crypto_fear_greed import fetch_fear_greed_snapshot
    from execution.crypto_regime import calculate_crypto_regime
    from execution.crypto_signals import build_crypto_operational_dashboard

    binance = fetch_binance_crypto_snapshot()
    coingecko = fetch_coingecko_crypto_snapshot()
    fear_greed = fetch_fear_greed_snapshot()
    defillama = fetch_defillama_crypto_snapshot()
    bgeometrics = fetch_bgeometrics_snapshot()
    regime = calculate_crypto_regime(binance, coingecko, fear_greed, defillama, bgeometrics)
    operational = build_crypto_operational_dashboard(binance, regime, coingecko)
    return {
        "binance": binance,
        "coingecko": coingecko,
        "fear_greed": fear_greed,
        "defillama": defillama,
        "bgeometrics": bgeometrics,
        "regime": regime,
        "operational": operational,
    }


@st.cache_data(ttl=21600, show_spinner=False)
def load_crypto_mvrv_history(refresh_key: int = 0):
    """Load MVRV history lazily so this optional chart never blocks app boot."""
    try:
        from execution.crypto_bgeometrics import fetch_bgeometrics_mvrv_zscore_history

        return fetch_bgeometrics_mvrv_zscore_history(days=6000)
    except Exception as exc:
        return {
            "source": "BGeometrics",
            "status": "error",
            "data": {"points": []},
            "warnings": [f"Historico MVRV indisponivel: {exc}"],
        }


@st.cache_data(ttl=86400, show_spinner=False)
def load_crypto_rainbow_chart(refresh_key: int = 0):
    """Load Bitcoin Rainbow Chart lazily to avoid slowing app boot."""
    try:
        from execution.crypto_bgeometrics import fetch_bgeometrics_rainbow_chart

        return fetch_bgeometrics_rainbow_chart()
    except Exception as exc:
        return {
            "source": "BGeometrics",
            "status": "error",
            "data": {"points": []},
            "warnings": [f"Rainbow Chart indisponivel: {exc}"],
        }


def update_usdt_dominance_flow(usdt_dom: float | None) -> dict[str, Any]:
    """Persist a lightweight USDT dominance trail and classify cash flow."""
    if usdt_dom is None:
        return {
            "status": "Sem dado",
            "cls": "neutral",
            "delta": None,
            "text": "Aguardando market cap de USDT e mercado cripto.",
            "history": [],
        }
    path = os.path.join(LOCAL_TMP_DIR, "crypto_usdt_dominance_history.json")
    os.makedirs(LOCAL_TMP_DIR, exist_ok=True)
    history: list[dict[str, Any]] = []
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            if isinstance(loaded, list):
                history = loaded
    except Exception:
        history = []

    now_ts = time.time()
    last = history[-1] if history else {}
    try:
        last_value = float(last.get("value")) if isinstance(last, dict) and last.get("value") is not None else None
    except Exception:
        last_value = None
    delta = None if last_value is None else float(usdt_dom) - float(last_value)
    if not history or abs(float(usdt_dom) - float(last_value or 0)) >= 0.001 or now_ts - float(last.get("ts") or 0) > 900:
        history.append({
            "ts": now_ts,
            "updated_at": datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%Y-%m-%d %H:%M:%S"),
            "value": round(float(usdt_dom), 4),
        })
        history = history[-3000:]
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(history, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    lookback_seconds = 15 * 24 * 60 * 60
    window_rows = [
        row for row in history
        if isinstance(row, dict) and now_ts - float(row.get("ts") or 0) <= lookback_seconds and row.get("value") is not None
    ]
    window_values = []
    for row in window_rows:
        try:
            window_values.append(float(row.get("value")))
        except Exception:
            continue
    high_15d = max(window_values) if window_values else float(usdt_dom)
    low_15d = min(window_values) if window_values else float(usdt_dom)
    distance_from_high = float(usdt_dom) - high_15d
    distance_from_low = float(usdt_dom) - low_15d
    range_15d = high_15d - low_15d
    near_high = abs(distance_from_high) <= 0.05
    near_low = abs(distance_from_low) <= 0.05
    flow_context = "Aguardando janela de 15 dias ganhar historico."
    if len(window_values) >= 2:
        if near_high and (delta or 0) > 0:
            flow_context = "USDT perto/rompendo maxima de 15d: realizacao em cripto e caixa aumentando."
        elif near_high:
            flow_context = "USDT perto da maxima de 15d: mercado ainda com caixa defensivo elevado."
        elif (delta or 0) < -0.05 and range_15d > 0:
            flow_context = "USDT afastando da maxima de 15d: caixa saindo de stable e voltando para cripto."
        elif near_low:
            flow_context = "USDT perto da minima de 15d: caixa em stable comprimido, apetite por cripto maior."
        elif (delta or 0) > 0.05:
            flow_context = "USDT subindo dentro da janela de 15d: fluxo marginal indo para stable."

    if delta is None:
        return {
            "status": "Monitorando",
            "cls": "neutral",
            "delta": None,
            "high_15d": high_15d,
            "low_15d": low_15d,
            "distance_from_high": distance_from_high,
            "flow_context": flow_context,
            "text": "Primeira leitura registrada. Proxima atualizacao define fluxo.",
            "history": history,
        }
    if delta >= 0.08:
        status, cls = "Caixa subindo", "bad"
        text = "USDT dominance subiu: dinheiro realizando cripto e indo para stable/caixa."
    elif delta <= -0.08:
        status, cls = "Caixa saindo", "good"
        text = "USDT dominance caiu: dinheiro saindo de stable e entrando em cripto."
    elif delta > 0:
        status, cls = "Leve defesa", "warn"
        text = "USDT dominance subiu pouco; fluxo marginal mais defensivo."
    elif delta < 0:
        status, cls = "Leve risco", "good"
        text = "USDT dominance caiu pouco; fluxo marginal favorece risco cripto."
    else:
        status, cls = "Estavel", "neutral"
        text = "USDT dominance sem mudanca relevante."
    return {
        "status": status,
        "cls": cls,
        "delta": delta,
        "high_15d": high_15d,
        "low_15d": low_15d,
        "distance_from_high": distance_from_high,
        "flow_context": flow_context,
        "text": f"{text} {flow_context}",
        "history": history,
    }


def _crypto_mini_chart_html(symbol: str, asset: dict[str, Any], uid: str) -> str:
    candles = asset.get("candles_1h") or []
    chart_candles = [
        {
            "time": int(row.get("time") or 0),
            "open": float(row.get("open") or 0),
            "high": float(row.get("high") or 0),
            "low": float(row.get("low") or 0),
            "close": float(row.get("close") or 0),
        }
        for row in candles
        if row.get("time") and row.get("open") and row.get("high") and row.get("low") and row.get("close")
    ][-72:]
    payload = json.dumps({"candles": chart_candles}, ensure_ascii=False)
    safe_symbol = html.escape(symbol.replace("USDT", ""))
    change = float(asset.get("change_pct_24h") or 0)
    change_cls = "pos" if change >= 0 else "neg"
    return f"""
    <div class="crypto-mini-card">
      <div class="crypto-mini-head"><b>{safe_symbol}</b><span class="{change_cls}">{change:+.2f}%</span><em>1h</em></div>
      <div id="crypto-mini-{uid}" class="crypto-mini-chart"></div>
    </div>
    <script>
    (function() {{
      const payload = {payload};
      const root = document.getElementById("crypto-mini-{uid}");
      if (!root || !window.LightweightCharts || !payload.candles || !payload.candles.length) return;
      const chart = LightweightCharts.createChart(root, {{
        layout: {{ background: {{ color: "#030712" }}, textColor: "#AAB7C4" }},
        grid: {{ vertLines: {{ color: "rgba(148,163,184,.08)" }}, horzLines: {{ color: "rgba(148,163,184,.08)" }} }},
        rightPriceScale: {{ borderVisible: false }},
        timeScale: {{ borderVisible: false, timeVisible: true, secondsVisible: false, rightOffset: 3 }},
        crosshair: {{ mode: 1 }},
        handleScroll: false,
        handleScale: false,
      }});
      let series;
      if (chart.addSeries && LightweightCharts.AreaSeries) {{
        series = chart.addSeries(LightweightCharts.AreaSeries, {{
          lineColor: "{'#00D084' if change >= 0 else '#FF5D5D'}",
          topColor: "{'rgba(0,208,132,.30)' if change >= 0 else 'rgba(255,93,93,.28)'}",
          bottomColor: "rgba(3,7,18,0)",
          lineWidth: 2,
        }});
      }} else {{
        series = chart.addAreaSeries({{
          lineColor: "{'#00D084' if change >= 0 else '#FF5D5D'}",
          topColor: "{'rgba(0,208,132,.30)' if change >= 0 else 'rgba(255,93,93,.28)'}",
          bottomColor: "rgba(3,7,18,0)",
          lineWidth: 2,
        }});
      }}
      series.setData(payload.candles.map(c => {{ return {{ time: c.time, value: c.close }}; }}));
      chart.timeScale().fitContent();
      new ResizeObserver(function() {{ chart.applyOptions({{ width: root.clientWidth }}); }}).observe(root);
    }})();
    </script>
    """


def _render_crypto_mvrv_chart(points: list[dict[str, Any]]) -> None:
    chart_points = [
        {
            "time": int(row.get("time") or 0),
            "btc_price": float(row.get("btc_price") or 0),
            "realized_price": float(row.get("realized_price") or 0),
        }
        for row in points
        if row.get("time") and row.get("btc_price") and row.get("realized_price")
    ][-730:]
    if not chart_points:
        st.markdown("<div class='crypto-empty'>Sem historico BTC / Realized Price disponivel agora.</div>", unsafe_allow_html=True)
        return

    import plotly.graph_objects as go

    df = pd.DataFrame(chart_points)
    df["date"] = pd.to_datetime(df["time"], unit="s", utc=True)
    latest_btc = float(chart_points[-1]["btc_price"] or 0)
    latest_realized = float(chart_points[-1]["realized_price"] or 0)
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df["date"],
        y=df["btc_price"],
        mode="lines",
        name="BTC Price",
        line={"color": "#F8FAFC", "width": 2},
        hovertemplate="%{x|%d/%m/%Y}<br>BTC: US$ %{y:,.0f}<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=df["date"],
        y=df["realized_price"],
        mode="lines",
        name="Realized Price",
        line={"color": "#22D3EE", "width": 2},
        hovertemplate="%{x|%d/%m/%Y}<br>Realized: US$ %{y:,.0f}<extra></extra>",
    ))
    fig.update_layout(
        height=470,
        margin={"l": 10, "r": 22, "t": 8, "b": 10},
        paper_bgcolor="#050B14",
        plot_bgcolor="#050B14",
        font={"color": "#AAB7C4"},
        legend={"orientation": "h", "y": -0.08, "x": 0.01, "font": {"size": 11}},
        xaxis={"gridcolor": "rgba(148,163,184,.08)", "zeroline": False},
    )
    fig.update_yaxes(
        type="log",
        title_text="BTC / Realized Price",
        gridcolor="rgba(148,163,184,.08)",
        zeroline=False,
        tickprefix="US$ ",
    )
    st.markdown(
        f"""
    <div class="crypto-mvrv-card">
      <div class="crypto-mvrv-head">
        <div><span class="crypto-label">Bitcoin: BTC vs Realized Price</span><b>US$ {latest_btc:,.0f}</b></div>
        <div class="crypto-mvrv-legend">
          <span class="green">BTC</span>
          <span class="yellow">Realized US$ {latest_realized:,.0f}</span>
        </div>
      </div>
    </div>
        """,
        unsafe_allow_html=True,
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


def _render_crypto_rainbow_chart(points: list[dict[str, Any]]) -> None:
    chart_points = [
        {
            "time": int(row.get("time") or 0),
            "price": float(row.get("price") or 0),
            "band_index": int(float(row.get("band_index") or 0)),
            "band_label": str(row.get("band_label") or "---"),
            **{f"band{i}": float(row.get(f"band{i}") or 0) for i in range(1, 10)},
        }
        for row in points
        if row.get("time") and row.get("price")
    ]
    if not chart_points:
        st.markdown("<div class='crypto-empty'>Sem historico Rainbow Chart disponivel agora.</div>", unsafe_allow_html=True)
        return

    import plotly.graph_objects as go

    df = pd.DataFrame(chart_points)
    df["date"] = pd.to_datetime(df["time"], unit="s", utc=True)
    latest = chart_points[-1]
    zone_label = sanitize_text(latest.get("band_label", "---"))
    zone_idx = int(latest.get("band_index") or 0)
    price = float(latest.get("price") or 0)
    band_colors = [
        "rgba(0,208,132,.16)",
        "rgba(45,255,170,.14)",
        "rgba(125,211,252,.13)",
        "rgba(255,203,107,.14)",
        "rgba(255,176,32,.15)",
        "rgba(255,138,76,.16)",
        "rgba(255,93,93,.17)",
        "rgba(190,24,93,.18)",
    ]
    fig = go.Figure()
    for idx in range(1, 9):
        lower = f"band{idx}"
        upper = f"band{idx + 1}"
        if not (df[lower] > 0).any() or not (df[upper] > 0).any():
            continue
        fig.add_trace(go.Scatter(
            x=df["date"],
            y=df[upper],
            mode="lines",
            line={"width": 0},
            showlegend=False,
            hoverinfo="skip",
        ))
        fig.add_trace(go.Scatter(
            x=df["date"],
            y=df[lower],
            mode="lines",
            line={"width": 0},
            fill="tonexty",
            fillcolor=band_colors[idx - 1],
            name=f"Banda {idx}",
            hoverinfo="skip",
        ))
    for idx, color in [(1, "#00D084"), (3, "#22D3EE"), (5, "#FFB020"), (7, "#FF5D5D"), (9, "#BE185D")]:
        col = f"band{idx}"
        if (df[col] > 0).any():
            fig.add_trace(go.Scatter(
                x=df["date"],
                y=df[col],
                mode="lines",
                name=col.upper(),
                line={"color": color, "width": 1, "dash": "dot"},
                hovertemplate="%{x|%d/%m/%Y}<br>%{y:,.0f}<extra></extra>",
            ))
    fig.add_trace(go.Scatter(
        x=df["date"],
        y=df["price"],
        mode="lines",
        name="BTC",
        line={"color": "#F8FAFC", "width": 2.2},
        hovertemplate="%{x|%d/%m/%Y}<br>BTC: US$ %{y:,.0f}<extra></extra>",
    ))
    fig.update_layout(
        height=500,
        margin={"l": 10, "r": 22, "t": 8, "b": 10},
        paper_bgcolor="#050B14",
        plot_bgcolor="#050B14",
        font={"color": "#AAB7C4"},
        legend={"orientation": "h", "y": -0.08, "x": 0.01, "font": {"size": 10}},
        xaxis={"gridcolor": "rgba(148,163,184,.08)", "zeroline": False},
        yaxis={
            "type": "log",
            "title": "BTC Rainbow",
            "gridcolor": "rgba(148,163,184,.08)",
            "zeroline": False,
            "tickprefix": "US$ ",
        },
    )
    st.markdown(
        f"""
    <div class="crypto-mvrv-card">
      <div class="crypto-mvrv-head">
        <div><span class="crypto-label">Bitcoin Rainbow Chart</span><b>US$ {price:,.0f}</b></div>
        <div class="crypto-mvrv-legend">
          <span class="green">zona {zone_idx}</span>
          <span class="yellow">{zone_label}</span>
          <span class="red">ciclo</span>
        </div>
      </div>
    </div>
        """,
        unsafe_allow_html=True,
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


def _render_crypto_mvrv_pricing_bands_chart(points: list[dict[str, Any]]) -> None:
    chart_points = [
        {
            "time": int(row.get("time") or 0),
            "btc_price": float(row.get("btc_price") or 0),
            "realized_price": float(row.get("realized_price") or 0),
            "mvrv": float(row.get("mvrv") or 0),
        }
        for row in points
        if row.get("time") and row.get("btc_price") and row.get("realized_price")
    ]
    if not chart_points:
        st.markdown("<div class='crypto-empty'>Sem dados suficientes para MVRV Pricing Bands agora.</div>", unsafe_allow_html=True)
        return

    import plotly.graph_objects as go

    df = pd.DataFrame(chart_points)
    df["date"] = pd.to_datetime(df["time"], unit="s", utc=True)
    if not (df["mvrv"] > 0).any():
        df["mvrv"] = df["btc_price"] / df["realized_price"].replace({0: pd.NA})

    valid_mvrv = df.loc[df["mvrv"] > 0, "mvrv"]
    if len(valid_mvrv) < 365:
        st.markdown("<div class='crypto-empty'>Historico insuficiente para calcular bandas estatisticas do MVRV.</div>", unsafe_allow_html=True)
        return
    mvrv_mean = float(valid_mvrv.mean())
    mvrv_std = float(valid_mvrv.std())
    if not mvrv_std or pd.isna(mvrv_std):
        st.markdown("<div class='crypto-empty'>Desvio do MVRV indisponivel para calcular bandas.</div>", unsafe_allow_html=True)
        return

    band_defs = [
        ("-1.0sd", -1.0, "#10B981", "solid"),
        ("-0.5sd", -0.5, "#34D399", "solid"),
        ("Mean", 0.0, "#F59E0B", "solid"),
        ("+0.5sd", 0.5, "#60A5FA", "solid"),
        ("+1.0sd", 1.0, "#8B5CF6", "solid"),
        ("+2.0sd", 2.0, "#FB7185", "solid"),
    ]
    for label, sigma, _color, _dash in band_defs:
        multiple = max(0.05, mvrv_mean + (mvrv_std * sigma))
        df[f"band_{label}"] = df["realized_price"] * multiple

    latest = chart_points[-1]
    latest_btc = float(latest.get("btc_price") or 0)
    latest_realized = float(latest.get("realized_price") or 0)
    latest_mvrv = float(latest.get("mvrv") or (latest_btc / latest_realized if latest_realized else 0))
    latest_sigma = (latest_mvrv - mvrv_mean) / mvrv_std if mvrv_std else 0
    if latest_sigma >= 2:
        zone, zone_cls, zone_text = "Euforia / risco alto", "red", "Preco esticado contra o realized price."
    elif latest_sigma >= 1:
        zone, zone_cls, zone_text = "Aquecido", "yellow", "Ciclo positivo, mas com assimetria menor."
    elif latest_sigma >= -0.5:
        zone, zone_cls, zone_text = "Neutro construtivo", "green", "Mercado acima do realized price."
    else:
        zone, zone_cls, zone_text = "Acumulacao / desconto", "green", "Preco abaixo ou perto do realized price."

    fig = go.Figure()
    fill_pairs = [
        ("band_-1.0sd", "band_-0.5sd", "rgba(16,185,129,.10)"),
        ("band_-0.5sd", "band_Mean", "rgba(52,211,153,.08)"),
        ("band_Mean", "band_+0.5sd", "rgba(245,158,11,.08)"),
        ("band_+0.5sd", "band_+1.0sd", "rgba(96,165,250,.09)"),
        ("band_+1.0sd", "band_+2.0sd", "rgba(244,63,94,.12)"),
    ]
    for lower, upper, fill_color in fill_pairs:
        fig.add_trace(go.Scatter(
            x=df["date"],
            y=df[upper],
            mode="lines",
            line={"width": 0},
            showlegend=False,
            hoverinfo="skip",
        ))
        fig.add_trace(go.Scatter(
            x=df["date"],
            y=df[lower],
            mode="lines",
            line={"width": 0},
            fill="tonexty",
            fillcolor=fill_color,
            showlegend=False,
            hoverinfo="skip",
        ))

    fig.add_trace(go.Scatter(
        x=df["date"],
        y=df["realized_price"],
        mode="lines",
        name="Realized Price",
        line={"color": "#22D3EE", "width": 2},
        hovertemplate="%{x|%d/%m/%Y}<br>Realized: US$ %{y:,.0f}<extra></extra>",
    ))
    for label, _sigma, color, dash in band_defs:
        fig.add_trace(go.Scatter(
            x=df["date"],
            y=df[f"band_{label}"],
            mode="lines",
            name=label,
            line={"color": color, "width": 1.5, "dash": dash},
            hovertemplate=f"%{{x|%d/%m/%Y}}<br>{label}: US$ %{{y:,.0f}}<extra></extra>",
        ))
    fig.add_trace(go.Scatter(
        x=df["date"],
        y=df["btc_price"],
        mode="lines",
        name="BTC Price",
        line={"color": "#F8FAFC", "width": 2.4},
        hovertemplate="%{x|%d/%m/%Y}<br>BTC: US$ %{y:,.0f}<extra></extra>",
    ))

    fig.update_layout(
        height=520,
        margin={"l": 10, "r": 22, "t": 8, "b": 10},
        paper_bgcolor="#050B14",
        plot_bgcolor="#050B14",
        font={"color": "#AAB7C4"},
        legend={"orientation": "h", "y": -0.08, "x": 0.01, "font": {"size": 10}},
        xaxis={"gridcolor": "rgba(148,163,184,.08)", "zeroline": False},
        yaxis={
            "type": "log",
            "title": "MVRV Pricing Bands",
            "gridcolor": "rgba(148,163,184,.08)",
            "zeroline": False,
            "tickprefix": "US$ ",
        },
    )
    st.markdown(
        f"""
    <div class="crypto-mvrv-card">
      <div class="crypto-mvrv-head">
        <div>
          <span class="crypto-label">Bitcoin MVRV Pricing Bands</span>
          <b>US$ {latest_btc:,.0f}</b>
          <small style="display:block;color:#8EA3B8;margin-top:5px;">MVRV {latest_mvrv:.2f}x | z {latest_sigma:+.2f}sd | Realized US$ {latest_realized:,.0f}</small>
        </div>
        <div class="crypto-mvrv-legend">
          <span class="{zone_cls}">{sanitize_text(zone)}</span>
          <span class="yellow">{sanitize_text(zone_text)}</span>
        </div>
      </div>
    </div>
        """,
        unsafe_allow_html=True,
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


def pagina_crypto_terminal():
    """Crypto Terminal phase 1: public crypto data, regime and source health."""
    st.title("Crypto Terminal")
    st.caption("Fase 1 adaptada: Binance, CoinGecko, Alternative.me e DefiLlama com leitura local de regime cripto.")

    col_refresh, col_note = st.columns([1, 4])
    with col_refresh:
        if st.button("Atualizar Crypto", type="primary", use_container_width=True, key="crypto_terminal_refresh"):
            st.session_state["crypto_terminal_refresh_key"] = st.session_state.get("crypto_terminal_refresh_key", 0) + 1
            load_crypto_terminal_payload.clear()
    with col_note:
        st.caption("Sem chaves privadas, sem wallet e sem dados sensiveis. Somente fontes publicas e cache local.")

    refresh_key = int(st.session_state.get("crypto_terminal_refresh_key", 0))
    try:
        payload = load_crypto_terminal_payload(refresh_key)
    except Exception as e:
        st.error(f"Nao foi possivel carregar o Crypto Terminal agora: {e}")
        return

    binance = payload.get("binance", {})
    coingecko = payload.get("coingecko", {})
    fear_greed = payload.get("fear_greed", {})
    defillama = payload.get("defillama", {})
    bgeometrics = payload.get("bgeometrics", {})
    regime = payload.get("regime", {})
    operational = payload.get("operational", {})
    binance_data = binance.get("data", {}) if isinstance(binance, dict) else {}
    coingecko_data = coingecko.get("data", {}) if isinstance(coingecko, dict) else {}
    defillama_data = defillama.get("data", {}) if isinstance(defillama, dict) else {}
    bgeometrics_data = bgeometrics.get("data", {}) if isinstance(bgeometrics, dict) else {}
    fear_data = fear_greed.get("data", {}) if isinstance(fear_greed, dict) else {}

    def _num(value, default=0.0):
        try:
            if value is None:
                return default
            return float(value)
        except Exception:
            return default

    symbols = {
        str(item.get("symbol")): item
        for item in (binance_data.get("assets") or [])
        if isinstance(item, dict) and item.get("symbol")
    }
    if not symbols:
        import time as _time
        now_ts = int(_time.time())
        for item in (coingecko_data.get("markets") or [])[:20]:
            if not isinstance(item, dict):
                continue
            symbol = str(item.get("symbol") or "").upper()
            if not symbol:
                continue
            spark = ((item.get("sparkline_in_7d") or {}).get("price") or [])[-80:]
            candles = []
            for idx, price in enumerate(spark):
                px = _num(price)
                if px <= 0:
                    continue
                candles.append({
                    "time": now_ts - (len(spark) - idx) * 3600,
                    "open": px,
                    "high": px,
                    "low": px,
                    "close": px,
                    "volume": 0,
                })
            symbols[f"{symbol}USDT"] = {
                "symbol": f"{symbol}USDT",
                "price": _num(item.get("current_price")),
                "change_pct_24h": _num(item.get("price_change_percentage_24h")),
                "quote_volume": _num(item.get("total_volume")),
                "high_24h": _num(item.get("high_24h")),
                "low_24h": _num(item.get("low_24h")),
                "funding_rate": 0,
                "candles_1h": candles,
                "trend_80h_pct": _num(item.get("price_change_percentage_7d_in_currency")),
            }
    coingecko_symbol_index = {}
    for item in coingecko_data.get("markets") or []:
        if not isinstance(item, dict):
            continue
        symbol_key = str(item.get("symbol") or "").upper()
        if symbol_key:
            coingecko_symbol_index[f"{symbol_key}USDT"] = item
    for symbol_key, asset in list(symbols.items()):
        cg_item = coingecko_symbol_index.get(symbol_key)
        if not isinstance(cg_item, dict):
            continue
        asset["change_pct_1h"] = _num(cg_item.get("price_change_percentage_1h_in_currency"), asset.get("change_pct_1h", 0))
        asset["change_pct_7d"] = _num(cg_item.get("price_change_percentage_7d_in_currency"), asset.get("change_pct_7d", asset.get("trend_80h_pct", 0)))
        asset["change_pct_30d"] = _num(cg_item.get("price_change_percentage_30d_in_currency"), asset.get("change_pct_30d", 0))
        asset["market_cap"] = _num(cg_item.get("market_cap"), asset.get("market_cap", 0))
        asset["total_volume"] = _num(cg_item.get("total_volume"), asset.get("quote_volume", 0))
    if not isinstance(operational, dict) or not operational.get("ranking"):
        try:
            from execution.crypto_signals import build_crypto_operational_dashboard
            operational = build_crypto_operational_dashboard(binance, regime, coingecko)
        except Exception:
            operational = {}

    def _money(value, decimals=2):
        value = _num(value)
        if abs(value) >= 1_000_000_000:
            return f"US$ {value / 1_000_000_000:.2f}B"
        if abs(value) >= 1_000_000:
            return f"US$ {value / 1_000_000:.2f}M"
        return f"US$ {value:,.{decimals}f}"

    def _pct(value):
        value = _num(value)
        cls = "pos" if value >= 0 else "neg"
        return f"<span class='{cls}'>{value:+.2f}%</span>"

    def _plain_pct(value):
        return f"{_num(value):+.2f}%"

    def _bias_class(value):
        text = str(value or "").lower()
        if "risk-on" in text or "comprador" in text or "momentum" in text:
            return "good"
        if "risk-off" in text or "vendedora" in text or "estresse" in text:
            return "bad"
        if "repique" in text or "correcao" in text or "alavancagem" in text:
            return "warn"
        return "neutral"

    def _score_class(value):
        score_value = _num(value)
        if score_value >= 68:
            return "good"
        if score_value <= 38:
            return "bad"
        return "warn"

    def _mvrv_zone(value):
        z = _num(value, None)
        if z is None or z == 0:
            return "Sem dado", "neutral", "Aguardando BGeometrics/cache."
        if z < 0:
            return "Acumulacao profunda", "good", "Bitcoin negociando abaixo do valor realizado ajustado."
        if z < 2:
            return "Saudavel", "good", "Zona historicamente construtiva para ciclo."
        if z < 4:
            return "Neutro / aquecendo", "warn", "Risco de ciclo ainda controlado, mas monitorar aceleracao."
        if z < 7:
            return "Risco de ciclo", "bad", "On-chain aquecido; reduzir complacencia em beta cripto."
        return "Euforia", "bad", "Zona historicamente associada a assimetria pior para novas compras."

    def _onchain_metric_zone(metric, value):
        v = _num(value, None)
        if v is None or (v == 0 and metric not in {"mvrv_z_score"}):
            return "Sem dado", "neutral", "Aguardando leitura."
        if metric == "mvrv_z_score":
            if v < 0:
                return "Acumulacao", "good", "Abaixo de 0 sugere assimetria historicamente favoravel."
            if v < 2:
                return "Saudavel", "good", "Ciclo construtivo, sem euforia on-chain."
            if v < 4:
                return "Aquecendo", "warn", "Risco ainda moderado, monitorar aceleracao."
            return "Risco", "bad", "Valuation de ciclo exige cautela."
        if metric == "mvrv":
            if v < 1:
                return "Desconto", "good", "Preco abaixo do realizado."
            if v < 1.8:
                return "Saudavel", "good", "Lucro agregado controlado."
            if v < 3:
                return "Aquecendo", "warn", "Lucro agregado ja relevante."
            return "Euforia", "bad", "Lucro agregado historicamente esticado."
        if metric == "mayer_multiple":
            if v < 0.8:
                return "Desconto", "good", "BTC abaixo da media longa; ciclo frio."
            if v < 1.7:
                return "Neutro", "neutral", "Preco dentro de faixa normal vs media longa."
            if v < 2.4:
                return "Aquecendo", "warn", "Preco acelerando contra media longa."
            return "Risco", "bad", "Zona historicamente esticada."
        if metric == "puell_multiple":
            if v < 0.6:
                return "Miner stress", "good", "Receita dos mineradores comprimida; zona de acumulacao."
            if v < 2:
                return "Normal", "neutral", "Sem excesso minerador relevante."
            if v < 3:
                return "Aquecendo", "warn", "Receita mineradora elevada."
            return "Euforia", "bad", "Excesso minerador de ciclo."
        if metric == "aviv":
            if v < 0.75:
                return "Desconto", "good", "Preco descontado vs valor ativo."
            if v < 1.5:
                return "Saudavel", "good", "Valuation ativo sem estresse."
            if v < 2:
                return "Aquecendo", "warn", "Valuation ativo em expansao."
            return "Risco", "bad", "Valuation ativo esticado."
        if metric == "fear_greed":
            if v <= 25:
                return "Medo", "warn", "Sentimento defensivo; pode indicar stress."
            if v < 55:
                return "Neutro", "neutral", "Sentimento sem excesso."
            if v < 75:
                return "Apetite", "good", "Sentimento favoravel ao risco."
            return "Ganancia", "bad", "Sentimento esticado."
        if metric in {"active_addresses", "hashrate"}:
            return "Online", "good", "Dado de rede carregado para contexto."
        return "Monitorado", "neutral", "Metrica acompanhada no painel."

    def _compact_number(value):
        v = _num(value, None)
        if v is None:
            return "---"
        if abs(v) >= 1_000_000_000:
            return f"{v / 1_000_000_000:.2f}B"
        if abs(v) >= 1_000_000:
            return f"{v / 1_000_000:.2f}M"
        if abs(v) >= 1_000:
            return f"{v / 1_000:.1f}K"
        return f"{v:.2f}"

    def _onchain_heatmap_html(data):
        metrics = [
            ("mvrv_z_score", "MVRV Z", "{:.2f}"),
            ("mvrv", "MVRV", "{:.2f}x"),
            ("mayer_multiple", "Mayer", "{:.2f}x"),
            ("puell_multiple", "Puell", "{:.2f}"),
            ("aviv", "AVIV", "{:.2f}"),
            ("fear_greed", "Fear & Greed", "{:.0f}"),
            ("active_addresses", "Enderecos ativos", None),
            ("hashrate", "Hashrate", None),
        ]
        cards = []
        for key, label, fmt in metrics:
            value = _num(data.get(key), None)
            zone, cls, text = _onchain_metric_zone(key, value)
            display = "---" if value is None else (_compact_number(value) if fmt is None else fmt.format(value))
            cards.append(
                "<div class='crypto-onchain-tile {cls}'>"
                "<div><span>{label}</span><b>{display}</b></div>"
                "<em>{zone}</em>"
                "<small>{text}</small>"
                "</div>".format(
                    cls=cls,
                    label=sanitize_text(label),
                    display=sanitize_text(display),
                    zone=sanitize_text(zone),
                    text=sanitize_text(text),
                )
            )
        return "".join(cards)

    def _rotation_cards_html(rows):
        cards = []
        for row in rows[:8]:
            bias = sanitize_text(row.get("vies", "Neutro"))
            change_cls = "pos" if _num(row.get("change_24h")) >= 0 else "neg"
            trend_cls = "pos" if _num(row.get("trend_80h")) >= 0 else "neg"
            cards.append(
                "<div class='crypto-class-card'>"
                "<div class='crypto-class-top'>"
                f"<b>{sanitize_text(row.get('classe', '---'))}</b>"
                f"<span class='crypto-chip {_bias_class(bias)}'>{bias}</span>"
                "</div>"
                f"<div class='crypto-class-score {_score_class(row.get('score'))}'>{_num(row.get('score')):.1f}</div>"
                "<div class='crypto-class-metrics'>"
                f"<span>24h <strong class='{change_cls}'>{_plain_pct(row.get('change_24h'))}</strong></span>"
                f"<span>Tend. <strong class='{trend_cls}'>{_plain_pct(row.get('trend_80h'))}</strong></span>"
                f"<span>Vol <strong>{_num(row.get('vol_realizada')):.2f}%</strong></span>"
                f"<span>Lider <strong>{sanitize_text(row.get('lider', '---'))}</strong></span>"
                "</div>"
                "</div>"
            )
        return "".join(cards)

    def _ranking_cards_html(rows, mode="leader"):
        cards = []
        for row in rows[:5]:
            score_cls = _score_class(row.get("score"))
            change_cls = "pos" if _num(row.get("change_24h")) >= 0 else "neg"
            trend_cls = "pos" if _num(row.get("trend_80h")) >= 0 else "neg"
            rel_cls = "pos" if _num(row.get("relative_to_btc_24h")) >= 0 else "neg"
            cards.append(
                "<div class='crypto-rank-card'>"
                "<div class='crypto-rank-head'>"
                "<div>"
                f"<b>{sanitize_text(row.get('symbol', '---'))}</b>"
                f"<small>{sanitize_text(row.get('subclass', '---'))} | {sanitize_text(row.get('theme', ''))}</small>"
                "</div>"
                f"<span class='crypto-rank-score {score_cls}'>{_num(row.get('score')):.0f}</span>"
                "</div>"
                f"<div class='crypto-chip {_bias_class(row.get('bias'))}'>{sanitize_text(row.get('bias', 'Neutro'))}</div>"
                "<div class='crypto-rank-grid'>"
                f"<span>24h <strong class='{change_cls}'>{_plain_pct(row.get('change_24h'))}</strong></span>"
                f"<span>Tend. <strong class='{trend_cls}'>{_plain_pct(row.get('trend_80h'))}</strong></span>"
                f"<span>vs BTC <strong class='{rel_cls}'>{_plain_pct(row.get('relative_to_btc_24h'))}</strong></span>"
                f"<span>Funding <strong>{_num(row.get('funding_annual_pct')):.2f}%</strong></span>"
                "</div>"
                "</div>"
            )
        empty = "Sem ranking de lideres agora." if mode == "leader" else "Sem ranking defensivo agora."
        return "".join(cards) or f"<div class='crypto-empty'>{empty}</div>"

    def _operational_group_html(groups):
        specs = [
            ("buy_strength", "Comprar forca", "Tendencia + score forte, sem funding excessivo.", "good"),
            ("alt_vs_btc", "Alts contra BTC", "Ativos batendo Bitcoin em forca relativa.", "good"),
            ("pullback_watch", "Monitorar pullback", "Tendencia viva, mas sem compra perseguida.", "warn"),
            ("avoid_defensive", "Evitar / defensivos", "Baixo score ou perda de forca contra BTC.", "bad"),
            ("leverage_risk", "Risco de alavancagem", "Funding/faixa esticados; cuidado com squeeze.", "warn"),
        ]
        panels = []
        for key, title, subtitle, cls in specs:
            rows = groups.get(key, []) if isinstance(groups, dict) else []
            panels.append(
                f"<div class='crypto-rank-panel crypto-op-panel {cls}'>"
                f"<h5>{sanitize_text(title)}</h5>"
                f"<p>{sanitize_text(subtitle)}</p>"
                f"<div class='crypto-rank-list'>{_ranking_cards_html(rows, key)}</div>"
                "</div>"
            )
        return "".join(panels)

    def _relative_btc_strength(symbols_map):
        btc_asset = symbols_map.get("BTCUSDT") or {}
        btc_price = _num(btc_asset.get("price"), None)
        if not btc_price:
            return [], "Aguardando BTC", "Sem preco de BTC para comparar as alts.", "neutral"

        def _ratio_change(asset, window):
            btc_change = _num(btc_asset.get(window), 0)
            alt_change = _num(asset.get(window), 0)
            btc_factor = 1 + btc_change / 100
            alt_factor = 1 + alt_change / 100
            if alt_factor <= 0:
                return 0.0
            return (btc_factor / alt_factor - 1) * 100

        rows = []
        for alt_symbol in ["ETHUSDT", "SOLUSDT", "AAVEUSDT", "LINKUSDT"]:
            asset = symbols_map.get(alt_symbol) or {}
            alt_price = _num(asset.get("price"), None)
            if not alt_price:
                continue
            ratio = btc_price / alt_price if alt_price else 0
            rel_1h = _num(asset.get("change_pct_1h"), 0) - _num(btc_asset.get("change_pct_1h"), 0)
            rel_24h = _num(asset.get("change_pct_24h"), 0) - _num(btc_asset.get("change_pct_24h"), 0)
            rel_7d = _num(asset.get("change_pct_7d"), 0) - _num(btc_asset.get("change_pct_7d"), 0)
            weighted_rel = rel_1h * 0.25 + rel_24h * 0.45 + rel_7d * 0.30
            score = max(0, min(100, 50 + weighted_rel * 3.0))
            ratio_24h = _ratio_change(asset, "change_pct_24h")
            if score >= 57:
                status, cls, winner = "ALT > BTC", "good", alt_symbol.replace("USDT", "")
            elif score <= 43:
                status, cls, winner = "BTC > ALT", "bad", "BTC"
            else:
                status, cls, winner = "Neutro", "neutral", "Misto"
            rows.append({
                "asset": alt_symbol.replace("USDT", ""),
                "ratio": ratio,
                "score": score,
                "status": status,
                "cls": cls,
                "winner": winner,
                "rel_1h": rel_1h,
                "rel_24h": rel_24h,
                "rel_7d": rel_7d,
                "ratio_24h": ratio_24h,
            })
        rows.sort(key=lambda row: row["score"], reverse=True)
        alt_wins = sum(1 for row in rows if row["score"] >= 57)
        btc_wins = sum(1 for row in rows if row["score"] <= 43)
        if alt_wins >= 3:
            title, text, cls = "Alts ganhando do BTC", "Rotacao pro-risco: alts monitoradas batem BTC em forca relativa.", "good"
        elif btc_wins >= 3:
            title, text, cls = "BTC dominante", "Fluxo defensivo dentro de cripto: BTC performa melhor que as alts monitoradas.", "bad"
        elif rows:
            title, text, cls = "Rotacao seletiva", "Forca mista: escolher apenas alts que batem BTC no ranking.", "warn"
        else:
            title, text, cls = "Aguardando alts", "Sem dados suficientes para comparar BTC contra ETH, SOL, AAVE e LINK.", "neutral"
        return rows, title, text, cls

    def _btc_strength_html(rows, title, text, cls):
        logo_meta = {
            "ETH": ("#627EEA", "ETH"),
            "SOL": ("#14F195", "SOL"),
            "AAVE": ("#B6509E", "AAVE"),
            "LINK": ("#2A5ADA", "LINK"),
        }
        if not rows:
            table = "<div class='crypto-empty'>Sem dados suficientes para forca BTC/Alts agora.</div>"
            leader = "---"
            weakest = "---"
        else:
            leader = rows[0]["asset"]
            weakest = rows[-1]["asset"]
            body = []
            for row in rows:
                rel_cls = "pos" if row["rel_24h"] >= 0 else "neg"
                ratio_cls = "neg" if row["ratio_24h"] >= 0 else "pos"
                logo_color, logo_text = logo_meta.get(row["asset"], ("#60A5FA", row["asset"][:4]))
                score_width = max(0, min(100, row["score"]))
                score_cls = sanitize_text(row["cls"])
                body.append(
                    f"<div class='crypto-btc-strength-token {score_cls}'>"
                    "<div class='crypto-btc-token-head'>"
                    f"<span class='crypto-token-logo' style='background:linear-gradient(135deg,{logo_color},#0B1220);'>{sanitize_text(logo_text)}</span>"
                    "<div>"
                    f"<b>{sanitize_text(row['asset'])}</b>"
                    f"<small>{sanitize_text(row['status'])}</small>"
                    "</div>"
                    f"<em>{row['score']:.0f}</em>"
                    "</div>"
                    "<div class='crypto-btc-scorebar'>"
                    f"<div class='{score_cls}' style='width:{score_width:.0f}%;'></div>"
                    "</div>"
                    "<div class='crypto-btc-token-grid'>"
                    f"<span>BTC/{sanitize_text(row['asset'])}<b>{row['ratio']:.4f}</b></span>"
                    f"<span>1h<b class='{rel_cls}'>{row['rel_1h']:+.2f}%</b></span>"
                    f"<span>24h<b class='{rel_cls}'>{row['rel_24h']:+.2f}%</b></span>"
                    f"<span>7d<b class='{rel_cls}'>{row['rel_7d']:+.2f}%</b></span>"
                    f"<span>Ratio 24h<b class='{ratio_cls}'>{row['ratio_24h']:+.2f}%</b></span>"
                    "</div>"
                    "</div>"
                )
            table = "".join(body)
        return (
            "<div class='crypto-btc-strength'>"
            f"<div class='crypto-btc-strength-head {sanitize_text(cls)}'>"
            "<div>"
            "<span class='crypto-label'>Forca Bitcoin vs Alts</span>"
            f"<strong>{sanitize_text(title)}</strong>"
            f"<p>{sanitize_text(text)}</p>"
            "</div>"
            "<div class='crypto-btc-strength-kpis'>"
            f"<span>Lider contra BTC <b>{sanitize_text(leader)}</b></span>"
            f"<span>Mais fraca vs BTC <b>{sanitize_text(weakest)}</b></span>"
            "</div>"
            "</div>"
            "<div class='crypto-btc-strength-grid'>"
            f"{table}"
            "</div>"
            "</div>"
        )

    btc = symbols.get("BTCUSDT", {})
    eth = symbols.get("ETHUSDT", {})
    btc_strength_rows, btc_strength_title, btc_strength_text, btc_strength_cls = _relative_btc_strength(symbols)
    btc_strength_html = _btc_strength_html(btc_strength_rows, btc_strength_title, btc_strength_text, btc_strength_cls)
    regime_name = sanitize_text(regime.get("regime", "Neutro"))
    regime_bias = sanitize_text(regime.get("bias", "Neutro"))
    score = _num(regime.get("score"), 50)
    regime_confidence = _num(regime.get("confidence"), 0)
    regime_summary = sanitize_text(regime.get("summary", "Leitura local indisponivel."))
    allocation = regime.get("allocation") if isinstance(regime.get("allocation"), dict) else {}
    allocation_action = sanitize_text(allocation.get("action", "NEUTRO"))
    allocation_bias = sanitize_text(allocation.get("bias", "Aguardar assimetria"))
    allocation_reason = sanitize_text(allocation.get("reason", "sem assimetria clara de ciclo"))
    allocation_risk_note = sanitize_text(allocation.get("risk_note", "sem risco extremo dominante"))
    allocation_condition = sanitize_text(allocation.get("condition", "Aguardar novo sinal de ciclo."))
    allocation_score = _num(allocation.get("score"), 50)
    allocation_crypto_pct = _num(allocation.get("crypto_pct"), 45)
    allocation_usdt_pct = _num(allocation.get("usdt_pct"), 55)
    allocation_btc_pct = _num(allocation.get("btc_pct"), 25)
    allocation_eth_pct = _num(allocation.get("eth_pct"), 10)
    allocation_alts_pct = _num(allocation.get("alts_pct"), 10)
    allocation_action_cls = "good" if "ACUMULAR" in allocation_action else "bad" if "CAIXA" in allocation_action or "REALIZAR" in allocation_action else "warn"
    positive_drivers = [sanitize_text(item) for item in (regime.get("drivers_positive") or [])[:3]]
    negative_drivers = [sanitize_text(item) for item in (regime.get("drivers_negative") or regime.get("alerts") or [])[:3]]
    positive_drivers_html = "".join(f"<li>{item}</li>" for item in positive_drivers) or "<li>Sem driver positivo forte.</li>"
    negative_drivers_html = "".join(f"<li>{item}</li>" for item in negative_drivers) or "<li>Sem alerta critico agora.</li>"
    fng_value = _num((fear_data.get("current") or {}).get("value"), 0)
    fng_label = sanitize_text((fear_data.get("current") or {}).get("classification", "---"))
    fng_gauge_value = max(0, min(100, fng_value))
    fng_bar_width = f"{fng_gauge_value:.0f}%"
    if fng_gauge_value <= 25:
        fng_gauge_cls = "fear"
        fng_gauge_text = "Medo"
    elif fng_gauge_value < 55:
        fng_gauge_cls = "neutral"
        fng_gauge_text = "Neutro"
    elif fng_gauge_value < 75:
        fng_gauge_cls = "greed"
        fng_gauge_text = "Apetite"
    else:
        fng_gauge_cls = "extreme"
        fng_gauge_text = "Ganancia"
    fng_badge_color = {
        "fear": "#FF8A8A",
        "neutral": "#FFCB6B",
        "greed": "#2DFFAA",
        "extreme": "#FF8A8A",
    }.get(fng_gauge_cls, "#CBD5E1")
    global_data = coingecko_data
    market_cap = _num(global_data.get("total_market_cap_usd"))
    volume_24h = _num(global_data.get("total_volume_usd"))
    btc_dom = _num(global_data.get("btc_dominance"))
    eth_dom = _num(global_data.get("eth_dominance"))
    tvl = _num(defillama_data.get("total_tvl_usd"))
    stable_cap = _num(defillama_data.get("stablecoin_market_cap_usd"))
    stablecoin_rows = defillama_data.get("stablecoins") if isinstance(defillama_data, dict) else []
    usdt_cap = 0.0
    for stable in stablecoin_rows or []:
        if not isinstance(stable, dict):
            continue
        stable_symbol = str(stable.get("symbol") or stable.get("gecko_id") or stable.get("name") or "").upper()
        stable_name = str(stable.get("name") or "").upper()
        if stable_symbol == "USDT" or "TETHER" in stable_name:
            usdt_cap = _num((stable.get("circulating") or {}).get("peggedUSD"), 0)
            break
    usdt_dom = (usdt_cap / market_cap * 100) if market_cap and usdt_cap else None
    usdt_flow = update_usdt_dominance_flow(usdt_dom)
    usdt_delta = usdt_flow.get("delta")
    usdt_delta_text = "---" if usdt_delta is None else f"{usdt_delta:+.3f} p.p."
    usdt_high_15d = usdt_flow.get("high_15d")
    usdt_distance_high = usdt_flow.get("distance_from_high")
    usdt_high_15d_text = "---" if usdt_high_15d is None else f"{float(usdt_high_15d):.2f}%"
    usdt_distance_high_text = "---" if usdt_distance_high is None else f"{float(usdt_distance_high):+.3f} p.p."
    mvrv_z = _num(bgeometrics_data.get("mvrv_z_score"), None)
    mvrv_ratio = _num(bgeometrics_data.get("mvrv"), None)
    mvrv_zone, mvrv_cls, mvrv_text = _mvrv_zone(mvrv_z)
    mvrv_status = sanitize_text(bgeometrics.get("status", "---") if isinstance(bgeometrics, dict) else "---")
    mvrv_date = sanitize_text(bgeometrics_data.get("date", "---"))
    mayer_multiple = _num(bgeometrics_data.get("mayer_multiple"), None)
    puell_multiple = _num(bgeometrics_data.get("puell_multiple"), None)
    aviv_value = _num(bgeometrics_data.get("aviv"), None)
    onchain_fear = _num(bgeometrics_data.get("fear_greed"), None)
    cycle_score_parts = []
    for metric_key in ("mvrv_z_score", "mvrv", "mayer_multiple", "puell_multiple", "aviv", "fear_greed"):
        _, cls, _ = _onchain_metric_zone(metric_key, bgeometrics_data.get(metric_key))
        if cls == "good":
            cycle_score_parts.append(1)
        elif cls == "bad":
            cycle_score_parts.append(-1)
        elif cls == "warn":
            cycle_score_parts.append(0)
    cycle_score = sum(cycle_score_parts)
    if cycle_score >= 3:
        cycle_regime, cycle_cls, cycle_text = "Acumulacao / saudavel", "good", "Valuation on-chain favorece paciencia compradora e menor risco de euforia."
    elif cycle_score <= -2:
        cycle_regime, cycle_cls, cycle_text = "Risco de ciclo", "bad", "Leitura on-chain exige cautela com beta cripto e alavancagem."
    elif cycle_score <= 0:
        cycle_regime, cycle_cls, cycle_text = "Neutro / aquecendo", "warn", "Ciclo sem capitulacao, mas tambem sem grande desconto agregado."
    else:
        cycle_regime, cycle_cls, cycle_text = "Construtivo", "good", "On-chain ainda construtivo, com risco de ciclo controlado."

    today_brt = datetime.now(ZoneInfo("America/Sao_Paulo")).date()
    last_halving = datetime(2024, 4, 20).date()
    next_halving = datetime(2028, 4, 15).date()
    days_since_halving = max(0, (today_brt - last_halving).days)
    days_to_next_halving = max(0, (next_halving - today_brt).days)
    cycle_progress = min(100, max(0, days_since_halving / max(1, (next_halving - last_halving).days) * 100))
    btc_change_24h = _num(btc.get("change_pct_24h"), 0)
    btc_trend_80h = _num(btc.get("trend_80h_pct"), 0)
    stable_dom = (stable_cap / market_cap * 100) if market_cap and stable_cap else None
    macro_liquidity_score = 0
    macro_liquidity_score += 1 if btc_change_24h >= 0 else -1
    macro_liquidity_score += 1 if btc_trend_80h >= 0 else -1
    if stable_dom is not None:
        macro_liquidity_score += -1 if stable_dom > 9 else 1
    if fng_gauge_value >= 75:
        macro_liquidity_score -= 1
    elif fng_gauge_value <= 25:
        macro_liquidity_score -= 1
    elif fng_gauge_value >= 55:
        macro_liquidity_score += 1
    if macro_liquidity_score >= 2:
        macro_liquidity, macro_cls = "Liquidez favoravel", "good"
        macro_text = "Preco e sentimento ainda sustentam apetite por risco cripto."
    elif macro_liquidity_score <= -2:
        macro_liquidity, macro_cls = "Liquidez restritiva", "bad"
        macro_text = "Fluxo curto sugere defesa; evitar alavancagem e compras emocionais."
    else:
        macro_liquidity, macro_cls = "Liquidez neutra", "warn"
        macro_text = "Confirmar com DXY, juros e Nasdaq antes de aumentar risco."

    if mvrv_z is not None and mvrv_z >= 4:
        btc_cycle_phase, btc_cycle_cls = "Distribuicao / euforia", "bad"
        btc_cycle_read = "On-chain em zona quente; foco em protecao de lucro."
    elif mvrv_z is not None and mvrv_z <= 0:
        btc_cycle_phase, btc_cycle_cls = "Capitulacao / acumulacao", "good"
        btc_cycle_read = "Valuation frio; historicamente boa assimetria para longo prazo."
    elif days_since_halving < 180:
        btc_cycle_phase, btc_cycle_cls = "Pos-halving inicial", "neutral"
        btc_cycle_read = "Oferta nova menor, mas ainda precisa de liquidez para tracao."
    elif days_since_halving < 540:
        btc_cycle_phase, btc_cycle_cls = "Expansao de ciclo", "good"
        btc_cycle_read = "Janela historicamente favoravel se liquidez e on-chain confirmarem."
    elif days_since_halving < 900:
        btc_cycle_phase, btc_cycle_cls = "Maturacao / pos-pico", "warn"
        btc_cycle_read = "Ciclo avancado; priorizar confirmacao por MVRV, DXY e fluxo."
    else:
        btc_cycle_phase, btc_cycle_cls = "Late-cycle / pre-halving", "warn"
        btc_cycle_read = "Aproximacao do proximo halving; buscar sinais de acumulacao."

    if btc_cycle_cls == "good" and macro_cls == "good":
        cycle_operational_bias, cycle_operational_cls = "Risk-on seletivo", "good"
    elif btc_cycle_cls == "bad" or macro_cls == "bad":
        cycle_operational_bias, cycle_operational_cls = "Defensivo", "bad"
    else:
        cycle_operational_bias, cycle_operational_cls = "Neutro com confirmacao", "warn"

    st.markdown(
        f"""
        <style>
          .crypto-hero {{display:grid; grid-template-columns:1.15fr .85fr; gap:14px; margin:12px 0 16px;}}
          .crypto-card {{background:#07111F; border:1px solid #1F334A; border-radius:8px; padding:14px; color:#E5E7EB;}}
          .crypto-card, .crypto-kpi, .crypto-asset, .crypto-rot-card, .crypto-class-card, .crypto-rank-card {{box-shadow:0 10px 28px rgba(0,0,0,.18);}}
          .crypto-regime {{border-left:5px solid #22D3EE; background:linear-gradient(135deg,#06101E 0%,#0B1727 58%,#0F2437 100%);}}
          .crypto-label {{display:block; color:#8FB6E8; font-size:.72rem; font-weight:900; text-transform:uppercase; letter-spacing:.02em;}}
          .crypto-regime strong {{display:block; color:#F8FAFC; font-size:2.35rem; line-height:1.05; margin-top:5px;}}
          .crypto-bias {{display:inline-flex; margin-top:9px; padding:5px 9px; border-radius:999px; background:#0B2440; color:#7DD3FC; font-weight:900;}}
          .crypto-regime-top {{display:flex; align-items:flex-start; justify-content:space-between; gap:14px;}}
          .crypto-regime-scorebox {{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:8px; min-width:210px;}}
          .crypto-regime-scorebox span {{display:block; background:#08111F; border:1px solid #203047; border-radius:8px; padding:8px; color:#8FA4BD; font-size:.66rem; font-weight:900; text-transform:uppercase;}}
          .crypto-regime-scorebox b {{display:block; color:#F8FAFC; font-size:1.15rem; margin-top:3px;}}
          .crypto-regime-summary {{margin:12px 0 0; color:#CBD5E1; font-weight:850; line-height:1.35;}}
          .crypto-regime-drivers {{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:10px; margin-top:12px;}}
          .crypto-regime-driver {{background:#08111F; border:1px solid #203047; border-radius:8px; padding:10px;}}
          .crypto-regime-driver.good {{border-color:rgba(0,208,132,.30);}}
          .crypto-regime-driver.bad {{border-color:rgba(255,75,75,.35);}}
          .crypto-regime-driver span {{display:block; color:#8FB6E8; font-size:.66rem; font-weight:950; text-transform:uppercase; margin-bottom:6px;}}
          .crypto-regime-driver ul {{margin:0; padding-left:16px; color:#CBD5E1; font-size:.76rem; line-height:1.35; font-weight:760;}}
          .crypto-regime-driver li {{margin-bottom:4px;}}
          .crypto-allocation-card {{margin-top:12px; background:linear-gradient(135deg,#08111F,#050B14); border:1px solid #203047; border-radius:8px; padding:12px;}}
          .crypto-allocation-top {{display:flex; align-items:flex-start; justify-content:space-between; gap:12px;}}
          .crypto-allocation-action {{display:inline-flex; border-radius:999px; padding:6px 10px; font-size:.76rem; font-weight:950; text-transform:uppercase; border:1px solid rgba(148,163,184,.25);}}
          .crypto-allocation-action.good {{background:rgba(0,208,132,.14); color:#2DFFAA; border-color:rgba(0,208,132,.40);}}
          .crypto-allocation-action.warn {{background:rgba(255,176,32,.14); color:#FFCB6B; border-color:rgba(255,176,32,.40);}}
          .crypto-allocation-action.bad {{background:rgba(255,75,75,.14); color:#FF8A8A; border-color:rgba(255,75,75,.40);}}
          .crypto-allocation-card strong {{display:block; color:#F8FAFC; font-size:1.35rem; margin-top:6px;}}
          .crypto-allocation-card p {{margin:8px 0 0; color:#CBD5E1; font-weight:800; line-height:1.35;}}
          .crypto-allocation-score {{text-align:right; color:#8FA4BD; font-size:.66rem; font-weight:900; text-transform:uppercase;}}
          .crypto-allocation-score b {{display:block; color:#F8FAFC; font-size:1.45rem; margin-top:3px;}}
          .crypto-allocation-meter {{height:12px; display:flex; overflow:hidden; border-radius:999px; background:#132338; margin:12px 0 8px; border:1px solid rgba(148,163,184,.18);}}
          .crypto-allocation-meter span {{display:block; height:100%;}}
          .crypto-allocation-meter .btc {{background:#F7931A; width:{allocation_btc_pct:.0f}%;}}
          .crypto-allocation-meter .eth {{background:#627EEA; width:{allocation_eth_pct:.0f}%;}}
          .crypto-allocation-meter .alts {{background:#14F195; width:{allocation_alts_pct:.0f}%;}}
          .crypto-allocation-meter .usdt {{background:#26A17B; width:{allocation_usdt_pct:.0f}%;}}
          .crypto-allocation-split {{display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:7px;}}
          .crypto-allocation-split span {{background:#07111F; border:1px solid #18283D; border-radius:7px; padding:7px; color:#8FA4BD; font-size:.64rem; font-weight:900; text-transform:uppercase;}}
          .crypto-allocation-split b {{display:block; color:#F8FAFC; font-size:.88rem; margin-top:3px;}}
          .crypto-score {{font-size:2.15rem; color:#FFB020; font-weight:950; text-align:right;}}
          .crypto-risk-card {{display:block;}}
          .crypto-fng-panel {{background:#08111F; border:1px solid #203047; border-radius:8px; padding:12px; min-height:118px;}}
          .crypto-fng-top {{display:flex; align-items:flex-end; justify-content:space-between; gap:10px; margin-top:6px;}}
          .crypto-fng-value {{color:#F8FAFC; font-size:2rem; line-height:1; font-weight:950;}}
          .crypto-fng-badge {{border-radius:999px; padding:4px 9px; font-size:.68rem; font-weight:950; text-transform:uppercase;}}
          .crypto-fng-badge.fear {{color:#FF8A8A; background:rgba(255,75,75,.15); border:1px solid rgba(255,75,75,.35);}}
          .crypto-fng-badge.neutral {{color:#FFCB6B; background:rgba(255,176,32,.14); border:1px solid rgba(255,176,32,.35);}}
          .crypto-fng-badge.greed {{color:#2DFFAA; background:rgba(0,208,132,.14); border:1px solid rgba(0,208,132,.35);}}
          .crypto-fng-badge.extreme {{color:#FF8A8A; background:rgba(255,75,75,.15); border:1px solid rgba(255,75,75,.35);}}
          .crypto-pressure-track {{height:18px; margin-top:14px; border-radius:999px; background:linear-gradient(90deg,#FF4B4B 0%,#FF4B4B 25%,#FFB020 25%,#FFB020 55%,#00D084 55%,#00D084 75%,#FF5D5D 75%,#FF5D5D 100%); padding:3px; box-shadow:inset 0 0 0 1px rgba(255,255,255,.08);}}
          .crypto-pressure-fill {{height:12px; width:{fng_bar_width}; border-radius:999px; background:rgba(248,250,252,.88); box-shadow:0 0 12px rgba(248,250,252,.35);}}
          .crypto-pressure-scale {{display:flex; justify-content:space-between; color:#64748B; font-size:.62rem; font-weight:900; margin-top:6px;}}
          .crypto-grid {{display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px; margin-bottom:16px;}}
          .crypto-kpi {{background:#0B1220; border:1px solid #1E293B; border-radius:8px; padding:12px; min-height:86px;}}
          .crypto-kpi span {{display:block; color:#94A3B8; font-size:.7rem; font-weight:900; text-transform:uppercase;}}
          .crypto-kpi strong {{display:block; color:#F8FAFC; font-size:1.22rem; margin-top:5px;}}
          .crypto-kpi em {{display:block; color:#94A3B8; font-size:.72rem; font-style:normal; margin-top:4px;}}
          .crypto-cycle-dashboard {{display:grid; grid-template-columns:1.1fr .85fr .9fr .9fr .9fr; gap:10px; margin:-2px 0 16px;}}
          .crypto-cycle-box {{background:#07111F; border:1px solid #1F334A; border-radius:8px; padding:13px; min-height:128px;}}
          .crypto-cycle-box.good {{border-left:5px solid #00D084; background:linear-gradient(180deg,rgba(0,208,132,.08),#07111F);}}
          .crypto-cycle-box.warn {{border-left:5px solid #FFB020; background:linear-gradient(180deg,rgba(255,176,32,.08),#07111F);}}
          .crypto-cycle-box.bad {{border-left:5px solid #FF4B4B; background:linear-gradient(180deg,rgba(255,75,75,.09),#07111F);}}
          .crypto-cycle-box.neutral {{border-left:5px solid #60A5FA;}}
          .crypto-cycle-box b {{display:block; color:#F8FAFC; font-size:1.3rem; line-height:1.05; margin-top:5px;}}
          .crypto-cycle-box p {{margin:8px 0 0; color:#CBD5E1; font-weight:800; font-size:.82rem; line-height:1.35;}}
          .crypto-cycle-meter {{height:10px; border-radius:999px; background:#132338; overflow:hidden; margin-top:12px; border:1px solid rgba(148,163,184,.18);}}
          .crypto-cycle-meter div {{height:100%; width:{cycle_progress:.0f}%; background:linear-gradient(90deg,#22D3EE,#FFB020,#FF4B4B); border-radius:999px;}}
          .crypto-cycle-mini {{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:8px; margin-top:10px;}}
          .crypto-cycle-mini span {{display:block; color:#8FA4BD; font-size:.66rem; font-weight:900; text-transform:uppercase;}}
          .crypto-cycle-mini strong {{display:block; color:#F8FAFC; font-size:.9rem;}}
          .crypto-btc-strength {{background:linear-gradient(135deg,#06101E,#0B1727); border:1px solid #28405E; border-radius:8px; padding:12px; margin:-2px 0 16px;}}
          .crypto-btc-strength-head {{display:flex; align-items:flex-start; justify-content:space-between; gap:14px; border-left:5px solid #60A5FA; padding-left:12px; margin-bottom:10px;}}
          .crypto-btc-strength-head.good {{border-left-color:#00D084;}}
          .crypto-btc-strength-head.warn {{border-left-color:#FFB020;}}
          .crypto-btc-strength-head.bad {{border-left-color:#FF4B4B;}}
          .crypto-btc-strength-head strong {{display:block; color:#F8FAFC; font-size:1.55rem; line-height:1.05; margin-top:4px;}}
          .crypto-btc-strength-head p {{margin:7px 0 0; color:#CBD5E1; font-weight:800;}}
          .crypto-btc-strength-kpis {{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:8px; min-width:310px;}}
          .crypto-btc-strength-kpis span {{background:#08111F; border:1px solid #203047; border-radius:8px; padding:8px; color:#8FA4BD; font-size:.68rem; font-weight:900; text-transform:uppercase;}}
          .crypto-btc-strength-kpis b {{display:block; color:#F8FAFC; font-size:1.05rem; margin-top:3px;}}
          .crypto-btc-strength-grid {{display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px;}}
          .crypto-btc-strength-token {{background:radial-gradient(circle at top left,rgba(96,165,250,.16),#08111F 44%,#050B14); border:1px solid #203047; border-radius:8px; padding:11px; min-height:178px;}}
          .crypto-btc-strength-token.good {{border-color:rgba(0,208,132,.45); box-shadow:inset 0 0 0 1px rgba(0,208,132,.08);}}
          .crypto-btc-strength-token.warn {{border-color:rgba(255,176,32,.45); box-shadow:inset 0 0 0 1px rgba(255,176,32,.08);}}
          .crypto-btc-strength-token.bad {{border-color:rgba(255,75,75,.48); box-shadow:inset 0 0 0 1px rgba(255,75,75,.08);}}
          .crypto-btc-token-head {{display:flex; align-items:center; gap:9px;}}
          .crypto-token-logo {{display:flex; align-items:center; justify-content:center; width:38px; height:38px; border-radius:999px; color:#FFFFFF; font-size:.62rem; font-weight:950; border:1px solid rgba(255,255,255,.18); box-shadow:0 0 18px rgba(34,211,238,.18);}}
          .crypto-btc-token-head b {{display:block; color:#F8FAFC; font-size:1.05rem; line-height:1;}}
          .crypto-btc-token-head small {{display:block; color:#8FA4BD; font-size:.66rem; font-weight:900; margin-top:4px; text-transform:uppercase;}}
          .crypto-btc-token-head em {{margin-left:auto; color:#F8FAFC; font-style:normal; font-size:1.45rem; font-weight:950;}}
          .crypto-btc-scorebar {{height:8px; border-radius:999px; background:#132338; overflow:hidden; margin:11px 0; border:1px solid rgba(148,163,184,.18);}}
          .crypto-btc-scorebar div {{height:100%; border-radius:999px;}}
          .crypto-btc-scorebar div.good {{background:linear-gradient(90deg,#0EA5E9,#00D084);}}
          .crypto-btc-scorebar div.warn {{background:linear-gradient(90deg,#64748B,#FFB020);}}
          .crypto-btc-scorebar div.bad {{background:linear-gradient(90deg,#FFB020,#FF4B4B);}}
          .crypto-btc-token-grid {{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:7px;}}
          .crypto-btc-token-grid span {{background:#07111F; border:1px solid #18283D; border-radius:7px; padding:7px; color:#8FA4BD; font-size:.66rem; font-weight:900; text-transform:uppercase;}}
          .crypto-btc-token-grid b {{display:block; color:#F8FAFC; font-size:.82rem; margin-top:3px; text-transform:none;}}
          .crypto-onchain {{display:grid; grid-template-columns:.85fr 1.15fr; gap:12px; margin:-2px 0 16px;}}
          .crypto-onchain-main {{background:linear-gradient(135deg,#07111F,#101728); border:1px solid #28405E; border-left:5px solid #22D3EE; border-radius:8px; padding:13px;}}
          .crypto-onchain-main strong {{display:block; color:#F8FAFC; font-size:2rem; line-height:1; margin-top:5px;}}
          .crypto-onchain-main p {{margin:8px 0 0; color:#CBD5E1; font-weight:800;}}
          .crypto-onchain-grid {{display:grid; grid-template-columns:repeat(5,minmax(0,1fr)); gap:8px;}}
          .crypto-onchain-grid div {{background:#08111F; border:1px solid #203047; border-radius:8px; padding:10px;}}
          .crypto-onchain-grid span {{display:block; color:#94A3B8; font-size:.68rem; font-weight:900; text-transform:uppercase;}}
          .crypto-onchain-grid b {{display:block; color:#F8FAFC; font-size:1rem; margin-top:4px;}}
          .crypto-onchain-cycle {{display:grid; grid-template-columns:.78fr 1.22fr; gap:12px; margin:-6px 0 18px;}}
          .crypto-cycle-card {{background:linear-gradient(135deg,#07111F,#0B1727); border:1px solid #28405E; border-radius:8px; padding:13px;}}
          .crypto-cycle-card.good {{border-left:5px solid #00D084;}}
          .crypto-cycle-card.warn {{border-left:5px solid #FFB020;}}
          .crypto-cycle-card.bad {{border-left:5px solid #FF4B4B;}}
          .crypto-cycle-card strong {{display:block; color:#F8FAFC; font-size:1.45rem; margin-top:5px;}}
          .crypto-cycle-card p {{color:#CBD5E1; font-weight:800; margin:8px 0 0;}}
          .crypto-onchain-heatmap {{display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:8px;}}
          .crypto-onchain-tile {{background:#08111F; border:1px solid #203047; border-radius:8px; padding:10px; min-height:96px;}}
          .crypto-onchain-tile.good {{border-color:rgba(0,208,132,.45); background:linear-gradient(180deg,rgba(0,208,132,.10),#08111F);}}
          .crypto-onchain-tile.warn {{border-color:rgba(255,176,32,.45); background:linear-gradient(180deg,rgba(255,176,32,.10),#08111F);}}
          .crypto-onchain-tile.bad {{border-color:rgba(255,75,75,.50); background:linear-gradient(180deg,rgba(255,75,75,.11),#08111F);}}
          .crypto-onchain-tile span {{display:block; color:#94A3B8; font-size:.66rem; font-weight:900; text-transform:uppercase;}}
          .crypto-onchain-tile b {{display:block; color:#F8FAFC; font-size:1.18rem; margin-top:3px;}}
          .crypto-onchain-tile em {{display:inline-flex; margin-top:7px; border-radius:999px; padding:2px 7px; color:#E5E7EB; background:rgba(148,163,184,.12); font-size:.66rem; font-style:normal; font-weight:950; text-transform:uppercase;}}
          .crypto-onchain-tile small {{display:block; color:#AAB7C4; font-size:.68rem; line-height:1.25; margin-top:7px;}}
          .crypto-mvrv-card {{background:#050B14; border:1px solid #1F334A; border-radius:8px; padding:12px; margin:-4px 0 18px;}}
          .crypto-mvrv-head {{display:flex; align-items:flex-start; justify-content:space-between; gap:12px; margin-bottom:8px;}}
          .crypto-mvrv-head b {{display:block; color:#F8FAFC; font-size:1.45rem; line-height:1; margin-top:4px;}}
          .crypto-mvrv-legend {{display:flex; flex-wrap:wrap; gap:6px; justify-content:flex-end;}}
          .crypto-mvrv-legend span {{border-radius:999px; padding:3px 8px; font-size:.68rem; font-weight:900; border:1px solid rgba(148,163,184,.25);}}
          .crypto-mvrv-legend .green {{color:#2DFFAA; background:rgba(0,208,132,.10);}}
          .crypto-mvrv-legend .yellow {{color:#FFCB6B; background:rgba(255,176,32,.10);}}
          .crypto-mvrv-legend .red {{color:#FF8A8A; background:rgba(255,75,75,.10);}}
          .crypto-mvrv-chart {{height:320px; width:100%;}}
          .crypto-assets {{display:grid; grid-template-columns:repeat(5,minmax(0,1fr)); gap:10px; margin-bottom:16px;}}
          .crypto-asset {{background:linear-gradient(180deg,#08111F,#050B14); border:1px solid #203047; border-radius:8px; padding:11px;}}
          .crypto-asset b {{display:block; color:#F8FAFC; font-size:1rem;}}
          .crypto-asset small {{display:block; color:#94A3B8; margin-top:3px;}}
          .crypto-alerts {{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:10px; margin:14px 0;}}
          .crypto-alert {{background:#1F1117; border:1px solid #7F1D1D; border-left:4px solid #FF4B4B; border-radius:8px; color:#FECACA; padding:10px 12px; font-weight:800;}}
          .crypto-rotation {{display:grid; grid-template-columns:1.3fr .85fr .85fr; gap:10px; margin:14px 0 16px;}}
          .crypto-rot-card {{background:#07111F; border:1px solid #1F334A; border-radius:8px; padding:12px; color:#E5E7EB;}}
          .crypto-rot-card strong {{display:block; color:#F8FAFC; font-size:1.08rem; margin-top:5px;}}
          .crypto-rot-card p {{margin:7px 0 0; color:#CBD5E1; font-weight:700;}}
          .crypto-class-grid {{display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px; margin:10px 0 16px;}}
          .crypto-class-card {{background:#07111F; border:1px solid #1F334A; border-radius:8px; padding:12px; min-height:142px;}}
          .crypto-class-top {{display:flex; align-items:flex-start; justify-content:space-between; gap:10px;}}
          .crypto-class-top b {{color:#F8FAFC; font-size:1rem;}}
          .crypto-chip {{display:inline-flex; width:max-content; border-radius:999px; padding:3px 8px; font-size:.68rem; font-weight:950; text-transform:uppercase;}}
          .crypto-chip.good {{background:rgba(0,208,132,.14); color:#2DFFAA; border:1px solid rgba(0,208,132,.35);}}
          .crypto-chip.bad {{background:rgba(255,75,75,.14); color:#FF8A8A; border:1px solid rgba(255,75,75,.35);}}
          .crypto-chip.warn {{background:rgba(255,176,32,.14); color:#FFCB6B; border:1px solid rgba(255,176,32,.35);}}
          .crypto-chip.neutral {{background:rgba(148,163,184,.13); color:#CBD5E1; border:1px solid rgba(148,163,184,.28);}}
          .crypto-class-score {{font-size:2rem; font-weight:950; margin:10px 0 8px;}}
          .crypto-class-score.good, .crypto-rank-score.good {{color:#00D084;}}
          .crypto-class-score.bad, .crypto-rank-score.bad {{color:#FF5D5D;}}
          .crypto-class-score.warn, .crypto-rank-score.warn {{color:#FFB020;}}
          .crypto-class-metrics, .crypto-rank-grid {{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:6px;}}
          .crypto-class-metrics span, .crypto-rank-grid span {{color:#8FA4BD; font-size:.74rem;}}
          .crypto-class-metrics strong, .crypto-rank-grid strong {{color:#F8FAFC;}}
          .crypto-rank-wrap {{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:12px; margin-top:8px;}}
          .crypto-rank-panel {{background:#050B14; border:1px solid #1A2B42; border-radius:8px; padding:12px;}}
          .crypto-op-wrap {{display:grid; grid-template-columns:repeat(5,minmax(0,1fr)); gap:10px; margin-top:8px;}}
          .crypto-op-panel.good {{border-top:3px solid #00D084;}}
          .crypto-op-panel.warn {{border-top:3px solid #FFB020;}}
          .crypto-op-panel.bad {{border-top:3px solid #FF4B4B;}}
          .crypto-rank-panel h5 {{margin:0 0 10px; color:#F8FAFC; font-size:.94rem;}}
          .crypto-rank-panel p {{margin:-4px 0 10px; color:#8FA4BD; font-size:.70rem; line-height:1.3; font-weight:800;}}
          .crypto-rank-list {{display:grid; gap:9px;}}
          .crypto-rank-card {{background:#08111F; border:1px solid #203047; border-radius:8px; padding:10px;}}
          .crypto-rank-head {{display:flex; justify-content:space-between; gap:10px; margin-bottom:7px;}}
          .crypto-rank-head b {{display:block; color:#F8FAFC; font-size:.95rem;}}
          .crypto-rank-head small {{display:block; color:#8FA4BD; margin-top:2px; font-size:.70rem;}}
          .crypto-rank-score {{font-size:1.25rem; font-weight:950;}}
          .crypto-empty {{background:#0B233A; border:1px solid #16466E; color:#7DD3FC; border-radius:8px; padding:12px; font-weight:800;}}
          .crypto-mini-wrap {{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:12px; margin:12px 0 18px;}}
          .crypto-mini-card {{background:#050C16; border:1px solid #1F334A; border-radius:8px; padding:9px;}}
          .crypto-mini-head {{display:flex; gap:8px; align-items:center; color:#E5E7EB; margin-bottom:6px;}}
          .crypto-mini-head b {{font-size:.92rem;}}
          .crypto-mini-head span {{font-size:.78rem; font-weight:900;}}
          .crypto-mini-head em {{margin-left:auto; font-style:normal; background:#132338; color:#93C5FD; border-radius:999px; padding:2px 7px; font-size:.66rem; font-weight:900;}}
          .crypto-mini-chart {{height:210px; width:100%;}}
          .pos {{color:#00D084; font-weight:900;}}
          .neg {{color:#FF5D5D; font-weight:900;}}
          .crypto-drivers li {{margin-bottom:7px;}}
          @media (max-width: 1200px) {{
            .crypto-hero {{grid-template-columns:1fr;}}
            .crypto-regime-top {{display:block;}}
            .crypto-regime-scorebox {{grid-template-columns:1fr 1fr; min-width:0; margin-top:10px;}}
            .crypto-regime-drivers {{grid-template-columns:1fr;}}
            .crypto-allocation-top {{display:block;}}
            .crypto-allocation-score {{text-align:left; margin-top:8px;}}
            .crypto-allocation-split {{grid-template-columns:repeat(2,minmax(0,1fr));}}
            .crypto-grid {{grid-template-columns:repeat(2,minmax(0,1fr));}}
            .crypto-cycle-dashboard {{grid-template-columns:1fr;}}
            .crypto-btc-strength-head {{display:block;}}
            .crypto-btc-strength-kpis {{grid-template-columns:1fr; min-width:0; margin-top:10px;}}
            .crypto-btc-strength-grid {{grid-template-columns:1fr;}}
            .crypto-assets {{grid-template-columns:repeat(2,minmax(0,1fr));}}
            .crypto-onchain {{grid-template-columns:1fr;}}
            .crypto-onchain-grid {{grid-template-columns:1fr;}}
            .crypto-onchain-cycle {{grid-template-columns:1fr;}}
            .crypto-onchain-heatmap {{grid-template-columns:repeat(2,minmax(0,1fr));}}
            .crypto-mvrv-head {{display:block;}}
            .crypto-mvrv-legend {{justify-content:flex-start; margin-top:8px;}}
            .crypto-alerts {{grid-template-columns:1fr;}}
            .crypto-rotation {{grid-template-columns:1fr;}}
            .crypto-class-grid {{grid-template-columns:repeat(2,minmax(0,1fr));}}
            .crypto-rank-wrap {{grid-template-columns:1fr;}}
            .crypto-op-wrap {{grid-template-columns:1fr;}}
            .crypto-mini-wrap {{grid-template-columns:1fr;}}
          }}
        </style>
        <div class="crypto-hero">
          <div class="crypto-card crypto-regime">
            <div class="crypto-regime-top">
              <div>
                <span class="crypto-label">Regime Cripto</span>
                <strong>{regime_name}</strong>
                <div class="crypto-bias">{regime_bias}</div>
              </div>
              <div class="crypto-regime-scorebox">
                <span>Score<b>{score:.0f}/100</b></span>
                <span>Confiança<b>{regime_confidence:.0f}%</b></span>
              </div>
            </div>
            <p class="crypto-regime-summary">{regime_summary}</p>
            <div class="crypto-allocation-card">
              <div class="crypto-allocation-top">
                <div>
                  <span class="crypto-allocation-action {allocation_action_cls}">{allocation_action}</span>
                  <strong>{allocation_bias}</strong>
                  <p><b>Motivo:</b> {allocation_reason}. <b>Risco:</b> {allocation_risk_note}.</p>
                </div>
                <div class="crypto-allocation-score">Score ciclo + momentum<b>{allocation_score:.0f}/100</b></div>
              </div>
              <div class="crypto-allocation-meter"><span class="btc"></span><span class="eth"></span><span class="alts"></span><span class="usdt"></span></div>
              <div class="crypto-allocation-split">
                <span>BTC<b>{allocation_btc_pct:.0f}%</b></span>
                <span>ETH<b>{allocation_eth_pct:.0f}%</b></span>
                <span>Alts<b>{allocation_alts_pct:.0f}%</b></span>
                <span>USDT<b>{allocation_usdt_pct:.0f}%</b></span>
              </div>
              <p><b>Condição de mudança:</b> {allocation_condition}</p>
            </div>
            <div class="crypto-regime-drivers">
              <div class="crypto-regime-driver good"><span>Suportes do regime</span><ul>{positive_drivers_html}</ul></div>
              <div class="crypto-regime-driver bad"><span>Riscos monitorados</span><ul>{negative_drivers_html}</ul></div>
            </div>
          </div>
          <div class="crypto-card crypto-risk-card">
            <span class="crypto-label">Score de risco</span>
            <div class="crypto-score">{score:.0f}/100</div>
            <div style="background:#08111F;border:1px solid #203047;border-radius:8px;padding:12px;min-height:110px;margin-top:12px;">
              <span style="display:block;color:#8FB6E8;font-size:.72rem;font-weight:900;text-transform:uppercase;">Fear & Greed</span>
              <div style="display:flex;align-items:flex-end;justify-content:space-between;gap:10px;margin-top:6px;">
                <div style="color:#F8FAFC;font-size:2rem;line-height:1;font-weight:950;">{fng_gauge_value:.0f}</div>
                <div style="color:{fng_badge_color};background:rgba(148,163,184,.12);border:1px solid {fng_badge_color}55;border-radius:999px;padding:4px 9px;font-size:.68rem;font-weight:950;text-transform:uppercase;">{sanitize_text(fng_label or fng_gauge_text)}</div>
              </div>
              <div style="height:18px;margin-top:14px;border-radius:999px;background:linear-gradient(90deg,#FF4B4B 0%,#FF4B4B 25%,#FFB020 25%,#FFB020 55%,#00D084 55%,#00D084 75%,#FF5D5D 75%,#FF5D5D 100%);padding:3px;box-shadow:inset 0 0 0 1px rgba(255,255,255,.08);">
                <div style="height:12px;width:{fng_bar_width};border-radius:999px;background:#F8FAFC;box-shadow:0 0 12px rgba(248,250,252,.35);"></div>
              </div>
              <div style="display:flex;justify-content:space-between;color:#64748B;font-size:.62rem;font-weight:900;margin-top:6px;"><span>0</span><span>25</span><span>50</span><span>75</span><span>100</span></div>
            </div>
            <div style="display:flex;justify-content:space-between;margin-top:12px;gap:10px;">
              <div><span class="crypto-label">Dominance</span><b>BTC {btc_dom:.2f}%</b><br><small>ETH {eth_dom:.2f}% | USDT {'---' if usdt_dom is None else f'{usdt_dom:.2f}%'}</small></div>
              <div style="text-align:right;"><span class="crypto-label">Fonte</span><b>{sanitize_text(fng_label or fng_gauge_text)}</b><br><small>sentimento</small></div>
            </div>
          </div>
        </div>
        <div class="crypto-cycle-dashboard">
          <div class="crypto-cycle-box {btc_cycle_cls}">
            <span class="crypto-label">Bitcoin 4-Year Cycle</span>
            <b>{sanitize_text(btc_cycle_phase)}</b>
            <p>{sanitize_text(btc_cycle_read)}</p>
            <div class="crypto-cycle-meter"><div></div></div>
            <p style="font-size:.72rem;color:#8FA4BD;margin-top:7px;">Progresso halving: {cycle_progress:.0f}%</p>
          </div>
          <div class="crypto-cycle-box neutral">
            <span class="crypto-label">Calendario do halving</span>
            <b>{days_since_halving} dias</b>
            <p>Desde o halving de 20/04/2024. Proximo estimado: {days_to_next_halving} dias.</p>
            <div class="crypto-cycle-mini">
              <div><span>Ultimo</span><strong>20/04/2024</strong></div>
              <div><span>Prox.</span><strong>Abr/2028</strong></div>
            </div>
          </div>
          <div class="crypto-cycle-box {macro_cls}">
            <span class="crypto-label">Liquidez / macro cripto</span>
            <b>{sanitize_text(macro_liquidity)}</b>
            <p>{sanitize_text(macro_text)}</p>
            <div class="crypto-cycle-mini">
              <div><span>BTC 24h</span><strong>{_plain_pct(btc_change_24h)}</strong></div>
              <div><span>Stables dom.</span><strong>{'---' if stable_dom is None else f'{stable_dom:.1f}%'}</strong></div>
            </div>
          </div>
          <div class="crypto-cycle-box {cycle_operational_cls}">
            <span class="crypto-label">Vies operacional</span>
            <b>{sanitize_text(cycle_operational_bias)}</b>
            <p>Combina ciclo, on-chain e fluxo curto. Use como filtro de contexto, nao como gatilho isolado.</p>
            <div class="crypto-cycle-mini">
              <div><span>MVRV Z</span><strong>{'---' if mvrv_z is None else f'{mvrv_z:.2f}'}</strong></div>
              <div><span>F&G</span><strong>{fng_gauge_value:.0f}</strong></div>
            </div>
          </div>
          <div class="crypto-cycle-box {sanitize_text(usdt_flow.get('cls', 'neutral'))}">
            <span class="crypto-label">Fluxo USDT</span>
            <b>{sanitize_text(usdt_flow.get('status', 'Monitorando'))}</b>
            <p>{sanitize_text(usdt_flow.get('text', 'Aguardando leitura.'))}</p>
            <div class="crypto-cycle-mini">
              <div><span>USDT dom.</span><strong>{'---' if usdt_dom is None else f'{usdt_dom:.2f}%'}</strong></div>
              <div><span>Delta</span><strong>{sanitize_text(usdt_delta_text)}</strong></div>
              <div><span>Max 15d</span><strong>{sanitize_text(usdt_high_15d_text)}</strong></div>
              <div><span>Dist. max</span><strong>{sanitize_text(usdt_distance_high_text)}</strong></div>
            </div>
          </div>
        </div>
        {btc_strength_html}
        <div class="crypto-grid">
          <div class="crypto-kpi"><span>BTC</span><strong>{_money(btc.get("price"), 2)}</strong><em>24h {_pct(btc.get("change_pct_24h"))}</em></div>
          <div class="crypto-kpi"><span>ETH</span><strong>{_money(eth.get("price"), 2)}</strong><em>24h {_pct(eth.get("change_pct_24h"))}</em></div>
          <div class="crypto-kpi"><span>Market Cap</span><strong>{_money(market_cap, 2)}</strong><em>Volume 24h {_money(volume_24h, 2)}</em></div>
          <div class="crypto-kpi"><span>DeFi + Stablecoins</span><strong>TVL {_money(tvl, 2)}</strong><em>Stables {_money(stable_cap, 2)}</em></div>
        </div>
        <div class="crypto-onchain">
          <div class="crypto-onchain-main">
            <span class="crypto-label">Bitcoin On-chain | MVRV Z-Score</span>
            <strong>{'---' if mvrv_z is None else f'{mvrv_z:.2f}'}</strong>
            <span class="crypto-chip {mvrv_cls}">{sanitize_text(mvrv_zone)}</span>
            <p>{sanitize_text(mvrv_text)}</p>
          </div>
          <div class="crypto-onchain-grid">
            <div><span>MVRV</span><b>{'---' if mvrv_ratio is None else f'{mvrv_ratio:.2f}x'}</b></div>
            <div><span>Mayer / Puell</span><b>{'---' if mayer_multiple is None else f'{mayer_multiple:.2f}x'} / {'---' if puell_multiple is None else f'{puell_multiple:.2f}'}</b></div>
            <div><span>AVIV / F&G</span><b>{'---' if aviv_value is None else f'{aviv_value:.2f}'} / {'---' if onchain_fear is None else f'{onchain_fear:.0f}'}</b></div>
            <div><span>Data on-chain</span><b>{mvrv_date}</b></div>
            <div><span>Fonte</span><b>BGeometrics · {mvrv_status}</b></div>
          </div>
        </div>
        <div class="crypto-onchain-cycle">
          <div class="crypto-cycle-card {cycle_cls}">
            <span class="crypto-label">Regime on-chain BTC</span>
            <strong>{sanitize_text(cycle_regime)}</strong>
            <p>{sanitize_text(cycle_text)}</p>
          </div>
          <div class="crypto-onchain-heatmap">
            {_onchain_heatmap_html(bgeometrics_data)}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    mvrv_refresh_key = int(st.session_state.get("crypto_terminal_refresh_key", 0))
    bgeometrics_mvrv_history = load_crypto_mvrv_history(mvrv_refresh_key)
    mvrv_history_data = bgeometrics_mvrv_history.get("data", {}) if isinstance(bgeometrics_mvrv_history, dict) else {}
    _render_crypto_mvrv_chart(mvrv_history_data.get("points") or [])
    bgeometrics_rainbow = load_crypto_rainbow_chart(mvrv_refresh_key)
    rainbow_data = bgeometrics_rainbow.get("data", {}) if isinstance(bgeometrics_rainbow, dict) else {}
    _render_crypto_rainbow_chart(rainbow_data.get("points") or [])
    _render_crypto_mvrv_pricing_bands_chart(mvrv_history_data.get("points") or [])

    st.markdown("#### Majors e Altcoins")
    top_symbols = ["BTCUSDT", "ETHUSDT", "SOLUSDT", "BNBUSDT", "XRPUSDT", "ADAUSDT", "DOGEUSDT", "LINKUSDT", "AAVEUSDT", "AVAXUSDT", "SUIUSDT"]
    asset_cards = []
    for sym in top_symbols:
        item = symbols.get(sym, {})
        if not item:
            continue
        asset_cards.append(
            "<div class='crypto-asset'>"
            f"<b>{sanitize_text(sym.replace('USDT', ''))}</b>"
            f"<small>{_money(item.get('price'), 4)} | 24h {_pct(item.get('change_pct_24h'))}</small>"
            f"<small>Vol {_money(item.get('quote_volume'), 1)} | Funding {_num(item.get('funding_rate')) * 100:+.4f}%</small>"
            f"<small>Range 24h: {_money(item.get('low_24h'), 4)} - {_money(item.get('high_24h'), 4)}</small>"
            "</div>"
        )
    cards_html = "".join(asset_cards) or '<div class="crypto-card">Sem dados de ativos agora.</div>'
    st.markdown(f"<div class='crypto-assets'>{cards_html}</div>", unsafe_allow_html=True)

    rotation = operational.get("rotation", {}) if isinstance(operational, dict) else {}
    leader_class = rotation.get("leader_class", {}) if isinstance(rotation, dict) else {}
    weakest_class = rotation.get("weakest_class", {}) if isinstance(rotation, dict) else {}
    st.markdown("#### Regime por subclasse")
    st.markdown(
        "<div class='crypto-rotation'>"
        "<div class='crypto-rot-card'>"
        "<span class='crypto-label'>Leitura IA local</span>"
        f"<strong>{sanitize_text(rotation.get('flow', 'Fluxo indefinido'))}</strong>"
        f"<p>{sanitize_text(rotation.get('ai_summary', 'Sem leitura suficiente agora.'))}</p>"
        "</div>"
        "<div class='crypto-rot-card'>"
        "<span class='crypto-label'>Classe lider</span>"
        f"<strong>{sanitize_text(leader_class.get('classe', '---'))}</strong>"
        f"<p>Score {leader_class.get('score', '---')} | Lider {sanitize_text(leader_class.get('lider', '---'))}</p>"
        "</div>"
        "<div class='crypto-rot-card'>"
        "<span class='crypto-label'>Classe mais fraca</span>"
        f"<strong>{sanitize_text(weakest_class.get('classe', '---'))}</strong>"
        f"<p>Score {weakest_class.get('score', '---')} | Lider {sanitize_text(weakest_class.get('lider', '---'))}</p>"
        "</div>"
        "</div>",
        unsafe_allow_html=True,
    )
    classes = rotation.get("classes", []) if isinstance(rotation, dict) else []
    if classes:
        st.markdown(f"<div class='crypto-class-grid'>{_rotation_cards_html(classes)}</div>", unsafe_allow_html=True)
    else:
        st.info("Sem dados suficientes para rotacao por subclasse agora.")

    st.markdown("#### Alertas e ranking operacional")
    alerts = operational.get("alerts", []) if isinstance(operational, dict) else []
    if alerts:
        st.markdown(
            "<div class='crypto-alerts'>" + "".join(f"<div class='crypto-alert'>{sanitize_text(item)}</div>" for item in alerts[:6]) + "</div>",
            unsafe_allow_html=True,
        )
    else:
        st.success("Sem alertas criticos de alavancagem, faixa extrema ou volatilidade agora.")

    operational_groups = operational.get("operational_groups", {}) if isinstance(operational, dict) else {}
    leaders = operational.get("leaders", []) if isinstance(operational, dict) else []
    laggards = operational.get("laggards", []) if isinstance(operational, dict) else []
    if operational_groups:
        ranking_html = f"<div class='crypto-op-wrap'>{_operational_group_html(operational_groups)}</div>"
    else:
        ranking_html = (
            "<div class='crypto-rank-wrap'>"
            "<div class='crypto-rank-panel'><h5>Lideres de risco</h5>"
            f"<div class='crypto-rank-list'>{_ranking_cards_html(leaders, 'leader')}</div></div>"
            "<div class='crypto-rank-panel'><h5>Mais fracos / defensivos</h5>"
            f"<div class='crypto-rank-list'>{_ranking_cards_html(laggards, 'laggard')}</div></div>"
            "</div>"
        )
    st.markdown(
        ranking_html,
        unsafe_allow_html=True,
    )

    st.markdown("#### Mini graficos intraday")
    chart_symbols = ["BTCUSDT", "ETHUSDT", "SOLUSDT", "BNBUSDT"]
    chart_html = []
    for idx, sym in enumerate(chart_symbols):
        item = symbols.get(sym, {})
        if item:
            chart_html.append(_crypto_mini_chart_html(sym, item, f"{idx}-{refresh_key}"))
    if chart_html:
        components.html(
            "<script src='https://unpkg.com/lightweight-charts/dist/lightweight-charts.standalone.production.js'></script>"
            "<style>body{margin:0;background:transparent;font-family:Inter,Segoe UI,Arial,sans-serif;}.crypto-mini-wrap{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}.crypto-mini-card{background:#050C16;border:1px solid #1F334A;border-radius:8px;padding:9px}.crypto-mini-head{display:flex;gap:8px;align-items:center;color:#E5E7EB;margin-bottom:6px}.crypto-mini-head b{font-size:.92rem}.crypto-mini-head span{font-size:.78rem;font-weight:900}.crypto-mini-head em{margin-left:auto;font-style:normal;background:#132338;color:#93C5FD;border-radius:999px;padding:2px 7px;font-size:.66rem;font-weight:900}.crypto-mini-chart{height:210px;width:100%}.pos{color:#00D084}.neg{color:#FF5D5D}@media(max-width:900px){.crypto-mini-wrap{grid-template-columns:1fr}}</style>"
            f"<div class='crypto-mini-wrap'>{''.join(chart_html)}</div>",
            height=470,
        )
    else:
        st.info("Sem candles suficientes para os mini graficos.")

    st.markdown("#### Drivers do regime")
    drivers = []
    if isinstance(regime, dict):
        drivers = list(regime.get("drivers_positive", []) or []) + list(regime.get("drivers_negative", []) or [])
    if drivers:
        st.markdown("<ul class='crypto-drivers'>" + "".join(f"<li>{sanitize_text(driver)}</li>" for driver in drivers[:10]) + "</ul>", unsafe_allow_html=True)
    else:
        st.info("Sem drivers suficientes para o regime agora.")

    with st.expander("Detalhes tecnicos das fontes cripto", expanded=False):
        c_left, c_right = st.columns([1, 1])
        with c_left:
            st.markdown("#### Scores por bloco")
            score_rows = regime.get("scores", {}) if isinstance(regime, dict) else {}
            if score_rows:
                st.dataframe(
                    pd.DataFrame([{"Bloco": key, "Score": value} for key, value in score_rows.items()]),
                    hide_index=True,
                    use_container_width=True,
                    height=260,
                )
        with c_right:
            st.markdown("#### Saude das fontes cripto")
            raw_health = get_source_health()
            health_iter = raw_health.values() if isinstance(raw_health, dict) else raw_health
            health_rows = []
            for item in health_iter or []:
                if not isinstance(item, dict):
                    continue
                if str(item.get("name", "")).lower().startswith("crypto"):
                    health_rows.append(item)
            if health_rows:
                st.dataframe(pd.DataFrame(health_rows), hide_index=True, use_container_width=True, height=260)
            else:
                st.info("Sem telemetria de fontes cripto registrada ainda.")

        c1, c2 = st.columns([1, 1])
        with c1:
            st.markdown("#### Top chains por TVL")
            chains = defillama_data.get("top_chains", []) if isinstance(defillama_data, dict) else []
            if chains:
                df_chains = pd.DataFrame(chains[:15])
                cols = [col for col in ["name", "symbol", "tvl", "change_1d", "change_7d"] if col in df_chains.columns]
                st.dataframe(df_chains[cols], hide_index=True, use_container_width=True, height=430)
            else:
                st.info("DefiLlama sem dados de chains agora.")
        with c2:
            st.markdown("#### Top mercado CoinGecko")
            markets = coingecko_data.get("markets", []) if isinstance(coingecko_data, dict) else []
            if markets:
                df_markets = pd.DataFrame(markets[:15])
                cols = [col for col in ["symbol", "name", "current_price", "price_change_percentage_24h", "market_cap", "total_volume"] if col in df_markets.columns]
                st.dataframe(df_markets[cols], hide_index=True, use_container_width=True, height=430)
            else:
                st.info("CoinGecko sem dados de mercado agora.")


@st.cache_data(ttl=240, show_spinner=False)
def get_market_moving_events_cached(refresh_nonce: int = 0):
    news_items, _sources, _warnings, _loaded_at = load_bloomberg_news_feed(refresh_nonce)
    calendar_events = get_calendar_data() or []
    from execution.market_moving import build_market_moving_events

    return build_market_moving_events(news_items, calendar_events=calendar_events, max_events=8)


def _market_moving_chart_html(chart: dict, uid: str) -> str:
    candles = chart.get("candles", [])
    metrics = chart.get("metrics", {})
    event_time = chart.get("event_time")
    marker_time = chart.get("marker_time") or event_time
    marker_label = html.escape(str(chart.get("marker_label") or "NEWS"))
    timeframe = html.escape(str(chart.get("timeframe") or "5m"))
    title = html.escape(str(chart.get("label") or chart.get("symbol") or "Ativo"))
    symbol = html.escape(str(chart.get("symbol") or ""))
    source = html.escape(str(chart.get("source") or ""))
    payload = json.dumps({
        "candles": candles,
        "eventTime": event_time,
        "markerTime": marker_time,
        "markerLabel": marker_label,
        "visibleStart": (marker_time - 30 * 60) if marker_time else None,
        "visibleEnd": candles[-1]["time"] if candles else None,
    }, ensure_ascii=False)
    metric_html = " ".join(
        f"<span>{label}: <b class='{('pos' if (value or 0) >= 0 else 'neg')}'>{value:+.2f}%</b></span>"
        for label, value in [
            ("5m", metrics.get("ret_5m")),
            ("15m", metrics.get("ret_15m")),
            ("30m", metrics.get("ret_30m")),
        ]
        if isinstance(value, (int, float))
    )
    if not metric_html:
        metric_html = "<span>Sem candle suficiente para medir reação.</span>"
    return f"""
    <div class="mm-chart-card">
      <div class="mm-chart-title"><span>{symbol}</span>{title}{f"<small>{source}</small>" if source else ""}<em>{timeframe}</em></div>
      <div id="mm-chart-{uid}" class="mm-chart"><div id="mm-event-{uid}" class="mm-event-marker"><div class="mm-event-ring"></div><div class="mm-event-label">{marker_label}</div></div></div>
      <div class="mm-metrics">{metric_html}</div>
    </div>
    <script>
    (function() {{
      const payload = {payload};
      const root = document.getElementById("mm-chart-{uid}");
      if (!root || !window.LightweightCharts || !payload.candles || !payload.candles.length) return;
      const chart = LightweightCharts.createChart(root, {{
        layout: {{ background: {{ color: "#030712" }}, textColor: "#D1D5DB" }},
        grid: {{ vertLines: {{ color: "rgba(148,163,184,.12)" }}, horzLines: {{ color: "rgba(148,163,184,.12)" }} }},
        rightPriceScale: {{ borderColor: "rgba(148,163,184,.20)" }},
        timeScale: {{ borderColor: "rgba(148,163,184,.20)", timeVisible: true, secondsVisible: false, rightOffset: 8, barSpacing: 7 }},
        handleScroll: {{ mouseWheel: true, pressedMouseMove: true, horzTouchDrag: true, vertTouchDrag: false }},
        handleScale: {{ axisPressedMouseMove: true, mouseWheel: true, pinch: true }},
        crosshair: {{ mode: 1 }},
      }});
      let series;
      if (chart.addSeries && LightweightCharts.CandlestickSeries) {{
        series = chart.addSeries(LightweightCharts.CandlestickSeries, {{
          upColor: "#00C896", downColor: "#FF3B30", borderVisible: false,
          wickUpColor: "#00C896", wickDownColor: "#FF3B30"
        }});
      }} else {{
        series = chart.addCandlestickSeries({{
          upColor: "#00C896", downColor: "#FF3B30", borderVisible: false,
          wickUpColor: "#00C896", wickDownColor: "#FF3B30"
        }});
      }}
      series.setData(payload.candles);
      const markerTime = payload.markerTime || payload.eventTime;
      function pickEventCandle(candles, eventTime) {{
        if (!candles || !candles.length || !eventTime) return null;
        const containing = candles.find(function(candle, idx) {{
          const next = candles[idx + 1];
          const end = next ? next.time : candle.time + 300;
          return candle.time <= eventTime && eventTime < end;
        }});
        if (containing) return containing;
        return candles.reduce((best, candle) => {{
          if (!best) return candle;
          return Math.abs(candle.time - eventTime) < Math.abs(best.time - eventTime) ? candle : best;
        }}, null);
      }}
      const eventCandle = pickEventCandle(payload.candles, markerTime);
      if (eventCandle && series.setMarkers) {{
        series.setMarkers([{{
          time: eventCandle.time, position: "aboveBar", color: "#22D3EE",
          shape: "circle", text: payload.markerLabel || "EVENTO"
        }}]);
      }}
      function positionEventMarker() {{
        if (!eventCandle) return;
        const marker = document.getElementById("mm-event-{uid}");
        if (!marker) return;
        const x = chart.timeScale().timeToCoordinate(eventCandle.time);
        const y = series.priceToCoordinate(eventCandle.high);
        if (x == null || y == null) {{
          marker.style.display = "none";
          return;
        }}
        marker.style.display = "block";
        marker.style.left = Math.max(12, Math.min(root.clientWidth - 62, x - 31)) + "px";
        marker.style.top = Math.max(6, y - 44) + "px";
      }}
      chart.timeScale().subscribeVisibleTimeRangeChange(positionEventMarker);
      if (payload.visibleStart && payload.visibleEnd && chart.timeScale().setVisibleRange) {{
        chart.timeScale().setVisibleRange({{ from: payload.visibleStart, to: payload.visibleEnd }});
      }} else {{
        chart.timeScale().fitContent();
      }}
      positionEventMarker();
      new ResizeObserver(function() {{
        chart.applyOptions({{ width: root.clientWidth }});
        setTimeout(positionEventMarker, 50);
      }}).observe(root);
    }})();
    </script>
    """


def pagina_market_moving():
    """Pagina Market Moving: noticias relevantes e reacao nos ativos."""
    st.title("Market Moving")
    st.caption("Notícias de alta relevância com marcação do horário no gráfico e reação intraday pós-evento.")

    c1, c2 = st.columns([1, 4])
    with c1:
        if st.button("Atualizar eventos", type="primary", use_container_width=True, key="market_moving_refresh"):
            st.session_state["market_moving_refresh_nonce"] = st.session_state.get("market_moving_refresh_nonce", 0) + 1
            get_market_moving_events_cached.clear()
    nonce = int(st.session_state.get("market_moving_refresh_nonce", 0))

    with st.spinner("Mapeando notícias e reação nos ativos..."):
        events = get_market_moving_events_cached(nonce)

    market_moving_css = """
        <style>
          body {margin:0; background:transparent; font-family:"Inter","Segoe UI",Arial,sans-serif;}
          .mm-card {background:#B80000; border:1px solid rgba(255,255,255,.20); border-radius:8px; padding:14px 14px 10px; margin:0 0 16px; color:#fff;}
          .mm-head {display:grid; grid-template-columns:28px 1fr; gap:10px; align-items:start;}
          .mm-dot {width:22px; height:22px; border:3px solid #22D3EE; border-radius:999px; margin-top:3px; box-shadow:0 0 18px rgba(34,211,238,.75);}
          .mm-title {font-size:1rem; font-weight:900; line-height:1.25;}
          .mm-sub {color:#FECACA; font-size:.74rem; font-weight:800; margin-top:5px;}
          .mm-tags {display:flex; flex-wrap:wrap; gap:6px; margin:9px 0 10px 38px;}
          .mm-tag {background:rgba(17,24,39,.55); color:#FECACA; border:1px solid rgba(255,255,255,.14); border-radius:999px; padding:3px 8px; font-size:.68rem; font-weight:900;}
          .mm-grid {display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:10px; margin-left:38px;}
          .mm-chart-card {background:#020617; border:1px solid rgba(255,255,255,.15); border-radius:6px; padding:8px; min-width:0;}
          .mm-chart-title {display:flex; gap:7px; align-items:center; color:#E5E7EB; font-size:.82rem; font-weight:900; margin-bottom:5px;}
          .mm-chart-title span {background:#111827; border-radius:999px; padding:2px 6px; color:#BFDBFE; font-size:.68rem;}
          .mm-chart-title small {color:#93C5FD; font-size:.58rem; font-weight:900; text-transform:uppercase; opacity:.88;}
          .mm-chart-title em {margin-left:auto; font-style:normal; background:#22D3EE; color:#00111A; border-radius:999px; padding:2px 7px; font-size:.62rem; font-weight:950;}
          .mm-chart {height:260px; width:100%; position:relative; overflow:hidden;}
          .mm-event-marker {display:none; position:absolute; z-index:9; width:62px; text-align:center; pointer-events:none; filter:drop-shadow(0 0 14px rgba(34,211,238,.95));}
          .mm-event-ring {width:48px; height:48px; margin:0 auto; border:5px solid #22D3EE; border-radius:999px; background:rgba(34,211,238,.12); box-shadow:0 0 0 7px rgba(34,211,238,.16), 0 0 24px rgba(34,211,238,.95);}
          .mm-event-label {display:inline-block; margin-top:3px; padding:3px 7px; border-radius:999px; background:#22D3EE; color:#00111A; font-size:.62rem; font-weight:950; letter-spacing:.04em;}
          .mm-metrics {display:flex; flex-wrap:wrap; gap:8px; margin-top:6px; color:#CBD5E1; font-size:.70rem; font-weight:800;}
          .mm-metrics .pos {color:#00FFA3;} .mm-metrics .neg {color:#FFB4A8;}
          @media(max-width:1100px){.mm-grid{grid-template-columns:1fr; margin-left:0}.mm-tags{margin-left:0}.mm-chart{height:300px}}
        </style>
    """

    if not events:
        st.info("Sem notícias de alto impacto com candles disponíveis no momento.")
        return

    ready_events = [event for event in events if event.get("charts")]
    pending_events = [event for event in events if not event.get("charts")]
    if pending_events:
        st.info(
            f"{len(pending_events)} evento(s) de alto impacto aguardando abertura/candle dos ativos mapeados. "
            "Eles aparecerão automaticamente quando houver candle pós-evento."
        )
    if not ready_events:
        st.warning("Nenhum evento com candle pós-notícia disponível agora. Atualize após a abertura do mercado dos ativos mapeados.")
        return

    for event_idx, event in enumerate(ready_events):
        title = html.escape(str(event.get("title") or "---"))
        source = html.escape(str(event.get("source") or ""))
        impact = html.escape(str(event.get("impact") or "ALTO IMPACTO"))
        event_dt = html.escape(str(event.get("event_dt") or ""))
        tags = event.get("tags") or ["Macro"]
        charts = event.get("charts") or []
        tags_html = "".join(f"<span class='mm-tag'>{html.escape(str(tag))}</span>" for tag in [impact, *tags])
        charts_html = "".join(_market_moving_chart_html(chart, f"{event_idx}-{chart_idx}") for chart_idx, chart in enumerate(charts))
        card_html = f"""
        {market_moving_css}
        <script src="https://unpkg.com/lightweight-charts/dist/lightweight-charts.standalone.production.js"></script>
        <div class="mm-card">
          <div class="mm-head">
            <div class="mm-dot"></div>
            <div>
              <div class="mm-title">{title}</div>
              <div class="mm-sub">{event_dt} | {source}</div>
            </div>
          </div>
          <div class="mm-tags">{tags_html}</div>
          <div class="mm-grid">{charts_html}</div>
        </div>
        """
        card_html = "".join(line.strip() for line in card_html.splitlines())
        components.html(card_html, height=420 if charts else 190, scrolling=False)


def pagina_graficos():
    """Página com integração TradingView Advanced Chart."""
    st.markdown("### 📊 Gráficos Avançados TradingView")
    from lightweight_chart_component import render_lightweight_chart_html

    st.markdown("#### Gráfico próprio - Lightweight Charts")
    st.caption("Candles OHLCV, volume, VWAP diária, bandas, médias móveis, oscilador e atualização via Binance WebSocket.")
    st.markdown("##### Motor de reversao")
    components.html(
        render_lightweight_chart_html(signal_mode="reversal", chart_title="Motor de reversao", instance_id="reversal_main"),
        height=1680,
        scrolling=False,
    )
    
    assets = {
        "MINI ÍNDICE (WIN)": "BRA50",
        "MINI DÓLAR (WDO)": "BMFBOVESPA:WDO1!",
        "IBOVESPA": "BMFBOVESPA:IBOV",
        "S&P 500 (Futuro)": "USA500",
        "NASDAQ (Futuro)": "ACTIVTRADES:USATEC",
        "VIX": "CBOE:VIX",
        "DXY (Dólar Index)": "CAPITALCOM:DXY",
        "USDBRL": "FX_IDC:USDBRL",
        "6L (Real CME)": "CME:6L1!",
        "US 10Y (Yield)": "TVC:US10Y",
        "US 30Y (Yield)": "TVC:US30Y",
        "US 02Y (Yield)": "TVC:US02Y",
        "EEM (Emerging Markets)": "AMEX:EEM",
        "EWZ (Brazil ETF)": "AMEX:EWZ",
        "PETR4": "BMFBOVESPA:PETR4",
        "VALE3": "BMFBOVESPA:VALE3",
        "ITUB4": "BMFBOVESPA:ITUB4",
        "SPY (S&P 500 ETF)": "AMEX:SPY",
        "XOP (Oil & Gas)": "AMEX:XOP",
        "XLE (Energy)": "AMEX:XLE",
        "XLK (Tech)": "AMEX:XLK",
        "XLP (Staples)": "AMEX:XLP",
        "XLB (Materials)": "AMEX:XLB",
        "XLI (Industrials)": "AMEX:XLI",
        "XLV (Health)": "AMEX:XLV",
        "XLRE (Real Estate)": "AMEX:XLRE",
        "XBI (Biotech)": "AMEX:XBI",
        "XLY (Consumer)": "AMEX:XLY",
        "XLC (Comm)": "AMEX:XLC",
        "BRENT OIL": "TVC:UKOIL",
        "WTI OIL": "TVC:USOIL",
        "GOLD": "TVC:GOLD",
        "SILVER": "TVC:SILVER",
        "BITCOIN": "BINANCE:BTCUSDT",
        "ETHEREUM": "BINANCE:ETHUSDT"
    }

    c_topo1, c_topo2, c_topo3 = st.columns([1, 1, 2])
    with c_topo3:
        chart_height = st.slider("Ajustar Altura dos Gráficos", min_value=300, max_value=1200, value=550, step=50)

    st.markdown("---")
    st.markdown("#### 📈 Gráfico 1")
    c1a, c1b = st.columns([2, 1])
    with c1a:
        sym1 = st.selectbox("Ativo 1", list(assets.keys()), index=0, key="sym1")
    with c1b:
        int1 = st.selectbox("Intervalo 1", ["1", "5", "15", "60", "D", "W"], index=1, key="int1")
        
    tv_html1 = f"""
    <div class="tradingview-widget-container" style="height: {chart_height}px; width: 100%;">
      <div id="tv_chart_1" style="height: 100%; width: 100%;"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
      <script type="text/javascript">
      new TradingView.widget({{
        "autosize": true,
        "symbol": "{assets[sym1]}",
        "interval": "{int1}",
        "timezone": "America/Sao_Paulo",
        "theme": "dark",
        "style": "1",
        "locale": "br",
        "toolbar_bg": "#f1f3f6",
        "enable_publishing": false,
        "hide_top_toolbar": false,
        "save_image": true,
        "hide_volume": true,
        "container_id": "tv_chart_1",
        "studies": [
          {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Session" }}, "plots": {{ "VWAP": {{ "color": "#FFD166" }} }} }},
          {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Week" }}, "plots": {{ "VWAP": {{ "color": "#06D6A0" }} }} }},
          {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Month" }}, "plots": {{ "VWAP": {{ "color": "#118AB2" }} }} }}
        ]
      }});
      </script>
    </div>
    """
    components.html(tv_html1, height=chart_height + 20)

    st.markdown("---")
    st.markdown("#### 📉 Gráfico 2")
    c2a, c2b = st.columns([2, 1])
    with c2a:
        sym2 = st.selectbox("Ativo 2", list(assets.keys()), index=3, key="sym2")
    with c2b:
        int2 = st.selectbox("Intervalo 2", ["1", "5", "15", "60", "D", "W"], index=1, key="int2")
        
    tv_html2 = f"""
    <div class="tradingview-widget-container" style="height: {chart_height}px; width: 100%;">
      <div id="tv_chart_2" style="height: 100%; width: 100%;"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
      <script type="text/javascript">
      new TradingView.widget({{
        "autosize": true,
        "symbol": "{assets[sym2]}",
        "interval": "{int2}",
        "timezone": "America/Sao_Paulo",
        "theme": "dark",
        "style": "1",
        "locale": "br",
        "toolbar_bg": "#f1f3f6",
        "enable_publishing": false,
        "hide_top_toolbar": false,
        "save_image": true,
        "hide_volume": true,
        "container_id": "tv_chart_2",
        "studies": [
          {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Session" }}, "plots": {{ "VWAP": {{ "color": "#FFD166" }} }} }},
          {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Week" }}, "plots": {{ "VWAP": {{ "color": "#06D6A0" }} }} }},
          {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Month" }}, "plots": {{ "VWAP": {{ "color": "#118AB2" }} }} }}
        ]
      }});
      </script>
    </div>
    """
    components.html(tv_html2, height=chart_height + 20)

def pagina_correlacao():
    """Página dedicada ao estudo de correlação unificada com visibilidade aprimorada."""
    st.markdown("### ⚖️ Painel de Correlação Macro")
    
    # Controles Gerais
    c1, c2, c3 = st.columns([1, 1, 2])
    with c1:
        interval = st.selectbox("Intervalo", ["5", "15", "60", "D", "W"], index=3)
    with c2:
        theme = st.selectbox("Tema", ["dark", "light"], index=0)
    with c3:
        c_height = st.slider("Altura do Gráfico", 400, 1200, 800, 50)

    # Legenda customizada com cores sugeridas para o usuário ajustar no widget
    st.info("💡 Dica: Nos gráficos abaixo, você pode clicar em cada ativo na legenda (canto superior esquerdo) para ajustar a cor e a espessura da linha para melhor visibilidade.")

    # Widget 1: Correlação Macro Tradicional
    st.markdown("#### 📊 Correlação Macro (USA500, Ouro, Petróleo, US10Y, US30Y)")
    tv_html = f"""
    <div class="tradingview-widget-container" style="height: {c_height}px; width: 100%;">
      <div id="tradingview_unified_v2" style="height: 100%; width: 100%;"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
      <script type="text/javascript">
      new TradingView.widget(
      {{
        "autosize": true,
        "symbol": "USA500",
        "interval": "{interval}",
        "timezone": "America/Sao_Paulo",
        "theme": "{theme}",
        "style": "2",
        "locale": "br",
        "toolbar_bg": "#f1f3f6",
        "enable_publishing": false,
        "hide_top_toolbar": false,
        "hide_side_toolbar": true,
        "allow_symbol_change": true,
        "save_image": true,
        "details": false,
        "hotlist": false,
        "calendar": false,
        "hide_volume": true,
        "container_id": "tradingview_unified_v2",
        "overrides": {{
            "mainSeriesProperties.lineStyle.color": "#FF00FF",
            "mainSeriesProperties.lineStyle.linewidth": 3
        }},
        "studies": [
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "TVC:GOLD" }}, "plots": {{ "Plot": {{ "color": "#FFFF00" }} }} }},
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "TVC:UKOIL" }}, "plots": {{ "Plot": {{ "color": "#006400" }} }} }},
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "OTCB:US10Y" }}, "plots": {{ "Plot": {{ "color": "#FF9800" }} }} }},
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "OTCB:US30Y" }}, "plots": {{ "Plot": {{ "color": "#00BFFF" }} }} }}
        ]
      }}
      );
      </script>
    </div>
    """
    components.html(tv_html, height=c_height + 20)

    # Widget 2: Comparativo de Moedas vs DXY
    st.markdown("---")
    st.markdown("#### 💱 Comparativo de Moedas vs DXY (Escala em Porcentagem)")
    st.markdown("<p style='color:#888; font-size:0.85rem;'>Gráfico comparando o DXY (base em Branco) com as taxas Spot de Real Brasileiro (BRLUSD - Verde), Euro (EURUSD - Azul Claro) e Iene Japonês (JPYUSD - Vermelho).</p>", unsafe_allow_html=True)
    
    tv_html_currencies = f"""
    <div class="tradingview-widget-container" style="height: {c_height}px; width: 100%;">
      <div id="tradingview_currencies_v2" style="height: 100%; width: 100%;"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
      <script type="text/javascript">
      new TradingView.widget(
      {{
        "autosize": true,
        "symbol": "CAPITALCOM:DXY",
        "interval": "{interval}",
        "timezone": "America/Sao_Paulo",
        "theme": "{theme}",
        "style": "2",
        "locale": "br",
        "toolbar_bg": "#f1f3f6",
        "enable_publishing": false,
        "hide_top_toolbar": false,
        "hide_side_toolbar": true,
        "allow_symbol_change": true,
        "save_image": true,
        "details": false,
        "hotlist": false,
        "calendar": false,
        "hide_volume": true,
        "container_id": "tradingview_currencies_v2",
        "overrides": {{
            "mainSeriesProperties.lineStyle.color": "#FFFFFF",
            "mainSeriesProperties.lineStyle.linewidth": 3,
            "scalesProperties.scaleMode": 2
        }},
        "studies": [
          {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Session" }}, "plots": {{ "VWAP": {{ "color": "#FFD166" }} }} }},
          {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Week" }}, "plots": {{ "VWAP": {{ "color": "#06D6A0" }} }} }},
          {{ "id": "VWAP@tv-basicstudies", "inputs": {{ "Anchor Period": "Month" }}, "plots": {{ "VWAP": {{ "color": "#118AB2" }} }} }},
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "FX_IDC:BRLUSD" }}, "plots": {{ "Plot": {{ "color": "#00FFA3" }} }} }},
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "FX:EURUSD" }}, "plots": {{ "Plot": {{ "color": "#00BFFF" }} }} }},
          {{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "FX_IDC:JPYUSD" }}, "plots": {{ "Plot": {{ "color": "#FF4B4B" }} }} }}
        ]
      }}
      );
      </script>
    </div>
    """
    components.html(tv_html_currencies, height=c_height + 20)

    def render_macro_class_chart(title, description, container_id, main_symbol, main_color, overlays):
        class_chart_height = max(1050, min(c_height + 250, 1400))
        class_interval = "5"
        studies = ",\n          ".join(
            [
                '{{ "id": "Overlay@tv-basicstudies", "inputs": {{ "symbol": "{}" }}, "plots": {{ "Plot": {{ "color": "{}" }} }} }}'.format(symbol, color)
                for symbol, color, _label in overlays
            ]
        )
        legend_items = "".join(
            [
                f"<span style='display:inline-flex; align-items:center; gap:6px; margin-right:14px; margin-bottom:6px; color:#CBD5E1; font-size:0.78rem;'><i style='width:10px; height:10px; border-radius:50%; background:{color}; display:inline-block;'></i>{label}</span>"
                for symbol, color, label in [(main_symbol, main_color, "Base")] + overlays
            ]
        )
        st.markdown(f"#### {title}")
        st.markdown(
            f"<div style='color:#94A3B8; font-size:0.84rem; margin-bottom:8px;'>{description}</div>"
            f"<div style='margin-bottom:10px;'>{legend_items}</div>",
            unsafe_allow_html=True,
        )
        tv_html_class = f"""
        <div class="tradingview-widget-container" style="height: {class_chart_height}px; width: 100%;">
          <div id="{container_id}" style="height: 100%; width: 100%;"></div>
          <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
          <script type="text/javascript">
          new TradingView.widget(
          {{
            "autosize": true,
            "symbol": "{main_symbol}",
            "interval": "{class_interval}",
            "timezone": "America/Sao_Paulo",
            "theme": "{theme}",
            "style": "2",
            "locale": "br",
            "toolbar_bg": "#f1f3f6",
            "enable_publishing": false,
            "hide_top_toolbar": false,
            "hide_side_toolbar": true,
            "allow_symbol_change": true,
            "save_image": true,
            "details": false,
            "hotlist": false,
            "calendar": false,
            "hide_volume": true,
            "container_id": "{container_id}",
            "overrides": {{
                "mainSeriesProperties.lineStyle.color": "{main_color}",
                "mainSeriesProperties.lineStyle.linewidth": 3,
                "scalesProperties.scaleMode": 2
            }},
            "studies": [
              {studies}
            ]
          }}
          );
          </script>
        </div>
        """
        components.html(tv_html_class, height=class_chart_height + 20)

    st.markdown("---")
    st.markdown("### Comparativos por Classe Macro")
    st.markdown(
        "<p style='color:#94A3B8; font-size:0.88rem;'>Quatro leituras separadas por classe para comparar energia/metais, moedas, bolsas e juros globais na mesma escala visual.</p>",
        unsafe_allow_html=True,
    )

    commodity_col, fx_col, equity_col, bonds_col = st.columns(4, gap="medium")
    with commodity_col:
        render_macro_class_chart(
            "Commodities",
            "Energia, metais industriais e metais preciosos pela fonte ActivTrades.",
            "tradingview_macro_commodities_v1",
            "ACTIVTRADES:BRENT",
            "#22C55E",
            [
                ("ACTIVTRADES:LCRUDE", "#F97316", "Petroleo WTI"),
                ("ACTIVTRADES:NGAS", "#60A5FA", "Gas natural"),
                ("ACTIVTRADES:COPPERN2026", "#D97706", "Cobre"),
                ("ACTIVTRADES:GOLD", "#FACC15", "Ouro"),
                ("ACTIVTRADES:SILVER", "#E2E8F0", "Prata"),
            ],
        )
    with fx_col:
        render_macro_class_chart(
            "FX",
            "Moedas de commodities, safe havens/majors, emergentes e carry.",
            "tradingview_macro_fx_v1",
            "CAPITALCOM:DXY",
            "#F8FAFC",
            [
                ("OANDA:AUDUSD", "#22C55E", "AUDUSD"),
                ("OANDA:USDCAD", "#F97316", "USDCAD"),
                ("OANDA:GBPUSD", "#A855F7", "GBPUSD"),
                ("OANDA:EURUSD", "#38BDF8", "EURUSD"),
                ("FX_IDC:USDBRL", "#FACC15", "USDBRL"),
                ("OANDA:USDJPY", "#EF4444", "USDJPY"),
            ],
        )
    with equity_col:
        render_macro_class_chart(
            "Equity",
            "Indices globais pela fonte ActivTrades: volatilidade, EUA, Brasil, Europa e Japao.",
            "tradingview_macro_equity_v1",
            "ACTIVTRADES:USA500",
            "#A855F7",
            [
                ("ACTIVTRADES:VXX.US", "#EF4444", "VIX/VXX"),
                ("ACTIVTRADES:JP225", "#38BDF8", "Nikkei"),
                ("ACTIVTRADES:BRA50", "#22C55E", "IBOV/BRA50"),
                ("ACTIVTRADES:EURO50", "#F97316", "EuroStoxx"),
                ("ACTIVTRADES:USARUS", "#FACC15", "RTY/Russell"),
                ("ACTIVTRADES:USATEC", "#60A5FA", "Nasdaq"),
            ],
        )
    with bonds_col:
        render_macro_class_chart(
            "Bonds",
            "Curvas globais: EUA, Brasil e Alemanha.",
            "tradingview_macro_bonds_v1",
            "OTCB:US10Y",
            "#FF9800",
            [
                ("OTCB:US02Y", "#FACC15", "2Y USA"),
                ("OTCB:US30Y", "#00BFFF", "30Y USA"),
                ("BMFBOVESPA:DI1F2029", "#22C55E", "DI1F2029 BR"),
                ("BMFBOVESPA:DI1F2032", "#14B8A6", "DI1F2032 BR"),
                ("BMFBOVESPA:DI1F2035", "#84CC16", "DI1F2035 BR"),
                ("OANDA:DE10YBEUR", "#38BDF8", "10Y Alemanha"),
            ],
        )


def render_tv_corr(container_id, main_sym, comp_sym, interval, height):
    """(Mantido para compatibilidade se necessário futuramente)"""
    pass

def pagina_painel_controle():
    import time
    import json
    st.markdown("### ⚙️ Central de Controle e Sincronização")
    st.write("Monitore o status das fontes de dados do seu terminal e force atualizações na nuvem de forma manual.")

    # 1. Busca dados do Supabase
    val_rtd, time_rtd, dt_rtd = fetch_app_state_with_time("dados_mercado")
    val_globais, time_globais, dt_globais = fetch_app_state_with_time("mercados_globais")
    val_ia, time_ia, dt_ia = fetch_app_state_with_time("ai_insight")
    val_report, time_report, dt_report = fetch_app_state_with_time("market_report")
    val_cal, time_cal, dt_cal = fetch_app_state_with_time("calendario_economico")
    val_fluxo, time_fluxo, dt_fluxo = fetch_app_state_with_time("fluxo_estrangeiro_b3")
    val_focus, time_focus, dt_focus = fetch_app_state_with_time("boletim_focus")

    # Calcula se os dados estão atualizados (B3 RTD - 30 segundos de tolerância)
    is_rtd_ok = False
    if dt_rtd:
        try:
            diff = (datetime.now(timezone.utc) - dt_rtd).total_seconds()
            if abs(diff) < 30:
                is_rtd_ok = True
        except Exception:
            pass

    # Status e cor dos badges
    def get_badge(status_ok, text_ok="🟢 ATUALIZADO", text_err="🔴 DESATUALIZADO / OFFLINE"):
        if status_ok:
            return f"<span style='color: #00FFA3; font-weight: bold;'>{text_ok}</span>"
        else:
            return f"<span style='color: #FF4B4B; font-weight: bold;'>{text_err}</span>"

    status_rtd = get_badge(is_rtd_ok, "🟢 EM OPERAÇÃO (TEMPO REAL)", "🔴 BRIDGE INDISPONÍVEL / OFFLINE")
    status_globais = get_badge(time_globais != "---" and "Erro" not in time_globais, "🟢 ATIVO", "🔴 INDISPONÍVEL")
    status_ia = get_badge(time_ia != "---" and "Erro" not in time_ia, "🟢 ATIVO", "🔴 INDISPONÍVEL")
    status_report = get_badge(time_report != "---" and "Erro" not in time_report, "🟢 ATIVO", "🔴 INDISPONÍVEL")
    status_cal = get_badge(time_cal != "---" and "Erro" not in time_cal, "🟢 ATIVO", "🔴 INDISPONÍVEL")
    status_fluxo = get_badge(time_fluxo != "---" and "Erro" not in time_fluxo, "🟢 ATIVO", "🔴 INDISPONÍVEL")
    status_focus = get_badge(time_focus != "---" and "Erro" not in time_focus, "🟢 ATIVO", "🔴 INDISPONÍVEL")

    # Exibe Grid de Status
    st.markdown("""
        <style>
        .status-table { width: 100%; border-collapse: collapse; margin-top: 15px; margin-bottom: 25px; }
        .status-table th, .status-table td { padding: 12px; text-align: left; border-bottom: 1px solid #222; }
        .status-table th { background-color: #111; color: #FF9800; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 1px; }
        .status-table td { font-size: 0.85rem; }
        </style>
    """, unsafe_allow_html=True)

    html = f"""
    <table class="status-table">
        <tr>
            <th>Fonte de Dados</th>
            <th>Status do Feed</th>
            <th>Última Sincronização (Brasília)</th>
            <th>Tipo</th>
        </tr>
        <tr>
            <td><b>ProfitChart B3 (RTD Local)</b></td>
            <td>{status_rtd}</td>
            <td>{time_rtd}</td>
            <td>Local COM (WIN/WDO)</td>
        </tr>
        <tr>
            <td><b>Mercados Globais (Yahoo Finance)</b></td>
            <td>{status_globais}</td>
            <td>{time_globais}</td>
            <td>Nuvem API (S&P 500, Moedas, Commodities)</td>
        </tr>
        <tr>
            <td><b>IA Analista (Gemini)</b></td>
            <td>{status_ia}</td>
            <td>{time_ia}</td>
            <td>Nuvem LLM</td>
        </tr>
        <tr>
            <td><b>Market Report (Gemini + RSS)</b></td>
            <td>{status_report}</td>
            <td>{time_report}</td>
            <td>Nuvem LLM / RSS Notícias</td>
        </tr>
        <tr>
            <td><b>Calendário Econômico (ForexFactory)</b></td>
            <td>{status_cal}</td>
            <td>{time_cal}</td>
            <td>Nuvem Scraping</td>
        </tr>
        <tr>
            <td><b>Fluxo Estrangeiro (B3 / DDM)</b></td>
            <td>{status_fluxo}</td>
            <td>{time_fluxo}</td>
            <td>Nuvem Scraping</td>
        </tr>
        <tr>
            <td><b>Boletim Focus (BCB)</b></td>
            <td>{status_focus}</td>
            <td>{time_focus}</td>
            <td>API Pública</td>
        </tr>
    </table>
    """
    st.markdown(html, unsafe_allow_html=True)

    st.warning("""
        ⚠️ **Nota sobre o ProfitChart B3 (RTD Local):** As cotações do Mini Índice (WIN), Mini Dólar (WDO) e a Escada de Níveis requerem que a sua plataforma ProfitChart esteja aberta localmente com o script `dashboard_bridge.py` em execução no seu computador. Essa conexão COM local não pode ser atualizada em nuvem sem a bridge física ativa.
    """)

    st.markdown("### 🔄 Sincronização Manual na Nuvem")
    st.write("Se o seu computador estiver desligado (sem a bridge rodando) ou você queira forçar a atualização imediata dos mercados globais, IA e notícias, use os botões abaixo:")

    c1, c2, c3 = st.columns(3)

    def run_update_task(name, func, key, success_msg):
        with st.spinner(f"Atualizando {name}..."):
            try:
                result = func()
                paths_map = {
                    "mercados_globais": ["mercados_globais.json", "execution/mercados_globais.json"],
                    "ai_insight": ["ai_insight.json", "execution/ai_insight.json"],
                    "market_report": ["market_report.json", "execution/market_report.json"],
                    "calendario_economico": ["calendario_economico.json", "execution/calendario_economico.json"],
                    "fluxo_estrangeiro_b3": ["fluxo_estrangeiro.json", "execution/fluxo_estrangeiro.json"],
                    "boletim_focus": ["focus_bcb.json", "execution/focus_bcb.json"]
                }
                
                if key == "mercados_globais":
                    for p in paths_map[key]:
                        if os.path.exists(p):
                            with open(p, "r", encoding="utf-8") as f:
                                data = json.load(f)
                                ok, warning = sync_app_state_value(key, data)
                                if not ok:
                                    st.warning(warning or f"Nao foi possivel sincronizar {name}.")
                                    return
                            break
                            
                elif key == "ai_insight":
                    for p in paths_map[key]:
                        if os.path.exists(p):
                            with open(p, "r", encoding="utf-8") as f:
                                new_insight = json.load(f)
                                ok, warning = sync_app_state_value(key, new_insight)
                                if not ok:
                                    st.warning(warning or f"Nao foi possivel sincronizar {name}.")
                                    return
                                
                                try:
                                    res = supabase.table("app_state").select("value").eq("key", "ai_insight_history").execute()
                                    history = res.data[0]["value"] if res.data else []
                                    if not isinstance(history, list): history = []
                                    
                                    history.append({
                                        "sentiment": new_insight.get("sentiment", "NEUTRO"),
                                        "updated_at": new_insight.get("updated_at", ""),
                                        "insight": new_insight.get("insight", ""),
                                        "macro_regime": new_insight.get("macro_regime", ""),
                                        "confidence": new_insight.get("confidence", ""),
                                        "macro_score": new_insight.get("macro_score", 0),
                                        "curve_regime": new_insight.get("curve_regime", ""),
                                        "curve_bias": new_insight.get("curve_bias", ""),
                                        "id": int(time.time())
                                    })
                                    history = history[-5:]
                                    ok, warning = sync_app_state_value("ai_insight_history", history)
                                    if not ok and warning:
                                        print(f"[WARN] ai_insight_history sync: {warning}")
                                except Exception as he:
                                    print(f"Erro histórico: {he}")
                            break
                            
                elif key == "market_report":
                    for p in paths_map[key]:
                        if os.path.exists(p):
                            with open(p, "r", encoding="utf-8") as f:
                                ok, warning = sync_app_state_value(key, json.load(f))
                                if not ok:
                                    st.warning(warning or f"Nao foi possivel sincronizar {name}.")
                                    return
                            break
                            
                elif key == "calendario_economico":
                    for p in paths_map[key]:
                        if os.path.exists(p):
                            with open(p, "r", encoding="utf-8") as f:
                                ok, warning = sync_app_state_value(key, json.load(f))
                                if not ok:
                                    st.warning(warning or f"Nao foi possivel sincronizar {name}.")
                                    return
                            break
                            
                elif key == "fluxo_estrangeiro_b3":
                    for p in paths_map[key]:
                        if os.path.exists(p):
                            with open(p, "r", encoding="utf-8") as f:
                                ok, warning = sync_app_state_value(key, json.load(f))
                                if not ok:
                                    st.warning(warning or f"Nao foi possivel sincronizar {name}.")
                                    return
                            break
                            
                elif key == "boletim_focus":
                    for p in paths_map[key]:
                        if os.path.exists(p):
                            with open(p, "r", encoding="utf-8") as f:
                                ok, warning = sync_app_state_value(key, json.load(f))
                                if not ok:
                                    st.warning(warning or f"Nao foi possivel sincronizar {name}.")
                                    return
                            break
                            
                st.success(success_msg)
                st.rerun()
            except Exception as e:
                st.error(f"Erro ao atualizar {name}: {e}")

    with c1:
        if st.button("📊 Atualizar Mercados Globais", use_container_width=True, help="Baixa cotações do Yahoo Finance em lote e sincroniza com o banco de dados."):
            import sys
            exec_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'execution'))
            if exec_path not in sys.path: sys.path.append(exec_path)
            from fetch_global_markets import fetch_global_data
            run_update_task("Mercados Globais", fetch_global_data, "mercados_globais", "✅ Mercados Globais atualizados com sucesso!")

    with c2:
        if st.button("🤖 Atualizar IA & Market Report", use_container_width=True, help="Executa a análise macro da IA com Gemini e o Market Report de notícias diário."):
            import sys
            exec_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'execution'))
            if exec_path not in sys.path: sys.path.append(exec_path)
            from ai_analyst import generate_macro_insight
            from market_report import generate_market_report
            
            with st.spinner("Atualizando IA Analista..."):
                try:
                    generate_macro_insight()
                    paths = ["ai_insight.json", "execution/ai_insight.json"]
                    for p in paths:
                        if os.path.exists(p):
                            with open(p, "r", encoding="utf-8") as f:
                                new_insight = json.load(f)
                                ok, warning = sync_app_state_value("ai_insight", new_insight)
                                if not ok:
                                    st.warning(warning or "Nao foi possivel sincronizar a analise da IA.")
                                    return
                                
                                try:
                                    res = supabase.table("app_state").select("value").eq("key", "ai_insight_history").execute()
                                    history = res.data[0]["value"] if res.data else []
                                    if not isinstance(history, list): history = []
                                    
                                    history.append({
                                        "sentiment": new_insight.get("sentiment", "NEUTRO"),
                                        "updated_at": new_insight.get("updated_at", ""),
                                        "insight": new_insight.get("insight", ""),
                                        "macro_regime": new_insight.get("macro_regime", ""),
                                        "confidence": new_insight.get("confidence", ""),
                                        "macro_score": new_insight.get("macro_score", 0),
                                        "curve_regime": new_insight.get("curve_regime", ""),
                                        "curve_bias": new_insight.get("curve_bias", ""),
                                        "id": int(time.time())
                                    })
                                    history = history[-5:]
                                    ok, warning = sync_app_state_value("ai_insight_history", history)
                                    if not ok and warning:
                                        print(f"[WARN] ai_insight_history sync: {warning}")
                                except Exception as he:
                                    print(f"Erro histórico: {he}")
                            break
                    st.success("✅ Análise da IA atualizada!")
                except Exception as e:
                    st.error(f"Erro na IA Analista: {e}")
                    
            with st.spinner("Atualizando Market Report..."):
                try:
                    generate_market_report(force=True)
                    paths = ["market_report.json", "execution/market_report.json"]
                    for p in paths:
                        if os.path.exists(p):
                            with open(p, "r", encoding="utf-8") as f:
                                ok, warning = sync_app_state_value("market_report", json.load(f))
                                if not ok:
                                    st.warning(warning or "Nao foi possivel sincronizar Market Report.")
                                    return
                            break
                    daily_paths = ["market_report_daily.json", "execution/market_report_daily.json"]
                    for p in daily_paths:
                        if os.path.exists(p):
                            with open(p, "r", encoding="utf-8") as f:
                                ok, warning = sync_app_state_value("market_report_daily", json.load(f))
                                if not ok:
                                    st.warning(warning or "Nao foi possivel sincronizar historico diario do Market Report.")
                                    return
                            break
                    st.success("✅ Market Report atualizado!")
                except Exception as e:
                    st.error(f"Erro no Market Report: {e}")
            st.rerun()

    with c3:
        if st.button("📅 Atualizar Dados (Calendário, Fluxo, Focus)", use_container_width=True, help="Busca o calendário, fluxo B3 e Focus."):
            import sys
            exec_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'execution'))
            if exec_path not in sys.path: sys.path.append(exec_path)
            
            from fetch_calendar import fetch_economic_calendar
            run_update_task("Calendário Econômico", fetch_economic_calendar, "calendario_economico", "✅ Calendário Econômico atualizado!")
            
            from fetch_foreign_flow import fetch_foreign_flow, save_flow_data
            def fetch_and_save_flow():
                records = fetch_foreign_flow()
                if records: save_flow_data(records, "fluxo_estrangeiro.json")
            run_update_task("Fluxo Estrangeiro", fetch_and_save_flow, "fluxo_estrangeiro_b3", "✅ Fluxo Estrangeiro atualizado!")
            
            from fetch_focus import fetch_focus_bcb, save_focus_data
            def fetch_and_save_focus():
                data = fetch_focus_bcb()
                if data: save_focus_data(data, "focus_bcb.json")
            run_update_task("Boletim Focus", fetch_and_save_focus, "boletim_focus", "✅ Boletim Focus atualizado!")

    st.markdown("---")
    st.markdown("### ☁️ Configurando Automação 24/7 Gratuita no GitHub")
    st.write("""
        Para que o seu site **nunca fique desatualizado** (mesmo com o seu computador desligado), configuramos um fluxo de trabalho automatizado (GitHub Actions) no seu repositório.
        
        Esse fluxo roda silenciosamente na nuvem a **cada hora**, buscando cotações globais, notícias e gerando relatórios de IA automaticamente.
        
        **Para ativá-lo, você só precisa cadastrar suas credenciais nas configurações do seu repositório no GitHub:**
        
        1. Acesse o seu repositório no **GitHub**.
        2. Vá em **Settings** (Configurações) > **Secrets and variables** > **Actions**.
        3. Clique em **New repository secret** (Novo segredo) e adicione as seguintes chaves:
           - Nome: `SUPABASE_URL` | Valor: *Sua URL do Supabase*
           - Nome: `SUPABASE_SERVICE_ROLE` | Valor: *Sua service_role key do Supabase*
           - Nome: `GEMINI_API_KEY` | Valor: *Sua API Key do Google Gemini*
        
        Pronto! Com isso cadastrado, o GitHub atualizará o seu site automaticamente 24 horas por dia, 7 dias por semana, sem que você precise deixar nenhum código rodando no seu computador!
    """)

def pagina_gestao_risco():
    """Página de Gestão de Risco com backtest manual por estratégia."""
    import json, base64, uuid
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    st.markdown("### 🛡️ Gestão de Risco & Backtest por Estratégia")
    st.markdown("<p style='color:#888; font-size:0.9rem;'>Registre suas operações por estratégia, acompanhe a curva de capital individual, compare estratégias e receba análise IA de performance.</p>", unsafe_allow_html=True)

    # ── CSS ──
    st.markdown("""
    <style>
    .risk-card { background: #111; border: 1px solid #222; border-radius: 8px; padding: 20px; margin-bottom: 16px; }
    .risk-metric { font-size: 2rem; font-weight: bold; color: #FFF; }
    .risk-label  { font-size: 0.7rem; color: #888; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 4px; }
    .ai-feedback  { background:#0A0A0A; border:1px solid #333; border-left:6px solid #FF9800; padding:20px; border-radius:8px; margin-top:12px; color:#E0E0E0; font-size:0.9rem; line-height:1.7; }
    </style>
    """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════
    # BLOCO 1 — Calculadora de Tamanho de Posição
    # ══════════════════════════════════════════════════
    with st.expander("📐 Calculadora de Tamanho de Posição", expanded=False):
        col_p, col_m = st.columns([1, 1])
        with col_p:
            capital_total   = st.number_input("Capital Total (R$)", min_value=1000.0, value=10000.0, step=500.0, format="%.2f", key="risk_capital")
            risco_por_trade = st.slider("Risco por Trade (%)", 0.5, 5.0, 1.0, 0.25, key="risk_pct", format="%.2f%%")
        with col_m:
            stop_pontos     = st.number_input("Stop Loss (pontos)", min_value=1.0, value=50.0, step=5.0, key="risk_stop")
            valor_por_ponto = st.number_input("Valor por Contrato/Ponto (R$)", min_value=0.1, value=0.20, step=0.05, format="%.2f", key="risk_vpp")

        risco_valor = capital_total * (risco_por_trade / 100.0)
        denom       = stop_pontos * valor_por_ponto
        tamanho_pos = risco_valor / denom if denom > 0 else 0
        alv_2r      = risco_valor * 2
        stop_total  = stop_pontos * valor_por_ponto * tamanho_pos

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Risco (R$)",      f"R$ {risco_valor:,.2f}")
        m2.metric("Contratos",       f"{tamanho_pos:.1f}")
        m3.metric("Stop Total (R$)", f"R$ {stop_total:,.2f}")
        m4.metric("Alvo 2:1 (R$)",  f"R$ {alv_2r:,.2f}")

    st.markdown("---")

    # ══════════════════════════════════════════════════
    # BLOCO 2 — Registro Manual de Trades por Estratégia
    # ══════════════════════════════════════════════════
    st.markdown("#### ✏️ Registro de Trades por Estratégia")

    # Carrega histórico persistido (Supabase → local)
    if "manual_trades" not in st.session_state:
        st.session_state["manual_trades"] = load_manual_trades()

    trades: list = st.session_state["manual_trades"]

    # Deriva lista de estratégias já cadastradas
    estrategias_existentes = sorted(set(t.get("estrategia", "") for t in trades if t.get("estrategia")))

    col_form, col_hist = st.columns([1, 1.4], gap="large")

    with col_form:
        st.markdown("<div style='background:#0d0d0d; border:1px solid #1e1e1e; border-radius:10px; padding:18px;'>", unsafe_allow_html=True)
        st.markdown("<div class='risk-label'>📋 Novo Trade</div>", unsafe_allow_html=True)

        with st.form("form_novo_trade", clear_on_submit=True):
            opcoes_estrategia = ["+ Nova Estratégia"] + estrategias_existentes
            sel = st.selectbox("Estratégia", opcoes_estrategia, key="form_sel_estrategia")
            nova_nome = ""
            if sel == "+ Nova Estratégia":
                nova_nome = st.text_input("Nome da nova estratégia",
                                          placeholder="Ex: Wyckoff M5, Filippo M15, ICT HTF...",
                                          key="form_nova_nome")
            estrategia_final = nova_nome.strip() if sel == "+ Nova Estratégia" else sel

            c1f, c2f = st.columns(2)
            with c1f:
                ativo    = st.text_input("Ativo", placeholder="WIN, WDO, EURUSD", key="form_ativo")
                direcao  = st.selectbox("Direção", ["Compra", "Venda"], key="form_dir")
            with c2f:
                resultado  = st.number_input("Resultado (R$)", value=0.0, step=10.0, format="%.2f", key="form_res")
                data_trade = st.date_input("Data do Trade", value=datetime.now().date(), key="form_data")

            obs = st.text_input("Observação (opcional)",
                                placeholder="Ex: Setup no teste de estrutura...", key="form_obs")

            submitted = st.form_submit_button("➕ Registrar Trade",
                                              use_container_width=True, type="primary")

        if submitted:
            if not estrategia_final:
                st.warning("⚠️ Informe o nome da estratégia.")
            else:
                novo_trade = {
                    "id":         str(uuid.uuid4())[:8],
                    "estrategia": estrategia_final,
                    "ativo":      ativo.upper().strip() if ativo else "—",
                    "direcao":    direcao,
                    "resultado":  resultado,
                    "data":       str(data_trade),
                    "obs":        obs,
                }
                trades.append(novo_trade)
                st.session_state["manual_trades"] = trades
                save_manual_trades(trades)
                st.success(f"✅ Trade registrado em **{estrategia_final}**: R$ {resultado:+.2f}")
                st.rerun()

    with col_hist:
        if trades:
            st.markdown(f"<div class='risk-label'>📚 Histórico — {len(trades)} operações</div>",
                        unsafe_allow_html=True)
            df_hist = pd.DataFrame(trades)
            df_hist["resultado"] = pd.to_numeric(df_hist["resultado"], errors="coerce").fillna(0)
            df_hist_show = df_hist[["data", "estrategia", "ativo", "direcao", "resultado", "obs"]].copy()
            df_hist_show.columns = ["Data", "Estratégia", "Ativo", "Dir.", "Resultado (R$)", "Obs."]
            df_hist_show = df_hist_show.sort_values("Data", ascending=False)

            def color_row(row):
                c = "#003318" if row["Resultado (R$)"] >= 0 else "#330000"
                return [f"background-color: {c}" for _ in row]

            styled = (
                df_hist_show.style
                .apply(color_row, axis=1)
                .format({"Resultado (R$)": "R$ {:+,.2f}"})
                .set_properties(**{"font-size": "0.78rem", "color": "#DDD"})
            )
            st.dataframe(styled, use_container_width=True, hide_index=True, height=260)

            # Exclusão por índice
            col_del1, col_del2 = st.columns([2, 1])
            with col_del1:
                labels_del = [
                    f"{t['data']} | {t['estrategia']} | R$ {float(t.get('resultado', 0)):+.2f}"
                    for t in trades
                ]
                idx_del = st.selectbox(
                    "Selecionar trade para excluir",
                    options=range(len(trades)),
                    format_func=lambda i: labels_del[i],
                    key="sel_del_trade"
                )
            with col_del2:
                st.markdown("<div style='margin-top:24px;'></div>", unsafe_allow_html=True)
                if st.button("🗑️ Excluir", key="btn_del_trade", use_container_width=True):
                    trades.pop(idx_del)
                    st.session_state["manual_trades"] = trades
                    save_manual_trades(trades)
                    st.success("Trade excluído.")
                    st.rerun()

            if st.button("🗑️ Limpar TODO o histórico", key="btn_clear_all", type="secondary"):
                st.session_state["manual_trades"] = []
                save_manual_trades([])
                st.rerun()
        else:
            st.info("📭 Nenhum trade registrado ainda. Use o formulário ao lado para começar.")

    st.markdown("---")

    # ══════════════════════════════════════════════════
    # BLOCO 3 — Gráficos e Métricas por Estratégia
    # ══════════════════════════════════════════════════
    st.markdown("#### 📊 Performance por Estratégia")

    if not trades:
        st.info("✏️ Registre operações para visualizar os gráficos de rentabilidade.")
    else:
        df_all = pd.DataFrame(trades)
        df_all["resultado"] = pd.to_numeric(df_all["resultado"], errors="coerce").fillna(0)
        df_all["data"]      = pd.to_datetime(df_all["data"], errors="coerce")
        df_all = df_all.sort_values("data").reset_index(drop=True)

        estrategias_list = sorted(df_all["estrategia"].unique().tolist())

        tab_todas, *tabs_ind = st.tabs(
            ["📈 Todas Consolidadas"] + [f"🎯 {e}" for e in estrategias_list]
        )

        # ── Aba Consolidada ──
        with tab_todas:
            _plot_rentabilidade_df(df_all, "Curva de Capital Consolidada — Todas as Estratégias")

            st.markdown("##### 📊 Comparativo entre Estratégias")
            resumo = []
            for est in estrategias_list:
                df_e  = df_all[df_all["estrategia"] == est]
                res   = df_e["resultado"].tolist()
                n_e   = len(res)
                wins  = sum(1 for r in res if r > 0)
                total = sum(res)
                wr_e  = wins / n_e * 100 if n_e else 0
                gross_w = sum(r for r in res if r > 0)
                gross_l = abs(sum(r for r in res if r < 0))
                pf_e = gross_w / gross_l if gross_l else float("inf")
                peak_e = s_e = max_dd_e = 0.0
                for r in res:
                    s_e += r
                    if s_e > peak_e: peak_e = s_e
                    dd_e = peak_e - s_e
                    if dd_e > max_dd_e: max_dd_e = dd_e
                resumo.append({
                    "Estratégia":    est,
                    "Trades":        n_e,
                    "Win Rate":      f"{wr_e:.0f}%",
                    "Total (R$)":    total,
                    "Profit Factor": f"{pf_e:.2f}" if pf_e != float("inf") else "∞",
                    "Max DD (R$)":   max_dd_e,
                })

            df_resumo = pd.DataFrame(resumo)
            cores_bar = ["#00FFA3" if v >= 0 else "#FF4B4B" for v in df_resumo["Total (R$)"]]

            fig_comp = go.Figure(go.Bar(
                x=df_resumo["Estratégia"],
                y=df_resumo["Total (R$)"],
                marker_color=cores_bar,
                text=[f"R$ {v:+,.0f}" for v in df_resumo["Total (R$)"]],
                textposition="outside",
                marker_line_width=0,
            ))
            fig_comp.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#E0E0E0", size=11),
                margin=dict(l=10, r=10, t=30, b=10), height=280,
                xaxis=dict(showgrid=False, color="#888"),
                yaxis=dict(showgrid=True, gridcolor="#1a1a1a", zeroline=True,
                           zerolinecolor="#444", tickprefix="R$ "),
                showlegend=False,
            )
            fig_comp.add_hline(y=0, line_dash="dot", line_color="#444")
            st.plotly_chart(fig_comp, use_container_width=True, config={"displayModeBar": False})

            def _color_total(val):
                return "color: #00FFA3; font-weight:bold" if val >= 0 else "color: #FF4B4B; font-weight:bold"

            resumo_styler = df_resumo.style
            if hasattr(resumo_styler, "map"):
                resumo_styler = resumo_styler.map(_color_total, subset=["Total (R$)"])
            else:
                resumo_styler = resumo_styler.applymap(_color_total, subset=["Total (R$)"])
            styled_resumo = (
                resumo_styler
                .format({"Total (R$)": "R$ {:+,.2f}", "Max DD (R$)": "R$ {:.2f}"})
                .set_properties(**{"font-size": "0.82rem", "color": "#DDD",
                                   "background-color": "#0d0d0d"})
            )
            st.dataframe(styled_resumo, use_container_width=True, hide_index=True)

        # ── Abas individuais ──
        for tab_e, est in zip(tabs_ind, estrategias_list):
            with tab_e:
                df_e = df_all[df_all["estrategia"] == est].reset_index(drop=True)
                _plot_rentabilidade_df(df_e, f"Curva de Capital — {est}")

    st.markdown("---")

    # ══════════════════════════════════════════════════
    # BLOCO 4 — Upload + Análise IA
    # ══════════════════════════════════════════════════
    with st.expander("🤖 Análise IA de Relatório (Upload CSV/Excel/Imagem)", expanded=False):
        st.markdown("<p style='color:#666; font-size:0.8rem;'>Aceita: Excel (.xlsx), CSV, ou imagem (print/foto)</p>",
                    unsafe_allow_html=True)
        col_up, col_p2 = st.columns([1.5, 1])
        with col_up:
            uploaded_file = st.file_uploader(
                "Selecione o arquivo",
                type=["xlsx", "csv", "png", "jpg", "jpeg", "webp"],
                label_visibility="collapsed", key="risk_upload"
            )
        with col_p2:
            cap_ia = st.number_input("Capital (R$)", min_value=1000.0, value=10000.0,
                                     step=500.0, format="%.2f", key="risk_cap_ia")
            rpc_ia = st.slider("Risco/Trade (%)", 0.5, 5.0, 1.0, 0.25,
                               key="risk_pct_ia", format="%.2f%%")
            stp_ia = st.number_input("Stop (pts)", min_value=1.0, value=50.0,
                                     step=5.0, key="risk_stop_ia")
            vpp_ia = st.number_input("Vlr/Ponto (R$)", min_value=0.1, value=0.20,
                                     step=0.05, format="%.2f", key="risk_vpp_ia")

        if uploaded_file:
            file_ext = uploaded_file.name.split(".")[-1].lower()
            df_trades_ia = None
            img_b64 = None

            if file_ext in ["csv", "xlsx"]:
                try:
                    df_trades_ia = (pd.read_csv(uploaded_file) if file_ext == "csv"
                                    else pd.read_excel(uploaded_file))
                    st.success(f"✅ {len(df_trades_ia)} registros carregados.")
                    with st.expander("Pré-visualização", expanded=False):
                        st.dataframe(df_trades_ia.head(30), use_container_width=True, hide_index=True)
                except Exception as e:
                    st.error(f"Erro ao ler arquivo: {e}")
            else:
                img_bytes = uploaded_file.read()
                img_b64 = base64.b64encode(img_bytes).decode()
                st.image(img_bytes, caption="Relatório enviado", use_column_width=True)

            risco_v = cap_ia * (rpc_ia / 100.0)
            den2    = stp_ia * vpp_ia
            tam_ia  = risco_v / den2 if den2 > 0 else 0

            if st.button("🤖 Analisar com IA", use_container_width=True,
                         type="primary", key="btn_ia_analise"):
                api_key = os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY")
                try:
                    api_key = api_key or st.secrets.get(
                        "GOOGLE_API_KEY", st.secrets.get("GEMINI_API_KEY", ""))
                except Exception:
                    pass
                if not api_key:
                    st.error("🔑 Configure GOOGLE_API_KEY nos segredos.")
                else:
                    import google.generativeai as genai
                    genai.configure(api_key=api_key)
                    prompt_base = f"""
Você é um Coach de Trading e Gestor de Risco profissional. Analise o relatório e forneça:

PARÂMETROS: Capital R$ {cap_ia:,.2f} | Risco {rpc_ia:.2f}% | Stop {stp_ia:.0f}pts | Vlr/ponto R$ {vpp_ia:.2f} | Sizing calculado: {tam_ia:.1f} contratos

RELATÓRIO: {df_trades_ia.to_string() if df_trades_ia is not None else '[Imagem enviada]'}

Forneça:
1. 📊 DIAGNÓSTICO (Win Rate, Profit Factor, Max Drawdown, R/R médio)
2. 📐 SIZING RECOMENDADO (manter, aumentar ou reduzir {tam_ia:.1f} contratos — justifique)
3. 🧠 ANÁLISE PSICOLÓGICA (overtrading, revenge, FOMO, cortar gains, deixar losses)
4. 🎯 TOP 3 AÇÕES para os próximos 30 dias
5. ⚠️ ALERTAS DE RISCO críticos
Seja direto e cirúrgico.
                    """
                    with st.spinner("🧠 Analisando..."):
                        try:
                            model = genai.GenerativeModel("gemini-2.0-flash")
                            if img_b64:
                                mime = f"image/{file_ext if file_ext != 'jpg' else 'jpeg'}"
                                response = model.generate_content(
                                    [prompt_base, {"mime_type": mime, "data": img_b64}])
                            else:
                                response = model.generate_content(prompt_base)
                            st.session_state["risco_ia_resultado"] = response.text
                        except Exception as e:
                            st.error(f"Erro na análise IA: {e}")

        if "risco_ia_resultado" in st.session_state:
            ia_text = st.session_state["risco_ia_resultado"]
            st.markdown(
                f"<div class='ai-feedback'>"
                f"{sanitize_text(ia_text).replace(chr(10), '<br>')}"
                f"</div>",
                unsafe_allow_html=True
            )


def _plot_rentabilidade_df(df: "pd.DataFrame", titulo: str = "Curva de Capital"):
    """Plota gráfico premium de rentabilidade acumulada a partir de um DataFrame de trades."""
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    if df.empty:
        st.info("Nenhum dado para exibir.")
        return

    df = df.copy()
    df["resultado"] = pd.to_numeric(df["resultado"], errors="coerce").fillna(0)
    resultados = df["resultado"].tolist()
    acumulado  = []
    soma = 0.0
    for r in resultados:
        soma += r
        acumulado.append(soma)

    n      = len(acumulado)
    trades = list(range(1, n + 1))
    ultimo = acumulado[-1]
    wins   = sum(1 for r in resultados if r > 0)
    wr     = wins / n * 100 if n else 0
    gross_w = sum(r for r in resultados if r > 0)
    gross_l = abs(sum(r for r in resultados if r < 0))
    pf      = gross_w / gross_l if gross_l else float("inf")
    peak = max_dd = 0.0
    for v in acumulado:
        if v > peak: peak = v
        dd = peak - v
        if dd > max_dd: max_dd = dd

    cores_barras = ["#00FFA3" if r >= 0 else "#FF4B4B" for r in resultados]
    color_tot    = "#00FFA3" if ultimo >= 0 else "#FF4B4B"
    pf_str = f"{pf:.2f}" if pf != float("inf") else "∞"

    fig = make_subplots(
        rows=2, cols=1, shared_xaxes=True,
        row_heights=[0.65, 0.35], vertical_spacing=0.06,
        subplot_titles=(titulo, "Resultado por Trade (R$)")
    )
    fig.add_trace(go.Scatter(
        x=trades, y=acumulado, mode="lines+markers",
        line=dict(color="#FF9800", width=2.5),
        marker=dict(size=5, color="#FF9800"),
        fill="tozeroy", fillcolor="rgba(255,152,0,0.08)",
        name="Capital Acumulado"
    ), row=1, col=1)
    fig.add_hline(y=0, line_dash="dot", line_color="#444", row=1, col=1)
    fig.add_trace(go.Bar(
        x=trades, y=resultados,
        marker_color=cores_barras, name="Resultado",
        marker_line_width=0, opacity=0.85
    ), row=2, col=1)
    fig.add_annotation(
        text=(f"Total: R$ {ultimo:+,.2f} | Win Rate: {wr:.0f}% | "
              f"Profit Factor: {pf_str} | Max DD: R$ {max_dd:,.2f}"),
        xref="paper", yref="paper", x=0.5, y=1.02,
        showarrow=False, font=dict(size=11, color=color_tot),
        bgcolor="rgba(0,0,0,0.5)"
    )
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family='"Roboto Mono", monospace', color="#E0E0E0", size=11),
        margin=dict(l=10, r=10, t=40, b=10), height=440,
        showlegend=False,
        xaxis2=dict(title="Nº do Trade", showgrid=False, color="#888"),
        yaxis=dict(showgrid=True, gridcolor="#1a1a1a", zeroline=True,
                   zerolinecolor="#444", tickprefix="R$ "),
        yaxis2=dict(showgrid=True, gridcolor="#1a1a1a", zeroline=True,
                    zerolinecolor="#444", tickprefix="R$ "),
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total Trades",  n)
    c2.metric("Win Rate",      f"{wr:.1f}%")
    c3.metric("Resultado",     f"R$ {ultimo:+,.2f}")
    c4.metric("Profit Factor", pf_str)
    c5.metric("Max Drawdown",  f"R$ {max_dd:,.2f}")


def _plot_rentabilidade(resultados: list, acumulado: list):
    """Wrapper legado — mantido para compatibilidade."""
    df_tmp = pd.DataFrame({"resultado": resultados})
    _plot_rentabilidade_df(df_tmp)


# Navegação na Barra Lateral

@st.fragment(run_every=30)
def sidebar_news():
    if "sidebar_news_translate" not in st.session_state:
        st.session_state.sidebar_news_translate = False
    if "sidebar_news_refresh_nonce" not in st.session_state:
        st.session_state.sidebar_news_refresh_nonce = 0
    if "sidebar_news_zoom" not in st.session_state:
        st.session_state.sidebar_news_zoom = 1.0
    if "bb_translation_cache" not in st.session_state:
        st.session_state.bb_translation_cache = {}
    if "sidebar_news_history" not in st.session_state:
        st.session_state.sidebar_news_history = []

    refresh_col, translate_col, zoom_out_col, zoom_in_col = st.columns([1.35, 1.35, 0.55, 0.55])
    with refresh_col:
        if st.button("Atualizar", use_container_width=True, key="sidebar_news_refresh"):
            load_bloomberg_news_feed.clear()
            st.session_state.sidebar_news_refresh_nonce += 1
    with translate_col:
        translate_label = "Ver EN" if st.session_state.sidebar_news_translate else "Traduzir"
        if st.button(translate_label, use_container_width=True, key="sidebar_news_translate_btn"):
            st.session_state.sidebar_news_translate = not st.session_state.sidebar_news_translate
    with zoom_out_col:
        if st.button("-", use_container_width=True, key="sidebar_news_zoom_out", help="Diminuir fonte do feed NEWS"):
            st.session_state.sidebar_news_zoom = max(0.85, round(st.session_state.sidebar_news_zoom - 0.1, 2))
    with zoom_in_col:
        if st.button("+", use_container_width=True, key="sidebar_news_zoom_in", help="Aumentar fonte do feed NEWS"):
            st.session_state.sidebar_news_zoom = min(1.45, round(st.session_state.sidebar_news_zoom + 0.1, 2))

    news_list, news_sources, news_warnings, feed_loaded_at = load_bloomberg_news_feed(
        st.session_state.sidebar_news_refresh_nonce
    )
    if news_list:
        st.session_state.sidebar_news_history = news_list[:10]
    elif st.session_state.sidebar_news_history:
        news_list = st.session_state.sidebar_news_history
        news_warnings = ["Fonte ao vivo carregando; exibindo historico da sessao."]
    if not news_list:
        st.info("Carregando noticias...")
        return

    def esc(value) -> str:
        return html.escape(str(value or ""), quote=True)

    def translate_sidebar_item(item: dict) -> dict:
        translated_item = dict(item)
        title_original = item.get("title_en") or item.get("title") or item.get("title_pt") or ""
        summary_original = item.get("summary") or item.get("description") or ""
        try:
            from execution.fetch_financial_juice import translate_text_google
            cache = st.session_state.bb_translation_cache
            if title_original:
                title_key = f"title::{title_original}"
                if title_key not in cache:
                    cache[title_key] = translate_text_google(title_original)
                translated_item["title_pt"] = cache[title_key]
            if summary_original and summary_original != title_original:
                summary_key = f"summary::{summary_original}"
                if summary_key not in cache:
                    cache[summary_key] = translate_text_google(summary_original)
                translated_item["summary_pt"] = cache[summary_key]
        except Exception:
            try:
                from execution.fetch_financial_juice import ensure_portuguese_fields
                ensure_portuguese_fields(translated_item)
            except Exception:
                pass
        return translated_item

    def news_text(item) -> str:
        return f"{item.get('title_en', '')} {item.get('title_pt', '')} {item.get('summary', '')}".lower()

    def compact_impact(item):
        text = news_text(item)
        source = str(item.get("source", "")).lower()
        score = 0
        rules = [
            (5, ["fed", "fomc", "powell", "ecb", "boj", "boe", "copom", "bcb", "interest rate", "juros"]),
            (5, ["cpi", "pce", "ppi", "inflation", "inflacao", "inflação"]),
            (4, ["treasury", "treasuries", "yield", "yields", "dxy", "dollar", "oil", "crude", "brent", "wti"]),
            (4, ["payroll", "jobs", "jobless", "gdp", "retail sales", "pmi", "ism"]),
            (4, ["iran", "israel", "china", "russia", "war", "guerra", "sanctions", "ataque"]),
            (3, ["s&p", "nasdaq", "dow", "stocks", "futuros", "bitcoin", "crypto"]),
        ]
        for weight, keywords in rules:
            if any(keyword in text for keyword in keywords):
                score += weight
        if any(word in text for word in ["breaking", "urgent", "alert", "unexpected", "surprise"]):
            score += 3
        if any(name in source for name in ["financial", "reuters", "bloomberg", "cnbc"]):
            score += 1
        if score >= 12:
            return "URGENTE", "#FF2D20"
        if score >= 8:
            return "ALTO", "#FF4B4B"
        if score >= 4:
            return "MEDIO", "#FF9800"
        return "BAIXO", "#94A3B8"

    def sort_key(item):
        impact_label, _ = compact_impact(item)
        impact_rank = {"URGENTE": 0, "ALTO": 1, "MEDIO": 2, "BAIXO": 3}
        try:
            ts = float(item.get("timestamp") or 0)
        except Exception:
            ts = 0
        return (impact_rank.get(impact_label, 9), -ts)

    filtered = sorted(news_list, key=sort_key)[:10]
    translate_enabled = bool(st.session_state.get("sidebar_news_translate", False))
    if translate_enabled:
        with st.spinner("Traduzindo..."):
            filtered = [translate_sidebar_item(item) for item in filtered]
    zoom = float(st.session_state.get("sidebar_news_zoom", 1.0))
    meta_font = 0.68 * zoom
    badge_font = 0.62 * zoom
    title_font = 0.78 * zoom
    line_height = max(1.25, 1.35 * zoom)
    st.markdown(
        f"<div style='text-align:right; font-size:0.65rem; color:#666; margin-bottom:10px;'>NEWS: {esc(feed_loaded_at)} | {len(news_list)} itens | {'PT-BR' if translate_enabled else 'EN'} | Zoom {zoom:.1f}x</div>",
        unsafe_allow_html=True,
    )
    for item in filtered:
        if translate_enabled:
            title_raw = item.get("title_pt") or item.get("title_en") or item.get("title") or "---"
        else:
            title_raw = item.get("title_en") or item.get("title") or item.get("title_pt") or "---"
        title = esc(title_raw)
        published = esc(item.get("published_str", "--:--"))
        source = esc(item.get("source", "Financial Juice"))
        link = safe_external_url(item.get("link"))
        impact_label, impact_color = compact_impact(item)
        st.markdown(
            f'''
            <div style="border-bottom:1px solid #1f2937; padding:9px 0;">
                <div style="display:flex; justify-content:space-between; gap:8px; align-items:center;">
                    <span style="font-size:{meta_font:.2f}rem; color:#94A3B8;">{published} | {source}</span>
                    <span style="font-size:{badge_font:.2f}rem; color:{impact_color}; border:1px solid {impact_color}66; border-radius:4px; padding:1px 5px; font-weight:900;">{impact_label}</span>
                </div>
                <a href="{link}" target="_blank" rel="noopener noreferrer" style="display:block; color:#E5E7EB; text-decoration:none; font-size:{title_font:.2f}rem; line-height:{line_height:.2f}; font-weight:700; margin-top:4px;">{title}</a>
            </div>
            ''',
            unsafe_allow_html=True,
        )

    for warning in news_warnings[:1]:
        st.caption(warning)


def sidebar_clock():
    components.html(
        """
        <div class="clock-card">
          <div class="clock-top">
            <span>Horario Brasilia</span>
            <i></i>
          </div>
          <div id="tts-clock-time" class="clock-time">--:--:--</div>
          <div id="tts-clock-date" class="clock-date">--</div>
          <div id="tts-clock-events" class="clock-events"></div>
        </div>
        <style>
          html, body { margin:0; padding:0; background:transparent; overflow:hidden; font-family:Inter, "Segoe UI", Arial, sans-serif; }
          .clock-card {
            box-sizing:border-box;
            width:100%;
            border:1px solid rgba(51,65,85,.95);
            border-radius:8px;
            padding:12px 13px;
            background:linear-gradient(135deg, rgba(15,23,42,.98), rgba(2,8,23,.98));
            box-shadow:inset 0 1px 0 rgba(148,163,184,.10), 0 10px 24px rgba(0,0,0,.22);
          }
          .clock-top { display:flex; justify-content:space-between; align-items:center; gap:8px; margin-bottom:7px; }
          .clock-top span { color:#94A3B8; font-size:.68rem; font-weight:900; letter-spacing:.08em; text-transform:uppercase; }
          .clock-top i { width:8px; height:8px; border-radius:999px; background:#22C55E; box-shadow:0 0 12px rgba(34,197,94,.9); display:inline-block; }
          .clock-time { color:#F8FAFC; font-size:2.28rem; line-height:.98; font-weight:950; letter-spacing:.02em; font-variant-numeric:tabular-nums; }
          .clock-date { color:#38BDF8; font-size:.78rem; font-weight:800; margin-top:7px; }
          .clock-events { border-top:1px solid rgba(51,65,85,.80); margin-top:10px; padding-top:5px; }
          .clock-event { display:flex; align-items:center; justify-content:space-between; gap:8px; padding:6px 7px; border-radius:6px; margin-top:6px; border:1px solid rgba(51,65,85,.70); background:rgba(15,23,42,.45); }
          .clock-event-main { min-width:0; overflow:hidden; white-space:nowrap; text-overflow:ellipsis; }
          .clock-event-time { color:#F8FAFC; font-size:.76rem; font-weight:950; font-variant-numeric:tabular-nums; }
          .clock-event-label { color:#CBD5E1; font-size:.70rem; font-weight:800; margin-left:5px; }
          .clock-event-status { font-size:.60rem; font-weight:950; letter-spacing:.04em; }
        </style>
        <script>
          const tz = "America/Sao_Paulo";
          const sessions = [
            { time: "09:00", label: "Abertura BR" },
            { time: "10:00", label: "Abertura acoes BR" },
            { time: "10:30", label: "Abertura NY" }
          ];
          const timeFmt = new Intl.DateTimeFormat("pt-BR", { timeZone: tz, hour: "2-digit", minute: "2-digit", second: "2-digit", hour12: false });
          const dateFmt = new Intl.DateTimeFormat("pt-BR", { timeZone: tz, weekday: "short", day: "2-digit", month: "2-digit", year: "numeric" });
          function partsInSaoPaulo(date) {
            const parts = new Intl.DateTimeFormat("en-CA", {
              timeZone: tz, year: "numeric", month: "2-digit", day: "2-digit",
              hour: "2-digit", minute: "2-digit", second: "2-digit", hour12: false
            }).formatToParts(date).reduce((acc, part) => {
              acc[part.type] = part.value;
              return acc;
            }, {});
            return {
              y: Number(parts.year), m: Number(parts.month), d: Number(parts.day),
              hour: Number(parts.hour), minute: Number(parts.minute), second: Number(parts.second)
            };
          }
          function minutesNow(p) { return p.hour * 60 + p.minute + p.second / 60; }
          function minutesOf(value) {
            const [h, m] = value.split(":").map(Number);
            return h * 60 + m;
          }
          function styleFor(status) {
            if (status === "AGORA") return { color:"#22C55E", bg:"rgba(34,197,94,.12)", border:"rgba(34,197,94,.50)" };
            if (status === "PROX") return { color:"#F59E0B", bg:"rgba(245,158,11,.12)", border:"rgba(245,158,11,.45)" };
            if (status === "HOJE") return { color:"#38BDF8", bg:"rgba(56,189,248,.08)", border:"rgba(56,189,248,.28)" };
            return { color:"#64748B", bg:"rgba(15,23,42,.45)", border:"rgba(51,65,85,.70)" };
          }
          function renderEvents(nowMin) {
            let nextIndex = sessions.findIndex((item) => minutesOf(item.time) > nowMin);
            return sessions.map((item, index) => {
              const eventMin = minutesOf(item.time);
              const diff = nowMin - eventMin;
              let status = "HOJE";
              if (diff >= 0 && diff < 30) status = "AGORA";
              else if (diff >= 30) status = "OK";
              else if (index === nextIndex) status = "PROX";
              const st = styleFor(status);
              return `<div class="clock-event" style="border-color:${st.border}; background:${st.bg};">
                <div class="clock-event-main"><span class="clock-event-time">${item.time}</span><span class="clock-event-label">${item.label}</span></div>
                <span class="clock-event-status" style="color:${st.color};">${status}</span>
              </div>`;
            }).join("");
          }
          function tickClock() {
            const now = new Date();
            const p = partsInSaoPaulo(now);
            document.getElementById("tts-clock-time").textContent = timeFmt.format(now);
            document.getElementById("tts-clock-date").textContent = dateFmt.format(now).replace(".", "");
            document.getElementById("tts-clock-events").innerHTML = renderEvents(minutesNow(p));
          }
          tickClock();
          setInterval(tickClock, 1000);
        </script>
        """,
        height=226,
    )

auth_user = require_authenticated_user()
auth_role = st.session_state.get("auth_role", "member")
auth_role_label = "Acesso publico" if not AUTH_REQUIRED else ("Administrador" if auth_role == "admin" else "Membro")
post_auth_loading_placeholder = start_post_auth_loading()


with st.sidebar:
    logo_path = os.path.join(os.path.dirname(__file__), "assets", "trading_strategy_logo.png")
    if os.path.exists(logo_path):
        st.image(logo_path, use_container_width=True)
        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
    st.markdown(
        f"""
            <div style="border:1px solid #263244; border-radius:8px; padding:10px 11px; background:#0B1220; margin-bottom:12px;">
                <div style="color:#94A3B8; font-size:.68rem; font-weight:900; letter-spacing:.06em; text-transform:uppercase;">Acesso</div>
                <div style="color:#F8FAFC; font-size:.78rem; font-weight:800; overflow:hidden; text-overflow:ellipsis; white-space:nowrap;">{"Dashboard liberado" if not AUTH_REQUIRED else html.escape(_auth_user_email(auth_user))}</div>
                <div style="color:#38BDF8; font-size:.70rem; font-weight:900; margin-top:4px;">{html.escape(auth_role_label)}</div>
            </div>
        """,
        unsafe_allow_html=True,
    )
    if AUTH_REQUIRED and st.button("Sair", use_container_width=True, key="auth_logout_btn"):
        try:
            if supabase:
                supabase.auth.sign_out()
        except Exception:
            pass
        st.session_state.pop("auth_user", None)
        st.session_state.pop("auth_session", None)
        st.session_state.pop("auth_role", None)
        st.session_state.pop("auth_loading_message", None)
        st.session_state.pop("auth_loading_until", None)
        _auth_rerun()
    st.markdown("### 🧭 Navegação")
    page = st.radio("Ir para:", ["📉 Terminal de Trading", "🌎 Terminal Global", "Crypto Terminal", "📺 Terminal Bloomberg", "📰 Market Report", "Market Moving", "WATCHLIST", "WATCHLIST QUANT", "📊 Gráficos Avançados", "⚖️ Painel de Correlação", "🛡️ Gestão de Risco", "⚙️ Painel de Controle"], index=1, label_visibility="collapsed")
    sidebar_clock()
    
    st.markdown("---")
    
    tab1, tab2, tab3 = st.tabs(["🌍 MERCADOS", "📅 CALENDÁRIO", "NEWS"])
    with tab1: sidebar_mercados()
    with tab2: sidebar_calendario()
    with tab3: sidebar_news()

render_high_impact_news_ticker()

# Roteamento de Páginas
if page == "📉 Terminal de Trading":
    pagina_terminal()
elif page == "🌎 Terminal Global":
    pagina_terminal_global()
elif page == "Crypto Terminal":
    pagina_crypto_terminal()
elif page == "📺 Terminal Bloomberg":
    pagina_terminal_bloomberg()
elif page == "📰 Market Report":
    pagina_market_report()
elif page == "Market Moving":
    pagina_market_moving()
elif page == "WATCHLIST":
    pagina_watchlist()
elif page == "WATCHLIST QUANT":
    pagina_watchlist_quant()
elif page == "📊 Gráficos Avançados":
    pagina_graficos()
elif page == "⚖️ Painel de Correlação":
    pagina_correlacao()
elif page == "🛡️ Gestão de Risco":
    pagina_gestao_risco()
elif page == "⚙️ Painel de Controle":
    pagina_painel_controle()

stop_post_auth_loading(post_auth_loading_placeholder)



